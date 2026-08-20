#!/usr/bin/env python3
"""A2Z proposal engine (pure-Python / fpdf2, no graphics runtime needed).
Public: generate(workbook_path, kind, out_path), kind in {'LTD','SA'}.
Fonts live in ./fonts next to this file. Logo embedded."""
import os, sys, base64, math, datetime, tempfile
try:
    import openpyxl
    from fpdf import FPDF
    from fpdf.enums import XPos, YPos
except ImportError:
    _m="A2Z needs two free libraries (fpdf2 and openpyxl).\n\nDouble-click 'Setup (run once).bat' once, then try again."
    try:
        import tkinter as _tk; from tkinter import messagebox as _mb
        _r=_tk.Tk(); _r.withdraw(); _mb.showerror("Setup needed", _m); _r.destroy()
    except Exception:
        print("\n"+_m+"\n")
        try: input("Press Enter to close...")
        except Exception: pass
    sys.exit(1)

HERE = os.path.dirname(os.path.abspath(__file__))
FONTS = os.path.join(HERE, "fonts")
LOGO_B64 = "/9j/4AAQSkZJRgABAQAAAQABAAD/4gHYSUNDX1BST0ZJTEUAAQEAAAHIAAAAAAQwAABtbnRyUkdCIFhZWiAH4AABAAEAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAACRyWFlaAAABFAAAABRnWFlaAAABKAAAABRiWFlaAAABPAAAABR3dHB0AAABUAAAABRyVFJDAAABZAAAAChnVFJDAAABZAAAAChiVFJDAAABZAAAAChjcHJ0AAABjAAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAAgAAAAcAHMAUgBHAEJYWVogAAAAAAAAb6IAADj1AAADkFhZWiAAAAAAAABimQAAt4UAABjaWFlaIAAAAAAAACSgAAAPhAAAts9YWVogAAAAAAAA9tYAAQAAAADTLXBhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABtbHVjAAAAAAAAAAEAAAAMZW5VUwAAACAAAAAcAEcAbwBvAGcAbABlACAASQBuAGMALgAgADIAMAAxADb/2wBDAAUDBAQEAwUEBAQFBQUGBwwIBwcHBw8LCwkMEQ8SEhEPERETFhwXExQaFRERGCEYGh0dHx8fExciJCIeJBweHx7/2wBDAQUFBQcGBw4ICA4eFBEUHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh7/wAARCABiASkDASIAAhEBAxEB/8QAHQABAAIDAQEBAQAAAAAAAAAAAAYHBAUIAQMJAv/EAEAQAAEDBAECAwYDBgMGBwAAAAECAwQABQYRBxIhEzFBCBQiUWFxFSMyQlJigZGhJDOCFhc4c5KxOWNydLO0tf/EABkBAQACAwAAAAAAAAAAAAAAAAABAwIEBf/EACoRAAICAgIBAwIGAwAAAAAAAAABAhEDEiExEwQiQVFhIzJCcZGhsdHw/9oADAMBAAIRAxEAPwDsulKUApSlAKUpQClKUApSlAKUpQClKUApSlAKUpQClKUApSlAKUpQClKUApSlAKUpQClKUApSlAKUpQClKUApSlAKUpQHh8602YZPYsRsb15yC4NQYbQ/Us/Es+iUp81KPoB3rJyO8W/H7HNvV1fDEKG0p15Z9Egeg9SfID1Nfn5zJyNd+SMocuU1x1qA0VJt8MnSY7ZPqB2KyNdR/lvQFbHp/TvK/sUZsyxr7lmcme09k12dch4VGRY4O9e8vIDkpY+Y3tCB9NE/UVTV6zLL704pd1ye8zCo90uzXCj/AKd6/tWiFe12MeGGP8qOXPLKXbM2HebxBdDsG7XCI4DsLYkrbV/UEVYeD89ckYxJa8S8rvUNJ+ONcvzOofRz9aT9d/cHyqrwNnQGzU7xLh/kjKGw9bsVmtRlAFMiYBHbUD5EdeiofVIIplWOveIb37OzsDhvmnF+R2vdWVfhd6QNuW+Q4CpQ/ebV+2Pn2BHqPImzq/PnkPjfIeKfwm4XK+Q2bxIe8SKzAcWXGQjRLhWQNEEpA1vez37arq32buVG+RMXVGuLiU5DbgEzU6CQ8k/peSB6HyIHkfoRvlZ8Cit4dHRwZ3L2z7LZoK8qE8n8o4rxyYQyN6UlU3rLSY7PiHSdbJ79h8QrWjFydJGw5JK2Tf1pWkwjKLRmONRchsjq3IMnq6C4npUClRSQR6HYqP8AJPLWEcfvIiX+5n35xPWmJHbLroT+8oDskfLqI36b0aRhKT1S5DkkrbJ3Sq14w5pw/kO/OWSwIuQmNx1SVB+P0pDaVJSTsE+qk/1r4Z5zxx3h94ds865PzZ7B6X2YTJd8JX7qldk9X03seuqy8M9ta5MfLCrvgtGlQfjblLFc+tNxudmdksRrcQJS5rYaSjYJ31bI0ANk77VDrr7TPGMG4mI3IuU1CVdJkR4u2j9QSQSPqBRYcjdUHlglbZdNK0GIZhj2W47+PY/cG5sMbC+kELbUBspUk90qHyPzB8iDWt405IxrkIXA46qWoW9aEP8Ajs+H3V1a137/AKTWOkqbroy3jxz2TGlQZfKeKo5THGxXM/HirpCfA/K34Hj/AK9/uf37VgSua8FichHBpcuUxdRKTE6nGCGfFVrpHX9SQN/M1KxyfSI8kfqWRSoVyfydi/HSrYnI1y0m5FwR/AY8TfR0dW+/b/MT/eo/mfPmA4nk8zHbs5c/foagl0NROpPdIUNHffsRSOOcukHkiu2WrSqdsvtH8b3e8wbVEduvvE6S3GZ64mk9a1BKdnflsipVjXKWLZDn9ywi3Lmm727xveAtgpbHhLCF6Vvv3UKmWKce0FlhLpk4pVcY1zVg1/zheGw5Utq6pedYSmQx0IW42T1JSrfc/Coj56rZ5ZybjGM5taMPui5Yul2LQihtgqQfEcLaepXp8QNQ8c06olTi/kmlKj+cZnjWFWk3PJbqxBYJ02lRJcdV8kIG1KP2Hb1qto3tNcZPzERw5d09awhKzCJGydDsDv8AtUxxTkrSIllhHhsumleIUFJCh5EbFe1WZilKUApSlAKUpQClKUBzH7c2XvxoFowuI6UIlkzZvSe6kJOm0H6dW1fdCa5RFW77Yk9yZzrc46ztMGLGjo+gLYd/7uGv79nfhhrk5Eq53G9+5WqFI8B5mMAZLiukKGioFKE6UPiIV5Ea9a7WFxw4FJ/9Zyct5Mriuynz5E10BiHA+Lt2S2T895Dt1rkXVlL0SLHfbSVIUNpPWs/Ee43oaB7bNQrmnhzI+N5ZfcSblYnFaZuDSNBPfsh1PfoV5fQ+h8wIBDjT7/dIdtQ4t+TJU1Dj+Koq6dkIQkb8kjYGvSrJS8kU4SorjHR1JWWlzBx3cOE8zsN6tdx/EIinxJhuuoCVIdaUlRbWB2IIKTseeyNDWzans9cm8hcocnOC5zGYdhtsVb70SIwEtrWSENpUsgrJ7qVrqAPQe3atn7THHmV5g3YLXaENotFhtzj8u4S3gA4vQSEhI2pSwlsk9gPjHeox7J+TYVgXHN3veTX+Db5Nym6DZJU+pppOk6bSCojqWvWh61qSl5MGz5kbCisebXpFX+1Dlqsq5gunhOFUK1n8PjD002T1q89d3Cvv8umtPwdlz2FcnWi8oWoRlvJizUb7LYcISr+adhQ+qRVlZlyLwwq03Sz4hxy5Nn3Bl1tM99lIcDiwfzElRU5vZ3+yap634Nm09vxLfiN/kpPkpq3Oq/7JrYxuLx6tUUzT32Ttn6UpO09W+xG65N5Ok2fkL2ol2K73SFAslkgOxXHpD6G0eJ4aiogrIBUHXEp1/wCWTXS1jmzGcChXKdClCYi1tvvRktKU8HA0FKQEeZVvY6fPfaub+DOF05hcMjyDlPG7kw/IlBbDEgvRj1qKlrUNFJUO6Rvy8/Xdc7BrDaTZv5tpUkjZ+xZk70XHsoxGatpx60uqms9DgWFJIKXEpI7FIUgEEefiGtD7KWP2/kHkHKsxyxtF2mRXWnG2pGloDjynSVlJ/dDYSkeQG9DsNbPEMHvfG3tN9Fkxm8P4hMT7qX22HHmm23WxrqX3/S4BsqPYbJrGmYNyZwnyFOyLALU7kWPz1HrjNpK1eGVFSW1tpPX1J2elxII1vetkVe9ZSnq+WlRRFNauS4Vl2cwJhYhxhkuQ2G2woFyYtq22pEdhLa0hRA80jfY6P3Aqs/Y3wLG38Ecy25W6PcLpNlOtBySgOeC2g66Ug7AJO1FXmdgelSXDsny/lVF6xXMOOp2MWaVa3UGW8HNl0qSlKR1pR30oq8j+mq0wxrmbgu4S7JGxF/LLC+4XGzFbUpBVrXWhSApTZOhtK067dvmaoR/DlC/dZZKXvUq4Jz7Ya2MZ4d91skONAReLszGl+7thvrSG3HO+tb34SR9qlnB+B4tD4eskVdjgSfxCCiRMW/HStT63E9SuokEnW9D5AADyqL3GDlXOfFV7teSYm9iU5iU29aUygsda0p81FSQrR2pJISOyvWofh+c814BiowqTxhPukmCCxAmttOLQE/s9RQFJcSN9iFJ7AA9wTU6t4tE+b+o2UZ7tcM1vs/MnFOf82w22vOKtPgzG0oUsns0seGSfVQSop39TW+9hAaZzP/3Mf1/5tST2a+Kr9jk+75pmywb7d+oe7lSVqaC19bilkEgrWddh2AH10IHYrZyZwTn1+Nlw+RlOPXU9TaoqFHYSpRbJKAotrSFqBBTpXYjy3VjlGanFPnj+iuMZQ1m+uT6Sf/EEa/54/wDyKhfJWJS8u5u5IYtqVquFvQqfHQk6K/D8LrSPr0FRH1AFWlwbg+ZZFzFP5ezq0OWZxXUYUVY6FlamgyD0n4glLW0/EASSD6VsuOcbv8T2q8vvkuyT2LVKZcDMxyOpLLp/K7JURo+R/pUrIoN12ooaOdccNlKcw5y5yBx5x7cZJCrnbnpkK4d/1OD3cpX/AK0jq+/UPSunPaFsVkVxLll0XaICp4ty1e8mOku7AGj1a3uueeeOHsntPIsgYjYrrcbDcXkzm0RGFOpjuEnrQekdtEqI/hUB31uuoOcIE26cQ5NAt0R6XLft60NMMoK1rUfIADuTWGWULhq+DLHGVS2XJXfse2WzS+IIc+VaYL8tFwfKH3I6VOJKV/DpRGxr0qEcE/8AGJnH/qun/wBpurV9k6zXaxcQxoF6tsu2yxMfWWJTKm1hJV2PSob0agfDWK5LbvaozC+T7Bc4trkquPgTHYy0sudclCk9KyNHYBI+eqjdb5OfgaPSHBSz1gucu7Z5lljU4m4YxfDN+AnYaL7vUofMpUlKj/CFfaprl+XM51zPxHlLLfhKk+4NyG9/5b6Jqg4kfTZ2PoRVj+zdil5gZ1yOrILDOiQblKUGlSo6kIkIU69vpJHxAhQ8vnVYI4my3FefLHDt9mu8/HIV/jSY0xEdS2mmC6hR6lgaBSlICj2/TvturlljJu311/BVpKMVS7/2bnIGmuTPbDVj2REu2i3urYaiqOkqQyz1lP8AqXsn5jt5Cp3yTynxdgGTv4lcePXZC4IaIci2+N4XdCVp6epQOxsenmKwed+MMxh8jR+UuNmRJuKVIclRElIX1pT0dYSSAtKkaCkgg/LezqF8tXvlrlXHLfjUvia7W99iUH1SEMupQtfQpOvzEhKB8ROys1glHJrzxX1osk5QUuOb/c62ss5u6WeFc2ULQ1LjtvoSvXUkLSFAHXbfesutVh8WRBxKzwZbfhSI8Bhp1HUD0rS2kEbHY6IPlW1rnvs3l0KUpUEilKUApSlAKUpQHDHtk25ULm+ZKKfhuEKPIB+yPCP/AMYqW+wncUC/5PYXSnplxGpAST3V4alJP9nBUt9tzC3rpjVvzGE0VuWpRYlhI7+A4Rpf+lQA+y9+lc/+z/mMbBeUrdfLg4tFt8N1iZ4aOpRbUg60PXSwg/yrrxfl9NS7r/BzJLx+ot9M6qtt8Y46w96ycx5fa7k1LkLjQmHWi88uIVFI8Ydy4OnRUSnQ8iVE1z7n0bj3jfmG03zD7o1kUCK7727bWHwsMOJ7oSH9FJSSR81J133sV5OxfJefuUbnk2N2d+22iU6lKpk5ZLTSUJSjzA7qITvoTvRPn5muiuOOFMC44gm73FLFyuDKOt243AJ6GderaT8KB9e6vr6VTcMCuTtvtGaTzdLr5KrQr2guZmupC/8AZTHZIOiCqKl1sj6bdcSR9kq38vKV2D2b+NsWgC45pd3boUDbq5D3ukff0SlXV/LqNbLkHnRqO45Aw+O3IKR0mc+k9G/4E+v3P9CO9Uhe7zdr3NMy73CRNeJJ6nV71v5DyA+g0K1peol+nhfY2Y4Ip7S5Zef+8nizDYxh4hjTLnSND3KGiOhZHqpZAUfvomo9c+fckkOlFts9uiJUdICyt5e/v8I/tVQHuKmPDWMLyjPYMdaT7nEUJUpWv2EHYT/qVofYk+lUOTfZakvguy1RuaZwQ5Kutgt7au56mPEUAfoBo/1qZY/bMnivBd5yaPcUaP5bVtSx318+tVSAUqCaPABvfrXtKUJFavJ7v+B2l65LgSZbLCFOPBgo2hCUlRUeojtoenetpWDkNvF3sNwtSnC0JkVyOVgb6etJTvX03QijChZNanVtx5cuPAnOa/wUiQ34yQpSUp2lKj+orRr59Q9TqvtJyGwxVNiReYDanSsIBkJ+IoJCwO/7JBB+RGq1dwxBMgznmpvhyZEiJJaWpoKDa4/QUgjYKkko7jY8zWNFwl1ly2Ofix8WHJefcebaLbjoddW4pvYVroJXrpIV5b7HvSiTIk5zak4izlEGPJuNtc11LjlG2tkD4gpQ0dnWvOsi15ba5N4cs0spttyQUBMWS82HHCpAX8ICj1aB76rEaw3XHCMNduJWhppDLcgNaV0IUkp2N+ekgE/2rInYqzLukmcuSQX50eWUhHceEjo6d+fcb7+m6Az05DZ3YE+ZDnxprdvQpUkRnUuKR0gkggHsexrXWrNrNLiyZMtZtSIxaDipi0IT+YnqRpQJSSR6b2PUVh43grFosk62CSlz3mGIaX+lQcDQSoDq2og66t6ASPPt3rXt8aoQyVfiaA+FsqShuOW4/wCW2tvZQlQV1KS4dqCgSQn0GqAkt5y/GrVGkPyrvDJjxveltNupU4Wz5KSkHZB2NfPY+dZN1vLEWxG8xWV3KKlHi7irQrbYBJWCVAEaHoai9y47Zmx5EVu4e4xn7V+HLaitKSlY6EpSpYUshXTo9PbYB1s1ImbddpVqukK8T2HDLSppksNdKWkFsJ3o9ySrqVok62Bvtugo+Nuy6zSIjb02Wxa3FspfDEyQ2lwNKG0rICjpJ79/pWbIv9jjyER37vAQ+t8x0tqfSFF3QPRre+rSk9vP4h8xUflYDGlQrhGdnL/xlqi24LDY234HWQsfPZUNj+Gv5Rg7/vyJrt3QqUbiZy5DcYtuAlLSVNoIV2QoNAEK6gRreyN0Io28TLLLOsLl6t0xmZFRJTGUpDyEgLLiUaJUQB+oEDeyCNb2N5tsvtjukp2JbrvAmSGQS4yzIStaADokpB2O/b71p3MSW7jD9ieuRU0u4JltOJZAUhIkJf6D3+I9QI327Edu3f3H8Pbs8+DKbmqcMRqW3otgFfjvB3Z7+mtUFGXCym3zLtd7XEZkvSLZrrCUD846+IN7PxdJ0k+WjWDJzaMzZ5lzFpuKm4Dy2Jrf5aVsLSEnR2vSthaddJO91hQcA/DJjNytN5kRrn4UlEmQ4C6l4vK6yrw1K6UkLAV2HfXfdBhNxTb7dEN6jOCFLVLPiQ1KTIcIGlujxNrUFdStk62R2+EUJJpEdU/FZfUy4ypxCVltzXUgkb6TrY2PLtX1r5xkvJjNJkOJceCAHFpT0hStdyBs6G/TZr6UApSlAKUpQClKUApSlAY9zhRblb5FvnMNyIsltTTzTg2laVDRBH2rnbFfZas0TOJk+93AzcfbeK4EFCiHFpOiA8v5J7jQ7q0CSO4PSNPWrIZZQTUX2Vzxxm05Ii+TX/HOPcYbcdbaixGUBqJCjpCSsjyShI0PufIetcy8jZ/e81nKVMcMe3oVtiE2r4EfIq/eV9T/ACAq4+YeKLplV3XfLZeS4+UpQIcvshAHo2oDt89EeZPeqfufGGeW9Sg5jct5IOgqP0ug/wDSSf7VXbM6IcBqlSSJgObynfDaxS7g+W3Yymx/VWhUzxbgzKbg6hV6ej2iN+2N+K79gEnX899vrQUVpZrbPvFyYttsjLky3ldKG0Dufr9B8yewrrTijCIuE48I3wOXGRpya+nyWr0SN9+lOzr7k+tZWCYPYMNiKZtEY+M4B40l09TrmvLZ9B9BoVJqEilKUApSlAKUpQClKUApSlAKUpQClKUApSlAKUpQClKUApSlAKUpQClKUApSlAKUpQClKUApSlAKUpQClKUApSlAKUpQClKUApSlAKUpQClKUApSlAKUpQClKUApSlAKUpQClKUApSlAKUpQClKUApSlAKUpQClKUApSlAKUpQClKUApSlAKUpQClKUApSlAKUpQClKUApSlAKUpQClKUApSlAKUpQH/2Q=="

NAVY=(22,55,90); GREEN=(30,107,71); GREY=(91,103,112); INK=(27,42,56)
LINE=(217,222,229); PALE=(244,246,249); MUTED=(138,148,157); A9BBD0=(169,187,208)
SOFTLINE=(236,239,243); WHITE=(255,255,255)

# ----- Real A2Z Google reviews (edit testimonials.txt to change; these are the fallback) -----
A2Z_REVIEWS = [
 ("Wiktoria Wykrota","4 months ago","I can't recommend A2Z highly enough! Shabbir is extremely attentive, genuinely client-focused, and has a great sense of business. From the start he's taken the time to understand my personal situation and always makes himself available to answer all of my questions, explaining everything clearly and in a way that's easy to understand. His team is quick, efficient, and a real pleasure to work with, which makes the whole process really smooth and stress-free. I wouldn't hesitate to recommend A2Z to anyone looking for a reliable, professional, and trustworthy accounting company!"),
 ("BlackGard Joinery","4 months ago","I've been working with A2Z for a while now, and they've been a life saver for my joinery business. Shabbir and his team are super friendly, easy to talk to, very quick to reply and they really know their stuff. They've taken all the stress out of bookkeeping and tax time, which means I can focus on what I do best. If you're in the trades and want someone reliable who actually understands your business, I highly recommend them."),
 ("Rona Tonge","a year ago","Have been dealing with A2Z Accounting Solutions now for 4 years and the service is exceptional. Shabbir and his team are so helpful and professional with all aspects of our business, from tax and corporation tax to PAYE. Always at the end of the phone if you need help or have any queries. Very reasonably priced for their superb service, and I hope to have them as my accountants for many years to come."),
 ("Matthew Foy","4 months ago","I've had a great experience with the A2Z Accounting team. They communicate well and are always happy to help. Shabbir himself is always on hand to answer any questions I might have. I would definitely recommend Shabbir and the A2Z team."),
 ("Amanda Nasser","a year ago","An accountancy company that is not only efficient and fantastic to work with, but also helps to take the overwhelm out of accounting, offering solutions to suit each client's needs and providing clarity when clarity is needed. They came recommended, so I'm recommending them to everyone who wants to keep their business on track and help it grow as well."),
 ("Lee Wood","4 months ago","Really good experience. Helped me get my fines down substantially and my outstanding tax returns completed. They couldn't have been more helpful. Mahfuza is excellent and a pleasure to deal with. Thank you again."),
 ("Andrew Clarkson","a year ago","I've been thoroughly impressed with the professionalism and expertise of A2Z. Their team is detail-oriented, friendly, and always responsive to my questions. They go above and beyond to ensure everything is accurate and transparent. Highly recommend for anyone seeking reliable accounting services!"),
 ("Mark Cowie","a year ago","I've been with A2Z accounting from the start and they have made running my business so much easier. Shabbir's vast knowledge is second to none and he always goes the extra step to make sure things run smoothly."),
]

def get_reviews():
    """Reviews for the proposal. Reads testimonials.txt if it has lines in the
    form  Name | time ago | review text ; otherwise uses the built-in A2Z list."""
    out=[]
    try:
        for line in open(os.path.join(HERE,"testimonials.txt"), encoding="utf-8"):
            line=line.strip()
            if not line or line.startswith("#") or line.count("|")<2: continue
            name,tm,txt=[s.strip() for s in line.split("|",2)]
            if name and txt: out.append((name,tm,txt))
    except Exception:
        pass
    return out if out else A2Z_REVIEWS

def num(v):
    try: return float(v)
    except (TypeError, ValueError): return None
def gbp(v):
    v=v or 0
    return f"£{v:,.0f}" if float(v).is_integer() else f"£{v:,.2f}"
def fmt_date(v):
    if isinstance(v,(datetime.date,datetime.datetime)): return v.strftime("%d %B %Y")
    return str(v) if v else datetime.date.today().strftime("%d %B %Y")
def ref_for(co=""):
    import os, time
    base="23456789ABCDEFGHJKLMNPQRSTUVWXYZ"
    n=max(int(time.time())-1735689600, 0)          # seconds since 2025-01-01
    code=""
    while n: code=base[n%32]+code; n//=32
    code=(code or "2").rjust(6,"2")[-6:]
    tag=base[os.urandom(1)[0]%32]+base[os.urandom(1)[0]%32]
    return f"A2Z-{code}-{tag}"                        # e.g. A2Z-3KP9Q7-HS (13 chars)
def load_testimonials(n=2):
    """Reads testimonials.txt (next to this file). One per line: Quote :: Attribution."""
    out=[]
    try:
        for line in open(os.path.join(HERE,"testimonials.txt"), encoding="utf-8"):
            line=line.strip()
            if not line or line.startswith("#") or "::" not in line: continue
            q,by=line.split("::",1); out.append((q.strip(),by.strip()))
            if len(out)>=n: break
    except Exception: pass
    if not out: out=[("[ Add a genuine client review in testimonials.txt ]","Client name, business, location")]
    return out

class PDF(FPDF):
    def __init__(self):
        super().__init__(format="A4", unit="mm")
        self.set_auto_page_break(True, margin=20)
        self.set_margins(18, 21, 18)
        for fam,st,fn in [("Cormorant","","Cormorant-SemiBold.ttf"),("Cormorant","B","Cormorant-Bold.ttf"),
                          ("Nunito","","Nunito-Regular.ttf"),("Nunito","B","Nunito-Bold.ttf"),
                          ("NunitoSemi","","Nunito-SemiBold.ttf")]:
            self.add_font(fam, st, os.path.join(FONTS, fn))
        self.CW=174; self.X0=18; self.XR=192; self.right_header="PROPOSAL OF SERVICES"
    def header(self):
        if self.page_no()==1: return
        self.set_xy(self.X0, 11); self.set_font("Nunito","B",6.8); self.set_text_color(*MUTED)
        self.cell(0,4," ".join("A2Z ACCOUNTING SOLUTIONS"))
        self.set_xy(self.X0,11); self.set_font("Nunito","",6.8); self.set_text_color(*MUTED)
        self.cell(self.CW,4,getattr(self,"right_header","PROPOSAL OF SERVICES"),align="R")
        self.set_draw_color(*LINE); self.set_line_width(0.2); self.line(self.X0,16,self.XR,16)
        self.set_xy(self.l_margin, self.t_margin)
    def footer(self):
        if self.page_no()==1: return
        self.set_draw_color(*LINE); self.set_line_width(0.2); self.line(self.X0,283,self.XR,283)
        self.set_y(-12); self.set_font("Nunito","",7); self.set_text_color(*MUTED)
        self.set_x(self.X0); self.cell(self.CW/2,4,"A2Z Accounting Solutions  ·  Regulated by ACCA")
        self.cell(self.CW/2,4,f"Page {self.page_no()} of {{nb}}",align="R")
    # ---- helpers ----
    def f(self, fam, st, sz, col=NAVY): self.set_font(fam,st,sz); self.set_text_color(*col)
    def need(self, h):
        if self.get_y()+h > 297-20: self.add_page()
    def klabel(self, t):
        self.f("Nunito","B",7.5,GREEN); self.cell(0,4," ".join(t.upper()),new_x=XPos.LMARGIN,new_y=YPos.NEXT)
        self.ln(1.5)
    def heading(self, t, sz=22, col=NAVY):
        self.f("Cormorant","B",sz,col)
        for i,ln in enumerate(t.split("\n")):
            self.cell(0,sz*0.42,ln,new_x=XPos.LMARGIN,new_y=YPos.NEXT)
        self.ln(1.5)
    def lead(self, t, w=152):
        self.f("Nunito","",10,(58,71,80)); self.set_x(self.X0)
        self.multi_cell(w,5.2,t,align="L",new_x=XPos.LMARGIN,new_y=YPos.NEXT,markdown=True); self.ln(2)
    def hline(self, col=LINE, w=None, x=None, lw=0.2):
        x=self.X0 if x is None else x; w=self.CW if w is None else w; y=self.get_y()
        self.set_draw_color(*col); self.set_line_width(lw); self.line(x,y,x+w,y)
    def tick(self, x, y, s=2.6, col=GREEN):
        self.set_draw_color(*col); self.set_line_width(0.5)
        self.line(x, y+0.55*s, x+0.36*s, y+0.92*s); self.line(x+0.36*s, y+0.92*s, x+s, y+0.08*s)
    def star(self, cx, cy, r, col=None):
        pts=[]
        for i in range(10):
            ang=-math.pi/2 + i*math.pi/5; rr=r if i%2==0 else r*0.45
            pts.append((cx+rr*math.cos(ang), cy+rr*math.sin(ang)))
        self.set_fill_color(*(col or GREEN)); self.polygon(pts, style="F")
    def rrect(self, x,y,w,h, r=1.6, style="D", fill=None, draw=LINE, lw=0.2):
        if fill: self.set_fill_color(*fill)
        self.set_draw_color(*draw); self.set_line_width(lw)
        self.rect(x,y,w,h,style=style,round_corners=True,corner_radius=r)

# ---------- shared page sections ----------
def cover(p, company, subtitle, date, ref, prepared_by="Shabbir Rahman FCCA"):
    p.add_page(); p.set_auto_page_break(False)
    p.set_fill_color(*NAVY); p.rect(0,0,210,297,style="F")
    p.set_fill_color(*GREEN); p.rect(0,0,210,3.2,style="F")
    # logo on a clean white chip
    logo=os.path.join(tempfile.gettempdir(),"a2z_logo.png")
    if LOGO_B64 and not os.path.exists(logo): open(logo,"wb").write(base64.b64decode(LOGO_B64))
    src=logo if os.path.exists(logo) else os.path.join(HERE,"logo.png")
    p.rrect(24,26,66,21, r=2.5, style="F", fill=WHITE, draw=WHITE)
    try: p.image(src, x=30, y=30.5, h=12)
    except Exception: pass
    p.f("Nunito","",7.5,A9BBD0); p.set_xy(110,33); p.cell(76,5," ".join("PRIVATE & CONFIDENTIAL"),align="R")
    # title block
    p.set_fill_color(*GREEN); p.rect(24,98,48,1.4,style="F")
    p.f("Cormorant","B",49,WHITE); p.set_xy(23,105)
    p.cell(0,18,"Proposal of",new_x=XPos.LEFT,new_y=YPos.NEXT); p.set_x(23); p.cell(0,18,"Services")
    p.f("Nunito","",8,A9BBD0); p.set_xy(24,164); p.cell(0,5," ".join("PREPARED EXCLUSIVELY FOR"),new_x=XPos.LEFT,new_y=YPos.NEXT)
    p.f("Cormorant","B",25,WHITE); p.set_xy(24,170)
    _csz=25
    while _csz>13 and p.get_string_width(company)>160: _csz-=0.5; p.set_font("Cormorant","B",_csz)
    p.cell(0,12,company,new_x=XPos.LEFT,new_y=YPos.NEXT)
    p.f("Nunito","",9.5,(150,170,194)); p.set_x(24); p.cell(0,6,subtitle)
    # meta strip
    p.set_draw_color(56,84,114); p.set_line_width(0.3); p.line(24,250,186,250)
    meta=[("DATE",date),("PREPARED BY",prepared_by or "Shabbir Rahman FCCA"),("REFERENCE",ref),("VALID FOR","30 days")]
    x=24
    for lab,val in meta:
        p.f("Nunito","",7,A9BBD0); p.set_xy(x,254); p.cell(40,4," ".join(lab))
        val=str(val); sz=9.5; p.set_font("NunitoSemi","",sz)
        while sz>6.8 and p.get_string_width(val)>37: sz-=0.4; p.set_font("NunitoSemi","",sz)
        p.set_text_color(*WHITE); p.set_xy(x,259); p.cell(40,5,val); x+=40.5
    p.f("Nunito","",7,A9BBD0); p.set_xy(24,268); p.cell(0,4,"FCCA: Fellow Chartered Certified Accountant, the highest grade of ACCA membership.")
    # contact
    p.f("Nunito","",8,(150,170,194)); p.set_xy(24,283); p.cell(0,5,"1st Floor, 499 Union Street, Aberdeen, AB11 6DB")
    p.set_xy(110,283); p.cell(76,5,"01224 042961  ·  info@a2zaccounting.co.uk",align="R")
    p.set_auto_page_break(True, margin=20)

def letter(p, contact, company, paras, date="", prepared_by=None):
    p.add_page(); p.klabel("A note from the founder")
    p.heading("We don't act as your accountant.\nWe act as your finance function.",22)
    p.ln(5)
    p.f("Nunito","",10,INK); p.set_x(p.X0); p.cell(0,5,f"Dear {contact},",new_x=XPos.LMARGIN,new_y=YPos.NEXT); p.ln(3)
    for para in paras:
        p.f("Nunito","",10,INK); p.set_x(p.X0)
        p.multi_cell(p.CW,5.6,para,align="L",new_x=XPos.LMARGIN,new_y=YPos.NEXT); p.ln(3.2)
    p.ln(3)
    p.f("Cormorant","B",17,NAVY); p.cell(0,7,"Shabbir Rahman FCCA",new_x=XPos.LMARGIN,new_y=YPos.NEXT)
    p.f("Nunito","",8.5,GREY); p.cell(0,4,"Founder & Chief Executive",new_x=XPos.LMARGIN,new_y=YPos.NEXT)
    p.ln(11); p.hline(SOFTLINE); p.ln(8)
    pts=[("Zero","HMRC penalties since we founded"),("Zero","missed filing deadlines"),("75+","five-star Google reviews"),("FCCA","sign-off on every output")]
    y=p.get_y(); cw=p.CW/4
    for i,(big,small) in enumerate(pts):
        cx=p.X0+i*cw
        p.f("Cormorant","B",25,NAVY); p.set_xy(cx,y); p.cell(cw-3,10,big)
        p.f("Nunito","",8,(74,86,96)); p.set_xy(cx,y+11.5); p.multi_cell(cw-4,3.8,small, align="L", new_x=XPos.LMARGIN,new_y=YPos.TOP)
    p.set_y(y+24)

def fee_table(p, rows, year=False, prices=True, unit=None):
    detw=73 if prices else 101
    p.f("Nunito","B",7.5,MUTED)
    p.cell(73,6.5," ".join("SERVICE")); p.cell(detw,6.5," ".join("WHAT IT COVERS"))
    if prices:
        unit=unit or ("£ / year" if year else "£ / month")
        p.cell(28,6.5," ".join(unit),align="R")
    p.ln()
    p.set_draw_color(*NAVY); p.set_line_width(0.3); y=p.get_y(); p.line(p.X0,y,p.XR,y); p.ln(1)
    for row in rows:
        svc,det,amt=row[0],row[1],row[2]
        p.set_font("NunitoSemi","",9); sl=p.multi_cell(73,4.4,svc,dry_run=True,output="LINES")
        p.set_font("Nunito","",9); dl=p.multi_cell(detw,4.4,det,dry_run=True,output="LINES")
        rh=max(len(sl),len(dl),1)*4.4
        p.need(rh+2.2); y0=p.get_y()
        p.f("NunitoSemi","",9,NAVY); p.set_xy(p.X0,y0); p.multi_cell(73,4.4,svc,align="L",new_x=XPos.RIGHT,new_y=YPos.TOP)
        p.f("Nunito","",9,GREY); p.set_xy(p.X0+73,y0); p.multi_cell(detw,4.4,det,align="L",new_x=XPos.RIGHT,new_y=YPos.TOP)
        if prices:
            p.f("NunitoSemi","",9.5,NAVY); p.set_xy(p.X0+146,y0); p.cell(28,4.4,gbp(amt),align="R")
        p.set_y(y0+rh+1.4); p.hline(SOFTLINE)
    p.ln(0.5)
def total_row(p, lab, val, grand=False):
    if grand:
        p.set_draw_color(*NAVY); p.set_line_width(0.4); y=p.get_y(); p.line(p.X0,y,p.XR,y); p.ln(2)
    p.f("NunitoSemi","",9.3, NAVY if grand else GREY); p.cell(150,5,lab,align="R")
    if grand: p.f("Cormorant","B",14,NAVY)
    else: p.f("NunitoSemi","",9.3,NAVY)
    p.cell(24,5,gbp(val),align="R",new_x=XPos.LMARGIN,new_y=YPos.NEXT); p.ln(0.5)

def difference(p):
    p.ln(10); p.klabel("The A2Z difference"); p.heading("Why clients stay",22); p.ln(3)
    diff=[("FCCA director sign-off","Every set of accounts is reviewed and signed off personally by a chartered certified director. Nothing leaves the firm unchecked."),
          ("A four-layer review","Preparer, senior, manager and director: four sets of eyes before anything reaches you or HMRC."),
          ("Same-day response","Email us before 4pm on a working day and you hear back the same day. No chasing, no silence."),
          ("A track record that holds up","No HMRC penalties and no missed deadlines since the firm was founded, alongside more than 75 five-star reviews.")]
    for i,(t,d) in enumerate(diff):
        p.need(20); y0=p.get_y()
        p.f("Cormorant","B",26,(201,211,222)); p.set_xy(p.X0,y0-3); p.cell(17,12,f"{i+1:02d}")
        p.f("Nunito","B",10.5,NAVY); p.set_xy(p.X0+21,y0); p.cell(0,5,t,new_x=XPos.LMARGIN,new_y=YPos.NEXT)
        p.f("Nunito","",9.3,(74,86,96)); p.set_xy(p.X0+21,y0+5.6); p.multi_cell(p.CW-21,4.7,d,align="L",new_x=XPos.LMARGIN,new_y=YPos.NEXT)
        p.set_y(max(p.get_y(),y0+15)); p.hline(SOFTLINE); p.ln(4.5)

def proof(p):
    p.add_page(); p.klabel("The proof"); p.heading("What our clients say",22); p.ln(1)
    p.lead("A selection of our verified five-star Google reviews.")
    # ---- 5.0 hero band ----
    y=p.get_y()+1; bh=25
    p.rrect(p.X0,y,p.CW,bh,r=3,fill=NAVY,draw=NAVY,style="F")
    p.f("Cormorant","B",27,WHITE); p.set_xy(p.X0+12,y+4.0); p.cell(30,11,"5.0")
    for i in range(5): p.star(p.X0+13.6+i*7.0, y+18.4, 2.7, col=(176,201,184))
    p.set_draw_color(39,72,106); p.set_line_width(0.5); p.line(p.X0+56,y+5.0,p.X0+56,y+bh-5.0)
    p.f("Cormorant","B",15.5,WHITE); p.set_xy(p.X0+64,y+6.6); p.cell(110,8,"Rated 5.0 on Google")
    p.f("Nunito","",8.7,(176,196,217)); p.set_xy(p.X0+64,y+15.8); p.cell(110,5,"From 75+ verified reviews by business owners across the UK")
    p.set_y(y+bh+8)

    revs=get_reviews()[:8]
    n=len(revs); rows_n=max(1,(n+1)//2)
    gap=5.0; cw=(p.CW-gap)/2; pad=5.4; lh=4.0; tsize=8.4
    avatar=8.4; head_block=pad+avatar+2.4          # card top -> divider
    text_top=head_block+2.8                        # card top -> first quote line

    top=p.get_y()
    note_h=10
    bottom=p.h-18-note_h
    cardh=(bottom-top-(rows_n-1)*gap)/rows_n
    max_lines=max(3,int((cardh-text_top-pad)//lh))

    # cap each review to max_lines with an ellipsis
    disp=[]
    for name,tm,txt in revs:
        p.set_font("Nunito","",tsize)
        ls=list(p.multi_cell(cw-2*pad,lh,txt,dry_run=True,output="LINES"))
        if len(ls)>max_lines:
            ls=ls[:max_lines]; last=ls[-1].rstrip().rstrip(".")
            while last and p.get_string_width(last+"\u2026")>cw-2*pad:
                last=last.rsplit(" ",1)[0] if " " in last else last[:-1]
            ls[-1]=last+"\u2026"
        disp.append((name,tm,ls))

    for idx,(name,tm,ls) in enumerate(disp):
        col=idx%2; row=idx//2
        cx=p.X0+col*(cw+gap); yy=top+row*(cardh+gap)
        p.rrect(cx,yy,cw,cardh,r=3,fill=(249,251,253),draw=LINE,style="DF")
        init=(name.strip()[:1].upper() or "A")
        p.set_fill_color(*NAVY); p.ellipse(cx+pad,yy+pad,avatar,avatar,style="F")
        p.f("NunitoSemi","",9,WHITE); p.set_xy(cx+pad,yy+pad+2.0); p.cell(avatar,4,init,align="C")
        p.f("NunitoSemi","",9.4,NAVY); p.set_xy(cx+pad+avatar+3.4,yy+pad-0.2); p.cell(cw-2*pad-avatar-3.4,4,name)
        sx=cx+pad+avatar+3.6; sy=yy+pad+5.4
        for i in range(5): p.star(sx+i*3.0, sy, 1.25, col=(214,170,46))
        p.f("Nunito","",6.7,MUTED); p.set_xy(sx+5*3.0+1.6,sy-1.6); p.cell(40,4,str(tm)+"  \u00b7  Google review")
        dy=yy+head_block; p.set_draw_color(*SOFTLINE); p.set_line_width(0.2); p.line(cx+pad,dy,cx+cw-pad,dy)
        p.f("Nunito","",tsize,(60,72,82)); p.set_xy(cx+pad,yy+text_top)
        p.multi_cell(cw-2*pad,lh,"\n".join(ls),align="L",new_x=XPos.LMARGIN,new_y=YPos.NEXT)

    p.set_y(top+rows_n*cardh+(rows_n-1)*gap+4)
    p.set_auto_page_break(False)   # keep this single line on the reviews page (no extra blank page)
    p.f("Nunito","",8.2,MUTED); p.set_x(p.X0)
    p.multi_cell(p.CW,4.3,"Search \u201cA2Z Accounting Solutions\u201d on Google to read all 75+ of our five-star reviews.",align="C",new_x=XPos.LMARGIN,new_y=YPos.NEXT)
    p.set_auto_page_break(True, margin=20)

def steps(p, items):
    p.ln(6); p.klabel("Getting started"); p.heading("Switching is effortless",22); p.ln(1)
    for i,(t,d) in enumerate(items):
        p.need(15); y0=p.get_y()
        p.set_fill_color(*NAVY); p.ellipse(p.X0,y0,9,9,style="F")
        p.f("Nunito","B",10,WHITE); p.set_xy(p.X0,y0+1.6); p.cell(9,5,str(i+1),align="C")
        p.f("Nunito","B",10,NAVY); p.set_xy(p.X0+13,y0); p.cell(0,5,t,new_x=XPos.LMARGIN,new_y=YPos.NEXT)
        p.f("Nunito","",9.2,(74,86,96)); p.set_xy(p.X0+13,y0+5); p.multi_cell(p.CW-13,4.5,d, align="L", new_x=XPos.LMARGIN,new_y=YPos.NEXT)
        p.set_y(max(p.get_y(),y0+11)); p.ln(1.5)

FORMS={'LTD':{'onboard':("company onboarding form","https://a2zaccounting.co.uk/company-on-boarding-form/"),
              'reg':("limited company registration form","https://a2zaccounting.co.uk/limited-company-registration-form/")},
       'SA':{'onboard':("sole-trader onboarding form","https://a2zaccounting.co.uk/sole-trader-onboarding/"),
             'reg':("personal tax registration form","https://a2zaccounting.co.uk/personal-tax-registration/")}}
# Registration form links - EASILY CORRECTABLE. NOTE: the 'company' link is a suspected duplicate of the
# sole-trader/personal-tax URL; replace it here when the correct company-registration form is confirmed.
REG_LINKS={
 'company':("Company onboarding form","https://a2zaccounting.co.uk/company-on-boarding-form/"),
 'paye':   ("PAYE registration form","https://a2zaccounting.co.uk/paye-registration/"),
 'vat':    ("VAT registration form","https://a2zaccounting.co.uk/vat-registration/"),
 'cis_sub':("CIS subcontractor registration form","https://a2zaccounting.co.uk/subcontractor-cis-registration/"),
 'cis_con':("CIS contractor registration form","https://a2zaccounting.co.uk/contractor-cis-registration/"),
 'sa':     ("Personal tax registration form","https://a2zaccounting.co.uk/personal-tax-registration/"),
 'other':  ("Onboarding form","https://a2zaccounting.co.uk/company-on-boarding-form/"),
}
# Paste your Direct Debit mandate Google Form link here (a real http link makes the step on "Getting started" clickable):
DD_FORM = "https://forms.gle/Hzz6pGn968qGuBt2A"

def package_for(sub):
    if sub <= 200: return ("Tier 1 Compliance","")
    if sub < 500:  return ("Tier 2 Growth","")
    return ("Tier 3 Strategic","PREMIER SERVICE")

def _onboarding_render(p, items):
    words={3:"Three",4:"Four",5:"Five",6:"Six"}.get(len(items),str(len(items)))
    p.add_page(); p.klabel("Getting started"); p.heading("Getting started is simple",22)
    p.lead(f"{words} quick steps, most taking only a few minutes of your time. We handle everything else.")
    p.ln(1)
    for i,(t,dsc,link,label) in enumerate(items):
        p.need(30 if link else 22); y0=p.get_y()
        p.set_fill_color(*NAVY); p.ellipse(p.X0,y0,10,10,style="F")
        p.f("Nunito","B",11,WHITE); p.set_xy(p.X0,y0+2.1); p.cell(10,6,str(i+1),align="C")
        p.f("Cormorant","B",14.5,NAVY); p.set_xy(p.X0+15,y0-0.3); p.cell(0,7,t,new_x=XPos.LMARGIN,new_y=YPos.NEXT)
        p.f("Nunito","",9.4,(74,86,96)); p.set_xy(p.X0+15,y0+6.8); p.multi_cell(p.CW-15,4.7,dsc,align="L",new_x=XPos.LMARGIN,new_y=YPos.NEXT)
        if link:
            ly=p.get_y()+1.5; ax=p.X0+15
            p.set_font("NunitoSemi","",9.2); bw=p.get_string_width(label)+12; bh=8.4
            p.rrect(ax,ly,bw,bh,r=1.8,fill=GREEN,draw=GREEN,style="F")
            p.set_text_color(*WHITE); p.set_xy(ax,ly+0.4); p.cell(bw,bh-0.8,label,align="C",link=link)
            p.set_y(ly+bh+3)
        p.set_y(max(p.get_y(),y0+12)+1.5); p.hline(SOFTLINE); p.ln(2.5)

def onboarding(p, kind, needs_reg=False, direct_debit=True):
    fname,furl=FORMS[kind.upper()]['reg' if needs_reg else 'onboard']
    if needs_reg:
        title2="Register and onboard in about 5 minutes"
        desc2=(f"Complete our {fname} (about 5 minutes). You'll upload your photo ID and a proof of address dated within the "
               f"last 3 months, and settle the one-off registration fee.")
    else:
        title2="Complete onboarding in about 5 minutes"
        desc2=(f"Complete our {fname} (about 5 minutes) and upload your photo ID and a proof of address dated within the "
               f"last 3 months, for example a utility bill, bank statement or council tax letter.")
    items=[
        ("Approve this proposal","Sign at the back of this document, or simply reply to our email to confirm. That's all it takes to get moving.",None,None),
        (title2,desc2,furl,"Open the "+fname),
        ("Sign your engagement letter","A quick electronic signature. It is a legal requirement before we can act for you.",None,None),
    ]
    if direct_debit:
        dd_link=DD_FORM if (DD_FORM and DD_FORM.startswith("http")) else None
        dd_desc=("Complete our short Direct Debit mandate so your fees collect automatically, with no invoices to chase." if dd_link
                 else "Provide your bank details so fees are collected automatically, with no invoices to chase.")
        items.append(("Set up your Direct Debit",dd_desc,dd_link,"Open the Direct Debit mandate"))
    items.append(("We set you up and get going","We configure your software and services, and your dedicated team takes it from here. Your first deliverables are scheduled straight away.",None,None))
    _onboarding_render(p, items)

def onboarding_partnership(p, needs_reg=True, direct_debit=True):
    # PAYE/VAT/CIS/self-assessment registrations use the same online forms as a limited company
    # (shown as buttons on the Setup & registration page). Registering the partnership itself is by email.
    reg_bit=(" Any PAYE, VAT, CIS or self-assessment registrations are completed through the secure online forms shown on your setup page." if needs_reg else "")
    items=[
        ("Approve this proposal","Sign at the back of this document, or simply reply to our email to confirm. That's all it takes to get moving.",None,None),
        ("Get set up online","Reply with your photo ID and a proof of address dated within the last 3 months."+reg_bit+" Registering the partnership itself with HMRC is the one step we handle for you by email, with no form to complete.",None,None),
        ("Sign your engagement letter","A quick electronic signature. It is a legal requirement before we can act for you.",None,None),
    ]
    if direct_debit:
        dd_link=DD_FORM if (DD_FORM and DD_FORM.startswith("http")) else None
        dd_desc=("Complete our short Direct Debit mandate so your fees collect automatically, with no invoices to chase." if dd_link
                 else "Provide your bank details so fees are collected automatically, with no invoices to chase.")
        items.append(("Set up your Direct Debit",dd_desc,dd_link,"Open the Direct Debit mandate"))
    items.append(("We set you up and get going","We configure your software and services, and your dedicated team takes it from here. Your first deliverables are scheduled straight away.",None,None))
    _onboarding_render(p, items)

def commit_accept(p, company):
    p.add_page(); p.klabel("Our commitments to you"); p.heading("The promises behind the fee",22); p.ln(1)
    p.f("Nunito","",9.5,(74,86,96)); p.set_x(p.X0); p.multi_cell(p.CW,5.0,"Whatever you choose, every A2Z engagement comes with the same four promises.", align="L", new_x=XPos.LMARGIN,new_y=YPos.NEXT); p.ln(4)
    for c in ["A fixed monthly fee, agreed in advance and never billed by surprise.","A 30-day rolling agreement, with no long tie-in.","A named team, with an FCCA director who signs off your work.","Your filing completed ahead of every deadline."]:
        y0=p.get_y(); p.tick(p.X0, y0+0.4, 3.0)
        p.f("Nunito","",9.6,(39,50,59)); p.set_xy(p.X0+6,y0); p.cell(0,5,c,new_x=XPos.LMARGIN,new_y=YPos.NEXT)
        p.hline(SOFTLINE); p.ln(2.2)
    p.ln(5); y=p.get_y(); h=56
    p.rrect(p.X0,y,p.CW,h,r=2.5,fill=NAVY,draw=NAVY,style="DF")
    p.set_fill_color(*GREEN); p.rect(p.X0,y,p.CW,2.2,style="F")
    p.f("NunitoSemi","",7.5,(169,187,208)); p.set_xy(p.X0+9,y+8); p.cell(0,4," ".join("TO PROCEED"))
    p.f("Cormorant","B",16,WHITE); p.set_xy(p.X0+9,y+12.5); p.cell(0,7,"Accept your proposal")
    p.f("Nunito","",9,(210,221,232)); p.set_xy(p.X0+9,y+20.5); p.multi_cell(p.CW-18,4.5,f"Sign below on behalf of {company}, or simply reply to our email to confirm. This proposal is valid for 30 days from the date shown.",align="L")
    ly=y+44; p.set_draw_color(120,140,165); p.set_line_width(0.3)
    inner=p.CW-18; gap=11; colw=(inner-2*gap)/3
    for i,lab in enumerate(["Signature","Name & position","Date"]):
        lx=p.X0+9+i*(colw+gap); p.line(lx,ly,lx+colw,ly)
        p.f("Nunito","",7,(169,187,208)); p.set_xy(lx,ly+1.6); p.cell(colw,4,lab.upper())
    p.set_y(y+h+6)
    p.f("Cormorant","B",14,NAVY); p.set_x(p.X0); p.cell(0,7,"We look forward to working with you.",new_x=XPos.LMARGIN,new_y=YPos.NEXT); p.ln(2)
    p.hline(LINE); p.ln(3)
    p.f("Nunito","B",8.6,NAVY); p.cell(p.CW/2,4,"A2Z Accounting Solutions")
    p.cell(p.CW/2,4,"01224 042961",align="R",new_x=XPos.LMARGIN,new_y=YPos.NEXT)
    p.f("Nunito","",8.6,GREY); p.cell(p.CW/2,4,"1st Floor, 499 Union Street, Aberdeen, AB11 6DB")
    p.cell(p.CW/2,4,"info@a2zaccounting.co.uk",align="R")

# ---------- tiers ----------
def compute_tiers(P, band):
    row=None
    for r in range(6,20):
        if str(P.cell(r,1).value).strip()==str(band).strip(): row=r; break
    if not row: return None
    g=lambda c:(num(P.cell(row,c).value) or 0)
    accts,qb,dext,paym,vatq,bkm,mgmtq=g(2),g(3),g(5),g(7),g(9),g(13),g(14); cs=5
    return [dict(name="Tier 1 Compliance",price=accts+cs,rec=False,
                 items=["Annual accounts & Corporation Tax","Confirmation statement (CS01)","Director's self-assessment","FCCA director sign-off"]),
            dict(name="Tier 2 Growth",price=accts+qb+dext+paym+vatq+bkm+cs,rec=True,
                 items=["Everything in Tier 1","Monthly bookkeeping","Quarterly VAT returns","Monthly payroll (RTI)","QuickBooks + Dext included"]),
            dict(name="Tier 3 Strategic",price=accts+qb+dext+paym+vatq+bkm+mgmtq+cs,rec=False,
                 items=["Everything in Tier 2","Quarterly management accounts","Plain-English commentary","Proactive tax planning"])]

def tiers_page(p, has_cis=False, has_contractor=False, assigned="", company="this business", client_items=None, client_fee=None, ma_tier=None, partnership=False):
    p.add_page(); p.klabel("Choose your level"); p.heading("Three ways to work with us",22)
    a=(assigned or "").strip()
    if a:
        p.lead(f"Based on {company}'s size and needs, we recommend the **{a}** level, marked below. Each level includes everything in the one before it, so the only question is how far ahead you want us to look. Your tailored quote follows on the next page.")
    else:
        p.lead("Choose the level of support that fits. Each level includes everything in the one before it, and your tailored quote follows.")
    REPORTS_2=[("1","Financial Health","T1"),("2","Business Performance","T2")]
    REPORTS_3=[("1","Financial Health","T1"),("2","Business Performance","T2"),("3","Strategic Advisory","T3")]
    t1_items=(["Year-end accounts & partnership tax return (SA800)","Partners' self-assessment returns","Basic payroll for the partners on fixed monthly wages","HMRC deadline monitoring","Same-day email support"] if partnership else ["Year-end accounts & Corporation Tax","Confirmation statement filed","Companies House filing fee covered","Basic payroll for up to two owners on fixed monthly wages","HMRC deadline monitoring","Same-day email support"])
    tiers=[
        dict(name="Tier 1 Compliance", price="\u00a3105-\u00a3200", badge="", style="plain",
             who="Clean, accurate, on-time compliance.", base="", turnover="Turnover up to \u00a390k",
             items=t1_items,
             reports=[]),
        dict(name="Tier 2 Growth", price="\u00a3250-\u00a3500", badge="", style="plain",
             who="Stay in control as you grow.", base="Everything in Tier 1, plus", turnover="Turnover \u00a390k-\u00a3299k",
             items=["Quarterly bookkeeping & VAT under MTD","Cloud accounting: QuickBooks + Dext","Payroll for your team"],
             reports=REPORTS_2),
        dict(name="Tier 3 Strategic", price="\u00a3500-\u00a33,500+", badge="", style="plain",
             who="A full finance function behind you.", base="Everything in Tier 2, plus", turnover="Turnover \u00a3300k-\u00a35m",
             items=["Proactive tax planning & profit extraction","Dedicated manager and reviews through the year"],
             reports=REPORTS_3),
    ]
    al=a.lower()
    for t in tiers:
        if t['name'].lower()==al: t['style']="hero"; t['badge']="YOUR LEVEL"
    gap=4.5; cw=(p.CW-2*gap)/3; y0=p.get_y()+5; ch=104
    for i,t in enumerate(tiers):
        x=p.X0+i*(cw+gap); hero=(t['style']=='hero'); tx=x+5.5; iw=cw-11
        if hero:
            lift=4
            p.rrect(x,y0-lift,cw,ch+lift+5, r=2.2, fill=NAVY, draw=NAVY, lw=0.3, style="DF")
            p.set_fill_color(*GREEN); p.rect(x,y0-lift,cw,2.4,style="F")
            txtcol=WHITE; subcol=(169,187,208); bodycol=(224,232,240); badgecol=(176,201,184); tickcol=(159,199,175); rule=(39,72,106); basecol=(176,201,184)
            oy=y0-lift+3
        else:
            p.rrect(x,y0,cw,ch, r=2.2, fill=WHITE, draw=LINE, lw=0.2, style="D")
            txtcol=NAVY; subcol=MUTED; bodycol=(74,86,96); badgecol=GREEN; tickcol=GREEN; rule=SOFTLINE; basecol=GREEN
            oy=y0
        p.f("NunitoSemi","",6.5, badgecol); p.set_xy(tx,oy+4.5); p.cell(iw,3.5,(" ".join(t['badge']) if t['badge'] else ""))
        p.f("Cormorant","B",13.5, txtcol); p.set_xy(tx,oy+9); p.multi_cell(iw,5.0,t['name'], align="L", new_x=XPos.LMARGIN,new_y=YPos.NEXT)
        p.f("NunitoSemi","",8.2, basecol); p.set_xy(tx,oy+20.3); p.cell(iw,4,t.get("turnover",""))
        ps=17.0; p.set_font("Cormorant","B",ps)
        while p.get_string_width(t['price'])>iw-1 and ps>11: ps-=0.5; p.set_font("Cormorant","B",ps)
        p.set_text_color(*txtcol); p.set_xy(tx,oy+25.6); p.cell(iw,8,t['price'])
        p.f("Nunito","",7, subcol); p.set_xy(tx,oy+33.6); p.cell(iw,4,"+ VAT / month")
        p.f("NunitoSemi","",7.4, (210,221,232) if hero else GREY); p.set_xy(tx,oy+38.5); p.multi_cell(iw,3.5,t['who'],align="L",new_x=XPos.LMARGIN,new_y=YPos.NEXT)
        yy=oy+47
        if t.get('base'):
            blab=t['base'].upper(); bs=7.6; p.set_font("NunitoSemi","",bs)
            while p.get_string_width(blab)>iw and bs>6.2: bs-=0.3; p.set_font("NunitoSemi","",bs)
            p.set_text_color(*basecol); p.set_xy(tx,yy); p.cell(iw,4,blab)
            yy+=5.4
        for it in t['items']:
            p.tick(tx, yy+1.6, 2.3, col=tickcol)
            p.f("Nunito","",8.0, bodycol); p.set_xy(tx+4.2,yy+0.5)
            p.multi_cell(iw-4.2, 3.8, it, align="L", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            yy=max(p.get_y()+0.8, yy+5.0)
        if t['reports']:
            yy+=2.8
            p.set_draw_color(*rule); p.set_line_width(0.3); p.line(tx,yy,tx+iw,yy); yy+=3.4
            p.f("NunitoSemi","",8.2, basecol); p.set_xy(tx,yy); p.cell(iw,3.6,"Management reports"); yy+=5.2
            d=4.9
            for ri,(rank,rname,rcode) in enumerate(t['reports']):
                sel=(ma_tier==rcode and hero); cy=yy
                if sel:
                    p.set_fill_color(*tickcol); p.ellipse(tx,cy,d,d,style="F")
                    p.f("NunitoSemi","",7.8,(NAVY if hero else WHITE)); p.set_xy(tx,cy+1.05); p.cell(d,3.0,rank,align="C")
                    p.f("NunitoSemi","",9.2,(WHITE if hero else NAVY)); p.set_xy(tx+d+3.0,cy+0.7); p.cell(iw-d-3.0,4,rname)
                else:
                    p.set_draw_color(*subcol); p.set_line_width(0.35); p.ellipse(tx,cy,d,d,style="D")
                    p.f("NunitoSemi","",7.8,subcol); p.set_xy(tx,cy+1.05); p.cell(d,3.0,rank,align="C")
                    p.f("Nunito","",9.0,bodycol); p.set_xy(tx+d+3.0,cy+0.7); p.cell(iw-d-3.0,4,rname)
                yy=cy+6.2
    p.set_y(y0+ch+6)
    p.f("Nunito","",8.3,MUTED); p.set_x(p.X0)
    p.multi_cell(p.CW,4.4,"Indicative monthly ranges, excluding VAT. Management reports are optional and priced as you choose; your selected report is highlighted above. Most clients start at the level that fits today and move up as they grow, and you can change level at any time.", align="L", new_x=XPos.LMARGIN,new_y=YPos.NEXT)
    p.ln(4); y=p.get_y()
    note=("The package features above are typical of each level and are shown for guidance only. "
          f"The services agreed for {company} are those set out on the next page, \u2018Your tailored quote\u2019, which is the "
          "definitive list and takes precedence over the summaries above. Some items shown here may not be included for "
          "you, and others may be added, to match exactly what has been agreed.")
    p.set_font("Nunito","",9); nl=p.multi_cell(p.CW-12,4.6,note,dry_run=True,output="LINES")
    h=7+max(len(nl),1)*4.6
    p.rrect(p.X0,y,p.CW,h,fill=PALE,draw=PALE,style="F"); p.set_fill_color(*NAVY); p.rect(p.X0,y,1.5,h,style="F")
    p.f("Nunito","",9,(33,44,54)); p.set_xy(p.X0+6,y+3); p.multi_cell(p.CW-12,4.6,note,align="L")
    p.set_y(y+h); p.ln(2.5); p.f("Nunito","",8.3,MUTED); p.set_x(p.X0)
    _foot="Every level is signed off by an FCCA director and billed as one fixed monthly fee. "+("Each partner's personal self-assessment is billed separately, per year." if partnership else "Director and shareholder personal tax returns are \u00a3120 + VAT each per year, charged separately.")
    p.multi_cell(p.CW,4.2,_foot,align="C",new_x=XPos.LMARGIN,new_y=YPos.NEXT)

def cards_2x2(p, cards):
    gap=6; cw=(p.CW-gap)/2; x=[p.X0,p.X0+cw+gap]; y0=p.get_y(); ch=42
    for i,(t,d) in enumerate(cards):
        col=i%2; row=i//2; cx=x[col]; cy=y0+row*(ch+gap)
        p.rrect(cx,cy,cw,ch, r=2.2, fill=PALE, draw=PALE, style="F")
        p.f("Cormorant","B",14.5,NAVY); p.set_xy(cx+8,cy+9); p.multi_cell(cw-16,5.8,t,align="L",new_x=XPos.LMARGIN,new_y=YPos.NEXT)
        p.f("Nunito","",9.2,(74,86,96)); p.set_xy(cx+8,cy+19); p.multi_cell(cw-16,4.7,d,align="L")
    p.set_y(y0+2*ch+gap+2)

# ---------- readers ----------
def read_ltd(wb):
    L=wb['LTD Proposal']; g=lambda c:L[c].value
    people=int(num(g('C35')) or 0); lines=[]
    def add(label, detail, amt):
        a=num(amt)
        if a and a>0: lines.append((label, detail, a))
    def _fq(base,cell):
        q=str(g(cell) or '').strip(); return f"{base} ({q.lower()})" if q and q.upper()!='NA' else base
    _f1=str(g('C39') or '').strip(); _f2=str(g('C40') or '').strip(); _f3=str(g('C41') or '').strip()
    _isf=lambda x: x in ('Quarterly','Monthly')
    if _isf(_f3): msel='T3 '+_f3; mafee=g('G41')
    elif _isf(_f2): msel='T2 '+_f2; mafee=g('G40')
    elif _isf(_f1): msel='T1 '+_f1; mafee=g('G39')
    else: msel='NA'; mafee=0
    MGMT={
        "T1 Quarterly":("Financial Health Report","quarterly","Sales, profit, cash, P&L and balance sheet, your dividends and director's loan, and your VAT, PAYE and tax position"),
        "T2 Quarterly":("Business Performance Report","quarterly","Report 1 plus margins and KPIs, a forward projection and cash flow forecast, and the trends and drivers behind your numbers"),
        "T2 Monthly":("Business Performance Report","monthly","Report 1 plus margins and KPIs, a forward projection and cash flow forecast, and the trends and drivers behind your numbers"),
        "T3 Quarterly":("Strategic Advisory Report","quarterly","Report 2 plus budget versus actual, customer and supplier analysis, a CT projection, a tax savings review, profit-extraction, DLA and VAT advice, recommended actions and reviews through the year"),
        "T3 Monthly":("Strategic Advisory Report","monthly","Report 2 plus budget versus actual, customer and supplier analysis, a CT projection, a tax savings review, profit-extraction, DLA and VAT advice, recommended actions and reviews through the year"),
    }
    # Compliance
    add("Annual accounts & Corporation Tax","Statutory accounts & CT600, filed to deadline", (num(g('G22')) or 0)+(num(g('G28')) or 0))
    add("Companies House","Annual confirmation statement (CS01), filed as your authorised agent", g('G23'))
    add("Address service","Registered office / address service", g('G24'))
    # Bookkeeping & VAT (contractor last in this group)
    add(_fq("Bookkeeping",'C27'), "Records maintained and fully reconciled", g('G27'))
    add(_fq("VAT returns",'C29'), "Prepared, checked and filed under MTD", g('G29'))
    add("Bookkeeping software", f"{g('C30')} licence, set-up and support", g('G30'))
    add("Document software (Dext)","Automated receipt & invoice capture", g('G31'))
    add("Contractor / project accounting","Job costing, WIP & project profitability", g('G32'))
    # Payroll
    pay=(num(g('G34')) or 0)+(num(g('G35')) or 0)
    if pay>0:
        pf=str(g('C34') or '').strip(); ppl=f"{people} {'person' if people==1 else 'people'}"
        plab="Payroll ("+(f"{pf.lower()}, {ppl}" if pf else ppl)+")"
        lines.append((plab, "RTI submissions, payslips and year-end forms", pay))
    add("CIS returns","CIS scheme, prepared and filed", g('G36'))
    # Advisory & growth (management report last)
    if msel in MGMT:
        nm,freq,desc=MGMT[msel]; add(f"{nm} ({freq})", desc, mafee)
    oneoffs=[]; oneoffs_full=[]
    for r in range(50,58):
        de=g(f'B{r}'); pr=num(g(f'F{r}'))
        if (de and str(de).strip()) or (pr and pr>0):
            dl=g(f'G{r}')
            oneoffs.append((str(de or '').strip(), "", pr or 0))
            oneoffs_full.append((str(de or '').strip(), str(g(f'D{r}') or '').strip(), str(g(f'E{r}') or '').strip(), pr or 0, (fmt_date(dl) if hasattr(dl,'year') else (str(dl).strip() if dl else '')), ""))
    # --- Setup & Registration block (rows 60-66) ---
    _REGDEFS=[('company','Company formation / registration',60,'company'),
              ('paye','PAYE registration',61,'paye'),
              ('vat','VAT registration',62,'vat'),
              ('cis_sub','CIS subcontractor registration',63,'cis_sub'),
              ('cis_con','CIS contractor registration',64,'cis_con'),
              ('sa','Self-assessment registration',65,'sa'),
              ('other','Other registration',66,'other')]
    _comp=str(g('C60') or '').strip().lower()=='required'
    _paye=str(g('C61') or '').strip().lower()=='required'
    regs=[]
    for key,label,row,lk in _REGDEFS:
        if str(g(f'C{row}') or '').strip().lower()!='required': continue
        fee=num(g(f'F{row}')) or 0; inc=(key=='paye' and _comp and _paye)
        if inc: fee=0
        _form=FORMS['LTD']['reg'] if key=='company' else REG_LINKS.get(lk)   # company reg uses the Limited Company Registration form
        regs.append(dict(key=key,label=label,fee=fee,included=inc,form=_form))
    needs_reg=len(regs)>0
    return dict(company=g('C10') or 'Your Company Ltd',contact=g('C11') or 'Director',date=fmt_date(g('C14')),prepared_by=str(g('C15') or 'Shabbir Rahman FCCA').strip(),email=str(g('C12') or '').strip(),phone=str(g('C13') or '').strip(),
                band=str(g('C18') or ''),lines=lines,sub=num(g('G44')) or 0,vat=num(g('G45')) or 0,gross=num(g('G46')) or 0,discount=num(g('K4')) or 0,
                oneoffs=oneoffs,oneoffs_full=oneoffs_full,osub=num(g('F67')) or 0,ovat=num(g('F68')) or 0,ogross=num(g('F69')) or 0,
                directors=int(num(g('C25')) or 0),reg=needs_reg,regs=regs,comp_reg=_comp,notes=str(g('B72') or '').strip(),
                internal_notes=str(g('I10') or '').strip(),ma_sel=msel,bk=str(g('C27') or '').strip(),
                vat_scheme=("Flat Rate" if 'flat rate' in str(g('C29') or '').lower() else ("Standard" if str(g('C29') or '').strip().upper() not in ('','NA','-','NO','NONE') else "")),
                source=str(g('J25') or '').strip(),referrer=str(g('J26') or '').strip())

def read_sa(wb):
    S=wb['SA Proposal']; g=lambda c:S[c].value
    ctype=str(g('C13') or ''); freq=str(g('C15') or 'Annually')
    who=prov=send=""
    try:
        P=wb['SA Pricing']
        for r in range(2,12):
            if str(P.cell(r,1).value or '').strip().lower()==ctype.strip().lower():
                who=P.cell(r,9).value; prov=P.cell(r,10).value; send=P.cell(r,11).value; break
    except Exception: pass
    defs=[("Self-assessment tax return","Prepared, reviewed and filed by an FCCA director"),
          ("Property tax returns","Rental income pages & computations"),
          ("MTD for Income Tax (quarterly filing)","Four quarterly updates plus the final declaration, replacing the annual return"),
          ("Bookkeeping","Records maintained & reconciled"),
          ("VAT","Returns prepared & filed"),
          ("Payroll","RTI payroll & payslips"),
          ("Software","Cloud accounting software")]
    lines=[]
    for i,(lab,det) in enumerate(defs):
        r=26+i; a=num(g(f'D{r}')) or 0; m=num(g(f'E{r}')) or 0
        if a>0: lines.append((lab,det,a,'yr'))
        elif m>0: lines.append((lab,det,m,'mo'))
    oneoffs=[]; oneoffs_full=[]
    for r in range(40,48):
        de=g(f'B{r}'); pr=num(g(f'E{r}'))
        if (de and str(de).strip()) or (pr and pr>0):
            note=str(g(f'G{r}') or '').strip(); dl=g(f'F{r}')
            oneoffs.append((str(de or '').strip(), note, pr or 0))
            oneoffs_full.append((str(de or '').strip(), str(g(f'C{r}') or '').strip(), str(g(f'D{r}') or '').strip(), pr or 0, (fmt_date(dl) if hasattr(dl,'year') else (str(dl).strip() if dl else '')), note))
    # --- Setup & Registration block (rows 49-54) ---
    _SAREG=[('sa','Sole trader registration',49,'sa'),('paye','PAYE registration',50,'paye'),
            ('vat','VAT registration',51,'vat'),('cis_sub','CIS subcontractor registration',52,'cis_sub'),
            ('cis_con','CIS contractor registration',53,'cis_con'),('other','Other registration',54,'other')]
    regs=[]
    for key,label,row,lk in _SAREG:
        if str(g(f'C{row}') or '').strip().lower()!='required': continue
        regs.append(dict(key=key,label=label,fee=num(g(f'E{row}')) or 0,included=False,form=REG_LINKS.get(lk)))
    needs_reg=len(regs)>0
    return dict(company=g('C6') or g('C7') or 'Client',contact=g('C7') or 'there',date=fmt_date(g('C10')),prepared_by=str(g('C5') or 'Shabbir Rahman FCCA').strip(),email=str(g('C8') or '').strip(),phone=str(g('C9') or '').strip(),
                ctype=ctype,freq=freq,lines=lines,
                annual=num(g('D34')) or 0,annual_gross=num(g('D36')) or 0,
                monthly=num(g('E34')) or 0,monthly_gross=num(g('E36')) or 0,
                discount_annual=num(g('H34')) or 0,discount_monthly=num(g('H35')) or 0,
                oneoffs=oneoffs,oneoffs_full=oneoffs_full,osub=num(g('E55')) or 0,ovat=num(g('E56')) or 0,ogross=num(g('E57')) or 0,notes=str(g('B60') or '').strip(),
                who=str(who or ''),provides=str(prov or ''),send=str(send or ''),reg=needs_reg,regs=regs,sa_reg=str(g('C49') or '').strip().lower()=='required',
                vat_scheme=("Standard" if str(g('C18') or '').strip().upper()=='YES' else ""),
                source=str(g('J30') or '').strip(),referrer=str(g('J31') or '').strip())

# ---------- builders ----------
def rate_card_page(p, wb, partnership=False):
    try: RC=wb['Rate Card']
    except Exception: return
    items=[]
    for r in range(7,40):
        svc=RC.cell(r,2).value; price=RC.cell(r,3).value; note=RC.cell(r,4).value
        if svc and str(svc).strip() and price and str(price).strip():
            _svc=str(svc).strip(); _note=str(note or '').strip()
            if partnership:
                if any(k in _svc.lower() for k in ('company formation','companies house','corporation tax','confirmation statement')): continue
                _note=_note.replace('company formation','partnership registration')
            items.append((_svc, str(price).strip(), _note))
    if not items: return
    p.add_page(); p.klabel("Additional services"); p.heading("Beyond your monthly fee",22)
    p.lead("If you ever need more than your monthly service covers, here is what we offer. Fixed-price items are the same for every client; specialist work is scoped to you. All prices exclude VAT.")
    p.ln(1); sw=p.CW*0.62
    for svc,price,note in items:
        p.set_font("Cormorant","B",12); sl=p.multi_cell(sw,5.2,svc,dry_run=True,output="LINES"); sh=max(len(sl),1)*5.2
        nh=0
        if note:
            p.set_font("Nunito","",8.2); nl=p.multi_cell(p.CW*0.92,3.6,note,dry_run=True,output="LINES"); nh=max(len(nl),1)*3.6
        p.need(sh+nh+3); y=p.get_y()
        p.f("Cormorant","B",12,NAVY); p.set_xy(p.X0,y); p.multi_cell(sw,5.2,svc, align="L", new_x=XPos.LMARGIN,new_y=YPos.NEXT)
        p.f("Cormorant","B",12.5,GREEN); p.set_xy(p.X0+sw,y); p.cell(p.CW-sw,5.2,price,align="R")
        if note:
            p.f("Nunito","",8.2,MUTED); p.set_xy(p.X0,y+sh+0.4); p.multi_cell(p.CW*0.92,3.6,note, align="L", new_x=XPos.LMARGIN,new_y=YPos.NEXT)
        yb=y+sh+nh+1.6
        p.set_draw_color(*SOFTLINE); p.set_line_width(0.2); p.line(p.X0,yb,p.X0+p.CW,yb); p.set_y(yb+1.8)
    groups=[
      ("Groups, structures and incorporation","Group structuring and reorganisation, HMRC clearance letters, group consolidation accounts and loss surrenders; joint ventures and partnerships; and switching to a limited company, with the legal steps handled."),
      ("Property and VAT","Moving property into a limited company; VAT registration as a going concern, to keep your existing VAT number; VAT on the conversion of property to residential; and option to tax, including recovering the VAT paid on a property purchase."),
      ("Tax specialisms","Tax investigation and enquiry work, with deep expertise across every area; R&D tax credit claims through a specialist partner; seafarers' tax returns; fee protection and HMRC investigation support; and other specialist tax advice."),
    ]
    p.ln(4); p.need(28); y=p.get_y()
    p.f("Cormorant","B",15,NAVY); p.set_xy(p.X0,y); p.cell(0,7,"Specialist and advisory services")
    p.f("Nunito","",8,MUTED); p.set_xy(p.X0,y+1.8); p.cell(p.CW,5,"by arrangement",align="R")
    p.ln(9); p.hline(SOFTLINE); p.ln(3)
    for title,desc in groups:
        p.need(17)
        p.f("Cormorant","B",12.5,NAVY); p.set_x(p.X0); p.cell(0,5.5,title,new_x=XPos.LMARGIN,new_y=YPos.NEXT)
        p.f("Nunito","",8.7,(74,86,96)); p.set_x(p.X0); p.multi_cell(p.CW,4.3,desc, align="L", new_x=XPos.LMARGIN,new_y=YPos.NEXT); p.ln(2.6)
    p.f("Nunito","",8.3,MUTED); p.set_x(p.X0)
    p.multi_cell(p.CW,4.2,"Each is scoped and quoted to your situation. Just ask, and an FCCA director will advise.", align="L", new_x=XPos.LMARGIN,new_y=YPos.NEXT)

def mgmt_accounts_page(p, d):
    p.add_page(); p.klabel("Management reports"); p.heading("Looking forward, not just back",22)
    p.lead("Year-end accounts tell you what already happened. Management reports show what is happening now, in time to act. Three levels, each building on the one before.")
    _pf=bool(d.get('is_partnership'))
    _r1=("Drawings and partners' current accounts" if _pf else "Dividends and director's loan position")
    _r3ct=("Partners' tax projection" if _pf else "Corporation tax projection")
    _r3pe=("Profit allocation and drawings planning" if _pf else "Dividend and profit-extraction planning")
    _r3dl=("Partners' current account and VAT planning" if _pf else "Director's loan and VAT planning")
    ma=str(d.get('ma_sel','')).strip()
    sel='T1' if ma.startswith('T1') else 'T2' if ma.startswith('T2') else 'T3' if ma.startswith('T3') else None
    freq='monthly' if 'monthly' in ma.lower() else ('quarterly' if 'quarterly' in ma.lower() else '')
    names={'T1':'Financial Health Report','T2':'Business Performance Report','T3':'Strategic Advisory Report'}
    p.ln(2); y=p.get_y(); h=15
    if sel:
        p.rrect(p.X0,y,p.CW,h,r=2.2,fill=GREEN,draw=GREEN,style="F")
        p.set_fill_color(159,199,175); p.rect(p.X0,y,p.CW,1.5,style="F")
        p.f("NunitoSemi","",7.0,(199,224,208)); p.set_xy(p.X0+10,y+3.6); p.cell(0,3.4," ".join("INCLUDED IN YOUR FIXED FEE"))
        p.f("Cormorant","B",15.5,WHITE); p.set_xy(p.X0+10,y+7.0); p.cell(0,6,names[sel]+((" ("+freq+")") if freq else ""))
    else:
        p.rrect(p.X0,y,p.CW,h,r=2.2,fill=PALE,draw=(214,221,228),lw=0.4,style="DF"); p.set_fill_color(*NAVY); p.rect(p.X0,y,2.2,h,style="F")
        p.f("NunitoSemi","",7.0,MUTED); p.set_xy(p.X0+10,y+3.6); p.cell(0,3.4," ".join("NOT INCLUDED IN YOUR CURRENT QUOTE"))
        p.f("Cormorant","B",15,NAVY); p.set_xy(p.X0+10,y+7.0); p.cell(0,6,"No management report is included yet")
    p.set_y(y+h+8)
    reports=[
        ("01","T1","Financial Health Report","",
         ["Sales, profit and cash position","Management P&L and balance sheet",_r1,"Your VAT, PAYE and tax position"]),
        ("02","T2","Business Performance Report","Everything in Report 1, plus",
         ["Margins and efficiency KPIs","This year against last","Forward full-year projection","Rolling cash flow forecast","The trends and drivers behind your numbers"]),
        ("03","T3","Strategic Advisory Report","Everything in Report 2, plus",
         ["Budget versus actual","Customer and supplier analysis",_r3ct,"Tax savings review",_r3pe,_r3dl,"Recommended actions each period","Reviews through the year"]),
    ]
    nx=p.X0+22; lineH=5.8; colw=(p.CW-28)/2
    for rk,code,name,plus,feats in reports:
        inc=(code==sel); rows=(len(feats)+1)//2; ch=15.5+rows*lineH+3
        p.need(ch+8); y0=p.get_y()
        if inc:
            p.rrect(p.X0-1.5,y0-2,p.CW+3,ch+4,r=2.6,fill=(242,247,244),draw=(206,228,214),lw=0.5,style="DF")
            p.set_fill_color(*GREEN); p.rect(p.X0-1.5,y0-2,2.0,ch+4,style="F")
        numcol=GREEN if inc else (205,214,224)
        p.f("Cormorant","B",31,numcol); p.set_xy(p.X0+1,y0-2.5); p.cell(18,14,rk)
        p.f("Cormorant","B",15.5,NAVY); p.set_xy(nx,y0+1.4); p.cell(118,7,name)
        if inc:
            p.f("NunitoSemi","",7.2,GREEN); p.set_xy(p.X0+p.CW-72,y0+2.7); p.cell(64,4,"INCLUDED IN YOUR FEE",align="R")
        elif plus:
            p.f("Nunito","",7.8,MUTED); p.set_xy(p.X0+p.CW-72,y0+2.7); p.cell(64,4,plus,align="R")
        ry=y0+10.6; p.set_draw_color(*((176,205,184) if inc else SOFTLINE)); p.set_line_width(0.3); p.line(nx,ry,p.X0+p.CW-6,ry)
        fy=y0+14.6
        for fi,ft in enumerate(feats):
            coli=fi%2; rowi=fi//2; fx=nx+coli*colw; ryy=fy+rowi*lineH
            p.tick(fx,ryy+0.6,2.0,col=GREEN)
            p.f("Nunito","",8.8,(52,64,74)); p.set_xy(fx+4.8,ryy-0.4); p.cell(colw-7,4,ft)
        p.set_y(y0+ch+9)
    bk=str(d.get('bk','')).strip().lower(); avail=bk in ("monthly","quarterly")
    p.ln(0.5); p.f("Nunito","",8.3,MUTED); p.set_x(p.X0)
    note=("Available with your monthly or quarterly bookkeeping. Management reports are offered to limited companies only."
          if avail else
          "Management reports need monthly or quarterly bookkeeping, which is not currently selected, and are offered to limited companies only. Add bookkeeping and we can include them.")
    p.multi_cell(p.CW,4.2,note, align="L", new_x=XPos.LMARGIN,new_y=YPos.NEXT)

def reg_support_page(p, d):
    regs=d.get('regs') or []
    if not regs: return
    p.add_page(); p.klabel("Getting you set up"); p.heading("Setup & registration support",22)
    p.lead("Based on your requirements, we will assist with the following setup and registration work, handled on your behalf as your authorised agent.")
    p.ln(1)
    for rg in regs:
        p.need(20); y0=p.get_y()
        p.tick(p.X0, y0+1.6, 2.4, col=GREEN)
        p.f("Cormorant","B",14,NAVY); p.set_xy(p.X0+6,y0-0.4); p.cell(p.CW-74,7,rg['label'])
        if rg['included']: tag,tcol="Included with company setup",GREEN
        elif rg['fee']>0:  tag,tcol=f"{gbp(rg['fee'])} + VAT",NAVY
        else:              tag,tcol="",NAVY
        if tag:
            p.f("NunitoSemi","",9,tcol); p.set_xy(p.X0+p.CW-74,y0+0.5); p.cell(74,5,tag,align="R")
        if rg.get('form'):
            fname,furl=rg['form']; ly=y0+8.6; ax=p.X0+6
            p.set_font("NunitoSemi","",8.7); bw=p.get_string_width("Open the "+fname)+11; bh=7.4
            p.rrect(ax,ly,bw,bh,r=1.6,fill=GREEN,draw=GREEN,style="F")
            p.set_text_color(*WHITE); p.set_xy(ax,ly+0.2); p.cell(bw,bh-0.4,"Open the "+fname,align="C",link=furl)
            ny=ly+bh+3
        else:
            ny=y0+10.5
        p.set_y(max(ny,y0+12)); p.hline(SOFTLINE); p.ln(2.6)
    p.ln(0.5); p.f("Nunito","",8.3,MUTED); p.set_x(p.X0)
    p.multi_cell(p.CW,4.3,"Any registration fees are one-off and included in your tailored quote. We complete each registration with HMRC or Companies House on your behalf.",align="L",new_x=XPos.LMARGIN,new_y=YPos.NEXT)

def read_partnership(wb):
    L=wb['Partnership Proposal']; g=lambda c:L[c].value
    people=int(num(g('C35')) or 0); lines=[]
    def add(label, detail, amt):
        a=num(amt)
        if a and a>0: lines.append((label, detail, a))
    def _fq(base,cell):
        q=str(g(cell) or '').strip(); return f"{base} ({q.lower()})" if q and q.upper()!='NA' else base
    _f1=str(g('C39') or '').strip(); _f2=str(g('C40') or '').strip(); _f3=str(g('C41') or '').strip()
    _isf=lambda x: x in ('Quarterly','Monthly')
    if _isf(_f3): msel='T3 '+_f3; mafee=g('G41')
    elif _isf(_f2): msel='T2 '+_f2; mafee=g('G40')
    elif _isf(_f1): msel='T1 '+_f1; mafee=g('G39')
    else: msel='NA'; mafee=0
    MGMT={
        "T1 Quarterly":("Financial Health Report","quarterly","Sales, profit, cash, P&L and balance sheet, drawings and current accounts, and your VAT, PAYE and tax position"),
        "T2 Quarterly":("Business Performance Report","quarterly","Report 1 plus margins and KPIs, a forward projection and cash flow forecast, and the trends and drivers behind your numbers"),
        "T2 Monthly":("Business Performance Report","monthly","Report 1 plus margins and KPIs, a forward projection and cash flow forecast, and the trends and drivers behind your numbers"),
        "T3 Quarterly":("Strategic Advisory Report","quarterly","Report 2 plus budget versus actual, customer and supplier analysis, a tax projection, a tax savings review, profit-extraction and VAT advice, recommended actions and reviews through the year"),
        "T3 Monthly":("Strategic Advisory Report","monthly","Report 2 plus budget versus actual, customer and supplier analysis, a tax projection, a tax savings review, profit-extraction and VAT advice, recommended actions and reviews through the year"),
    }
    # Compliance  (NO confirmation statement - not applicable to a partnership)
    add("Annual accounts & partnership tax return","Partnership accounts and the SA800 partnership return, filed to deadline", (num(g('G22')) or 0)+(num(g('G28')) or 0))
    add("Address service","Registered office / address service", g('G24'))
    add(_fq("Bookkeeping",'C27'), "Records maintained and fully reconciled", g('G27'))
    add(_fq("VAT returns",'C29'), "Prepared, checked and filed under MTD", g('G29'))
    add("Bookkeeping software", f"{g('C30')} licence, set-up and support", g('G30'))
    add("Document software (Dext)","Automated receipt & invoice capture", g('G31'))
    add("Contractor / project accounting","Job costing, WIP & project profitability", g('G32'))
    pay=(num(g('G34')) or 0)+(num(g('G35')) or 0)
    if pay>0:
        pf=str(g('C34') or '').strip(); ppl=f"{people} {'person' if people==1 else 'people'}"
        plab="Payroll ("+(f"{pf.lower()}, {ppl}" if pf else ppl)+")"
        lines.append((plab, "RTI submissions, payslips and year-end forms", pay))
    add("CIS returns","CIS scheme, prepared and filed", g('G36'))
    if msel in MGMT:
        nm,freq,desc=MGMT[msel]; add(f"{nm} ({freq})", desc, mafee)
    oneoffs=[]; oneoffs_full=[]
    for r in range(50,58):
        de=g(f'B{r}'); pr=num(g(f'F{r}'))
        if (de and str(de).strip()) or (pr and pr>0):
            dl=g(f'G{r}')
            oneoffs.append((str(de or '').strip(), "", pr or 0))
            oneoffs_full.append((str(de or '').strip(), str(g(f'D{r}') or '').strip(), str(g(f'E{r}') or '').strip(), pr or 0, (fmt_date(dl) if hasattr(dl,'year') else (str(dl).strip() if dl else '')), ""))
    # Registrations rows 60-66. PAYE/VAT/CIS/self-assessment use the same online forms as a limited company.
    # The partnership registration itself ('company') is the one step handled by email -> no form link.
    _REGDEFS=[('company','Partnership registration',60),
              ('paye','PAYE registration',61),
              ('vat','VAT registration',62),
              ('cis_sub','CIS subcontractor registration',63),
              ('cis_con','CIS contractor registration',64),
              ('sa',"Partners' self-assessment registration",65),
              ('other','Other registration',66)]
    regs=[]
    for key,label,row in _REGDEFS:
        if str(g(f'C{row}') or '').strip().lower()!='required': continue
        _form=None if key=='company' else REG_LINKS.get(key)   # partnership reg is email-only; all others link like Ltd
        regs.append(dict(key=key,label=label,fee=num(g(f'F{row}')) or 0,included=False,form=_form))
    needs_reg=len(regs)>0
    # Partner self-assessment tiers (row 47, directly below the monthly totals: C47 Low, D47 Medium, E47 High)
    plow=int(num(g('C47')) or 0); pmed=int(num(g('D47')) or 0); phigh=int(num(g('E47')) or 0)
    ptot=120*plow+300*pmed+500*phigh
    parts=[]
    if plow:  parts.append(f"{plow} at {gbp(120)} + VAT")
    if pmed:  parts.append(f"{pmed} at {gbp(300)} + VAT")
    if phigh: parts.append(f"{phigh} at {gbp(500)} + VAT")
    pitem=(" and ".join(parts)+f" = {gbp(ptot)} + VAT a year") if parts else ""
    return dict(company=g('C10') or 'The Partnership',contact=g('C11') or 'there',date=fmt_date(g('C14')),prepared_by=str(g('C15') or 'Shabbir Rahman FCCA').strip(),email=str(g('C12') or '').strip(),phone=str(g('C13') or '').strip(),
                band=str(g('C18') or ''),lines=lines,sub=num(g('G44')) or 0,vat=num(g('G45')) or 0,gross=num(g('G46')) or 0,discount=num(g('K4')) or 0,
                oneoffs=oneoffs,oneoffs_full=oneoffs_full,osub=num(g('F67')) or 0,ovat=num(g('F68')) or 0,ogross=num(g('F69')) or 0,
                directors=0,reg=needs_reg,regs=regs,is_partnership=True,notes=str(g('B72') or '').strip(),
                internal_notes=str(g('I10') or '').strip(),ma_sel=msel,bk=str(g('C27') or '').strip(),
                partner_low=plow,partner_med=pmed,partner_high=phigh,partner_count=plow+pmed+phigh,partner_sa_total=ptot,partner_sa_itemised=pitem,
                vat_scheme=("Flat Rate" if 'flat rate' in str(g('C29') or '').lower() else ("Standard" if str(g('C29') or '').strip().upper() not in ('','NA','-','NO','NONE') else "")),
                source=str(g('J25') or '').strip(),referrer=str(g('J26') or '').strip())


def build_ltd(wb, out, include_tiers=True, ref=None):
    d=read_ltd(wb); ref=ref or ref_for(d['company']); d['ref']=ref; p=PDF(); p.alias_nb_pages()
    cover(p,d['company'],"Limited company accounting & tax",d['date'],ref,prepared_by=d.get('prepared_by'))
    letter(p,d['contact'],d['company'],[
        "Thank you for considering A2Z. Most firms send a price list. We would rather set out exactly what you get, who stands behind it, and why ambitious businesses across the UK trust us with their numbers.",
        "Our approach is simple. Every figure we produce passes a four-layer review and is signed off personally by a chartered certified (FCCA) director, so what reaches you, and HMRC, is right. You get a named team, a fixed monthly fee, and a reply the same working day whenever you need us.",
        "We also look forward, not only back. Where there is tax to save or a decision to weigh, we raise it in time for you to act, rather than reporting it once the year has closed.",
        f"The proposal that follows is tailored to {d['company']}. If anything should be added, removed or adjusted, just tell us. It is your engagement, and we will shape it around you."],date=d['date'])
    if include_tiers:
        has_cis=any(str(l[0]).startswith("CIS") for l in d['lines'])
        has_contractor=any("Contractor" in str(l[0]) for l in d['lines'])
        _ma=str(d.get('ma_sel','')).strip(); _mt=('T1' if _ma.startswith('T1') else 'T2' if _ma.startswith('T2') else 'T3' if _ma.startswith('T3') else None)
        _rec=package_for(d['sub'])[0]
        if _mt=='T3': _rec="Tier 3 Strategic"
        elif _mt in ('T1','T2') and _rec=="Tier 1 Compliance": _rec="Tier 2 Growth"
        tiers_page(p, has_cis, has_contractor, _rec, d['company'], client_items=[str(l[0]) for l in d['lines']], client_fee=d['sub'], ma_tier=_mt)
    p.add_page(); p.klabel("Your engagement"); p.heading("Your tailored quote",22)
    p.lead(f"These are the services agreed for {d['company']}, delivered as one fixed monthly fee by a named team. This page is the definitive list of what is included; nothing else is added unless you ask us to.")
    y=p.get_y(); bh=27; p.rrect(p.X0,y,p.CW,bh,fill=NAVY,style="F")
    p.f("Nunito","B",7.5,A9BBD0); p.set_xy(p.X0+8,y+6.5); p.cell(0,4," ".join("YOUR MONTHLY FEE"))
    p.f("Cormorant","B",17,WHITE); p.set_xy(p.X0+8,y+12); p.cell(0,8,"Your finance function")
    p.f("Cormorant","B",29,WHITE); p.set_xy(p.X0+86,y+4.5); p.cell(p.CW-94,11,gbp(d['sub']),align="R")
    p.f("Nunito","",8,A9BBD0); p.set_xy(p.X0+86,y+18.6); p.cell(p.CW-94,4,f"+ VAT per month  ·  {gbp(d['gross'])} gross",align="R")
    p.set_y(y+bh+2)
    fee_table(p,d['lines'],prices=False)
    total_row(p,"Monthly subtotal",d['sub']); total_row(p,"VAT @ 20%",d['vat']); total_row(p,"Gross monthly",d['gross'],grand=True)
    if (num(d.get('discount',0)) or 0)>0:
        _dz=d['discount']; p.ln(2); p.need(16); _yd=p.get_y(); _hd=15
        p.set_fill_color(*PALE); p.rect(p.X0,_yd,p.CW,_hd,style="F"); p.set_fill_color(*NAVY); p.rect(p.X0,_yd,1.4,_hd,style="F")
        p.f("Nunito","",9,(39,50,59)); p.set_xy(p.X0+6,_yd+2.5)
        p.multi_cell(p.CW-12,4.6,f"Our fees are fixed. As an exceptional act of discretion, we have applied a goodwill discount of {gbp(_dz)} + VAT per month. Your fee above already reflects this reduction.",align="L")
        p.set_y(_yd+_hd)
    if d['directors']>0:
        p.ln(2); p.need(14); y=p.get_y(); h=13
        p.set_fill_color(*PALE); p.rect(p.X0,y,p.CW,h,style="F"); p.set_fill_color(*NAVY); p.rect(p.X0,y,1.4,h,style="F")
        p.f("Nunito","",9,(39,50,59)); p.set_xy(p.X0+6,y+2.5)
        p.multi_cell(p.CW-12,4.6,f"Directors' self-assessment: {d['directors']} × £120 + VAT per year. Our fixed fee covers income as a company director; any other personal income is quoted separately.",align="L",markdown=True)
        p.set_y(y+h)
    p.ln(2); p.f("Nunito","",8.3,MUTED); p.set_x(p.X0)
    p.multi_cell(p.CW,4.4,"Your fee is fixed. It only changes if the scope of work changes materially, and is never billed by the hour or by surprise.",align="L",new_x=XPos.LMARGIN,new_y=YPos.NEXT)
    _oo=[(a,b,c) for a,b,c in d['oneoffs']] if d.get('oneoffs') else []
    _extra=[]
    if (num(d.get('catchup',0)) or 0)>0: _extra.append(("Catch-up / backdated work","",num(d['catchup'])))
    for _a in (d.get('adhocs') or []):
        if isinstance(_a,dict): _extra.append((str(_a.get('label') or "Ad-hoc"), str(_a.get('detail') or ""), num(_a.get('amount',0)) or 0))
        else:
            try: _extra.append((str(_a[0]), str(_a[1]) if len(_a)>1 else "", num(_a[2] if len(_a)>2 else 0) or 0))
            except Exception: pass
    _oo=_oo+_extra
    if _oo:
        if _extra:
            _os=sum((c or 0) for a,b,c in _oo); _ov=_os*0.2; _og=_os*1.2
        else:
            _os=d.get('osub',0); _ov=d.get('ovat',0); _og=d.get('ogross',0)
        p.ln(3); p.need(26); p.f("Cormorant","B",13,NAVY); p.cell(0,7,"One-off, on starting",new_x=XPos.LMARGIN,new_y=YPos.NEXT); p.ln(1)
        fee_table(p,_oo,unit="£")
        total_row(p,"One-off subtotal",_os); total_row(p,"VAT @ 20%",_ov); total_row(p,"One-off inc VAT",_og,grand=True)
    _cnote=(d.get('client_notes') or '').strip() or (d.get('notes') or '').strip()
    if _cnote:
        p.ln(3); bw=p.CW-11; lh=4.6
        p.set_font("Nunito","",9.3); nl=p.multi_cell(bw,lh,_cnote,dry_run=True,output="LINES")
        h=8.6+max(len(nl),1)*lh+2; p.need(h+2); y=p.get_y()
        p.rrect(p.X0,y,p.CW,h, r=2.2, fill=PALE, draw=PALE, style="F")
        p.f("Cormorant","B",13,NAVY); p.set_xy(p.X0+6,y+2.6); p.cell(0,6,"A note on your engagement")
        p.f("Nunito","",9.3,(39,50,59)); p.set_xy(p.X0+6,y+8.8); p.multi_cell(bw,lh,_cnote,align="L")
        p.set_y(y+h+2)
    reg_support_page(p, d)
    mgmt_accounts_page(p, d)
    onboarding(p, "LTD", needs_reg=bool(d.get('comp_reg')))   # reg form iff company registration is required, else onboarding form
    commit_accept(p,d['company'])
    rate_card_page(p, wb)
    proof(p)
    p.output(out); return d

def build_sa(wb, out, ref=None):
    d=read_sa(wb); ref=ref or ref_for(d['company']); d['ref']=ref; p=PDF(); p.alias_nb_pages()
    cover(p,d['company'],"Self-assessment & sole-trader accounting",d['date'],ref,prepared_by=d.get('prepared_by'))
    letter(p,d['contact'],d['company'],[
        "Thank you for considering A2Z. Whether you're a sole trader, landlord or company director, your self-assessment should be accurate, filed early, and never a last-minute scramble. It should also come with someone who actually picks up the phone.",
        "Every return we prepare is reviewed and signed off by a chartered certified (FCCA) director, at a fixed fee agreed up front with no surprises. We tell you exactly what we need and when, and aim to file well ahead of the 31 January deadline. Since we founded the firm we've recorded zero HMRC penalties and zero missed deadlines.",
        "The proposal below is tailored to you. If anything should change, just say, and we'll shape it around your circumstances."],date=d['date'])
    p.add_page(); p.klabel("Your engagement"); p.heading("A clear, fixed fee",22)
    p.lead(f"Your **{d['ctype']}** engagement, with everything set out up front.")
    y=p.get_y(); p.rrect(p.X0,y,p.CW,24,fill=NAVY,style="F")
    p.f("Nunito","",7.5,A9BBD0); p.set_xy(p.X0+8,y+5); p.cell(0,4," ".join("YOUR FEE"))
    p.f("Cormorant","B",14,WHITE); p.set_xy(p.X0+8,y+11); p.cell(80,8,d['ctype'][:34])
    if d['annual']>0: big=gbp(d['annual']); suff="+ VAT / year"
    else: big=gbp(d['monthly']); suff="+ VAT / month"
    p.f("Cormorant","B",27,WHITE); p.set_xy(p.X0+92,y+5); p.cell(p.CW-100,12,big,align="R")
    if d['annual']>0 and d['monthly']>0: suff=f"+ VAT / yr  ·  plus {gbp(d['monthly'])}/mo"
    p.f("Nunito","",8,A9BBD0); p.set_xy(p.X0+50,y+17.5); p.cell(p.CW-58,4,suff,align="R")
    p.set_y(y+30)
    p.f("NunitoSemi","",7.5,MUTED); p.set_x(p.X0); p.cell(0,6," ".join("WHAT'S INCLUDED"),new_x=XPos.LMARGIN,new_y=YPos.NEXT); p.hline(LINE)
    for lab,det,amt,per in d['lines']:
        y0=p.get_y()
        p.f("NunitoSemi","",10,NAVY); p.set_xy(p.X0,y0+1.6); p.cell(0,5,lab,new_x=XPos.LMARGIN,new_y=YPos.NEXT)
        p.f("Nunito","",8.6,(120,130,140)); p.set_xy(p.X0,y0+6.8); p.cell(0,4,det,new_x=XPos.LMARGIN,new_y=YPos.NEXT)
        p.set_y(y0+12); p.hline(SOFTLINE)
    p.ln(1.5)
    if d['annual']>0:
        total_row(p,"Total per year",d['annual']); total_row(p,"Per year + VAT",round(d['annual_gross'],2),grand=True)
    if d['monthly']>0:
        total_row(p,"Total per month",d['monthly']); total_row(p,"Per month + VAT",round(d['monthly_gross'],2),grand=True)
    _da=num(d.get('discount_annual',0)) or 0; _dm=num(d.get('discount_monthly',0)) or 0
    if _da>0 or _dm>0:
        _pp=[]
        if _da>0: _pp.append(f"{gbp(_da)} + VAT per year")
        if _dm>0: _pp.append(f"{gbp(_dm)} + VAT per month")
        p.ln(2); p.need(16); _yd=p.get_y(); _hd=15
        p.set_fill_color(*PALE); p.rect(p.X0,_yd,p.CW,_hd,style="F"); p.set_fill_color(*NAVY); p.rect(p.X0,_yd,1.4,_hd,style="F")
        p.f("Nunito","",9,(39,50,59)); p.set_xy(p.X0+6,_yd+2.5)
        p.multi_cell(p.CW-12,4.6,f"Our fees are fixed. As an exceptional act of discretion, we have applied a goodwill discount of {' and '.join(_pp)}. Your fee above already reflects this reduction.",align="L")
        p.set_y(_yd+_hd)
    p.ln(2); p.f("Nunito","",8.3,MUTED); p.set_x(p.X0)
    p.multi_cell(p.CW,4.4,"Fees are fixed and agreed up front. They only change if the scope of work changes materially. Pay annually, or spread the cost monthly.",align="L",new_x=XPos.LMARGIN,new_y=YPos.NEXT)
    _oo=[(a,b,c) for a,b,c in d['oneoffs']] if d.get('oneoffs') else []
    _extra=[]
    if (num(d.get('catchup',0)) or 0)>0: _extra.append(("Catch-up / backdated work","",num(d['catchup'])))
    for _a in (d.get('adhocs') or []):
        if isinstance(_a,dict): _extra.append((str(_a.get('label') or "Ad-hoc"), str(_a.get('detail') or ""), num(_a.get('amount',0)) or 0))
        else:
            try: _extra.append((str(_a[0]), str(_a[1]) if len(_a)>1 else "", num(_a[2] if len(_a)>2 else 0) or 0))
            except Exception: pass
    _oo=_oo+_extra
    if _oo:
        if _extra:
            _os=sum((c or 0) for a,b,c in _oo); _ov=_os*0.2; _og=_os*1.2
        else:
            _os=d.get('osub',0); _ov=d.get('ovat',0); _og=d.get('ogross',0)
        p.ln(3); p.need(26); p.f("Cormorant","B",13,NAVY); p.set_x(p.X0); p.cell(0,7,"One-off / catch-up fees",new_x=XPos.LMARGIN,new_y=YPos.NEXT); p.ln(1)
        fee_table(p,_oo,unit="£")
        total_row(p,"One-off subtotal",_os); total_row(p,"VAT @ 20%",_ov); total_row(p,"One-off inc VAT",_og,grand=True)
    _cnote=(d.get('client_notes') or '').strip() or (d.get('notes') or '').strip()
    if _cnote:
        p.ln(3); bw=p.CW-12; lh=4.6
        p.set_font("Nunito","",9.3); nl=p.multi_cell(bw,lh,_cnote,dry_run=True,output="LINES")
        h=8.6+max(len(nl),1)*lh+2; p.need(h+2); y=p.get_y()
        p.rrect(p.X0,y,p.CW,h, r=2.2, fill=PALE, draw=PALE, style="F")
        p.f("Cormorant","B",13,NAVY); p.set_xy(p.X0+6,y+2.6); p.cell(0,6,"A note on your engagement")
        p.f("Nunito","",9.3,(39,50,59)); p.set_xy(p.X0+6,y+8.8); p.multi_cell(bw,lh,_cnote,align="L")
        p.set_y(y+h+2)
    who=(d['who'] or "").strip()
    if who:
        p.ln(3); bw=p.CW-12; lh=4.5
        p.set_font("Nunito","",9.2); nl=p.multi_cell(bw,lh,who,dry_run=True,output="LINES")
        h=12+max(len(nl),1)*lh+3; p.need(h+2); y=p.get_y()
        p.rrect(p.X0,y,p.CW,h, r=2.2, fill=PALE, draw=PALE, style="F")
        p.f("Cormorant","B",13.5,NAVY); p.set_xy(p.X0+6,y+5); p.cell(0,6,"Who this is for")
        p.f("Nunito","",9.2,(74,86,96)); p.set_xy(p.X0+6,y+11.5); p.multi_cell(bw,lh,who,align="L")
        p.set_y(y+h+1)
    prov=(d['provides'] or "").strip()
    send=(d['send'] or "Email your records and supporting documents to accounts@a2zaccounting.co.uk.").strip()
    items=[(t,x) for t,x in [("What we need from you",prov),("How to send us your records",send)] if x]
    if items:
        bw=p.CW-14; lh=4.8; blocks=[]
        for t,x in items:
            p.set_font("Nunito","",9.7); nl=p.multi_cell(bw,lh,x,dry_run=True,output="LINES"); blocks.append((t,x,max(len(nl),1)))
        H=13+sum(6+n*lh+5 for _,_,n in blocks)
        p.need(H+6); p.ln(5); y=p.get_y()
        p.set_draw_color(*NAVY); p.set_line_width(0.6); p.rect(p.X0,y,p.CW,H,style="D")
        p.set_fill_color(*NAVY); p.rect(p.X0,y,p.CW,9,style="F")
        p.f("NunitoSemi","",8.5,WHITE); p.set_xy(p.X0+6,y+2.5); p.cell(0,5,"IMPORTANT  ·  YOUR RECORDS")
        yy=y+13.5
        for t,x,n in blocks:
            p.f("Cormorant","B",13,NAVY); p.set_xy(p.X0+7,yy); p.cell(0,5.5,t,new_x=XPos.LMARGIN,new_y=YPos.NEXT)
            p.f("Nunito","",9.7,(33,44,54)); p.set_xy(p.X0+7,yy+6); p.multi_cell(bw,lh,x,align="L")
            yy+=6+n*lh+5
        p.set_y(y+H+2)
    reg_support_page(p, d)
    onboarding(p, "SA", needs_reg=bool(d.get('sa_reg')), direct_debit=(d["freq"]=="Monthly"))   # reg form iff self-assessment registration required
    commit_accept(p,d['company'])
    rate_card_page(p, wb)
    proof(p)
    p.output(out); return d

def build_partnership(wb, out, include_tiers=True, ref=None):
    d=read_partnership(wb); ref=ref or ref_for(d['company']); d['ref']=ref; p=PDF(); p.alias_nb_pages()
    cover(p,d['company'],"Partnership accounting & tax",d['date'],ref,prepared_by=d.get('prepared_by'))
    letter(p,d['contact'],d['company'],[
        "Thank you for considering A2Z. Most firms send a price list. We would rather set out exactly what you get, who stands behind it, and why ambitious businesses across the UK trust us with their numbers.",
        "Our approach is simple. Every figure we produce passes a four-layer review and is signed off personally by a chartered certified (FCCA) director, so what reaches you, the partners, and HMRC, is right. You get a named team, a fixed monthly fee, and a reply the same working day whenever you need us.",
        "We also look forward, not only back. Where there is tax to save or a decision to weigh, we raise it in time for you to act, rather than reporting it once the year has closed.",
        f"The proposal that follows is tailored to {d['company']}. If anything should be added, removed or adjusted, just tell us. It is your engagement, and we will shape it around you."],date=d['date'])
    if include_tiers:
        has_cis=any(str(l[0]).startswith("CIS") for l in d['lines'])
        has_contractor=any("Contractor" in str(l[0]) for l in d['lines'])
        _ma=str(d.get('ma_sel','')).strip(); _mt=('T1' if _ma.startswith('T1') else 'T2' if _ma.startswith('T2') else 'T3' if _ma.startswith('T3') else None)
        _rec=package_for(d['sub'])[0]
        if _mt=='T3': _rec="Tier 3 Strategic"
        elif _mt in ('T1','T2') and _rec=="Tier 1 Compliance": _rec="Tier 2 Growth"
        tiers_page(p, has_cis, has_contractor, _rec, d['company'], client_items=[str(l[0]) for l in d['lines']], client_fee=d['sub'], ma_tier=_mt, partnership=True)
    p.add_page(); p.klabel("Your engagement"); p.heading("Your tailored quote",22)
    p.lead(f"These are the services agreed for {d['company']}, delivered as one fixed monthly fee by a named team. This page is the definitive list of what is included; nothing else is added unless you ask us to.")
    y=p.get_y(); bh=27; p.rrect(p.X0,y,p.CW,bh,fill=NAVY,style="F")
    p.f("Nunito","B",7.5,A9BBD0); p.set_xy(p.X0+8,y+6.5); p.cell(0,4," ".join("YOUR MONTHLY FEE"))
    p.f("Cormorant","B",17,WHITE); p.set_xy(p.X0+8,y+12); p.cell(0,8,"Your finance function")
    p.f("Cormorant","B",29,WHITE); p.set_xy(p.X0+86,y+4.5); p.cell(p.CW-94,11,gbp(d['sub']),align="R")
    p.f("Nunito","",8,A9BBD0); p.set_xy(p.X0+86,y+18.6); p.cell(p.CW-94,4,f"+ VAT per month  \u00b7  {gbp(d['gross'])} gross",align="R")
    p.set_y(y+bh+2)
    fee_table(p,d['lines'],prices=False)
    total_row(p,"Monthly subtotal",d['sub']); total_row(p,"VAT @ 20%",d['vat']); total_row(p,"Gross monthly",d['gross'],grand=True)
    if (num(d.get('discount',0)) or 0)>0:
        _dz=d['discount']; p.ln(2); p.need(16); _yd=p.get_y(); _hd=15
        p.set_fill_color(*PALE); p.rect(p.X0,_yd,p.CW,_hd,style="F"); p.set_fill_color(*NAVY); p.rect(p.X0,_yd,1.4,_hd,style="F")
        p.f("Nunito","",9,(39,50,59)); p.set_xy(p.X0+6,_yd+2.5)
        p.multi_cell(p.CW-12,4.6,f"Our fees are fixed. As an exceptional act of discretion, we have applied a goodwill discount of {gbp(_dz)} + VAT per month. Your fee above already reflects this reduction.",align="L")
        p.set_y(_yd+_hd)
    if d.get('partner_count',0)>0:
        p.ln(2); p.need(16); y=p.get_y(); h=15
        p.set_fill_color(*PALE); p.rect(p.X0,y,p.CW,h,style="F"); p.set_fill_color(*NAVY); p.rect(p.X0,y,1.4,h,style="F")
        p.f("Nunito","",9,(39,50,59)); p.set_xy(p.X0+6,y+2.5)
        p.multi_cell(p.CW-12,4.6,f"Partners' self-assessment: {d['partner_sa_itemised']}. Each partner's personal return is billed separately; our fixed fee covers the partnership itself.",align="L",markdown=True)
        p.set_y(y+h)
    p.ln(2); p.f("Nunito","",8.3,MUTED); p.set_x(p.X0)
    p.multi_cell(p.CW,4.4,"Your fee is fixed. It only changes if the scope of work changes materially, and is never billed by the hour or by surprise.",align="L",new_x=XPos.LMARGIN,new_y=YPos.NEXT)
    _oo=[(a,b,c) for a,b,c in d['oneoffs']] if d.get('oneoffs') else []
    _extra=[]
    if (num(d.get('catchup',0)) or 0)>0: _extra.append(("Catch-up / backdated work","",num(d['catchup'])))
    for _a in (d.get('adhocs') or []):
        if isinstance(_a,dict): _extra.append((str(_a.get('label') or "Ad-hoc"), str(_a.get('detail') or ""), num(_a.get('amount',0)) or 0))
        else:
            try: _extra.append((str(_a[0]), str(_a[1]) if len(_a)>1 else "", num(_a[2] if len(_a)>2 else 0) or 0))
            except Exception: pass
    _oo=_oo+_extra
    if _oo:
        if _extra:
            _os=sum((c or 0) for a,b,c in _oo); _ov=_os*0.2; _og=_os*1.2
        else:
            _os=d.get('osub',0); _ov=d.get('ovat',0); _og=d.get('ogross',0)
        p.ln(3); p.need(26); p.f("Cormorant","B",13,NAVY); p.cell(0,7,"One-off, on starting",new_x=XPos.LMARGIN,new_y=YPos.NEXT); p.ln(1)
        fee_table(p,_oo,unit="\u00a3")
        total_row(p,"One-off subtotal",_os); total_row(p,"VAT @ 20%",_ov); total_row(p,"One-off inc VAT",_og,grand=True)
    _cnote=(d.get('client_notes') or '').strip() or (d.get('notes') or '').strip()
    if _cnote:
        p.ln(3); bw=p.CW-11; lh=4.6
        p.set_font("Nunito","",9.3); nl=p.multi_cell(bw,lh,_cnote,dry_run=True,output="LINES")
        h=8.6+max(len(nl),1)*lh+2; p.need(h+2); y=p.get_y()
        p.rrect(p.X0,y,p.CW,h, r=2.2, fill=PALE, draw=PALE, style="F")
        p.f("Cormorant","B",13,NAVY); p.set_xy(p.X0+6,y+2.6); p.cell(0,6,"A note on your engagement")
        p.f("Nunito","",9.3,(39,50,59)); p.set_xy(p.X0+6,y+8.8); p.multi_cell(bw,lh,_cnote,align="L")
        p.set_y(y+h+2)
    reg_support_page(p, d)
    mgmt_accounts_page(p, d)
    onboarding_partnership(p, needs_reg=d['reg'])
    commit_accept(p,d['company'])
    rate_card_page(p, wb, partnership=True)
    proof(p)
    p.output(out); return d


def build_email(d, kind):
    K=kind.upper()
    name=(str(d.get('contact') or 'there').strip()) or 'there'
    company=(str(d.get('company') or 'your business').strip()) or 'your business'
    subject=f"Your A2Z proposal for {company}"

    svc=[str(r[0]).strip() for r in d.get('lines',[]) if str(r[0]).strip()]
    regs=d.get('regs') or []
    keys={r.get('key') for r in regs}

    L=[f"Hi {name},","",
       (f"Thank you for the opportunity to look after {company} - it's a pleasure to put this proposal "
        f"together for you. I've attached the full proposal, and here's a quick summary."),"",
       "What we'd take care of for you:"]
    L+=[f"\u2022 {x}" for x in (svc or ["the services we discussed"])]
    L+=[""]

    # ---- monthly/role lines ----
    if K=='LTD':
        L+=[f"Your fee would be {gbp(d.get('sub',0))} + VAT a month - a single fixed amount, with no surprise bills along the way."]
        if (num(d.get('discount',0)) or 0)>0:
            L+=[f"As an exceptional act of discretion, a goodwill discount of {gbp(d['discount'])} + VAT per month has been applied. Your fee above already reflects this."]
        nd=int(num(d.get('directors',0)) or 0)
        if nd==1:
            L+=["Each director's personal tax return (self-assessment) is \u00a3120 + VAT a year, billed separately."]
        elif nd>1:
            L+=[f"Each director's personal tax return (self-assessment) is \u00a3120 + VAT a year, billed separately - for {nd} directors that comes to {gbp(nd*120)} + VAT a year."]
    elif K=='PARTNERSHIP':
        L+=[f"Your fee would be {gbp(d.get('sub',0))} + VAT a month - a single fixed amount covering the partnership itself, with no surprise bills along the way."]
        if (num(d.get('discount',0)) or 0)>0:
            L+=[f"As an exceptional act of discretion, a goodwill discount of {gbp(d['discount'])} + VAT per month has been applied. Your fee above already reflects this."]
        item=str(d.get('partner_sa_itemised') or '').strip()
        if item:
            L+=[f"Each partner's personal tax return (self-assessment) is billed separately: {item}."]
    else:
        m=num(d.get('monthly',0)) or 0; a=num(d.get('annual',0)) or 0
        if m>0:   L+=[f"Your fee would be {gbp(d['monthly'])} + VAT a month - a single fixed amount, with no surprise bills along the way."]
        elif a>0: L+=[f"Your fee would be {gbp(d['annual'])} + VAT a year - a single fixed amount, with no surprise bills along the way."]
        else:     L+=["Your fee is set out in the attached proposal."]
        _da=num(d.get('discount_annual',0)) or 0; _dm=num(d.get('discount_monthly',0)) or 0
        if _da>0 or _dm>0:
            _pp=[]
            if _da>0: _pp.append(f"{gbp(_da)} + VAT per year")
            if _dm>0: _pp.append(f"{gbp(_dm)} + VAT per month")
            L+=[f"As an exceptional act of discretion, a goodwill discount of {' and '.join(_pp)} has been applied. Your fee above already reflects this."]

    osub=num(d.get('osub',0)) or 0

    # ===================== PARTNERSHIP: email-based, NO form links =====================
    if K=='PARTNERSHIP':
        reg_lines=[]
        for r in regs:
            fee=num(r.get('fee')) or 0
            if r.get('included'): reg_lines.append(f"{r['label']} - included")
            elif fee>0:          reg_lines.append(f"{r['label']} - {gbp(fee)} + VAT (one-off)")
            else:                reg_lines.append(f"{r['label']} - one-off")
        if reg_lines:
            L+=["","We'd also handle these one-off registrations for you, as your authorised agent:"]
            L+=[f"\u2022 {x}" for x in reg_lines]
        reg_total=sum((num(r.get('fee')) or 0) for r in regs if not r.get('included'))
        other=osub-reg_total
        if other>0.5:
            L+=["",f"There's also a one-off setup of {gbp(other)} + VAT to get everything in place at the start."]
        formlines=[]; seen=set()
        for r in regs:
            if r.get('key')=='company': continue   # partnership registration: handled by email, no form
            fm=r.get('form')
            if fm and fm[1] not in seen:
                lab=fm[0][:-5] if str(fm[0]).lower().endswith(" form") else fm[0]
                formlines.append(f"{lab}: {fm[1]}"); seen.add(fm[1])
        L+=[""]
        if formlines:
            L+=["When you're ready to go ahead, you can complete these short online forms (a few minutes each):"]
            L+=formlines
        dd=DD_FORM if (DD_FORM and str(DD_FORM).startswith("http")) else None
        if dd:
            L+=["",f"Direct Debit mandate (so your fees collect automatically): {dd}"]
        if any(r.get('key')=='company' for r in regs):
            L+=["","Registering the partnership itself with HMRC is the one step we take care of for you by email, so there is no form for that part."]
        L+=["","If you have any questions, or would like to talk anything through, just let me know. I'd be glad to help."]
        return subject, "\n".join(L)

    # ===================== LTD / SA: form-based onboarding (unchanged flow) =====================
    if osub>0:
        L+=[f"There's also a one-off setup of {gbp(osub)} + VAT to get everything in place at the start."]
    L+=[""]

    primary_key='company' if K=='LTD' else 'sa'
    unregistered=primary_key in keys
    pname,purl=FORMS[K]['reg'] if unregistered else FORMS[K]['onboard']
    primary_fee=0
    for r in regs:
        if r.get('key')==primary_key:
            primary_fee=num(r.get('fee')) or 0; break
    feebit=""
    if K=='LTD' and unregistered and primary_fee>0:
        feebit=f" (the company formation fee is {gbp(primary_fee)} + VAT, charged once)"

    L+=[f"Whenever you're ready to go ahead, the next step is simply to complete our {pname}{feebit}, which only takes a few minutes:",
        purl]

    extra=[]; seen=set()
    for r in regs:
        if r.get('key')==primary_key: continue
        fm=r.get('form')
        if fm and fm[1] not in seen: extra.append(fm); seen.add(fm[1])
    if extra:
        L+=["","We'd also take care of a couple of registrations for you - you can complete those here as well:"]
        for nm_,u in extra:
            lab=nm_[:-5] if nm_.lower().endswith(" form") else nm_
            L+=[f"{lab}: {u}"]
    L+=["",
        "If you have any questions, or would like to talk anything through, just let me know - I'd be glad to help."]
    return subject, "\n".join(L)
def safe_load_workbook(path):
    """Copy to a temp file first, so a workbook open in Excel or stored online-only still opens."""
    import shutil, tempfile
    ext=os.path.splitext(path)[1].lower() or ".xlsx"
    tmp=os.path.join(tempfile.gettempdir(),"a2z_workbook_copy"+ext)
    try:
        shutil.copy2(path,tmp); src=tmp
    except Exception:
        src=path
    return openpyxl.load_workbook(src, data_only=True)

def build_internal(wb, kind, out, ref=None):
    kind=kind.upper()
    d = read_ltd(wb) if kind=='LTD' else (read_partnership(wb) if kind=='PARTNERSHIP' else read_sa(wb))
    ref = ref or ref_for(d.get('company',''))
    p=PDF(); p.right_header="INTERNAL WORK BRIEF"; p.alias_nb_pages(); p.add_page()
    X0,CW,XR=p.X0,p.CW,p.XR
    def band(title):
        p.ln(2.5); p.need(16); y=p.get_y()
        p.set_fill_color(*NAVY); p.rect(X0,y,CW,7,style="F")
        p.f("NunitoSemi","",8.3,WHITE); p.set_xy(X0+4,y+1.7); p.cell(0,4," ".join(title))
        p.set_y(y+7); p.ln(2.5)
    p.set_fill_color(*GREEN); p.rect(0,0,210,3.0,style="F")
    p.f("Nunito","B",7.5,GREEN); p.set_xy(X0,15); p.cell(0,4," ".join("FOR THE A2Z TEAM  \u00b7  CONFIDENTIAL"),new_x=XPos.LMARGIN,new_y=YPos.NEXT)
    p.f("Cormorant","B",26,NAVY); p.set_x(X0); p.cell(0,12,"Internal Work Brief",new_x=XPos.LMARGIN,new_y=YPos.NEXT); p.ln(3)
    eng=("Limited company  ·  "+(d.get('band') or '')) if kind=='LTD' else (("Partnership  ·  "+(d.get('band') or '')) if kind=='PARTNERSHIP' else ("Self-assessment  ·  "+(d.get('ctype') or '')))
    L=[("Client",d.get('company','')),("Contact",d.get('contact',''))]
    R=[("Engagement",eng),("Prepared by",d.get('prepared_by') or 'Shabbir Rahman FCCA'),("Date",d.get('date','')),("Reference",ref)]
    rowh=8.6; boxh=5+rowh*4; y=p.get_y(); p.rrect(X0,y,CW,boxh,r=2,fill=PALE,draw=PALE,style="F")
    for i,(lab,val) in enumerate(L):
        yy=y+3.5+i*rowh
        p.f("Nunito","",7,MUTED); p.set_xy(X0+5,yy); p.cell(0,3.4," ".join(lab.upper()))
        p.f("NunitoSemi","",9.5,NAVY); p.set_xy(X0+5,yy+3.3); p.cell(CW/2-8,4.6,str(val))
    for i,(lab,val) in enumerate(R):
        yy=y+3.5+i*rowh; xx=X0+CW/2+3
        p.f("Nunito","",7,MUTED); p.set_xy(xx,yy); p.cell(0,3.4," ".join(lab.upper()))
        p.f("NunitoSemi","",9.5,NAVY); p.set_xy(xx,yy+3.3); p.cell(CW/2-8,4.6,str(val))
    p.set_y(y+boxh)
    src=(d.get('source') or '').strip(); refn=(d.get('referrer') or '').strip()
    if src or refn:
        p.ln(1); yy=p.get_y(); p.set_fill_color(*PALE); p.rect(X0,yy,CW,8.5,style="F"); p.set_fill_color(*GREEN); p.rect(X0,yy,1.4,8.5,style="F")
        info="How they found us:  "+(src or "not recorded")
        if refn: info+="      Referrer:  "+refn
        p.f("NunitoSemi","",9,NAVY); p.set_xy(X0+5,yy+2.5); p.cell(0,4,info)
        p.set_y(yy+8.5)
    facts=[]
    _vs=str(d.get('vat_scheme') or '').strip()
    if _vs: facts.append("VAT scheme:  "+_vs)
    _ms=str(d.get('ma_sel') or '').strip()
    if _ms and _ms.upper()!='NA':
        _MN={'T1':'Financial Health','T2':'Business Performance','T3':'Strategic Advisory'}
        _t=_ms.split(' ',1)[0]; _fr=_ms.split(' ',1)[1] if ' ' in _ms else ''
        facts.append("Mgmt report:  "+_MN.get(_t,_t)+((" ("+_fr.lower()+")") if _fr else ""))
    _nd=int(d.get('directors',0) or 0)
    if kind=='LTD' and _nd>0: facts.append("Director SA:  "+str(_nd)+(" directors" if _nd!=1 else " director"))
    _np=int(d.get('partner_count',0) or 0)
    if kind=='PARTNERSHIP' and _np>0: facts.append("Partner SA:  "+str(_np)+(" partners" if _np!=1 else " partner"))
    if facts:
        p.ln(1); yy=p.get_y(); p.set_fill_color(*PALE); p.rect(X0,yy,CW,8.5,style="F"); p.set_fill_color(*NAVY); p.rect(X0,yy,1.4,8.5,style="F")
        p.f("NunitoSemi","",8.5,NAVY); p.set_xy(X0+5,yy+2.7); p.cell(0,4,"        ".join(facts))
        p.set_y(yy+8.5)
    # ---------- WORK BY DEPARTMENT ----------
    DEPTS=["Admin","Payroll","Year End","Advisory"]
    DEPTCOL={"Admin":(90,100,112),"Payroll":(38,110,150),"Year End":(30,107,71),"Advisory":(120,80,150)}
    def dept_of(label):
        l=str(label).lower()
        if any(k in l for k in ('payroll','cis','pension')): return 'Payroll'
        if any(k in l for k in ('corporation tax','annual accounts','self-assessment','self assessment','companies house','confirmation statement','address service')): return 'Year End'
        if any(k in l for k in ('health report','performance report','advisory report','bookkeeping','vat','contractor','project accounting','management')): return 'Advisory'
        return 'Admin'
    def map_oneoff_dept(tag):
        t=str(tag or '').lower()
        if 'payroll' in t: return 'Payroll'
        if any(k in t for k in ('year','ye','account','tax','self')): return 'Year End'
        if any(k in t for k in ('advis','book','vat','manage')) or t=='ma': return 'Advisory'
        return 'Admin'
    buckets={k:[] for k in DEPTS}
    for rg in (d.get('regs') or []):
        if rg['included']: tag="Included \u00b7 \u00a30"
        elif rg['fee']>0: tag=gbp(rg['fee'])+" + VAT"
        else: tag="Required"
        extra=("Send: "+rg['form'][0]) if rg.get('form') else ""
        if rg['key']=='paye' and rg['included']: extra=(extra+"   " if extra else "")+"PAYE reg still required after company setup."
        buckets['Admin'].append(dict(kind='setup',label="Setup: "+rg['label'],detail=extra,right=tag))
    for row in (d['lines'] or []):
        lab=row[0]; det=row[1] if len(row)>1 else ""; amt=row[2] if len(row)>2 else 0
        unit="/yr" if (len(row)>3 and row[3]=='yr') else "/mo"
        right=(gbp(amt)+" + VAT "+unit) if amt else ""
        buckets[dept_of(lab)].append(dict(kind='recurring',label=lab,detail=det,right=right))
    for desc,dpt,prio,price,deadline,note in (d.get('oneoffs_full') or []):
        buckets[map_oneoff_dept(dpt)].append(dict(kind='oneoff',label=(desc or "(item)"),detail=note,right=(gbp(price)+" + VAT" if price else ""),prio=prio,deadline=deadline))
    band("WORK BY DEPARTMENT")
    PRIO={"High":(192,57,57),"Medium":(198,128,40),"Low":(30,107,71)}
    for dp in DEPTS:
        items=buckets[dp]; col=DEPTCOL[dp]
        p.need(15); ys=p.get_y()
        p.set_fill_color(*col); p.rect(X0,ys,2.2,6.4,style="F")
        p.f("NunitoSemi","",10.5,NAVY); p.set_xy(X0+5.5,ys+1.0); p.cell(0,4.6," ".join(dp.upper()))
        p.f("Nunito","",8,MUTED); p.set_xy(X0+CW-34,ys+1.5); p.cell(34,3.6,((str(len(items))+(" item" if len(items)==1 else " items")) if items else "no items"),align="R")
        p.set_y(ys+6.4); p.hline(SOFTLINE); p.ln(1.6)
        if not items:
            p.f("Nunito","",8.6,GREY); p.set_x(X0+5.5); p.cell(0,4.6,"-  nothing assigned to this team for this client",new_x=XPos.LMARGIN,new_y=YPos.NEXT); p.ln(2.2); continue
        for it in items:
            p.need(10); y0=p.get_y(); rightw=36
            p.f("NunitoSemi","",9.3,NAVY); p.set_xy(X0+5.5,y0); p.cell(4,5,"\u2022"); p.cell(CW-9.5-rightw,5,str(it['label']))
            if it['right']:
                p.f("NunitoSemi","",8.7,NAVY); p.set_xy(X0+CW-rightw,y0); p.cell(rightw,5,it['right'],align="R")
            p.set_y(y0+5)
            meta=("Due: "+it['deadline']) if it.get('deadline') else ""
            if it.get('prio'):
                pc=PRIO.get(it['prio'],(120,130,140)); bx=X0+9.5; by=p.get_y()
                p.rrect(bx,by,17,4.6,r=1.1,fill=pc,draw=pc,style="F")
                p.f("NunitoSemi","",7.1,WHITE); p.set_xy(bx,by+1.0); p.cell(17,2.8,it['prio'].upper(),align="C")
                if meta: p.f("Nunito","",8.1,GREY); p.set_xy(bx+19,by+0.4); p.cell(0,3.8,meta,new_x=XPos.LMARGIN,new_y=YPos.NEXT)
                else: p.set_y(by+4.6)
            elif meta:
                p.f("Nunito","",8.1,GREY); p.set_x(X0+9.5); p.cell(0,4.2,meta,new_x=XPos.LMARGIN,new_y=YPos.NEXT)
            if it['detail']:
                p.f("Nunito","",8.2,GREY); p.set_x(X0+9.5); p.multi_cell(CW-9.5,4.2,str(it['detail']),align="L",new_x=XPos.LMARGIN,new_y=YPos.NEXT)
            p.ln(0.6)
        p.ln(2.2)
    frows=[]
    if kind in ('LTD','PARTNERSHIP'):
        frows.append(("Recurring fee", f"{gbp(d.get('sub',0))} + VAT / month   ({gbp(d.get('gross',0))} inc VAT)"))
        if kind=='LTD':
            _nd2=int(d.get('directors',0) or 0)
            if _nd2>0: frows.append((f"Director self-assessment  ({_nd2})", f"{gbp(120*_nd2)} + VAT / year   (£120 each)"))
        else:
            if (d.get('partner_sa_total') or 0)>0:
                frows.append(("Partner self-assessment", f"{gbp(d.get('partner_sa_total'))} + VAT / year   (billed separately)"))
    else:
        if d.get('annual',0)>0: frows.append(("Recurring fee", f"{gbp(d['annual'])} + VAT / year"))
        if d.get('monthly',0)>0: frows.append(("Recurring fee", f"{gbp(d['monthly'])} + VAT / month"))
    if (d.get('osub') or 0)>0: frows.append(("One-off / catch-up", f"{gbp(d['osub'])} + VAT   ({gbp(d.get('ogross',0))} inc VAT)"))
    _rt=sum((rg.get('fee') or 0) for rg in (d.get('regs') or []) if not rg.get('included'))
    if _rt>0: frows.append(("Registrations", f"{gbp(_rt)} + VAT"))
    if not frows: frows.append(("Fee","as set out in the proposal"))
    band("FEE SUMMARY")
    for lab,val in frows:
        p.need(7); y0=p.get_y()
        p.f("Nunito","",9,MUTED); p.set_xy(X0,y0); p.cell(CW*0.42,5,lab)
        p.f("NunitoSemi","",9.5,NAVY); p.set_xy(X0+CW*0.42,y0); p.cell(CW*0.58,5,val,align="R")
        p.set_y(y0+5.4)
    band("RECORDS FROM THE CLIENT")
    prov=(d.get('provides') or '').strip(); send=(d.get('send') or '').strip()
    if kind=='SA':
        if prov:
            p.f("NunitoSemi","",9,NAVY); p.set_x(X0); p.cell(0,5,"What they provide",new_x=XPos.LMARGIN,new_y=YPos.NEXT)
            p.f("Nunito","",9,(58,71,80)); p.set_x(X0); p.multi_cell(CW,4.6,prov, align="L", new_x=XPos.LMARGIN,new_y=YPos.NEXT); p.ln(1.5)
        p.f("NunitoSemi","",9,NAVY); p.set_x(X0); p.cell(0,5,"How they send their records",new_x=XPos.LMARGIN,new_y=YPos.NEXT)
        p.f("Nunito","",9,(58,71,80)); p.set_x(X0); p.multi_cell(CW,4.6,send or "Email records to accounts@a2zaccounting.co.uk.", align="L", new_x=XPos.LMARGIN,new_y=YPos.NEXT)
    else:
        p.f("Nunito","",9,(58,71,80)); p.set_x(X0)
        p.multi_cell(CW,4.6,"Records are received through the agreed bookkeeping software (QuickBooks / Xero / Dext) and the onboarding process. Any documents should be emailed to accounts@a2zaccounting.co.uk.", align="L", new_x=XPos.LMARGIN,new_y=YPos.NEXT)
    if (d.get('internal_notes') or '').strip():
        band("INTERNAL NOTES  \u00b7  TEAM ONLY")
        p.f("Nunito","",9.5,(58,71,80)); p.set_x(X0); p.multi_cell(CW,5,d['internal_notes'], align="L", new_x=XPos.LMARGIN,new_y=YPos.NEXT)
    if (d.get('notes') or '').strip():
        band("NOTE GIVEN TO THE CLIENT")
        p.f("Nunito","",9.5,(58,71,80)); p.set_x(X0); p.multi_cell(CW,5,d['notes'], align="L", new_x=XPos.LMARGIN,new_y=YPos.NEXT)
    p.output(out); return d

def generate(workbook_path, kind, out_path):
    wb=safe_load_workbook(workbook_path)
    # unified reference: use the register code stamped on the proposal sheet
    # (LTD C16 / SA C11) if present, otherwise fall back to a generated one
    try:
        _qc = wb['LTD Proposal']['C16'] if kind.upper()=='LTD' else (wb['Partnership Proposal']['C16'] if kind.upper()=='PARTNERSHIP' else wb['SA Proposal']['C11'])
        ref = str(_qc.value or '').strip() or ref_for("")
    except Exception:
        ref = ref_for("")
    d=build_ltd(wb,out_path,ref=ref) if kind.upper()=='LTD' else (build_partnership(wb,out_path,ref=ref) if kind.upper()=='PARTNERSHIP' else build_sa(wb,out_path,ref=ref))
    subject,body=build_email(d,kind)
    email_path=os.path.splitext(out_path)[0]+" - email.txt"
    try:
        with open(email_path,"w",encoding="utf-8") as fh: fh.write("Subject: "+subject+"\n\n"+body)
    except Exception: email_path=None
    internal_path=os.path.splitext(out_path)[0]+" - INTERNAL.pdf"
    try: build_internal(wb, kind, internal_path, ref=ref)
    except Exception: internal_path=None
    return {"pdf":out_path,"email_path":email_path,"subject":subject,"body":body,"internal_path":internal_path}

def _find_workbook(folder):
    import glob
    c=[f for f in glob.glob(os.path.join(folder,"*.xlsx"))+glob.glob(os.path.join(folder,"*.xlsm"))
       if "proposal" not in os.path.basename(f).lower() and "INTERNAL" not in os.path.basename(f)
       and "Weekly Onboarding Report" not in os.path.basename(f) and not os.path.basename(f).startswith("~$")]
    return max(c,key=os.path.getmtime) if c else None

def _company_name(src, kind):
    try:
        wbx=safe_load_workbook(src)
        if kind=="LTD":
            co=str(wbx['LTD Proposal']['C10'].value or "Client").strip()
        elif kind=="PARTNERSHIP":
            co=str(wbx['Partnership Proposal']['C10'].value or "Client").strip()
        else:
            co=str((wbx['SA Proposal']['C6'].value or wbx['SA Proposal']['C7'].value) or "Client").strip()
    except Exception:
        co="Client"
    safe=" ".join(str(co).split())
    for ch in '\\/:*?"<>|': safe=safe.replace(ch,"-")
    return safe or "Client"

def run_gui():
    """Desktop window: pick Ltd or Sole Trader, build, then SHOW the drafted email to copy."""
    import threading, traceback
    try:
        import tkinter as tk
        from tkinter import filedialog, messagebox
    except Exception:
        return False

    HERE_=os.path.dirname(os.path.abspath(__file__))
    NAVYX="#16375A"; GREENX="#1E6B47"; INKX="#1B2A38"; GREYX="#5B6770"; PAPER="#F4F6F9"; CARD="#FFFFFF"; HAIR="#D9DEE5"
    state={"src":_find_workbook(HERE_)}

    root=tk.Tk(); root.title("A2Z Proposal Builder"); root.configure(bg=PAPER)
    def centre(w,h):
        root.update_idletasks(); x=(root.winfo_screenwidth()-w)//2; y=max(20,(root.winfo_screenheight()-h)//3)
        root.geometry(f"{w}x{h}+{x}+{y}"); root.minsize(w,h)

    band=tk.Frame(root,bg=NAVYX,height=78); band.pack(fill="x"); band.pack_propagate(False)
    tk.Label(band,text="A 2 Z   A C C O U N T I N G   S O L U T I O N S",bg=NAVYX,fg="#A9BBD0",font=("Segoe UI",8,"bold")).pack(anchor="w",padx=22,pady=(16,0))
    tk.Label(band,text="Proposal Builder",bg=NAVYX,fg="white",font=("Georgia",19,"bold")).pack(anchor="w",padx=20,pady=(0,12))
    container=tk.Frame(root,bg=PAPER); container.pack(fill="both",expand=True)
    def clear():
        for w in container.winfo_children(): w.destroy()
    def _open(path):
        try: os.startfile(path)
        except Exception:
            try: os.startfile(HERE_)
            except Exception: pass
    def btn(parent,text,cmd,primary=False):
        bg=NAVYX if primary else "#E7ECF2"; fg="white" if primary else INKX
        b=tk.Button(parent,text=text,command=cmd,relief="flat",bd=0,cursor="hand2",
                    bg=bg,fg=fg,activebackground=(GREENX if primary else "#DCE3EB"),activeforeground=fg,
                    font=("Segoe UI",9,"bold"),padx=14,pady=7)
        return b

    working={"on":False}

    # ---------------- home view ----------------
    def show_home():
        centre(560,420); clear()
        body=tk.Frame(container,bg=PAPER); body.pack(fill="both",expand=True,padx=22,pady=14)
        wbrow=tk.Frame(body,bg=PAPER); wbrow.pack(fill="x"); wb_var=tk.StringVar()
        def refresh():
            sx=state["src"]; wb_var.set(("Workbook:  "+os.path.basename(sx)) if sx else "No workbook found in this folder.")
        refresh()
        tk.Label(wbrow,textvariable=wb_var,bg=PAPER,fg=GREYX,font=("Segoe UI",9)).pack(side="left")
        def choose():
            f=filedialog.askopenfilename(title="Choose the A2Z Growth Engine workbook",filetypes=[("Excel workbook","*.xlsm *.xlsx")])
            if f: state["src"]=f; refresh()
        tk.Button(wbrow,text="Choose\u2026",command=choose,relief="flat",bg=PAPER,fg=NAVYX,font=("Segoe UI",9,"underline"),cursor="hand2",bd=0,activebackground=PAPER).pack(side="right")
        tk.Label(body,text="Which proposal would you like to build?",bg=PAPER,fg=INKX,font=("Segoe UI",12,"bold")).pack(anchor="w",pady=(18,2))
        tk.Label(body,text="Choose the client type. The workbook in this folder is read but never changed.",bg=PAPER,fg=GREYX,font=("Segoe UI",9)).pack(anchor="w",pady=(0,14))
        btns=tk.Frame(body,bg=PAPER); btns.pack(fill="x")
        status=tk.StringVar(value=""); statuslbl=tk.Label(body,textvariable=status,bg=PAPER,fg=GREENX,font=("Segoe UI",9,"bold")); statuslbl.pack(anchor="w",pady=(16,0))
        def big(parent,title,sub,kind):
            fr=tk.Frame(parent,bg=NAVYX,cursor="hand2")
            t=tk.Label(fr,text=title,bg=NAVYX,fg="white",font=("Segoe UI",12,"bold")); s2=tk.Label(fr,text=sub,bg=NAVYX,fg="#A9BBD0",font=("Segoe UI",8))
            t.pack(pady=(16,0)); s2.pack(pady=(2,16))
            def on(e=None): _go(kind,status,statuslbl)
            for w in (fr,t,s2): w.bind("<Button-1>",on)
            fr.bind("<Enter>",lambda e:[w.configure(bg=GREENX) for w in (fr,t,s2)] or s2.configure(fg="#CDE6D7"))
            fr.bind("<Leave>",lambda e:[w.configure(bg=NAVYX) for w in (fr,t,s2)] or s2.configure(fg="#A9BBD0"))
            return fr
        b1=big(btns,"Limited Company","Ltd · accounts & CT","LTD")
        b2=big(btns,"Partnership","Partnership · SA800 & SA","PARTNERSHIP")
        b3=big(btns,"Sole Trader","Self-assessment / SA","SA")
        b1.pack(side="left",expand=True,fill="both",padx=(0,5)); b2.pack(side="left",expand=True,fill="both",padx=(5,5)); b3.pack(side="left",expand=True,fill="both",padx=(5,0))

    def _go(kind,status,statuslbl):
        if working["on"]: return
        src=state["src"]
        if not src:
            messagebox.showwarning("No workbook","Put your A2Z Growth Engine workbook (.xlsm or .xlsx) in this folder, or use Choose\u2026"); return
        working["on"]=True; status.set("Building the "+("Limited Company" if kind=="LTD" else ("Partnership" if kind=="PARTNERSHIP" else "Sole Trader"))+" proposal\u2026 please wait"); statuslbl.configure(fg=GREENX); root.update_idletasks()
        def work():
            try:
                out=os.path.join(HERE_,f"{_company_name(src,kind)} - proposal.pdf")
                d=generate(src,kind,out); root.after(0,lambda:show_result(d))
            except Exception as ex:
                tb=traceback.format_exc(); root.after(0,lambda:(_fail(status,statuslbl,ex,tb)))
        threading.Thread(target=work,daemon=True).start()
    def _fail(status,statuslbl,ex,tb):
        working["on"]=False; status.set("Couldn't build the proposal."); statuslbl.configure(fg="#C03E39")
        messagebox.showerror("Problem building the proposal", str(ex)+"\n\n"+tb)

    # ---------------- result view (shows the drafted email) ----------------
    def show_result(d):
        working["on"]=False
        full="Subject: "+str(d.get("subject",""))+"\n\n"+str(d.get("body",""))
        try:
            _render_result(d, full)
        except Exception:
            # never leave a blank screen: at least copy the email and show it
            try:
                root.clipboard_clear(); root.clipboard_append(full); root.update()
            except Exception: pass
            messagebox.showinfo("Proposal ready",
                "Your files are in this folder.\n\nThe email has been copied to your clipboard - paste it into Outlook and attach the proposal PDF.\n\n"+full)

    def _render_result(d, full):
        centre(640,690); clear()
        f=tk.Frame(container,bg=PAPER); f.pack(fill="both",expand=True,padx=22,pady=(12,14))
        tk.Label(f,text="Proposal ready",bg=PAPER,fg=NAVYX,font=("Georgia",16,"bold")).pack(anchor="w")
        files=[os.path.basename(d["pdf"])+"  (client proposal)"]
        if d.get("internal_path"): files.append(os.path.basename(d["internal_path"])+"  (internal brief)")
        if d.get("email_path"):    files.append(os.path.basename(d["email_path"])+"  (email .txt)")
        tk.Label(f,text="Created in this folder:  "+"   \u00b7   ".join(files),bg=PAPER,fg=GREYX,
                 font=("Segoe UI",8),wraplength=580,justify="left").pack(anchor="w",pady=(2,8))

        msg=tk.StringVar(value="")
        def copy_email():
            try:
                root.clipboard_clear(); root.clipboard_append(full); root.update()
                msg.set("Email copied to clipboard \u2713   -   paste into Outlook (Ctrl+V)")
            except Exception:
                msg.set("Select the text below and press Ctrl+C.")

        # pin status + buttons to the BOTTOM first, so they are always visible
        tk.Label(f,textvariable=msg,bg=PAPER,fg=GREENX,font=("Segoe UI",9,"bold")).pack(side="bottom",anchor="w",pady=(7,0))
        bar=tk.Frame(f,bg=PAPER); bar.pack(side="bottom",fill="x",pady=(10,0))
        btn(bar,"Copy email",copy_email,primary=True).pack(side="left")
        btn(bar,"Open proposal",lambda:_open(d.get("pdf"))).pack(side="left",padx=6)
        btn(bar,"Open folder",lambda:_open(HERE_)).pack(side="left")
        btn(bar,"Build another",show_home).pack(side="right")

        tk.Label(f,text="Your email - ready to paste into Outlook (attach the proposal PDF):",
                 bg=PAPER,fg=INKX,font=("Segoe UI",10,"bold")).pack(anchor="w",pady=(2,4))
        wrap=tk.Frame(f,bg=HAIR); wrap.pack(fill="both",expand=True)
        sb=tk.Scrollbar(wrap); sb.pack(side="right",fill="y")
        txt=tk.Text(wrap,wrap="word",font=("Segoe UI",10),bg=CARD,fg=INKX,bd=0,padx=14,pady=12,
                    spacing1=2,spacing2=1,spacing3=2,yscrollcommand=sb.set); txt.pack(side="left",fill="both",expand=True,padx=1,pady=1)
        sb.config(command=txt.yview)
        txt.insert("1.0",full); txt.configure(state="disabled")
        # auto-copy so it's ready to paste straight away
        try:
            root.clipboard_clear(); root.clipboard_append(full); root.update()
            msg.set("Email copied to clipboard \u2713   -   paste into Outlook (Ctrl+V)")
        except Exception:
            pass

    show_home(); root.mainloop(); return True

if __name__=='__main__':
    # command-line still works:  a2z_proposals_fpdf.py <workbook> <LTD|SA> <output>
    if len(sys.argv)>1:
        src=sys.argv[1]; kind=(sys.argv[2] if len(sys.argv)>2 else 'LTD').upper()
        out=(sys.argv[3] if len(sys.argv)>3 else f"A2Z_{kind}_Proposal.pdf")
        d=generate(src,kind,out); print(f"Generated {out}  (email: {d.get('email_path')})"); sys.exit(0)
    # normal double-click: open the app window; if Tk is unavailable, fall back to a simple prompt
    try:
        if run_gui(): sys.exit(0)
    except Exception:
        pass
    import glob
    HERE_=os.path.dirname(os.path.abspath(__file__))
    src=_find_workbook(HERE_)
    if not src:
        print("\nNo workbook found. Put your A2Z Growth Engine workbook here, then run again.\n")
        input("Press Enter to close..."); sys.exit(1)
    print("Reading:", os.path.basename(src))
    kind=""
    while kind.upper() not in ("LTD","SA","PARTNERSHIP"):
        kind=input("\nWhich proposal?  Type  LTD,  SA  or  PARTNERSHIP  then Enter:  ").strip()
    kind=kind.upper(); out=os.path.join(HERE_, f"{_company_name(src,kind)} - proposal.pdf")
    try:
        d=generate(src, kind, out); print("\nCreated:", os.path.basename(d["pdf"]))
    except Exception as ex:
        import traceback; print("\nCouldn't build the proposal:\n  ", ex); traceback.print_exc()
    input("\nPress Enter to close...")




# ================= UNIVERSAL ENGAGEMENT AGREEMENT (dynamic, legal-form aware) =================
def _eng_para(p, text, col=(39,50,59), w=None, lh=5.0, justify=True):
    w = w or p.CW
    p.f("Nunito","",9.5,col); p.set_x(p.X0)
    p.multi_cell(w,lh,text,align=("J" if justify else "L"),new_x=XPos.LMARGIN,new_y=YPos.NEXT,markdown=True); p.ln(1.8)

def _eng_sublabel(p, t):
    p.f("Nunito","B",7.5,GREEN); p.set_x(p.X0)
    p.cell(0,4," ".join(t.upper()),new_x=XPos.LMARGIN,new_y=YPos.NEXT); p.ln(1.6)

def _eng_ticklist(p, items, lh=4.6):
    for it in items:
        p.set_font("Nunito","",9.3)
        lines = p.multi_cell(p.CW-7,lh,it,dry_run=True,output="LINES",markdown=True)
        h = max(len(lines),1)*lh
        p.need(h+1.5); y0=p.get_y(); p.tick(p.X0, y0+0.6, 2.8)
        p.f("Nunito","",9.3,(39,50,59)); p.set_xy(p.X0+7,y0)
        p.multi_cell(p.CW-7,lh,it,align="L",new_x=XPos.LMARGIN,new_y=YPos.NEXT,markdown=True)
        p.set_y(y0+h+0.9); p.hline(SOFTLINE); p.ln(0.9)

_ENG_BASE = dict(entity="organisation", entity_the="the organisation", officer_word="responsible officer",
                 officer_plural="responsible officers", officers="the responsible officers",
                 authorities="HMRC", has_number=False, number_label="registration number", ch=False)
def _eng_form(lf):
    lf = (lf or "").strip().lower().replace(" ", "_").replace("-", "_")
    A = {
      "sole_trader": dict(entity="business", entity_the="the business", officer_word="proprietor",
          officer_plural="proprietors", officers="you as the business owner", authorities="HMRC",
          has_number=False, number_label=None, ch=False),
      "ltd": dict(entity="company", entity_the="the company", officer_word="director",
          officer_plural="directors", officers="the directors", authorities="Companies House and HMRC",
          has_number=True, number_label="company number", ch=True),
      "cic": dict(entity="community interest company (CIC)", entity_the="the company", officer_word="director",
          officer_plural="directors", officers="the directors", authorities="Companies House, the CIC Regulator and HMRC",
          has_number=True, number_label="company number", ch=True),
      "charitable_company": dict(entity="charitable company", entity_the="the company",
          officer_word="director and charity trustee", officer_plural="directors and charity trustees",
          officers="the directors and charity trustees", authorities="Companies House, the charity regulator and HMRC",
          has_number=True, number_label="company number", ch=True),
      "llp": dict(entity="limited liability partnership (LLP)", entity_the="the LLP", officer_word="member",
          officer_plural="members", officers="the members", authorities="Companies House and HMRC",
          has_number=True, number_label="LLP number", ch=True),
      "partnership": dict(entity="partnership", entity_the="the partnership", officer_word="partner",
          officer_plural="partners", officers="the partners", authorities="HMRC",
          has_number=False, number_label=None, ch=False),
      "charity": dict(entity="charity", entity_the="the charity", officer_word="trustee",
          officer_plural="trustees", officers="the trustees", authorities="the charity regulator and HMRC",
          has_number=True, number_label="charity number", ch=False),
      "scio": dict(entity="SCIO", entity_the="the SCIO", officer_word="charity trustee",
          officer_plural="charity trustees", officers="the charity trustees", authorities="OSCR and HMRC",
          has_number=True, number_label="Scottish charity number", ch=False),
      "trust": dict(entity="trust", entity_the="the trust", officer_word="trustee",
          officer_plural="trustees", officers="the trustees", authorities="HMRC",
          has_number=False, number_label=None, ch=False),
    }
    A["limited"]=A["ltd"]; A["limited_company"]=A["ltd"]; A["company"]=A["ltd"]
    A["community_interest_company"]=A["cic"]; A["partnership_general"]=A["partnership"]
    f=dict(_ENG_BASE); f.update(A.get(lf, {})); return f

def _eng_form_from_kind(kind):
    k=(kind or "").upper()
    return {"LTD":"ltd","CIC":"cic","CHARITY":"charity","PARTNERSHIP":"partnership","SA":"sole_trader"}.get(k,"ltd")

def _eng_people_phrase(eng, form):
    ppl = eng.get("responsible_persons") or []
    names=[str(x.get("name","")).strip() for x in ppl if isinstance(x,dict) and x.get("name")]
    if not names:
        return None, form["officers"]
    if len(names)==1:
        return names[0], f"{form['officer_word']} {names[0]}"
    joined=", ".join(names[:-1])+" and "+names[-1]
    return joined, f"{form['officer_plural']} {joined}"

def _eng_flags(lines):
    L=" | ".join(str(x[0]).lower() for x in (lines or []))
    has=lambda *ks: any(k in L for k in ks)
    return dict(
        accounts=has("annual accounts","accounts & corporation","corporation tax","partnership tax"),
        sa=has("self assessment","self-assessment"),
        vat=has("vat"),
        payroll=has("payroll"),
        pension=has("pension","auto-enrolment","auto enrolment"),
        cis=has("cis"),
        book=has("bookkeeping"),
        cs01=has("confirmation statement","companies house","cs01"),
        mgmt=has("financial health","business performance","strategic advisory","management report","advisory report"),
        software=has("software","quickbooks","xero","dext"),
        address=has("address service","registered office"),
    )

_BRANDS=[("Dext","document capture software"),("QuickBooks","cloud accounting software"),("Quickbooks","cloud accounting software"),("Xero","cloud accounting software"),("Sage","cloud accounting software"),("FreeAgent","cloud accounting software"),("Free Agent","cloud accounting software")]
def _degen(s):
    s=str(s)
    for b,g in _BRANDS: s=s.replace(b,g)
    s=s.replace("Document software (document capture software)","Document capture software")
    s=s.replace("Bookkeeping software","Cloud accounting software")
    if s.startswith("cloud "): s="C"+s[1:]
    return s

def _accounts_label(form):
    ow=form["officer_word"]
    if "proprietor" in ow: return ("Annual accounts & Self Assessment","Accounts and tax return, prepared and filed to deadline")
    if "member" in ow: return ("Annual accounts & Partnership Tax Return","LLP accounts and partnership return, filed to deadline")
    if "partner" in ow: return ("Annual accounts & Partnership Tax Return","Accounts and partnership return, filed to deadline")
    if "trustee" in ow: return ("Annual accounts","Charity accounts, prepared and filed to the regulator")
    return ("Annual accounts & Corporation Tax","Statutory accounts & CT600, filed to deadline")

def _eng_relabel_lines(lines, form):
    lab,det=_accounts_label(form); out=[]
    for l in lines:
        L=str(l[0]); D=str(l[1]) if len(l)>1 else ""; A=l[2] if len(l)>2 else 0; low=L.lower()
        if "annual accounts" in low or "corporation tax" in low:
            L=lab; D=det
        if ("companies house" in low or "confirmation statement" in low) and not form["ch"]: continue
        if ("registered office" in low or "address service" in low) and not form["ch"]: continue
        out.append((_degen(L),_degen(D),A))
    return out

def cover_engagement(p, company, subtitle, date, ref, lead_name="A2Z Accounting Solutions", status="For acceptance"):
    p.add_page(); p.set_auto_page_break(False)
    p.set_fill_color(*NAVY); p.rect(0,0,210,297,style="F")
    p.set_fill_color(*GREEN); p.rect(0,0,210,3.2,style="F")
    logo=os.path.join(tempfile.gettempdir(),"a2z_logo.png")
    if LOGO_B64 and not os.path.exists(logo): open(logo,"wb").write(base64.b64decode(LOGO_B64))
    src=logo if os.path.exists(logo) else os.path.join(HERE,"logo.png")
    p.rrect(24,26,66,21, r=2.5, style="F", fill=WHITE, draw=WHITE)
    try: p.image(src, x=30, y=30.5, h=12)
    except Exception: pass
    p.f("Nunito","",7.5,A9BBD0); p.set_xy(110,33); p.cell(76,5," ".join("PRIVATE & CONFIDENTIAL"),align="R")
    p.set_fill_color(*GREEN); p.rect(24,98,48,1.4,style="F")
    p.f("Cormorant","B",47,WHITE); p.set_xy(23,105)
    p.cell(0,17,"Engagement",new_x=XPos.LEFT,new_y=YPos.NEXT); p.set_x(23); p.cell(0,17,"Agreement")
    p.f("Nunito","",8,A9BBD0); p.set_xy(24,164); p.cell(0,5," ".join("AGREED BETWEEN A2Z AND"),new_x=XPos.LEFT,new_y=YPos.NEXT)
    p.f("Cormorant","B",25,WHITE); p.set_xy(24,170)
    _csz=25
    while _csz>13 and p.get_string_width(company)>160: _csz-=0.5; p.set_font("Cormorant","B",_csz)
    p.cell(0,12,company,new_x=XPos.LEFT,new_y=YPos.NEXT)
    p.f("Nunito","",9.5,(150,170,194)); p.set_x(24); p.cell(0,6,subtitle)
    p.set_draw_color(56,84,114); p.set_line_width(0.3); p.line(24,250,186,250)
    meta=[("DATE",date),("ENGAGEMENT LEAD",lead_name or "A2Z Accounting Solutions"),("REFERENCE",ref),("STATUS",status)]
    x=24
    for lab,val in meta:
        p.f("Nunito","",7,A9BBD0); p.set_xy(x,254); p.cell(40,4," ".join(lab))
        val=str(val); sz=9.5; p.set_font("NunitoSemi","",sz)
        while sz>6.6 and p.get_string_width(val)>37: sz-=0.4; p.set_font("NunitoSemi","",sz)
        p.set_text_color(*WHITE); p.set_xy(x,259); p.cell(40,5,val); x+=40.5
    p.f("Nunito","",7,A9BBD0); p.set_xy(24,268); p.cell(0,4,"Chartered Certified Accountants, regulated by ACCA.")
    p.f("Nunito","",8,(150,170,194)); p.set_xy(24,283); p.cell(0,5,"1st Floor, 499 Union Street, Aberdeen, AB11 6DB")
    p.set_xy(110,283); p.cell(76,5,"01224 042961  \u00b7  info@a2zaccounting.co.uk",align="R")
    p.set_auto_page_break(True, margin=20)

def engagement_accept(p, company, form, signatory, acceptance=None, ref="", version="v1", sec_no=99):
    sig_name=(signatory or {}).get("name") or "A2Z Accounting Solutions Ltd"
    sig_title=(signatory or {}).get("title") or "Director, for and on behalf of the firm"
    _esec(p, sec_no, "Signature & acceptance", "How this agreement is signed", reserve=40)
    _eng_para(p, f"This agreement is signed and accepted **electronically**. Selecting **I agree** on the secure link we send you is a deliberate act of acceptance and creates a binding engagement between {company} and A2Z Accounting Solutions Ltd, in the same way as a handwritten signature. The person accepting confirms they are authorised to accept it on behalf of {form['entity_the']}.")
    _eng_para(p, "The signature block below is completed at the moment of acceptance. We record the name and stated position of the person accepting, the exact date and time, the device network address (IP), and the version and reference of this agreement - your evidence of exactly what was agreed.")
    p.need(54); y0=p.get_y(); gap=8; colw=(p.CW-gap)/2; h=47
    cols=[(p.X0,"ACCEPTED BY THE CLIENT",bool(acceptance)),(p.X0+colw+gap,"FOR A2Z ACCOUNTING SOLUTIONS LTD",True)]
    for cx,ctitle,done in cols:
        p.rrect(cx,y0,colw,h,r=2.2,draw=LINE,style="D")
        p.set_fill_color(*(GREEN if done else (206,214,222))); p.rect(cx,y0,colw,1.8,style="F")
        p.f("Nunito","B",7,GREEN if done else MUTED); p.set_xy(cx+7,y0+5.5); p.cell(0,4," ".join(ctitle))
    def field(cx,ln,label,val,filled):
        yy=y0+12.5+ln*10.5
        p.f("Nunito","",6.6,MUTED); p.set_xy(cx+7,yy); p.cell(0,3.4," ".join(label.upper()))
        p.f("NunitoSemi","",9.3,(39,50,59) if filled else (176,185,194)); p.set_xy(cx+7,yy+3.8)
        p.multi_cell(colw-14,4.4,val,align="L")
        p.set_draw_color(*SOFTLINE); p.set_line_width(0.2); p.line(cx+7,yy+9.6,cx+colw-7,yy+9.6)
    if acceptance:
        field(p.X0,0,"Name",acceptance.get("name") or "-",True)
        field(p.X0,1,"Position",acceptance.get("position") or "-",True)
        field(p.X0,2,"Date and time",acceptance.get("when") or "-",True)
    else:
        for i,l in enumerate(["Name","Position","Date and time"]): field(p.X0,i,l,"Completed on acceptance",False)
    ax=p.X0+colw+gap
    field(ax,0,"Signatory",sig_name,True)
    field(ax,1,"Title",sig_title,True)
    field(ax,2,"Date",(acceptance.get("when") if acceptance else "On acceptance"),bool(acceptance))
    p.set_y(y0+h+4)
    if acceptance:
        ev=[]
        if acceptance.get("ip"): ev.append("IP "+str(acceptance.get("ip")))
        if ref: ev.append("Ref "+str(ref))
        if version: ev.append("Version "+str(version))
        ev.append("accepted electronically, authority to bind confirmed")
        p.f("Nunito","",8,MUTED); p.set_x(p.X0); p.multi_cell(p.CW,4.3,"Signature evidence:  "+"  \u00b7  ".join(ev),align="L")
    else:
        p.f("Nunito","",8,MUTED); p.set_x(p.X0); p.multi_cell(p.CW,4.3,"Not yet accepted - the client completes the left-hand block via the secure link we send, which records the date, time and IP as the electronic signature.",align="L")
    p.ln(1)

class EngagementPDF(PDF):
    def header(self):
        if self.page_no()==1: return
        self.set_fill_color(*GREEN); self.rect(0,0,210,2.0,style="F")
        self.set_xy(self.X0,8.2); self.set_font("Cormorant","B",11); self.set_text_color(*NAVY)
        self.cell(0,5,"A2Z Accounting Solutions")
        self.set_xy(self.X0,9); self.set_font("Nunito","",7); self.set_text_color(*MUTED)
        self.cell(self.CW,5, getattr(self,"doc_title","Engagement Agreement"), align="R")
        self.set_draw_color(*LINE); self.set_line_width(0.2); self.line(self.X0,15.6,self.XR,15.6)
        self.set_xy(self.l_margin, self.t_margin)
    def footer(self):
        if self.page_no()==1: return
        self.set_fill_color(*NAVY); self.rect(0,285,210,12,style="F")
        self.set_fill_color(*GREEN); self.rect(0,285,210,0.8,style="F")
        self.set_y(288.6); self.set_font("Nunito","",7); self.set_text_color(210,221,232)
        self.set_x(self.X0); self.cell(self.CW/2,4,"A2Z Accounting Solutions  \u00b7  Regulated by ACCA")
        self.cell(self.CW/2,4,f"Page {self.page_no()} of {{nb}}",align="R")

def _esec(p, n, klab, title, reserve=24):
    p.ln(3.2); p.need(reserve)
    p.f("Nunito","B",7.5,GREEN); p.set_x(p.X0); p.cell(0,4," ".join(klab.upper()),new_x=XPos.LMARGIN,new_y=YPos.NEXT); p.ln(1.4)
    p.f("Cormorant","B",15.5,GREEN); p.set_x(p.X0)
    p.cell(8.5,7,f"{n}.",new_x=XPos.RIGHT,new_y=YPos.TOP)
    p.f("Cormorant","B",15.5,NAVY)
    p.multi_cell(p.CW-8.5,7,title,new_x=XPos.LMARGIN,new_y=YPos.NEXT); p.ln(2.3)

def _eng_note(p, text):
    p.ln(1); p.need(14); y=p.get_y()
    p.set_font("Nunito","",9); nl=p.multi_cell(p.CW-12,4.4,text,dry_run=True,output="LINES",markdown=True)
    h=5+max(len(nl),1)*4.4
    p.set_fill_color(*PALE); p.rect(p.X0,y,p.CW,h,style="F"); p.set_fill_color(*NAVY); p.rect(p.X0,y,1.4,h,style="F")
    p.f("Nunito","",9,(39,50,59)); p.set_xy(p.X0+6,y+2.4); p.multi_cell(p.CW-12,4.4,text,align="L",markdown=True)
    p.set_y(y+h+1)

def build_engagement(wb, out, ref=None, acceptance=None, eng=None):
    d=read_ltd(wb); ref=ref or ref_for(d["company"]); d["ref"]=ref
    eng=eng or {}
    version=str(eng.get("version") or "v1.0")
    form=_eng_form(eng.get("legal_form") or _eng_form_from_kind(eng.get("kind")))
    d["lines"]=_eng_relabel_lines(d["lines"], form)
    fl=_eng_flags(d["lines"])
    company=eng.get("client_legal_name") or d["company"]
    lead=eng.get("a2z_lead") or {}; lead_name=lead.get("name") or ""
    signatory=eng.get("a2z_signatory") or lead or {}
    _pn, people_are=_eng_people_phrase(eng, form)
    addr=eng.get("address") or ""; number=eng.get("entity_number") or ""
    numlab=eng.get("entity_number_label") or form["number_label"]
    sw=[]
    for l in d["lines"]:
        low=str(l[0]).lower()
        if any(k in low for k in ("software","quickbooks","xero","dext")): sw.append(str(l[0]))
    sw=list(dict.fromkeys(sw))

    p=EngagementPDF(); p.alias_nb_pages(); p.right_header="ENGAGEMENT AGREEMENT"; p.doc_title="Engagement Agreement"
    cover_engagement(p, company, "Accounting & tax services - terms of engagement", d["date"], ref,
                     lead_name=lead_name or "A2Z Accounting Solutions", status=("Accepted" if acceptance else "For acceptance"))
    p.add_page()

    _esec(p,1,"Welcome to A2Z","Welcome, and thank you")
    idnum=f", {numlab} {number}" if (form["has_number"] and number and numlab) else ""
    idaddr=f", of {addr}" if addr else ""
    _eng_para(p, f"Thank you for choosing A2Z Accounting Solutions to look after {company}. This agreement sets out what we will do for you, how often, and what it costs, in plain English. You do not need to know the accounting or tax answer - that is exactly what we are here for. Your part is simple: keep good records, send us what we ask for when we ask for it, tell us anything relevant, and flag anything you are unsure about.")
    _eng_para(p, f"This agreement is made between **A2Z Accounting Solutions Ltd** (company number SC618668, 1st Floor, 499 Union Street, Aberdeen, AB11 6DB) - \u201cwe\u201d, \u201cus\u201d, \u201cthe firm\u201d - and **{company}**{idnum}{idaddr} - \u201cyou\u201d, \u201cthe client\u201d. We are Chartered Certified Accountants regulated by the ACCA and follow its Code of Ethics and Conduct. It takes effect on the date you accept it and continues until ended by either of us in writing.")

    _esec(p,2,"Your service package","The services we will provide")
    _eng_para(p, "These are the services covered by your fee, with the frequency of each. This schedule is the definitive list of what is included; anything outside it is agreed and quoted separately.")
    fee_table(p, d["lines"], prices=False)

    if sw:
        _esec(p,3,"Included software","The tools you will use")
        _eng_para(p, "Your fee includes the following software, set up and supported by us. Licences are for use while you are engaged with us, under the providers' own terms, and end with the engagement unless you arrange to take them over.")
        _eng_ticklist(p, sw)
        _n=4
    else:
        _n=3

    _esec(p,_n,"The fees","Your fees", reserve=58); _fee_n=_n
    p.need(31); y=p.get_y(); bh=25; p.rrect(p.X0,y,p.CW,bh,fill=NAVY,style="F")
    p.f("Nunito","B",7.5,A9BBD0); p.set_xy(p.X0+8,y+5.5); p.cell(0,4," ".join("YOUR PROFESSIONAL FEE"))
    p.f("Cormorant","B",16,WHITE); p.set_xy(p.X0+8,y+10.5); p.cell(0,8,"Your finance function")
    p.f("Cormorant","B",27,WHITE); p.set_xy(p.X0+86,y+4), p.cell(p.CW-94,11,gbp(d["sub"]),align="R")
    p.f("Nunito","",8,A9BBD0); p.set_xy(p.X0+86,y+17.4); p.cell(p.CW-94,4,f"+ VAT per month  \u00b7  {gbp(d['gross'])} gross",align="R")
    p.set_y(y+bh+2.5)
    total_row(p,"Monthly instalment (subtotal)",d["sub"]); total_row(p,"VAT @ 20%",d["vat"]); total_row(p,"Gross monthly instalment",d["gross"],grand=True)
    p.ln(0.5); _eng_para(p,"Your fee is an **annual professional fee**, spread for your convenience into equal monthly instalments collected by Direct Debit in advance. It is one fixed fee for the year's service, not twelve separate monthly purchases. It changes only if the scope of work changes materially and is agreed with you in advance, and it is not billed by the hour.")
    def _rows_from(seq, default_label="Item"):
        out=[]
        for x in (seq or []):
            if isinstance(x,dict): out.append((_degen(x.get("label") or default_label), _degen(x.get("detail") or ""), num(x.get("amount",0)) or 0))
            elif isinstance(x,(list,tuple)): out.append((_degen(x[0]), _degen(x[1]) if len(x)>1 else "", num(x[2] if len(x)>2 else 0) or 0))
        return out
    if (eng.get("oneoffs") is not None) or (eng.get("adhocs") is not None) or (eng.get("catchup") is not None):
        one_rows=_rows_from(eng.get("oneoffs")); catch=num(eng.get("catchup",0)) or 0; adhoc_rows=_rows_from(eng.get("adhocs"),"Ad-hoc work")
    else:
        one_rows=[(_degen(a),_degen(b),c) for a,b,c in (d.get("oneoffs") or [])]; catch=0; adhoc_rows=[]
    def _minihead(t):
        p.ln(1.4); p.need(24); p.f("Cormorant","B",12.5,NAVY); p.set_x(p.X0); p.cell(0,6,t,new_x=XPos.LMARGIN,new_y=YPos.NEXT); p.ln(0.8)
    if one_rows:
        _minihead("Registration & one-off fees, on starting")
        fee_table(p,one_rows,unit="\u00a3")
        _os=sum(r[2] for r in one_rows); total_row(p,"One-off subtotal",_os); total_row(p,"VAT @ 20%",_os*0.2); total_row(p,"One-off inc VAT",_os*1.2,grand=True)
    if catch>0:
        _minihead("Catch-up / backdated work")
        fee_table(p,[("Catch-up / backdated work","Bringing prior periods up to date; billed on starting",catch)],unit="\u00a3")
        total_row(p,"Catch-up inc VAT",catch*1.2,grand=True)
    if adhoc_rows:
        _minihead("Ad-hoc & additional work")
        _eng_para(p,"The items below sit outside your fixed fee and are quoted and billed as the work arises. Any other ad-hoc or specialist work is agreed with you before we begin.")
        fee_table(p,[(a,(b or "Quoted and billed as the work arises"),c) for a,b,c in adhoc_rows],prices=False)
    _dep=0
    try: _dep=float(eng.get("deposit") or 0)
    except Exception: _dep=0
    if _dep>0: _eng_note(p, f"A deposit / advance of {gbp(_dep)} is payable on acceptance to reserve capacity and begin your onboarding.")
    if d["directors"]>0 and form["ch"]:
        _eng_note(p, f"Personal tax returns (Self Assessment) for {form['officer_plural']} and shareholders are charged separately at \u00a3120 + VAT each per year.")

    _esec(p,_fee_n+1,"Working together","How we will work together")
    _eng_para(p, "The way this works is simple. You run your business; we look after the numbers and the filings behind it. We tell you what we need and by when, ask questions where something is not clear, work to have your filings ready ahead of their deadlines, and raise anything worth acting on while there is still time to act.")
    _eng_sublabel(p,"What you can expect from us")
    _eng_ticklist(p,[
        "Your agreed work planned and prepared to be ready ahead of its statutory deadlines",
        "Your work reviewed in stages and signed off by a chartered certified (FCCA) director",
        "Acting as your authorised agent with "+form["authorities"],
        "Our aim to reply the same working day when you contact us before 4pm",
        "A fixed fee agreed with you in advance"])

    _esec(p,_fee_n+2,"Your responsibilities","What stays with you")
    _eng_para(p, f"Engaging us helps you meet your obligations - it does not transfer them. Responsibility for the underlying affairs of {form['entity_the']} always remains with {people_are}. We prepare and file from what you give us; the business and its records remain yours.")
    _eng_ticklist(p,[
        "**Keep proper records** and provide complete, accurate information and documents.",
        "**Meet the dates we set.** Provide records and information by the deadline **we communicate to you** - set to allow proper preparation and review - not merely before the statutory filing deadline.",
        "**Answer our questions** fully and promptly, and tell us about anything unusual, missing or uncertain.",
        "**Flag anything you are unsure about** - in particular anything you think may not comply with tax, accounting, VAT, payroll, CIS, company, employment or other requirements. If in doubt, tell us.",
        "**Do not assume we can see what has not been disclosed.** We work from the records and information you provide; we cannot identify facts we have not been told.",
        "**Silence from us is not verification.** The absence of a query does not mean a transaction has been independently checked, verified or approved.",
        "**Review and approve** your accounts, returns, payroll and submissions where approval is required, before we file them.",
        "**Paying tax and other liabilities** on time remains your responsibility; we will tell you what is due and when."])

    st=[]
    if fl["accounts"]: st.append("**Accounts & tax** - prepared from your records and filed once approved. This is not an audit and does not verify the completeness of the underlying records.")
    if fl["sa"]: st.append("**Self Assessment** - prepared and filed from the information you provide by the dates we set; the accuracy of what you disclose remains yours.")
    if fl["vat"]: st.append("**VAT** - prepared, checked and filed under Making Tax Digital from your records; the validity of transactions and paying VAT due remain yours.")
    if fl["payroll"]: st.append("**Payroll** - payslips issued and RTI submissions made from the pay data you approve each period, by the cut-off we agree.")
    if fl["pension"]: st.append("**Pensions / auto-enrolment** - administered alongside payroll; the statutory employer duties remain yours.")
    if fl["cis"]: st.append("**CIS** - returns prepared and filed from the details you provide; verifying subcontractors and deductions remains yours.")
    if fl["cs01"] and form["ch"]: st.append("**Confirmation statement** - filed as your agent from the information you confirm to us.")
    if fl["mgmt"]: st.append("**Management & advisory reports** - prepared for internal decision-making from the data available; not audited, and figures may be provisional.")
    if fl["software"]: st.append("**Software** - provided for use during the engagement under the providers' own terms; we are not responsible for third-party outages or provider changes.")
    if st:
        _esec(p,_fee_n+3,"Your services in detail","How each service works")
        _eng_ticklist(p, st); _svc_used=1
    else: _svc_used=0
    _k=_fee_n+3+_svc_used

    _esec(p,_k,"Fees, cancellation & deposits","If things change")
    _eng_ticklist(p,[
        "**An annual fee, paid monthly.** Your recurring fee is an annual professional fee paid in monthly Direct Debit instalments. Work is not performed evenly across the year - much is weighted to your year end and filing periods.",
        "**Instalments already collected are not refunded** simply because you leave part-way through the annual cycle, and there is no automatic pro-rata refund based on months elapsed or work done in a given month.",
        "**Future instalments stop** from the effective date your engagement ends, subject to any payment already in the banking or Direct Debit collection process.",
        "**Work beyond fees paid.** If work already undertaken, started or committed for the year exceeds the instalments collected, we may charge a reasonable final amount for it.",
        "**Outstanding invoices remain payable**, and any licences or filings we have paid for on your behalf remain chargeable.",
        "**Deposits and advances.** Once you accept this engagement and we begin onboarding, setup, capacity allocation, preliminary work or incur costs in reliance on your acceptance, any deposit or advance is non-refundable if you later change your mind or cancel. Nothing here removes rights that cannot lawfully be excluded."])

    _esec(p,_k+1,"Anti-money-laundering","Identity checks and our legal duties")
    _eng_para(p, "As a firm supervised by the ACCA for anti-money-laundering, we are required by the Money Laundering Regulations and the Proceeds of Crime Act 2002 to carry out identity and background checks before and during our engagement. This protects you as well as us.")
    _eng_ticklist(p,[
        "We verify the identity of "+form["entity_the"]+" and of "+form["officer_plural"]+" and any beneficial owners, may use electronic verification, and may ask for documents. We keep these records for at least five years after the engagement ends.",
        "We may be unable to start or to continue acting, and may have to suspend work, until we have completed checks we are satisfied with.",
        "We are required to report any knowledge or suspicion of money laundering to the National Crime Agency, and the law may prohibit us from telling you that a report has been made or the reason for it.",
        "We will not be liable for any loss arising from any action we take, or any work we are unable to do, in order to meet these legal obligations."])

    _esec(p,_k+2,"The detail","Our terms of engagement")
    _eng_ticklist(p,[
        "**Scope & out-of-scope work** - we provide the services listed in your schedule. Anything outside it (enquiries, one-off projects, or new services) is agreed and quoted separately before we start.",
        "**Late or incomplete information** - if records arrive late, incomplete or inaccurate, we cannot guarantee a deadline, and additional work to correct or reconstruct records may be charged.",
        "**Advice reflects the facts and law at the time** - based on the information given and the law and HMRC practice then in force. We are not obliged to revisit past advice if the law or your circumstances later change, unless you engage us to do so.",
        "**Estimates and forecasts** are based on assumptions and are not guarantees of outcome.",
        "**Third-party and system delays** - we are not responsible for delays or errors caused by HMRC, Companies House, other authorities, banks or software providers.",
        "**Confidentiality & working papers** - we keep your information confidential except where the law or our regulator requires disclosure. Our working papers remain our property; your records remain yours.",
        "**Reliance** - our work is for you and may not be relied on by any third party unless we agree in writing.",
        "**Liability** - limited to the extent permitted by law and as set out in our Terms of Business, which form part of this engagement; we do not exclude liability that cannot lawfully be excluded.",
        "**Unpaid fees & suspension** - we may suspend work or withhold documents where fees are overdue, having given you notice.",
        "**Termination** - either of us may end this engagement in writing; we will act professionally on handover to a new adviser once our fees are settled.",
        "**Professional rules, complaints & governing law** - we are regulated by ACCA and follow its Code. Any complaint may be raised with the engagement director and, if unresolved, with ACCA. This engagement is governed by the law of Scotland."])

    _esec(p,_k+3,"Data protection","How we handle your data (UK GDPR)")
    _eng_para(p,"A2Z Accounting Solutions Ltd is the **data controller** for the personal data we hold to run our engagement with you. Where we process the personal data of your employees, subcontractors or others on your behalf - for example when we run your payroll - **you are the controller and we act as your data processor**, following your instructions and this agreement.")
    _eng_sublabel(p,"What we process, and our lawful bases")
    _eng_ticklist(p,[
        "Identification and contact details, financial and accounting records, tax information, and where relevant payroll, employee and officer details",
        "**Performance of this contract**, **legal obligation** (AML, tax, companies and charity law and our regulator's rules), and **legitimate interests** in running and securing our practice"])
    _eng_sublabel(p,"Sharing, international processing, retention and your rights")
    _eng_para(p,"We share data only as needed to deliver your services: "+form["authorities"]+"; pension, payroll and software providers where relevant; and our own regulated delivery team, **including our offshore processing team, who work strictly under our instruction and a written data-processing agreement** with appropriate safeguards for any processing outside the UK. We never sell your data. We keep records for as long as we act for you and at least six years afterwards, then delete them securely. If a personal-data breach affects you we will act promptly and notify you and the ICO where the law requires. You may access, correct, erase, restrict, port or object to the processing of your data, and complain to the Information Commissioner's Office (ico.org.uk). To exercise any of these, or for our full privacy notice, contact us at info@a2zaccounting.co.uk.")
    _eng_para(p,"You confirm you have the right to share with us any personal data you provide about other people, such as your employees or officers, and that you will help us respond to any request they make about data held for your engagement.")
    _ct="**Your consent.** By accepting this agreement you confirm you have read this section and consent to us, and our processors, handling your personal data - and, where you provide it, the personal data of your officers and employees - as described, and to us acting as your authorised agent with "+form["authorities"]+"."
    _eng_note(p, _ct)

    engagement_accept(p, company, form, signatory, acceptance=acceptance, ref=ref, version=version, sec_no=_k+4)

    _esec(p,_k+5,"Finally","We are here to help")
    _eng_para(p, "That is the detail done. What sits behind it matters more: our aim is to keep you clear on where you stand, to agree our fees with you up front rather than spring them on you, and to have your work handled by a named team and reviewed by a chartered certified (FCCA) director.")
    _eng_para(p, "You do not need to know the answer - just keep good records, send us what we ask for when we ask for it, tell us anything relevant, and flag anything you are unsure about. The rest is ours to look after, and we are really looking forward to working with you.")
    p.ln(1.5); p.f("Cormorant","B",15,NAVY); p.set_x(p.X0); p.cell(0,8,"Welcome to A2Z.",new_x=XPos.LMARGIN,new_y=YPos.NEXT)

    p.output(out); return d
