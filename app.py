"""
A2Z Proposal PDF service
-------------------------
A thin web wrapper around Shabbir's own a2z_proposals_fpdf.py generator.
The platform POSTs proposal data as JSON; this builds the in-memory workbook
his read_ltd / read_sa / read_partnership expect, then calls his UNCHANGED
build_ltd / build_sa / build_partnership and returns the byte-identical PDF.

His generator code is never modified - it is imported and called as-is.
"""
import io, os, tempfile, datetime, traceback, shutil, glob
from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import openpyxl

# --- Font self-heal ------------------------------------------------------
# His generator loads fonts from a "fonts/" subfolder next to it. When the .ttf
# files get uploaded loose at the repo root (GitHub drag-drop drops folders),
# they end up beside this file instead of inside fonts/. Make it robust: ensure
# a fonts/ folder exists and holds the 5 ttf, copying any that sit at the top
# level. Runs before importing his module so the path is ready. His file is
# NOT modified.
_HERE = os.path.dirname(os.path.abspath(__file__))
_FONTS_DIR = os.path.join(_HERE, "fonts")
_NEEDED = ["Cormorant-Bold.ttf", "Cormorant-SemiBold.ttf",
           "Nunito-Regular.ttf", "Nunito-SemiBold.ttf", "Nunito-Bold.ttf"]
try:
    os.makedirs(_FONTS_DIR, exist_ok=True)
    # index every .ttf anywhere under the project root, by filename
    found = {}
    for path in glob.glob(os.path.join(_HERE, "**", "*.ttf"), recursive=True):
        found.setdefault(os.path.basename(path), path)
    for fn in _NEEDED:
        dest = os.path.join(_FONTS_DIR, fn)
        if not os.path.exists(dest) and fn in found and os.path.abspath(found[fn]) != os.path.abspath(dest):
            shutil.copyfile(found[fn], dest)
    missing = [fn for fn in _NEEDED if not os.path.exists(os.path.join(_FONTS_DIR, fn))]
    if missing:
        print("FONT SELF-HEAL: still missing %s (looked under %s)" % (missing, _HERE), flush=True)
    else:
        print("FONT SELF-HEAL: all 5 fonts present in %s" % _FONTS_DIR, flush=True)
except Exception as _e:
    print("FONT SELF-HEAL failed: %s" % _e, flush=True)
# ------------------------------------------------------------------------

import a2z_proposals_fpdf as GEN  # his file, untouched

app = Flask(__name__)
# Allow the platform (and local testing) to call this endpoint from the browser.
CORS(app, resources={r"/*": {"origins": "*"}})


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _set(ws, cell, val):
    ws[cell] = val


# ---- Build the 'LTD Proposal' sheet his read_ltd() reads ----
def _fill_ltd_sheet(ws, d):
    _set(ws, "C10", d.get("company") or "Your Company Ltd")
    _set(ws, "C11", d.get("contact") or "Director")
    _set(ws, "C12", d.get("email") or "")
    _set(ws, "C13", d.get("phone") or "")
    _set(ws, "C14", d.get("date") or datetime.date.today().strftime("%d %B %Y"))
    _set(ws, "C15", d.get("prepared_by") or "Shabbir Rahman FCCA")
    _set(ws, "C18", d.get("band") or "")
    _set(ws, "C25", int(_num(d.get("directors"))))
    _set(ws, "C35", int(_num(d.get("people"))))
    _set(ws, "K4", _num(d.get("discount")))
    _set(ws, "B72", d.get("client_notes") or "")
    _set(ws, "I10", d.get("internal_notes") or d.get("notes") or "")
    _set(ws, "J25", d.get("source") or "")
    _set(ws, "J26", d.get("referrer") or "")

    svc = d.get("services", {})
    # accounts+CT (his read_ltd sums G22+G28)
    _set(ws, "G22", _num(svc.get("accounts")))
    _set(ws, "G28", _num(svc.get("ct")))
    _set(ws, "G23", _num(svc.get("cs01")))
    _set(ws, "G24", _num(svc.get("address")))
    _set(ws, "C27", svc.get("book_freq") or "")
    _set(ws, "G27", _num(svc.get("book")))
    _set(ws, "C29", svc.get("vat_freq") or "")
    _set(ws, "G29", _num(svc.get("vat")))
    _set(ws, "C30", svc.get("software_name") or "")
    _set(ws, "G30", _num(svc.get("software")))
    _set(ws, "G31", _num(svc.get("dext")))
    _set(ws, "G32", _num(svc.get("contractor")))
    _set(ws, "C34", svc.get("payroll_freq") or "")
    _set(ws, "G34", _num(svc.get("payroll")))
    _set(ws, "G35", _num(svc.get("payroll_extra")))
    _set(ws, "G36", _num(svc.get("cis")))
    # management report tier: one of C39/C40/C41 = Quarterly/Monthly, fee in G39/40/41
    mt = (d.get("mgmt_tier") or "").upper()
    mf = _num(d.get("mgmt_freq_fee"))
    freq = d.get("mgmt_freq") or "Quarterly"
    if mt == "T1":
        _set(ws, "C39", freq); _set(ws, "G39", mf)
    elif mt == "T2":
        _set(ws, "C40", freq); _set(ws, "G40", mf)
    elif mt == "T3":
        _set(ws, "C41", freq); _set(ws, "G41", mf)

    # totals
    _set(ws, "G44", _num(d.get("sub")))
    _set(ws, "G45", _num(d.get("vat")))
    _set(ws, "G46", _num(d.get("gross")))

    # one-offs rows 50-57 - fold in catch-up and ad-hoc so they print
    _oo = list(d.get("oneoffs", []))
    _cu = _num(d.get("catchup"))
    if _cu > 0:
        _oo.append({"label": "Catch-up / backdated work", "detail": "", "amount": _cu})
    _ad = 0.0
    for _a in (d.get("adhocs") or []):
        if isinstance(_a, dict):
            _amt = _num(_a.get("amount"))
            if _amt or _a.get("label"):
                _oo.append({"label": _a.get("label") or "Ad-hoc", "detail": _a.get("detail") or "", "amount": _amt})
                _ad += _amt
    for i, o in enumerate(_oo[:8]):
        r = 50 + i
        _set(ws, "B%d" % r, o.get("label") or "")
        _set(ws, "D%d" % r, o.get("detail") or "")
        _set(ws, "F%d" % r, _num(o.get("amount")))
    _osub = _num(d.get("osub")) + _cu + _ad
    _set(ws, "F67", _osub)
    _set(ws, "F68", _osub * 0.2)
    _set(ws, "F69", _osub * 1.2)

    # setup/registration rows 60-66 (C=Required flag, F=fee)
    regs = d.get("registrations", {})
    rowmap = {"company": 60, "paye": 61, "vat": 62, "cis_sub": 63,
              "cis_con": 64, "sa": 65, "other": 66}
    for key, row in rowmap.items():
        if regs.get(key):
            _set(ws, "C%d" % row, "Required")
            _set(ws, "F%d" % row, _num((regs.get(key) or {}).get("fee")))


def _fill_rate_card(wb, ratecard):
    """Optional: platform can pass its rate card so page 8 renders. Skipped if absent."""
    if not ratecard:
        return
    ws = wb.create_sheet("Rate Card")
    r = 7
    for item in ratecard[:33]:
        ws.cell(r, 2, item.get("service") or "")
        ws.cell(r, 3, item.get("price") or "")
        ws.cell(r, 4, item.get("note") or "")
        r += 1


def _build_workbook_ltd(d):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LTD Proposal"
    _fill_ltd_sheet(ws, d)
    _fill_rate_card(wb, d.get("ratecard"))
    return wb


# ---- Build the 'SA Proposal' sheet his read_sa() reads (sole trader / self-assessment) ----
def _fill_sa_sheet(ws, d):
    _set(ws, "C5", d.get("prepared_by") or "Shabbir Rahman FCCA")
    _set(ws, "C6", d.get("company") or d.get("contact") or "Client")
    _set(ws, "C7", d.get("contact") or "there")
    _set(ws, "C8", d.get("email") or "")
    _set(ws, "C9", d.get("phone") or "")
    _set(ws, "C10", d.get("date") or datetime.date.today().strftime("%d %B %Y"))
    _set(ws, "C13", d.get("ctype") or "")
    _set(ws, "C15", d.get("freq") or "Annually")
    _set(ws, "J30", d.get("source") or "")
    _set(ws, "J31", d.get("referrer") or "")
    _set(ws, "B60", d.get("client_notes") or "")

    # service lines - his read_sa reads 7 fixed rows (26-32); the platform's SA proposal is
    # a package + optional extras, so map by service name into the right fixed row.
    sa = d.get("services", {})
    monthly = (d.get("freq") or "Annually").lower() == "monthly"
    col = "E" if monthly else "D"   # E = monthly, D = annual
    rowmap = {"sa_return": 26, "property": 27, "mtd": 28,
              "book": 29, "vat": 30, "payroll": 31, "software": 32}
    for key, row in rowmap.items():
        v = _num(sa.get(key))
        if v > 0:
            _set(ws, "%s%d" % (col, row), v)

    # totals: his read_sa reads annual D34/D36, monthly E34/E36
    fee = _num(d.get("sub"))
    if monthly:
        _set(ws, "E34", fee); _set(ws, "E36", fee * 1.2)
    else:
        _set(ws, "D34", fee); _set(ws, "D36", fee * 1.2)
    _set(ws, "H34", _num(d.get("discount_annual")))
    _set(ws, "H35", _num(d.get("discount_monthly")))

    # one-offs rows 40-47 - fold in catch-up and ad-hoc so they print
    _oo = list(d.get("oneoffs", []))
    _cu = _num(d.get("catchup"))
    if _cu > 0:
        _oo.append({"label": "Catch-up / backdated work", "detail": "", "amount": _cu})
    _ad = 0.0
    for _a in (d.get("adhocs") or []):
        if isinstance(_a, dict):
            _amt = _num(_a.get("amount"))
            if _amt or _a.get("label"):
                _oo.append({"label": _a.get("label") or "Ad-hoc", "detail": _a.get("detail") or "", "amount": _amt})
                _ad += _amt
    for i, o in enumerate(_oo[:8]):
        r = 40 + i
        _set(ws, "B%d" % r, o.get("label") or "")
        _set(ws, "E%d" % r, _num(o.get("amount")))
    _osub = _num(d.get("osub")) + _cu + _ad
    _set(ws, "E55", _osub)
    _set(ws, "E56", _osub * 0.2)
    _set(ws, "E57", _osub * 1.2)

    # registration block rows 49-54 (C=Required, E=fee): sa/paye/vat/cis_sub/cis_con/other
    regs = d.get("registrations", {})
    sarow = {"sa": 49, "paye": 50, "vat": 51, "cis_sub": 52, "cis_con": 53, "other": 54}
    for key, row in sarow.items():
        if regs.get(key):
            _set(ws, "C%d" % row, "Required")
            _set(ws, "E%d" % row, _num((regs.get(key) or {}).get("fee")))


def _build_workbook_sa(d):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SA Proposal"
    _fill_sa_sheet(ws, d)
    _fill_rate_card(wb, d.get("ratecard"))
    return wb


@app.route("/health")
def health():
    return jsonify(ok=True, service="a2z-proposal-pdf")


@app.route("/proposal", methods=["POST", "OPTIONS"])
def proposal():
    if request.method == "OPTIONS":
        return ("", 204)
    d = request.get_json(force=True, silent=True) or {}
    kind = (d.get("kind") or "LTD").upper()

    tmpdir = tempfile.mkdtemp()
    wb_path = os.path.join(tmpdir, "wb.xlsx")
    out_path = os.path.join(tmpdir, "proposal.pdf")
    try:
        if kind in ("LTD", "CIC", "CHARITY"):
            wb = _build_workbook_ltd(d)
            wb.save(wb_path)
            wb2 = GEN.safe_load_workbook(wb_path)
            GEN.build_ltd(wb2, out_path, ref=d.get("ref"))
        elif kind == "PARTNERSHIP":
            # partnership sheet uses the same layout under a different name
            wb = _build_workbook_ltd(d)
            wb.active.title = "Partnership Proposal"
            wb.save(wb_path)
            wb2 = GEN.safe_load_workbook(wb_path)
            GEN.build_partnership(wb2, out_path, ref=d.get("ref"))
        elif kind == "SA":
            wb = _build_workbook_sa(d)
            wb.save(wb_path)
            wb2 = GEN.safe_load_workbook(wb_path)
            GEN.build_sa(wb2, out_path, ref=d.get("ref"))
        else:
            return jsonify(error="Unknown kind: %s" % kind), 400
    except Exception as e:
        print("=== PROPOSAL PDF BUILD ERROR ===", flush=True)
        traceback.print_exc()
        print("=== END ERROR (kind=%s) ===" % (d.get("kind")), flush=True)
        return jsonify(error="PDF build failed", detail=str(e)), 500

    fname = (d.get("company") or "Proposal").replace("/", " ").replace("\\", " ")
    fname = " ".join(fname.split()) + " - proposal.pdf"
    return send_file(out_path, mimetype="application/pdf",
                     as_attachment=True, download_name=fname)


# ------------------------------------------------------------------
#  Direct email send via Microsoft 365 Graph (reuses the firm's
#  existing Entra app - Mail.Send application permission, admin
#  consent already granted for the Academy). Set these env vars on
#  the Render service (copy the values from the Academy setup):
#     MS_TENANT_ID, MS_CLIENT_ID, MS_CLIENT_SECRET
#     MAIL_FROM        (or INVITE_FROM) - the sending mailbox, e.g. info@a2zaccounting.co.uk
#     SEND_TOKEN       (optional but recommended) - a shared secret the
#                      platform must send in the X-A2Z-Token header
# ------------------------------------------------------------------
import base64 as _b64
import requests as _rq


def _mail_from():
    return os.environ.get("MAIL_FROM") or os.environ.get("INVITE_FROM") or ""


import re as _re


def _text_to_html(text):
    """Turn his plain-text email into simple, email-safe HTML:
    URLs become clickable links, bullet lines render as bullets, blank lines
    become paragraph breaks. Keeps his wording exactly."""
    def esc(s):
        return (s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))

    url_re = _re.compile(r"(https?://[^\s]+)")
    out = []
    out.append('<div style="font-family:Arial,Helvetica,sans-serif;font-size:14px;'
               'color:#1b2a38;line-height:1.6">')
    for raw in text.split("\n"):
        line = raw.rstrip()
        if line == "":
            out.append("<br>")
            continue
        safe = esc(line)
        safe = url_re.sub(lambda m: '<a href="%s" style="color:#1e6b47">%s</a>'
                          % (m.group(1), m.group(1)), safe)
        if line.startswith("\u2022 "):
            out.append('<div style="margin:2px 0 2px 8px">&bull; %s</div>' % safe[2:])
        else:
            out.append('<div>%s</div>' % safe)
    out.append("</div>")
    return "".join(out)


def _graph_token():
    tid = os.environ.get("MS_TENANT_ID", "")
    cid = os.environ.get("MS_CLIENT_ID", "")
    sec = os.environ.get("MS_CLIENT_SECRET", "")
    if not (tid and cid and sec):
        raise RuntimeError("Microsoft 365 not configured (MS_TENANT_ID / MS_CLIENT_ID / MS_CLIENT_SECRET missing).")
    r = _rq.post(
        "https://login.microsoftonline.com/%s/oauth2/v2.0/token" % tid,
        data={
            "client_id": cid,
            "client_secret": sec,
            "scope": "https://graph.microsoft.com/.default",
            "grant_type": "client_credentials",
        }, timeout=30)
    r.raise_for_status()
    return r.json()["access_token"]


def _pdf_bytes_for(d):
    """Build the proposal PDF (same path as /proposal) and return raw bytes."""
    kind = (d.get("kind") or "LTD").upper()
    tmpdir = tempfile.mkdtemp()
    wb_path = os.path.join(tmpdir, "wb.xlsx")
    out_path = os.path.join(tmpdir, "proposal.pdf")
    if kind == "SA":
        wb = _build_workbook_sa(d)
    else:
        wb = _build_workbook_ltd(d)
        if kind == "PARTNERSHIP":
            wb.active.title = "Partnership Proposal"
    wb.save(wb_path)
    wb2 = GEN.safe_load_workbook(wb_path)
    if kind == "SA":
        GEN.build_sa(wb2, out_path, ref=d.get("ref"))
    elif kind == "PARTNERSHIP":
        GEN.build_partnership(wb2, out_path, ref=d.get("ref"))
    else:
        GEN.build_ltd(wb2, out_path, ref=d.get("ref"))
    with open(out_path, "rb") as f:
        return f.read()


def _enhanced_email(d, kind, scenario):
    """Returns (subject, text_body, html_body). Text is his exact wording; html is a
    branded design (navy + GREEN, matching the proposal PDF, so it is clearly distinct
    from the welcome email's gold). His a2z_proposals_fpdf.py is NOT modified."""
    K = kind.upper()
    gbp = GEN.gbp
    num = GEN.num
    FORMS = GEN.FORMS

    def esc(s):
        return (str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                .replace("\u00a3", "&pound;").replace(" - ", " - ")
                .replace("-", "-").replace("\u2019", "&rsquo;")
                .replace("\u2018", "&lsquo;").replace("\u2026", "&hellip;"))

    name = (str(d.get("contact") or "there").strip()) or "there"
    company = (str(d.get("company") or "your business").strip()) or "your business"
    subject = "Your A2Z proposal for %s" % company
    svc = [str(r[0]).strip() for r in d.get("lines", []) if str(r[0]).strip()]
    regs = d.get("regs") or []
    keys = {r.get("key") for r in regs}
    intro = ("Thank you for the opportunity to look after %s - it's a pleasure to put this "
             "proposal together for you. I've attached the full proposal, and here's a quick summary." % company)

    L = ["Hi %s," % name, "", intro, "", "What we'd take care of for you:"]
    L += ["\u2022 %s" % x for x in (svc or ["the services we discussed"])]
    L += [""]

    # ---- fee ----
    fee_lines = []
    fee_display = ""
    fee_sub = "a single fixed amount, with no surprise bills along the way"
    if K == "LTD":
        fee_display = "%s + VAT" % gbp(d.get("sub", 0)); fee_period = "per month"
        fee_lines.append("Your fee would be %s + VAT a month - a single fixed amount, with no surprise bills along the way." % gbp(d.get("sub", 0)))
        if (num(d.get("discount", 0)) or 0) > 0:
            fee_lines.append("As an exceptional act of discretion, a goodwill discount of %s + VAT per month has been applied. Your fee above already reflects this." % gbp(d["discount"]))
        nd = int(num(d.get("directors", 0)) or 0)
        if nd == 1:
            fee_lines.append("Each director's personal tax return (self-assessment) is \u00a3120 + VAT a year, billed separately.")
        elif nd > 1:
            fee_lines.append("Each director's personal tax return (self-assessment) is \u00a3120 + VAT a year, billed separately - for %d directors that comes to %s + VAT a year." % (nd, gbp(nd * 120)))
    else:
        m = num(d.get("monthly", 0)) or 0
        a = num(d.get("annual", 0)) or 0
        if m > 0:
            fee_display = "%s + VAT" % gbp(d["monthly"]); fee_period = "per month"
            fee_lines.append("Your fee would be %s + VAT a month - a single fixed amount, with no surprise bills along the way." % gbp(d["monthly"]))
        elif a > 0:
            fee_display = "%s + VAT" % gbp(d["annual"]); fee_period = "per year"
            fee_lines.append("Your fee would be %s + VAT a year - a single fixed amount, with no surprise bills along the way." % gbp(d["annual"]))
        else:
            fee_period = ""
            fee_lines.append("Your fee is set out in the attached proposal.")
    for fl in fee_lines:
        L += [fl]
    L += [""]

    # ---- itemised setup ----
    setup_rows = []   # (label, value)
    setup_total = 0
    reg_labels = set()
    for r in regs:
        lab = str(r.get("label") or "").strip()
        reg_labels.add(lab.lower())
        fee = num(r.get("fee")) or 0
        if r.get("included") or fee <= 0:
            setup_rows.append((lab, "included"))
        else:
            setup_rows.append((lab, "%s + VAT" % gbp(fee))); setup_total += fee
    for o in d.get("oneoffs", []):
        desc = str(o[0]).strip() if o and len(o) > 0 else ""
        price = num(o[2]) if o and len(o) > 2 else 0
        if not desc:
            continue
        dl = desc.lower()
        if any(dl == rl or dl in rl or rl in dl for rl in reg_labels):
            continue
        if price and price > 0:
            setup_rows.append((desc, "%s + VAT" % gbp(price))); setup_total += price
        else:
            setup_rows.append((desc, "included"))
    osub = num(d.get("osub", 0)) or 0
    if osub > setup_total:
        setup_total = osub
    setup_total_line = ("One-off total: %s + VAT" % gbp(setup_total)) if setup_total > 0 else "There's no charge for your setup - it's all included."
    if setup_rows:
        L += ["To get you set up, here's the one-off work at the start:"]
        L += ["\u2022 %s - %s" % (lab, val) for lab, val in setup_rows]
        L += [setup_total_line, ""]
    elif osub > 0:
        L += ["There's also a one-off setup of %s + VAT to get everything in place at the start." % gbp(osub), ""]

    # ---- onboarding scenario ----
    primary_key = "company" if K == "LTD" else "sa"
    primary_fee = 0
    for r in regs:
        if r.get("key") == primary_key:
            primary_fee = num(r.get("fee")) or 0
            break
    onboard_name, onboard_url = FORMS[K]["onboard"]
    reg_name, reg_url = FORMS[K]["reg"]
    sc = (scenario or "").lower()
    if sc not in ("newco", "switcher", "existing"):
        sc = "newco" if primary_key in keys else "existing"
    thing = "company" if K == "LTD" else "business"
    if sc == "newco":
        if K == "LTD":
            feebit = (" (the company formation fee is %s + VAT, charged once)" % gbp(primary_fee)) if primary_fee > 0 else ""
            onboard_text = "As we're forming %s for you, the first step is to complete our %s%s, which only takes a few minutes:" % (company, reg_name, feebit)
        else:
            onboard_text = "To get you registered, the first step is to complete our %s, which only takes a few minutes:" % reg_name
        cta_label, cta_url = "Open the " + reg_name, reg_url
    elif sc == "switcher":
        onboard_text = "You're moving to us from another accountant, so there's nothing for you to chase - just complete our %s and we'll write to your current accountant for professional clearance and handle the whole handover for you:" % onboard_name
        cta_label, cta_url = "Open the " + onboard_name, onboard_url
    else:
        if K == "LTD":
            onboard_text = "As your %s is already set up and you've not had an accountant before, getting started is simply completing our %s, which only takes a few minutes:" % (thing, onboard_name)
        else:
            onboard_text = "As you're already trading and you've not had an accountant before, getting started is simply completing our %s, which only takes a few minutes:" % onboard_name
        cta_label, cta_url = "Open the " + onboard_name, onboard_url
    L += [onboard_text, cta_url]

    # ---- extra registration links ----
    extra = []
    seen = set()
    for r in regs:
        if r.get("key") == primary_key:
            continue
        fm = r.get("form")
        if fm and fm[1] not in seen:
            lab = fm[0][:-5] if str(fm[0]).lower().endswith(" form") else fm[0]
            extra.append((lab, fm[1]))
            seen.add(fm[1])
    if extra:
        L += ["", "We'd also take care of a couple of registrations for you - you can complete those here as well:"]
        for lab, u in extra:
            L += ["%s: %s" % (lab, u)]
    closing = "If you have any questions, or would like to talk anything through, just let me know - I'd be glad to help."
    L += ["", closing]
    text_body = "\n".join(L)

    # ---------- branded HTML (navy + green) ----------
    NAVY = "#0D2B42"; GREEN = "#1E6B47"; CREAM = "#FBF8F3"; INK = "#243a4d"; MUTE = "#9fb3c8"
    svc_rows = "".join(
        '<tr><td valign="top" style="padding:3px 10px 3px 0;color:%s;font-weight:bold;">&#10003;</td>'
        '<td style="padding:3px 0;color:%s;font-size:15px;">%s</td></tr>' % (GREEN, INK, esc(x))
        for x in (svc or ["the services we discussed"]))
    setup_html = ""
    if setup_rows:
        rows = "".join(
            '<tr><td style="padding:7px 0;border-bottom:1px solid #eee6d8;color:%s;font-size:14px;">%s</td>'
            '<td align="right" style="padding:7px 0;border-bottom:1px solid #eee6d8;color:%s;font-size:14px;white-space:nowrap;">%s</td></tr>'
            % (INK, esc(lab), (GREEN if val == "included" else INK), esc(val))
            for lab, val in setup_rows)
        if setup_total <= 0:
            total_html = '<tr><td style="padding:9px 0 0;font-weight:bold;color:%s;font-size:14px;">%s</td><td></td></tr>' % (GREEN, esc(setup_total_line))
        elif len(setup_rows) > 1:
            total_html = '<tr><td style="padding:9px 0 0;font-weight:bold;color:%s;">One-off total</td><td align="right" style="padding:9px 0 0;font-weight:bold;color:%s;white-space:nowrap;">%s + VAT</td></tr>' % (GREEN, GREEN, esc(gbp(setup_total)))
        else:
            total_html = ''
        setup_html = (
            '<div style="font-size:12px;letter-spacing:1.5px;text-transform:uppercase;color:%s;font-weight:bold;margin:26px 0 10px;">To get you set up</div>'
            '<table role="presentation" width="100%%" cellpadding="0" cellspacing="0" style="border:1px solid #eee6d8;border-radius:8px;padding:6px 16px;background:#FCFAF5;">%s%s</table>'
            % (GREEN, rows, total_html))
    fee_lines_html = "".join('<p style="margin:12px 0 0;color:%s;font-size:14px;line-height:1.6;">%s</p>' % (INK, esc(fl)) for fl in fee_lines[1:])
    extra_html = ""
    if extra:
        links = "".join('<div style="margin:6px 0;"><a href="%s" style="color:%s;font-weight:bold;text-decoration:none;font-size:14px;">%s &rarr;</a></div>' % (u, GREEN, esc(lab)) for lab, u in extra)
        extra_html = ('<p style="margin:22px 0 6px;color:%s;font-size:15px;">We\'d also take care of a couple of registrations for you - you can complete those here as well:</p>%s' % (INK, links))
    fee_box = ""
    if fee_display:
        fee_box = (
            '<table role="presentation" width="100%%" cellpadding="0" cellspacing="0" style="margin:22px 0;"><tr>'
            '<td style="background:%s;border-radius:8px;padding:20px 24px;">'
            '<div style="color:%s;font-size:11px;letter-spacing:1.5px;text-transform:uppercase;">Your fee</div>'
            '<div style="color:#ffffff;font-size:30px;font-family:Georgia,serif;margin:4px 0 2px;">%s</div>'
            '<div style="color:#cfe3d8;font-size:13px;">%s &middot; %s</div>'
            '</td></tr></table>' % (NAVY, MUTE, esc(fee_display), esc(fee_period), esc(fee_sub)))
    html_body = (
        '<div style="margin:0;padding:0;background:%s;">'
        '<table role="presentation" width="100%%" cellpadding="0" cellspacing="0" style="background:%s;"><tr><td align="center" style="padding:24px 12px;">'
        '<table role="presentation" width="600" cellpadding="0" cellspacing="0" style="max-width:600px;width:100%%;">'
        # header
        '<tr><td style="background:%s;padding:26px 32px;border-radius:10px 10px 0 0;">'
        '<div style="font-family:Georgia,serif;color:#EAF0F7;font-size:13px;letter-spacing:3px;">A2Z ACCOUNTING SOLUTIONS</div>'
        '<div style="height:3px;width:46px;background:%s;margin:14px 0;"></div>'
        '<div style="color:%s;font-size:11px;letter-spacing:2px;text-transform:uppercase;">Proposal of services</div>'
        '<div style="color:#ffffff;font-size:23px;font-family:Georgia,serif;margin-top:3px;">%s</div>'
        '</td></tr>'
        # body
        '<tr><td style="background:#ffffff;padding:30px 32px;font-family:Arial,Helvetica,sans-serif;">'
        '<p style="margin:0 0 16px;color:%s;font-size:15px;">Hi %s,</p>'
        '<p style="margin:0 0 20px;color:%s;font-size:15px;line-height:1.65;">%s</p>'
        '<div style="font-size:12px;letter-spacing:1.5px;text-transform:uppercase;color:%s;font-weight:bold;margin:0 0 10px;">What we\'d take care of for you</div>'
        '<table role="presentation" cellpadding="0" cellspacing="0">%s</table>'
        '%s'   # fee box
        '%s'   # fee lines (directors etc)
        '%s'   # setup
        '<p style="margin:24px 0 14px;color:%s;font-size:15px;line-height:1.65;">%s</p>'
        '<a href="%s" style="display:inline-block;background:%s;color:#ffffff;text-decoration:none;padding:13px 28px;border-radius:6px;font-weight:bold;font-size:14px;">%s</a>'
        '%s'   # extra links
        '<p style="margin:24px 0 0;color:%s;font-size:15px;line-height:1.65;">%s</p>'
        '</td></tr>'
        # footer
        '<tr><td style="background:%s;padding:18px 32px;border-radius:0 0 10px 10px;color:%s;font-size:12px;line-height:1.7;">'
        'A2Z Accounting Solutions &middot; 01224 042961 &middot; info@a2zaccounting.co.uk<br>'
        '1st Floor, 499 Union Street, Aberdeen, AB11 6DB &middot; Regulated by ACCA'
        '</td></tr>'
        '</table></td></tr></table></div>'
    ) % (CREAM, CREAM, NAVY, GREEN, MUTE, esc(company),
         INK, esc(name), INK, esc(intro), GREEN, svc_rows,
         fee_box, fee_lines_html, setup_html,
         INK, esc(onboard_text), cta_url, GREEN, esc(cta_label),
         extra_html, INK, esc(closing), NAVY, MUTE)

    return subject, text_body, html_body


def _email_for(d):
    """Produce the proposal email. LTD/SA use _enhanced_email (his wording + itemised
    setup + 3 onboarding scenarios); PARTNERSHIP keeps his original build_email."""
    kind = (d.get("kind") or "LTD").upper()
    scenario = d.get("scenario") or ""
    tmpdir = tempfile.mkdtemp()
    wb_path = os.path.join(tmpdir, "wb.xlsx")
    if kind == "SA":
        wb = _build_workbook_sa(d)
    else:
        wb = _build_workbook_ltd(d)
        if kind == "PARTNERSHIP":
            wb.active.title = "Partnership Proposal"
    wb.save(wb_path)
    wb2 = GEN.safe_load_workbook(wb_path)
    if kind == "SA":
        return _enhanced_email(GEN.read_sa(wb2), "SA", scenario)
    if kind == "PARTNERSHIP":
        psub, pbody = GEN.build_email(GEN.read_partnership(wb2), "PARTNERSHIP")
        return psub, pbody, _text_to_html(pbody)
    return _enhanced_email(GEN.read_ltd(wb2), "LTD", scenario)


@app.route("/email", methods=["POST", "OPTIONS"])
def email():
    """Return his exact proposal email as {subject, body} for the given contract."""
    if request.method == "OPTIONS":
        return ("", 204)
    d = request.get_json(force=True, silent=True) or {}
    try:
        subject, body, html = _email_for(d)
        return jsonify(ok=True, subject=subject, body=body, html=html)
    except Exception as e:
        return jsonify(error="Could not build the email", detail=str(e)), 500



# ------------------------------------------------------------------
#  INVOICE PDF + inline logo (added Aug 2026)
#  The platform posts invoice data with a send; the service builds a
#  branded A2Z invoice PDF (navy/green, Nunito + Cormorant from the
#  same fonts/ folder) and attaches it. If the platform also sends
#  its embedded logo (logo_b64), the logo is used on the PDF AND
#  attached inline to the email as cid:a2zlogo, replacing the text
#  header of the standard email shell. His generator is untouched.
# ------------------------------------------------------------------
from fpdf import FPDF as _FPDF


def _logo_tmp(logo_b64):
    """Accepts a raw base64 PNG or a data URL; returns a temp file path or None."""
    if not logo_b64:
        return None
    b = str(logo_b64).strip()
    if b.lower().startswith("data:"):
        b = b.split(",", 1)[-1]
    try:
        raw = _b64.b64decode(b)
        p = os.path.join(tempfile.mkdtemp(), "logo.png")
        with open(p, "wb") as f:
            f.write(raw)
        return p
    except Exception:
        return None


def _invoice_pdf_bytes(inv, logo_path=None):
    NAVY = (22, 55, 90); GREEN = (30, 107, 71); GOLD = (196, 144, 90)
    INK = (35, 39, 43); MUTE = (110, 116, 122); LINE = (226, 210, 168)

    def n2(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0

    def gbp2(v):
        return u"\u00a3{:,.2f}".format(n2(v))

    pdf = _FPDF(format="A4")
    pdf.set_auto_page_break(True, margin=20)
    try:
        pdf.add_font("Nunito", "", os.path.join(_FONTS_DIR, "Nunito-Regular.ttf"))
        pdf.add_font("Nunito", "B", os.path.join(_FONTS_DIR, "Nunito-Bold.ttf"))
        pdf.add_font("CormorantB", "", os.path.join(_FONTS_DIR, "Cormorant-Bold.ttf"))
        BODY, HEAD = "Nunito", "CormorantB"
    except Exception:
        BODY, HEAD = "Helvetica", "Helvetica"
    pdf.add_page()
    M, W = 18, 210 - 36

    # header: logo (or wordmark) left, INVOICE + number right
    y0 = 16
    if logo_path:
        try:
            pdf.image(logo_path, x=M, y=y0, h=14)
        except Exception:
            logo_path = None
    if not logo_path:
        pdf.set_font(HEAD, "", 22)
        pdf.set_text_color(*NAVY)
        pdf.set_xy(M, y0)
        pdf.cell(90, 10, "A2Z Accounting")
    pdf.set_font(HEAD, "", 24)
    pdf.set_text_color(*NAVY)
    pdf.set_xy(M, y0)
    pdf.cell(W, 10, "INVOICE", align="R")
    pdf.set_font(BODY, "B", 11)
    pdf.set_text_color(*GOLD)
    pdf.set_xy(M, y0 + 10)
    pdf.cell(W, 6, str(inv.get("no") or ""), align="R")
    # invoice meta: labelled issue / due / VAT number (readable, right-aligned)
    my = y0 + 18
    for lab, val in (("Issue date", inv.get("issue") or ""), ("Due date", inv.get("due") or ""), ("VAT No", "435116127")):
        pdf.set_xy(M, my)
        pdf.set_font(BODY, "", 9)
        pdf.set_text_color(*MUTE)
        pdf.cell(W - 34, 5, lab, align="R")
        pdf.set_font(BODY, "B", 9)
        pdf.set_text_color(*INK)
        pdf.cell(34, 5, str(val), align="R")
        my += 5

    # gold rule
    pdf.set_draw_color(*GOLD)
    pdf.set_line_width(0.8)
    pdf.line(M, 52, 210 - M, 52)

    # parties: billed to (left) + our details (right)
    ytop = 58
    pdf.set_xy(M, ytop)
    pdf.set_font(BODY, "B", 8)
    pdf.set_text_color(*GREEN)
    pdf.cell(60, 5, "BILLED TO")
    pdf.set_xy(M, ytop + 6)
    pdf.set_font(BODY, "B", 12)
    pdf.set_text_color(*INK)
    pdf.cell(110, 6, str(inv.get("client") or ""))
    ly = ytop + 13
    addr = inv.get("address") or []
    if isinstance(addr, str):
        addr = [a.strip() for a in addr.split(",") if a.strip()]
    pdf.set_font(BODY, "", 9.5)
    pdf.set_text_color(*INK)
    for a in list(addr)[:5]:
        pdf.set_xy(M, ly)
        pdf.cell(110, 5, str(a))
        ly += 5
    fx = M + W - 72
    pdf.set_xy(fx, ytop)
    pdf.set_font(BODY, "B", 8)
    pdf.set_text_color(*GREEN)
    pdf.cell(72, 5, "FROM")
    pdf.set_xy(fx, ytop + 6)
    pdf.set_font(BODY, "B", 10)
    pdf.set_text_color(*INK)
    pdf.cell(72, 5, "A2Z Accounting Solutions Limited")
    ry = ytop + 12
    pdf.set_font(BODY, "", 9)
    pdf.set_text_color(*MUTE)
    for rline in ("First Floor", "499 Union Street", "Aberdeen", "AB11 6DB", "01224 042961", "info@a2zaccounting.co.uk"):
        pdf.set_xy(fx, ry)
        pdf.cell(72, 4.8, rline)
        ry += 4.8
    ytab = max(ly, ry)

    # lines table
    lines = inv.get("lines") or []
    if not lines:
        lines = [{"desc": "Professional services", "net": inv.get("sub") or 0, "vatRate": 20}]
    colD, colN, colV, colT = W - 96, 32, 26, 38
    pdf.set_xy(M, ytab + 4)
    pdf.set_fill_color(*NAVY)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font(BODY, "B", 9)
    pdf.cell(colD, 9, "  DESCRIPTION", fill=True)
    pdf.cell(colN, 9, "NET  ", align="R", fill=True)
    pdf.cell(colV, 9, "VAT  ", align="R", fill=True)
    pdf.cell(colT, 9, "TOTAL  ", align="R", fill=True)
    pdf.ln(9)
    net_sum = 0.0
    vat_sum = 0.0
    pdf.set_font(BODY, "", 10)
    for l in lines:
        netv = n2(l.get("net"))
        rate = l.get("vatRate")
        rate = 20.0 if rate is None else n2(rate)
        vatv = netv * rate / 100.0
        net_sum += netv
        vat_sum += vatv
        pdf.set_x(M)
        pdf.set_text_color(*INK)
        pdf.set_draw_color(*LINE)
        pdf.set_line_width(0.2)
        pdf.cell(colD, 9, "  " + str(l.get("desc") or ""), border="B")
        pdf.cell(colN, 9, gbp2(netv) + "  ", align="R", border="B")
        pdf.set_text_color(*MUTE)
        pdf.cell(colV, 9, ("%g%%  " % rate) if rate > 0 else "No VAT  ", align="R", border="B")
        pdf.set_text_color(*INK)
        pdf.cell(colT, 9, gbp2(netv + vatv) + "  ", align="R", border="B")
        pdf.ln(9)
    sub = n2(inv.get("sub")) or net_sum
    vat = n2(inv.get("vat")) or vat_sum
    gross = n2(inv.get("gross")) or (sub + vat)

    # totals (right)
    tx = M + W - 76
    y = pdf.get_y() + 6
    pdf.set_font(BODY, "", 10)
    pdf.set_text_color(*MUTE)
    pdf.set_xy(tx, y)
    pdf.cell(40, 7, "Sub Total")
    pdf.set_text_color(*INK)
    pdf.cell(36, 7, gbp2(sub), align="R")
    pdf.set_xy(tx, y + 7)
    pdf.set_text_color(*MUTE)
    pdf.cell(40, 7, "VAT")
    pdf.set_text_color(*INK)
    pdf.cell(36, 7, gbp2(vat), align="R")
    pdf.set_xy(tx, y + 16)
    pdf.set_fill_color(*NAVY)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font(BODY, "B", 11)
    pdf.cell(40, 10, "  To Pay", fill=True)
    pdf.cell(36, 10, gbp2(gross) + "  ", align="R", fill=True)

    # bank details (left, same block as the email)
    pdf.set_xy(M, y + 2)
    pdf.set_font(BODY, "B", 10)
    pdf.set_text_color(*NAVY)
    pdf.cell(90, 6, "Payment by bank transfer")
    pdf.set_font(BODY, "", 10)
    pdf.set_text_color(*INK)
    rows = [
        "Name: A2Z Accounting Solutions Limited",
        "Sort code: 20-29-24",
        "Account: 23875458",
        "Reference: %s" % (inv.get("no") or ""),
    ]
    yy = y + 8
    for rline in rows:
        pdf.set_xy(M, yy)
        pdf.cell(110, 5.5, rline)
        yy += 5.5

    # footer (page break off so the last line never spills onto a blank page 2)
    pdf.set_auto_page_break(False)
    pdf.set_y(-26)
    pdf.set_font(BODY, "", 8)
    pdf.set_text_color(*MUTE)
    pdf.cell(0, 5, u"A2Z Accounting Solutions Limited  \u00b7  First Floor, 499 Union Street, Aberdeen, AB11 6DB", align="C")
    pdf.set_y(-21)
    pdf.cell(0, 5, u"01224 042961  \u00b7  info@a2zaccounting.co.uk", align="C")

    out = pdf.output()
    return bytes(out)


@app.route("/invoice", methods=["POST", "OPTIONS"])
def invoice():
    """Return a branded invoice PDF for the posted invoice data (download)."""
    if request.method == "OPTIONS":
        return ("", 204)
    d = request.get_json(force=True, silent=True) or {}
    inv = d.get("invoice") or d
    try:
        pdf = _invoice_pdf_bytes(inv, _logo_tmp(d.get("logo_b64") or inv.get("logo_b64")))
    except Exception as e:
        print("=== INVOICE PDF BUILD ERROR ===", flush=True)
        traceback.print_exc()
        print("=== END ERROR ===", flush=True)
        return jsonify(error="Invoice PDF build failed", detail=str(e)), 500
    fname = "Invoice %s - A2Z Accounting Solutions.pdf" % (inv.get("no") or "")
    return send_file(io.BytesIO(pdf), mimetype="application/pdf",
                     as_attachment=True, download_name=fname)


# The text header inside the platform's standard email shell (emailShell in
# index.html). When a logo is supplied with a send, this exact block is swapped
# for the real logo image referenced as an inline cid attachment.
_SHELL_TEXT_HEADER = ('<div style="font-family:Cormorant Garamond,Georgia,serif;'
                      'font-size:1.6rem;color:#fff;font-weight:700;letter-spacing:.02em">'
                      'A2Z Accounting</div>')
_SHELL_LOGO_HEADER = ('<div style="display:inline-block;background:#ffffff;border-radius:8px;'
                      'padding:8px 18px;line-height:0"><img src="cid:a2zlogo" alt="A2Z Accounting" '
                      'style="height:30px"></div>')

_ENG_LOGO_ACC = "iVBORw0KGgoAAAANSUhEUgAAAqgAAADGCAIAAACRu+UIAAC0NElEQVR4nOxddXxUx9qemXPOnnWJJ4SEAAnuLqV4WyrU3dvbr3Zrt7fu7nrrLrSUChVKKVDcPRA0EEKAuKzLkZnvjzfZLsnuZhM2QNt9Lr9eSI7MOWdmXn9ezBhDCSSQQAIJJJDAPwPkeA8ggQQSSCCBBBI4duCP9wBOCDCEEGOMIfgfQk3/FxUYYYQQxghhDH/DHTvMBBJIIIEEEjha4H+gq5+BjGeMIYYRJgTHS2RTxhhtvCzGGOOEMpBAAgkkkMCJhX+E4G8U84whhMKKeVWlDU633emptzvrHW6Hy+NweVwen9cXCMiyLCuUMoQRR7DA81pRo9eKRqPOajJYTQabxZRiM1vMBovJQFpcmTJGKW3UA1r+OoEEEkgggQSOLf7Ogp/SRmHPcUekMrg8voMVNXsPlO89UFFSVlFWWVNV01DvcDndvoAkS5KsUtro/EcIsSaffmNE4E8XP0aI5ziNhteKGovRYLMaM1OTcjuldeuc2S0ns3tuVqeMZL1WDL21qlKEECYJHSCBBBJIIIHjg7+b4GcMUUYZYzzHBX+oKOresoqtu0o2bd+3vfhAycHK6jq72+tXVJVgzHEcz3E8TzjCERDJGKE/xX2EGzUmBiDGGKVUVamiqoqqKgpFiPE8bzLoMlJs3XIz+xV0GdKne98euV2yM4LynjGmUkowxoQkVIAEEkgggQSOGf4mgp8hRClF7Ajjvri0fG3hrhUbdmzesXf/oSqn20spE3hOo+EFnucIwRg35vMxxpos/Hag0QGAEbj0EUKMMUWlsqJIkqyoKsdxNrOxa07G0L75Y4f2Gd6/ICcrLXg66B+EJCosEkgggQQS6HD85QU/+POD8j4gy5uK9i5aXbhoTeH24rIGhwshrBUFUSNwHMEIQ/rdMXhq0AYIxhhjxpiiqgFJDkgyxjg1ydK/R5cJowZMGjVwQK+uQTeAqtI4ZhomkEACCSSQQEv8VQU/uPQhJx8hpFK6dsuuOYvXL1y5eff+Qz5/QKMRdKKG5zmEEKOMHu/HxE0JfggxWVZ9gYAsK0aDvm9+zpSxg88YP2xQn+5wpEopQohLOAASSCCBBBLoAPz1BD9jjNI/Tfzi0vLZ81f9vGjNtt2lPn9ApxW1ooYjhMJxx3esEYCbcvxVqnp9UkCSjAbdkL7550wZddakkdkZKXBYwgGQQAIJJJBA3PFXEvyhIl9WlN+Xb5rx8+Ila7bW2Z2iqNFrRY4QSulxN+7bBEIwwUSlqscbkBQlM9U2dezgy86aMH5EfzggIf4TSCCBBBKII/4agj9U5NfbXd/MXfb57D8Kd+1njBn1WoHnKaMnqnkfKwghBGNJll1ev4bnhw8ouPrcKedOHW3Qa1FC/CeQQAIJJBAnnOiCP1TkH66s/ei7+V/9vKTkUKVWIxh0WoQRVf/iAv9IQH0fY8zt9cuy2rt756vPnXzF2ZOSrCaUEP8JJJBAAgkcNU5cwc8QoiptFPlVde9+NfeLn/44XFVn0uu0ouZvYOJHByEEY+TzBbz+QLeczGsvOOW6C6YmWRrFfzNKogQSSCCBBBKIESeo4FcphbT22gbnO1/9+tG3v5dX1ZmMelEQVEpPzDF3BKC+3xeQPF5f99ysmy8749oLpuq1IjAQJ0r/E0gggQQSaCtOOMEfFGmKon703fzXPpm9r6ziHyjyQ4Ex5gjxBQIer39Az653X3/eRaePQwnPfwIJJJBAAm3HiSX4g07sBSs3P/nW12s27zLoRZ0o/mNFfihA/Lt9fkmSJ48e+Mi/Lx3evwdKeP4TSCCBBBJoC04UwQ/t7wjGBytqHn9zxsw5yzBGJoM+IfKbAUx8p8sjajTXX3jKfTdemGQxUUob2wAnkEACCSSQQFScEII/aLN+9O38p97+uryqzmYxIaDfTyAcOI5QShsc7oK87CfvvOKcKaNRwvRPIIEEEkggBhxnwQ/M+Rwh+8oq/vv8R3MWrzPptaJGo6jqcRzVXwU8z3l9gYAkX3bWhGf+c1VaslVVKZQDJJBAAgkkkEBYHE/BH0zd/2z2wodf+aKmwWEzGxO+/TYBPP/1dlfXnMyX7rvu9PHDEUKUskSz3wQSSCCBBMLiuAl+RVV5jrM73Xc/99EXP/5hMug0gpAw9NsHMP1lRbn1irOeuP1yUSMk3P4JJJBAAgmExXEQ/EEyvjVbdt/86P+K9hxIsZkThv5RAkz8ugbXuGF933ni1oK8TqBaHe9xJZBAAgkkcGLhWAt+yhiEoD+YNe++Fz6RFcVo0ClKwtCPDwSea3B6Uqym/z1281mTRlLKoB3w8R5XAgkkkEACJwqOqeCHoL6sqP997sO3Z/xqNRk4jqiJ1P24guO4gCQFAvKDN1/8wE0XoUTIP4EEEkgggRAcO8EPnufqOvs1974yf8XmlCQzpSzh3u8IEIIRQ3V212XTJ77z+M06rRjMo0wggQQSSOAfjmMk+EHq79hbdumdz+8qOZRsNckJ934Hg+e5mnrHycP7ffnSfzNSbYl0vwQSSCCBBNCxEfwg9Zdv2H75f16st7tMxkRQ/xhB4Ll6h7ugS9bM1+/v1a1zIt0vgQQSSCCBDhf8IGx+WbT2mntfVVVVqxXVRM3eMQTPcy63N9lm/ub1+0cM6JGQ/QkkkEAC/3B0rO8XxMzMOUsvu+sFhJhW1CSk/jGGoqhGg87u9Ey/8YlFqwt5jkuQJSSQQAIJ/JPRgYIfpP5nsxded/9rokbD83wigf+4QFWpTtRIknzBbc/MX7GJ57hEqCWBBBJI4B+LjhL8IPU/n/3HTQ//z6DXcgQnOu4cR6iUihoBMXTxHc/NX76J5xN2fwIJJJDAPxQdEuNXFJXnua9+WfKvB1436LUEY5oo2zsBwBEiyQrG6Ns3H5wwsn8i3p9AAgkk8A9E/AU/iJOf/1hz2V0v6ESRkITUP4HAERKQZQ3P//TeoyMG9EjU+CWQQAIJ/NMQZ8GvqJTnyNJ128656UlCCM8TShNS/8QCR4gvIFnNht8+erJXt84Jbp8EEkgggX8U4rnjq5TyHCnaU3rpXS8ghHieS0j9ExAqpTqtWNfguvC2Zypq6jmSUM4SSCCBBP5BiJvgp5RxhFTWNFx8x/Mut1cUhUQ23wkLVVVNRt3eAxWX3/Wizx9AGCW4kxNIIIEE/iGIj+BnjCHE/AHpiv++tK+swqjXqWpC6p/QUBQ1yWJctqHopkffIhjTRCJGAgkkkMA/A3EQ/AwhlVJCyJ1Pv79kzdYkizFRKnZcgDHmCOE5jiMkFgteVtTUJMuXPy565p1vOI4kuJUSSCCBBP4JiENyHyT0vT1jzu1PvpuabE2QwxwzYIwxRhhjhBClVJKUgCTLqipwnMVkoCym5oeEEIfLM/O1+6ZPHplI8k8ggQQS+NvjaAU/5ISv2LD99H89qhWFY9nn958JjDHBGGPMGJMVNSDJkiwjhAw6bVZ6UkFedt+C3JEDe27dVfrk/74yG/WtsiUSgmVZ1YqaxTOeK+jSiVJGCD4mj/J3RqPOFU71whgh9KfG1rFjYAj+F2YMOPifY3frxqfu6CdPIIG/I8LL1natJv5oxkEZIxjX1DtueOgNjBDBJEHK2xEICnvKqCQrgYAsKwrPc0lWc+/8nAE984b2zR/Qs2v3LplGvQ5OeferuTFenFImagS7y/2vB17//ZOnNALP2Am6M6sqDSPEIgACHx06nmagjDHKGGIcIRg3yvfop8ATEQxf+GgHwBhijFLG4Nlx0yCinsKgpgNjfJQKH1yKIcYRrtVbq5QihjCJx2MjRGnrOSrwQtpxcZXS6MYMRji6n6zVKwTRvkFSxqJnUre8bKundBAIxuTIVckQajXI2PKsZogxuNzql4qE6F+w3ZeNDthOGGMYYVgpkeZG8EiEEBzW6ixqv4HOEKIq5Thy6Z3Pf//7qhSrSU4EieOHZsLe75dkVRU1QkaKrWe3zkP75g/rX9CvIDcnKy30LFlRMUJrCnedes3DRoM29jo9nudq6hx3XHP2i/del2D0axNA4AVlPcDj89fWO2vtTrvD7fH5A5KMMdYIgl6rsZgNyVZzksVkNRtCr6OqtN2CEARfs92ntsFZ2+CobXA6XV6vP6CqlBCsFTUGvdZmNqbYzElWk0GnbfEgbR4CqPuhala9w1Vda6+qsztcHn9AYgiJgmAy6FKSLOkp1rRka/AeYd9eAgl0EBhjJ/hMAy02rAro9QX8AUlSFEoZwZjnOVEj6HViSwuHUkYZJZhE0ubbb/GD1H9/5m/fzluRmmSRE6H9eIAQTDBhjEmK4vdLsqKIGiEzLalvQZcRA3oMH9CjX0Fuis0SekqI1YgJwRwhn/2wUFZUjDGK2T5WFDXFZv7fF79MGNF/2vhhJ2aw/+Pv5pdX1QkCH11bJZj4A1J+XqdLzjiZsVat7vaDUsYY4zjCcRghdLCiZv22PesK9xTtKS2rqKltcPr8kiTLqtpoLmCMOY5oeF4UNVazISs1qXuXrIG9ug7rXzCgZ1eN0J7FqFIatIdkWdm6u3Rt4a5N2/ft2X+4orrO7vIEJFlW1KB5hzHmeU7D8zqtmGwzZWek9O6WM6Rv9+EDenTLyYQHacsboEHPSm2Dc9WmnSs2Fm3ZUVJ6uLrB4fIHZFlRmj4WFnii0WjMRl2ntOTe+TmjB/ceN6xv8Kbtm3LgdJy/YtOqTTv0Wi1lYaxYmA8Xn3FyQV6nNu37cPCsuct37TuoFTUtL44RVik16LQ3XHyaTqthLVwccIWv5ywp3l8e9grN8H8XTwN1MMZBUkoJIcs3bP918TqDLszjY4z9AenSM8f3LegCCha8sU3b9/38xxq9Vmx1SPECwcTrD4wc2OPUcUNhDPBy6h2uD775nTEadrcimPgCgX4FXc49ZUzLbwc/sTvd738zj6oU44j7HcE4IMl5nTMunz6xHbL/w1m/7z9UKWqEFjsPVlXVYjL8+8qzNALfcgLEDtCAOY6QJudGZW3Dzr0Hd+4r21NafrCipqrW7nJ7fQFJlhXKKMZY4DlRozHoRJvFlJFi65yV2iOvU8+unbvlZtrMRoI4FFnRaafgVynlOLJz38GHXv3cajIkiveOBkEvq6JSjzcQkGSeJ2nJ1lGDeo0e1Gv04F79e+YlWUzB4xljKqVB/09wuwQ98VBl7dylG0zGNldUMsa0ouaOp94b0i8/LckCizOOj9luwNw9WFFz1zMfuDxenuNacb1iLElydmbqaScPtZoMHaHjB1cpQri8uv6XRWvnLFq7ece+2gYnpUzgOUHgBZ7TCLyoEUJvzhhijKmqWlVrP1RZu3LTzk/ZQoNO7N4l69SThl5xzsT83E4IxTRgSinGBITuhqLi2fNXLVy5eU9pucfrJwQLPC8IHM9xeq2IMUZHjIExxgKyXFZes6+sYuHKLRijJItpQK+u504dfcG0k6wmY6tjgIvAHrVsfdFXPy/+Y3XhocoaVWUagddoeJ7jDHoR41CPAqKUujy+bXtKNxTt/Xz2H0lW0/D+PS4+4+RzpozSihqGEGtjigmllHDcr4vXvfHON7okq6IqLY/hCOd3ugf0yivI60Qpi125Aa3x89l//DpvhdZsVGlz2wZjLMlKerL1irMn6rQa1ELNBEnw6fcL5i9cG/YKfw6S4/wOl8vje/quq1SVxjhIyhhBaOXG7c+/8rkuOczjw2X798jrW9CFMYYQhje2buvux5//SGszH7NaHp7jffWOG/917qnjhsIY4G3VNTgfff1LVVUIxi1XNcdxfrvrvLMnhhf8CGGE6u2ux16fISsyITiaPx4xhHFBl07DB/SIna4UBvnJ9wtWrS0UDQZ65BfEmEiSlJ2Z9n8Xn6YR+JYTILZbNG4m8NG37CyZv3zj0nVFO/aW1TQ4JElGCPMc4XjQChr9+AzWIGUqZZRSRVVBC9dpxYxUW9/83PEj+p06bmj33KywN22P4Id3q6r09qfe83j9ZpM+IfjbATDuKWMBSfb5A4wxq9k4qHe3UYN6njSs75C+3TNTk4IHg+sm6P8P64qH5fT1L0uq6+ypNktbiyopYzpRc6C8+r4XPvnk+TupSlEb7b8Ogkopz3HL1hcFJLlTWrISw2TjONLgcK/etPO0k4e2aa+PaTwqhVW6q+Tg+zN/mz1/9eHqOoHjdFrRZjFihBtj/Y3CMcxWRDAO6gRgNe4rq3zq5U98gcBL912vqDR6qAWuDzvXnEXr3ps5d+XGHW6vX6fVaEWNzmZmjbdGTVkHYcbAEcxpBK2oIRgzxBRFXbFx++/LNvbOzxk7pI9KGRd5C4N9E2O8dN22lz/6YcnarQFJNui1VrMRI0wbnxuyB5rfmuc4gecMei1CWFGUBSs3z1u24ZWeef++4qwrz5mECW6H6W/Q63RJlhSbOeyc5wix85xGENp0zSAsRr0xyWI1GVomMGGMZVlJspqiK0kWo8GUZLGEu0LItRAzGT74Zt6V50wu6NIJTPkYR6jXitoIj89xxM4RUdP82XWiRpdkSbaajtnWzXNcLULBJKQ/R0hIis2sqmoEwU8aCDEb9VGuDFeQFSWq4EccRxqc7sfenPHrB4+3dTuwmgxmm8Vk0DdLjAAbI8lqbLdpEdxM7C7P9/NWzpyzZNP2fS6PVxB4ragx6rXEoEeNWcIRcvtgHBhhhBlilLLqWvvc8upfFq199I0ZF5w69n+P3dwycNAewQ9O/jc+/3nx6sLUpDYLmH8ygpJbUVUw7jUaPiczdWjf/AmjBowe3KtHXnbw4EbLHmOCMSEYXDeRwBjiOc4fkGb+ukyvFduXuaOoapLV9PWcJWdOGnHu1NEniMMfHA/zl29ijMmqGuNWJcnKvOUbTzt5aBxH0ihxOVJVZ3/xg++++HFRg9NtMuhSrGbGGGU0xrE1Jfz/uY51okZKto4b1hchhKP6C0EkcBgvX1/0zLvfLFm7DWNsMuhSbGaIDsa4HoNVBzBijiNajaZLp/QBPfNQ0zsPC5gVtfWOh1/7csbPixRVNRv1Br2WUhbL4zc9d2MiksWoRxjt2X/4+gde/+qXJS/ee12/Hl3aOvHA4oE/Ye+oqGq7k5nUpouHFfxKDBMyyhVCwXOc0+196q2vv3jp7jYNFj562MdnKPyzB085ljYbWKXNfsgQUlQ1kuCH8Ud/b3AFRVUJiyb4FVU1G3R/rC78/veV5586tk1zLPgFWwr+dr9DUI45jrg9vvdm/vbxd/OLS8s1Am/Qiyk2C21SnSlqZTnDOg79iSDwokbgea663lFeXc9xHKOs2YJus+CnlHIcKT5Q/sw7M61mQxTnVQJBBJ35sqy4/JKiKBaTYVDvruOG95s4asCQPt0tpj/zvGAagUsn9iQ7SlWO4xas3Lxjb1lY6yRGMMp0Ws0DL3168vC+NrPpuDv8wZ9sd7pXb9ml14kxpitSSvU6cdm6Il9A0omauET6IUTKYTxr7vKHXv289FClxWxMsZphUziaK2OMArJsNRkG9+mOEMKRfd2wW7k9vkff+PKDb+apKrWaDYihOIwBYY/P37cgF8yaSOYmDGDpum03P/pWcelhMHZVlcaeTRIK1pQYqNNqDDrtsvXbJl/1wPP/vebq86ZQymLITf5bQVFVq9nww/yV162dOn5E/0T3rLiDUqbVCE+/PfO0cUN1WvE4JvoFfWbf/77yqbe+3l58wKDTptjMUG1xlGuZMaYyhlXMKJs2fhhGSGW0md3YZsEP6/uBlz51uLxWcyK6Hw2N8p6hgCx7fQHEWFqK7aShfSePGTh+RP/e3XOC0w6yshuz89plZMOlvvjxD4SOIsMEHP5acd/BiqfemvnqgzeolHZgdlws46GM4/DqzbsOVdZYjLEqNJQxrSjsLSvftH3vmMG9KaMcPqo9FKR+QJLvfu7D92fO0+vElCSLqoQ3MdsKjLE/EBjQM69TejKLrGlBtcXW3aX/euD1Tdv3JttMGOF4LUCwXUYP6oWaIsctoaoqx3Gf/bDw9qfeY4ylJFkURW2fyG8GsGwsJoOsqDc89GbJoconbr+CUtroxPzHACOMMX78za/GDu1DMO7Q1NR/IChjBp22aM+Bd7+e+5/rzo09lyK+AO25wen+73MfffnTIlEjpCRZVPVo5X3zu1Bq0IsjBvRA4Rx4bRP8oKfMXrD6l0VrbRZjQupHAscRxJBfkn3+AMa4c0bKyAm9Th03ZNywvtkZKcHDVJUi3HqVaqsAE23P/sNL1m4zGY62UYKiqEkW00ffzb/4jJNHtCURpuMwf8UmRaVt2gQJJoGANH/5pjGDex8lqRS83pp6x+X/efGP1VtSkyyMsTgyVBJMJEkZ1q8ArOewm5GqUp7j5i3bcM19r7rcvrRkq6wocRG6jdenVK8VRw7qhSL4+UHteOvLOf955n2zycAREssb+FNsx/ANVJVyhCTbTE+/PdPnl1689zooQfznCD+VUrNBv3Ljjq9+XnLlOZNOkFjb3wkqpRaT/vXPfrz49JOz0pOOvUcTiG637i69+p6Xt+0pTbGZ47uZAAjG/oDUNSezV7fOKFyRSBtmFTAJuD2+x9+coQ1T2JDAn2hwuB1ub2aq7cpzJn396r2rv33l8xf/c+mZ47MzUiDiCIQjHEfazSsSCvgWX89Z0uB0x6UEn2Csquojr32h0vBlNscGEAMLSPLy9UU6URPWzx/p7VFGRVGzaHWhqlKetP+dgNSvqKk/84bHlqzbmp5sU1UaS8QBY0wI4TjCc1zoH/jooV+dMUY4PGpQr0iXUlSV48js+asuuv25gKSYjXpZCZO+3hJQ9NFsAMExBIcAOUo5Wal98nNQuFeqqirPcZ98v+CuZ963mo0Y4+iuF0IIz3HgRZBlRVYUyEbkuFZmO2OMUpqeYnvl4x+effcbjiP/tGAiGGrPvfet3ekhJEGEGmcwxjSCUFlrf+69WRgf69erqCrPkSVrt5567UPFpeVpSZYYNxMUsp80//Nnpn/IwQT7AvKQvt21ogYSxZpdrQ0WP5QcvDVjzvY9BxI5fZGAEaKMXT594jlTRo0d2icYvAdnflsj97GAMcRxnNcfmD1/lUGnjQt5okqpxWhYvHbrrF+XXXLm+ONleVDGOIy37CwpPnBYK4ph2dkkWRb4MNOYUqYXNTv2lu3YW9avR5c2pUkHESwUPv/WZwp3laRYLa1K3GB8R5KVgCSDzEMMYdyYgkMIJoRoBF4j8DzPYYRlRUmxmof07YbCBfjB1p+/YtPV974iCDzPca0uPSgYUanqC0iSrACRAFDpMcbAycRznEbDawQBaoScLs/AXt0MOm3LF0Up5Thu8ZrC255412IysKjWOyEEIeb2+CRZMehEs8mgEzUqpW6Pz+HyqJQa9FqtRhPlERhDqqqmJVkff/OrHl07nzhJpscGjDGdViw+cPiNz3965NZLO8gdDZQSsbzVILdjFMTCvwS3OxHowBVVtZmNX/y06Jrzpwzu0/2YeTRhIS9Zu/X8W59RKTUZtK2S3wSTw1SVwmYSJAUJHsARwvOcwPM8zxGMKWOMUsjwP2loHxRhtcYq+CHGebiq7q0vfrGY9P80NTx2YEK8Hu+1508dPbgXaqImhb2+g+4I6ZYLVm7eVXIojvEXSqleKz733qwzJ44w6LXHJREGpuyClZu9fsmg07WUFhjjkQN7bdxeHNYpwXFcvdO9YOWmfj26RIpbR707opRhjK5/4PX1W3enJllblfo8x0my4nD5CMYZaUndczK75WRmZ6ZYTUaB5xRVdXt8lbX2w1W1JQcrD5bX1NvdPEcoYwN6du2UntIywE+bCDOuvvcVsN2j12sQjDHBXl/A5w8YDbpuOZn5XTp1yU7PTLHp9VqO4IAkNzjch6vryg5XlxysrKiu9/gCOq1GoXTMkN6oRYAfCIAraxpueOhNMC+iaJY8x7m9PsbY2KG9z5w4csSAHp3Skw16rUppbb1zV8nBhau2zFm09mBlrc1sBBU50ptniBl02tuffHdwn265Wen/qBYSqqpaTYZ3Zvx6+fSJednp7dNZoyMgyX67qwG3niOiEXjIg4t0AMbY4w3EsjT8dpfHF2jPcOMNQrAky4+/+dVP7z5ybO4IC3nb7tJL73pBpVQrCtHLkkGiS7LsdPkpY1aTITcrLSs9KTXJYjLoBJ4nhEiy4nR76+2uytr6ypqGBqdblhVB4PVakUPMZjaOGNATIUTCpTfFKvgZYgSTlz/6obK2IaXtNeL/HGCMZUX9bPbCkYN6UtpKQXYktInHFA75+pelOApzVdtBGdNrxR17D34w6/c7rzn7uCTCQH/hRasLRY3QTEhgcIrqtLdfNf3GR950eXw8R5rtTpQxjcAvWLnlrmvPbYdSD2v12XdnzV6wOiPFFn1rI4QwRuvsztQky9lTRk2fPHJ4/4L0FFuk42VZKT1cvW7r7kVrCmfPXzWodzdwjIdOGNhtvf7Avx543enymo366OuO5zivP+APSP175p07dfTUsYN7d8/RippIxztcnt0lh5Zv2L5g5ea1hbuG9y9ALQL8UFVxzwsflZVXJ9vMUYKRHMfV2Z2D+3R/7LbLTzlpcLPf2szG/C5ZZ04ccd8NF7z+2U/vfPUrx3EagY+kx1DGRI1QU++478VPZr52H5DPRHn2vxMYQwLP19ldT78986Nn72i1DUGbAAth8uiBM95+OBwV3Z8AU3jJ2m2ffr/AZNCFVfgIwW6v//8uPm3UoF7RTWeMcUCSoVz5uOcMqSq1mAy/L9/488I1Z3V8V1KYvXan+8p7Xna5fSZDGBsmFBxHJElp8LgzUm2nnTx06tghw/sXdM5K1WvFsMfLslJRU797/+H1W/es2Li9cGdJeXnNkIE9uudmgZu55SkxCX5KKcFkT+nhL39aZDUbE43bo4BSatCJy9cXuT0+s1Efu6EMtGWUNTKwtoG6i5D9hyqXrttq1Gvj23hDpdRk1L09Y84VZ09MtpqOsdEPqs+e/YeL9pS2ZCYA0rRuOUnjR/TrnJGyZed+Qd/cLgGnxZYd+/YfqsrLTm9TIg9I/fXb9jz33iygB4lyMMcRry9AML75sjNuveLMbjmZ8HPoZ9GssRA0pxMEPr9LVn6XrMvOmvD4bZfDwJttiBAXf/adb9YU7kpPtkZ3DHIcqbM7e3TNvudf519w2klBzhaIMbUcAyHYYjIMH9Bj+IAe/7nu3B17y7p2zkCNvvpGwIb46+J1s+YuT7ZGk/o8x9XZnZdPn/jGIzcadFrQXEP78AX5RzLTkp777zUnD+93/QOvef2SqBEiTVpFVW0W048LVs9dun7ayScojXQHAdzRs+Yuu+6CqaMH946jOxq+SLeczOAsjQ5JVt77em4kCh3IDjlpWL+zJ49s6xiOLxhjGg3/5NtfTxk7WNQIHc3tzXHknuc/3r6nNDq9PUaIcKTB4c5Itd15zdlXnDMpN6QbSxMf1xEnEIwFgc/JSsvJSpsyZhBCaO+B8m/mLtOJGo4jUBbb8kax0RYihDF69ZMfHS4Pz3GJbJMoYIxpNZr9h6pWbNjOEGo1PAYsPapKMUKEYKBldLq981ds2ldWgVrLhoZN8/vfV9Y2OAWej++nARLf0kNVH3zzG8Y49pY/cQGwiC9aU2gPl7GICZYkuXf3zlpRM6BX14Akh/VoCTxf53AtWr0FNb2rmMAQQlhR1Ptf/ERR1ejqAs9xDpcnLzv95/cfffXBG7rlZMIHZYxhhMIm9xFCII6gqpRSmp2R0jkzBR25IVJKOUK27Nj3vy9+Sbaaom0WGGOM6+2ua86fuvSrFy6fPlHUCEA2AvZ62DFAZlOwaLilb4CxxgrGp96eqRH4KHYnSP2rz53y0bN3GHRaVVUhhNyUdYSDdaocRxhjsqKedvLQ7996SBR4pbGpRITvwBjPcy988J2iqn9jV3/YF4Axoow9/uZXlLLonE7tAKQYqyqN8icgyapKPd5AdDmNMfZ4fcHjo/yBpOb4PkgsCDt8SplRr9uyo+SDWfMIwR3XqxDo7ecsXvf5j38k28zRFzLCuN7uuuj0ccu+euGhWy7JzUqjlDYt5caum80y+zDGjCHKGJB5MMa652Y9eNPFd117LkLhzX0Ui+CHCNPu/Ye++225xWRIOPlbBcaYUvrzH2sxiuiepKxx0w9+S0rZ9uID78/87cr/vjTsnNsvvfMFkDdR5D5DiCOcrKiz56/SxdACpB1QVWoy6j76dn5tg/MY5xiD6blgxWaeD0POD9SwwHgzalAvgsO1nYe6AELmr9iEIsS6wkKlKiH423krlq3fHp08gOe4Bqd71KBe8z99+qShfSH3Jpb0dYwRyEJCSLClZvNHROipt2f6JanVEK/L7X3m7qvffeJWm9moqCpDCDTIVrdsSL9H4TRUmJw/zF+1cfteo14XaWcEPXXs0D7/e+xmSoF1PFp4C2Ms8JysKCMH9nzniVu9fn+UQVJKTQbd2i27F6zYjGMISP8VQRlTWxCrISjtMxqWrN0667dlhMT52SHFOEyKeIs/sehb4bPNW/zhOe64aG+RoumqSs0m/asfz66saYD2RXG/NUhrrz/w2Btfgl8hEkAR93j9L9x73Wcv/CcnKw2kOLzb6GFfjBGBLD+OAwutVarKGLI6EcIIvfXlHKfbm+jWGgsg9rx03VaHywNR6uCvgp+ENJlEDQ73gpWb73/p05Mv/e/4y+699Yl3vv995b6DFaeOG5LXOYNSGmWpUEoxRmu27Nq6a79e14YmvLEDjP4Dh6s/+X7BsTT64RVV1jRsKNqj04bJ56eUakVxcN/uCKHh/QusFpOiqC3fFGQqrCvcXV1nj11xgcSZNz//SSsKUdQpjhCnxzu4T7cf3nooLdkKHPvtKAtuWY2jUkoIXrV55+8rNlqiNsHiOGJ3up+488r/XHsuTC2+XbkYLacZIUSl9INvfotu7quU6rTim4/cJPAcQ7Gm4Ak8ryjqOVNHX33ulAaHK5oPnyGE0CffL0Anhos4jiAEe3z+SaMGdOmU7g/ILZ+OUqrTis++843L48OJ0r62A4jMz5s6JsJvmSgIh6tqX/jgO4wx64DNjVJKMP7q5yWFu/ZDyUykIzFCbq//zUdvuv2q6WDigxRvx03Bcxz93FYEP5gvZeU1389bYTLqlQ7zhxx34Kaetn8WR7a3RzhjTNQIpYerl68vAlLSP+V90ycpK6+Z8fPiq+55edQFd5190xMvf/TDll37McYpVlOS2cRz3Onjh7HIac+h+Pa35ZKsdBwNhapSo1776fcLnG4vOVKP6TiAkb18Q1FVrV3Tog8vBPgzUq29u+cghLp2zszvkuUPyC1r4SCSV1HTsHx9UfCyrdxapRjj+Ss2bdq+Dyjowx6GMZYVxWY2fv7i3SCb+fhFoOExPvxmniwrOLKjgudIg8N96VkT7r7uXEVR48IJAYDa33WFezZsK46SO8JznMPpvvb8Kb275yiq2qY4NOEIY+yhWy5JT7HJshJp4CqlRr12+Yai0kNVhBzreFOHAmPsD8jdc7IunHaSy+Nr+fYYJNjuO/j2jDnkmMfa/gbgCHa6vdOnjOpbkOv1hYlZKKpqNZs+/WHB1t37OS5axUr7QAgnyfIH38zTidE6IPMcV+9wP3DTRdecN0VWVNLxdY+tCX5KEUIffze/ut6h4flofue/IEhIMausKB5fwOH22p0eu9PjdHu9vgBEFtvhocIYUcp+WbQONZVug7zfs//wO1/9Ov3GJ8Zc9J9r73v1m1+XVdfZTQZ9is1s0IkIIZVSbyCQmZo0afRAHDX9FZzYdqf792Ub41W+H+lGOq1YXHr4u3krMI5Jdh49YInOX7GJsTBNawjGfknu2bWzzWxUFJXjyMiBPSOF+THCjKHfl29EsZmMcMxXvyyJfiwh2OXxPXbb5d1yMoFgJ+aHawXg3ztcVbdg5WajQUcjlM6C2OicmfrivdcxxkhrwYU2DgIhhH6Yv9IfCP9WEUIYI1lRkm3mmy49PQrTcCSAJMtKS7ri7IlOtzfKVBd4vq7BCV/wmPWPPzbgOFJrd14xfWKy1SSHi6KqVLWYDG9+/vPBipq/md5zDIAJ9gcko1579blTvP5A2DnGcdjnl55486u4311VKcZo0erCbbv3GyL3GeEIcbg8E0cNePCmi4DX7xj4taLtVuA2tLs8M39datKHL+f4KwLiWxhjX0Cqt7saHG5VpSk284CeeVPGDDpnyuizp4yaOGpAn/xcs0Hv8QVqGxwebwD8ATHeQqXMoBcXry2UFYXnuB3FZa9/9tNp1z0y7tL/3v7kewtWbvb5pSSryWYxCgLf1E2ZIYQIIV6vf/TgXmnJVhqOcSkIOH7Bys0HyqtFsWOJFIEF75PvFoBZ2XE3AoBO4/L4Vm3aqddpwrdDVZShffNRkxE/blhfSBxreTWVUr1OXLFxh8vj5VrzWFDGCMHl1XUrNhQZdLpIa5UQ4vb4RwzocdW5k4Dfpp2PGg7wRL8t3VBVa9fwEeOCHCEur+/Oa85OsZlVSuPo8gkSJi5ZU6jTaqJE910e35Qxg7tkp4Oy0tYbYYwYY1eePcliMigqjfQAjDGO4/5YXYiitg3864EhjpA6uzMzLemMCSOcbm9L9ZExpOH5mnrHs+/OwhiHTWRJIAowwbX1jgtOG9s5MzUgSeFYKanVZPh1yfrflm7gCIlnLgVGCKFvf1uhRq0ZoIwJAv/0XVdBxO/YxLOilfNBH/QfF6wqKauM1Ov6rwWolPMHpAavX6fVFOR1Gjmw16iBPfoU5GZnpCZbTaEHU8aq6+zF+w+v3LRj3rKNG4uKKWNmox4SLKPfiDEm8HxtvfP2p94tr6pfuWmH0+UBagUgZ4bkvrDnUsamNfn5o2ylMD2+/33lMZgolDKjTrtp+96l67dNGjWwowuroDHP+q27Sw9XteyBjWCp8PyIgT1Rk1NkxIAeGSk2p8fLc80zASHycuBw9Zotu6eMGQQXj3xrSjhu+fqiqlp7UuRu5RhjSZZvvGQa8Oi1p7l1ZIBsW7hqM8eRSBs9qK3dc7Mumz4B9KQ4DgAyknbsLdtbVqkVmzMoNMP0SSMYY5S1hf27CRA56tE1e3j/giXrtpkMehZOyaCMaUXNtt2lTre3TSWyJz4wxoGAzBj7v0tO+37+yrBhZijtm/Hz4mvOnzKsX8GJ0DvjrwKMMEa4wenWacWLTx/36iezw5LQMMR4nnvif19NHDVA4Pm4lPbBqnS4PCs37TBE7pMOOTrnTh0zpO+x4xBE0S1+jhDK2Jc/LdZoomX3/CVAMOY54gtItQ3O1CTLbVed9euHj6+Y+dKbj9x46VkTBvTsClIfCpwa66AwzkixnTSs733/d+HiL5/74e2HTx7Wr97uipFIizEmCPyn3y9cuHIzwTjFZjEZ9AT6N1MaVnXAGMmykpFimzRqID6yoroZoHz/YEXNig3bDTotPQbZzhiplH72w0IUoTwmjgBpN3/FZjlc7gLGGN7S4D7dEEKEI5SxtGTrkL7doZi+5QUJRrKigK84eucBCCss37Aj2jEYBwJSl+z0aeOHo3izkYDp7HB5tuwoidSeAG7q8frOnDDcYjREKtVtN2Cxr9+6x+P1cRHaHMBcTUu2jh7cO0gs2g6Ae2PiqIFR8lQYYxqBr6ipLy4tRzG1+/lrgDWGS1SM8ciBPSeO7O/0hA95YIJlRX3iza/C0hj9bdSgDoIkKwyhf114apI1PCEHpcyk120oKv7k+wXxKu2Dlbtp+75DFbUaTfMspT/BEMb4irMnMnZMvTkR9yzI7tmwbc/6rXvizgxzLAGOfX9AqrO78rtkvfbQDStnvfzSfdePHdJH1AhQYEqbJDEUOEFeH2sqsldUFWM8ZcyguR898frDN2KM/QEplu2eMWY1G4D7QomhGwMhxOMLDB/QIyPVRqM6b+FzzF2yvrrOrhHiXL4f/o4qNRl0C1duOXC4mhDSgbFGhniOkxV16bpt2nB+ZoKxLxAY2LtrapIFOHngmEmjB8KXCjN4ynRacenabZIsc5GJKKBtEqW0cFeJRogYPSEEe/2BsUN6W82GsA0wjgYgdHeVHKqoqRdaZDWGHiYI/CnjhnSEEATtZ8vOkii+ZYyJX5IL8jplpiW1I8Afch2MEBo5sKdWjBhTQAgRgn3+wO79h1BsGa9/FWCEZblRFN182RmRDlNVajHpF6zcPGfRWtLCHc1x5B/DatgeSLKCEereJevcqaMj1aaplJr0ulc+/qHe7jqy/AcisG1+v7Bw1m/dLclRsmSwX5LysjPGDu0bi/YMzB9AExLrnyYagGaXaqWK5utflvr80l83rsZxnKqqNQ2OvM4Zbz9+y/KvX7z5sjNSbGZ4I4wFW0eESYzCwRYIHIcgOZ+xGy+Z9sv7jyVbTZFSRZoBbhTjaDHCKqWnjhuKWtvd4NY/LVwj8MfIGcMQEni+psHx/fyVqCOtLsjeKtpTuqvkkC5cYx4gRT55eD/UpADB/Dx5eD+LMTzPBGVMJ2r2lB7esqMkeFYYMIYQqqpzlJXXaDR8pEfECFNKRw/uHWSjiyPggjv3lvkCUqSNAAINmalJA3t1Az6A+I6BcAQhtPdAObg9wx+DsSwrvbp2RjGwVEW7F8YIofwuWSk2sxyZzAcjzBgDSqu/ERhkqyCEGGOTRw8cM7i3y+ML6+oDt8ez786SZAU4W4JoXwHnPwZMkhrf8C2Xn2HUh6fLhbrl/Yeq3vzil5Z1yxzhEGubcgUTe+vuUkIiBuwIwT6/NLhPd6NeG4vfrpEFKzbWhGB5WlgyhvDRScjucbi985ZtMOi16l8wlRQUqAaHKzXJcs8NF9x0yekmow4hBB2+2xGiBlkrK8qIAT1+fOeRadc/4vH5BT6yD6ftkBUl2WqaNGoAiso2A1y2u0sOrd+2x3AMnTGUMa1G8+P8VbdfOb0Dew4xRhBauGqLx+vThQvIQX3XScP6oiZ7EQbTIy+7Z7fsrTv36/VhEmgJIV5/YP6KTcMH9Ij0ySB4fKiyxuF0i6Im0mEqpVqt2LegS8dl4hQfKEcsIl8bxliSlLzOGR3Bo9xIOeLzl1fXh6VOCj2ye5cshMJTJ8UIGHyKzZKZmlTX4BQELuwNGWIY4/KqOoTCVHn8dQGufkoZQowQctuV05dv2B6FaW7DtuKvflly9bmTQ/NsBJ5Diby/yAjIMkJIVtQ++blnThoxc87SsM3MoCXpB9/8du35U7IzUmCbhXi/EHUhhAW4RUsPV0WREWDpDezVFSFEGSUoYo4wrMrCXSUV1fVtEjqgWWakJg3s1TU0dyG84IcEqEWrt+w/VBXHhm/HDNDkwOPzn3fqmCduvwJYqVWVAgvS0VxZ4HlZUfv16PLeU7dd+O9nNEJEk6jNYybE6fGePLxfU450lCxQShD3y+J1dpcn9Rg2TKKU6nVi4a79G7fvHd6/o5KM4Jp/rNoihOONAR05v0tWn+45KEQ9gvY244b1XVe422jQUdT8nVBGtRph4eotD9x0caRhw90OV9YFJEWnE1U1bB4GhuZp2enJqAOSzEEQHq6si2IoYIwURc3tlIaaOhfHcQAQRa53uKMTdjGEMCGd0pPRUUtimO0ZKdYtO1WMcPgkDIYIIXV2F+r4FJNjBwbTiaqqKgg8pey0k4eOHNhzY1GxIRxVIqVUpxNf+2T2BaeO1ev+7Eyh0Qh/m1cSdzCEJAimMIYQuv2q6T8tWB3WRwU+leo6+2uf/vTy/ddTRhECdQoLAs/aYvKDnHa6vTV1DoEPr8sihBhiHCHdczNRa4sIlvmz78765qfFJmsbJDLHEZfdff4Z4799837GaJAUJGKFLkLoxwVrYrz6CQWCsd3pNhp0Hz1zx4yX74Eya/BhxGXLEHhOUdRpJw+94uyJDQ53vPLbgZQGuixEjw5whFDK5i5ZJ2qi8cp1BDhCfP7Az3+sQR3j7Qd/1/5DVYW7SvRasWWONyRYjBzYE/Izgh8Uls2kUQPDqguoKcxftOfA7pKDGIfP3wFBW1PvUKgaZR2qKjUb9VaLsb1PGQ0wR2saHBxp3mwwCDAUMlNsqCO+AmMIIYfL6/UFonAdMkYFnkuymGBARwPYha1mk6rSSJdiTTx36O+Vy4YxUlUV1jtllOe526+eHinkQRkzaMUdxWWfzV4I7miYpTqN5u/0TuIOyKIgHFEpHdS72xkTRzhcnrAluIqqWs2GGT8t2nugIpQxSStGY9ttCTjW4fK4vD6OI5ESiillGoFPS7YiFNMi0okanVFvMrTtj86o12ubt+gMH0wihNQ1OFdsKDJ0DBFsxwFjHJCVyaMHLfri2cumT4AM/XZzH0YClCHdff15VrMxXga3SlWTQTdp1EAU1Y6E4v6iPaVbdpYYtBFJIToIlFKdKM5fsUmS5Y5o1wR6zJI1W+saXJF6DmGMx4/ojxAKXU7gIBnaL79zZqokhWE/RU3ddBau2oKiplA4XJ4ocwXD/qsTtRpN0w/iCRi52+MjJILt2zQOi8kQ31sD4JY+vz86HSRjiCPEoA/fJ7Qd0Ositg8G/JkH93cRcgwhhLBCG3OvoIrqzIkjRw7s6fZ4w/r8VEoNeu3bM351ub1BtUwranBcW3L/nYARatYX565rz4mSScpzfL3D/cZnPzVmtjKEENKJmrYp2IwhhFweb8uEjCOPYhqBN+p1KLZ9hDYVnbUZLV2nLa8O6ufS9dsOV9W1ZEs9wUEw9vkDN116erecTEmSIUO/1bNY9BqvlnchmDHULSdz6tjBLrfv6I1+8GD37Nq5d34Oaq2QDyE0d8l6l8cXX96YWEAZ02k1u0sObY6eJddegKSZv3ITx4WxNTFCsqymJllGDuyJEMIhbwmMeIvJMLx/D68/EPajM8Y0PN/YsCfyG/YFpGhDxI0kDY23iKsMgidWVSpFZeqF22qE+NIHHAFJVluNGROMBZ5HcXoHHCExrsG/idhHCCGEG/N/GyUTo4znyF3XnhPJ6Acazd37D33x0+Kg48qgFwnGf6Vt+pgCg22GEeYIoZQO6t3t7CmjHC5PWI5tRVUtJsOsucv2HaggmMAb1unENooIhBAKSDIN8UqGHxzGPA/DOKbzOlzNKMYIoXnLNjL2l/Sq8Rx55ZPZiqoKUXdG2tSYEkECP0IIFqEaU5o8ZZQB0w4LW1vbNhBM/AH55OH9gBAmypEcISql85Zv1Go6pB1fq2jMklu+EcXbzwyuptoG57rC3fpwjXkwIb5AoG9Bl+yMlJYlZFCROWnUABYhLQ5CpJu27ztYUUNwxGZcMfH5R2BiiAtY+GZ9zdGhjTNimdAsrpqfoobpsXTk7ZjAc+hvVMePEEIY0yaLHzVWk7IzJgw/aWhfl9sbVj2llBp02ne/+tXt9fE8jxDSaUWOI38zPvW4gCGGMVJbdML97/XnGXTaSCtI4Ll6p/vNL34GZkmEkFGnDcsdHh1NXTejncUYa3JIHNPPF6YtBEeIx+dfvXmnXhuGLfUEh0qpyaBftXHHolVbIvXxBOcHaWpMiZrkPUKN1X3AtB/9RgQTjPGIAT2SrCZFidhiJEZQxgSBmzR6IIqa5QF+/u17DmzdvV9/zP38jWNgTNQIi9dupYxFYndp55UpRQit2rSjvLpOowlTRk8wliRlzODeYBM3q1hlDEGVndVkkBWl5UtkCGl4rqbBsXjNVhRZaAlR/SiMNao+oRlDcQNGqKnJafSNgDHkdvvieesjhoA0GiG6EYmhr4RfQke/Y2GEEIre9x03ZWmgeL/y4wvcQokE9ffeG86nEZLJoOHkrpKD3/y6DNxOep0YvYPiPxyhUgx6TvbJz734jJMdTk+Emn7VbNR/+9vysvIasB6NBl071M2mplkRTyQYS7Li9vjRMRb7LQU/zJ4tO0pKD3c4A3zHAWP8+uc/oxZ5QEDIAxQ9+w9Vfjhr3r8efH3KVQ+OvODOURfcOfmqB/714OuzflsuK2qr/TCgEVxOZmpuVlpAVo4msRljLElyTmbq8P4FKCpZRKOff+mGuMQX2gdKqU7U7Nh7cP/Bylg0pDagqTFPMGupGSDAefaUURxHRI3QrGJV4DmOkPwuWcMH9PD6AuGLoREiBAOFX6SCSZNRF/2ROEJcbq/D5Wn7E7YC3KhYYJ0IWl0EJjvECMHVdXbUIbVtGCFk0IliON3rz4MIVhTV7nTDgI4G4Lmpd7o4QiJeCiPGGKQ1/EU3pfDAiFIWuoiAQmrymEGnjB3scHvCLnPKqFYrvj9znj8gIYQMOi3EZP+C/tljgCbzLyQRmDF293XnJllNcjibDfojVNc7PvhmHkgQs1HfxncLi0jLR07pRwhhgiVZqa5rQCimRYQxJuH+tOO7N3eGw6Jaum5bICCZ9GHKok58qJSajPqla7cuWbt1/Ij+wYJXKD/jMN5QVPzm5z8vWLG51u7EGAtBu5/SFRu2fzb7jxH9e7zz5K1983OjsPOCCcLzXG6ntMJd+7FOROGqv2IBMMEN619gNuqj8wFD+s+CFZvEqH3iOxo8z9XbXas27eyWkxm9/DR2QLKYzx9YsWG7TiuGdTVBPeGrn8zWipqw2xx0ka6ud0TJ7ddrtWsLd9c1OJNt5rBF8EkWUxRjF8pD7C5PVa09PcXWQWX0NotRZVGa1iCe40rLq1G7OMVaAUYIIYvZoNeKLq+vZe+DpqOwqqrl1fXo6Or4UWOFJK2qtXORq6WBNCkzNSns7Y5jRJIhFGMiUSS07P0B/7j/posWrSkMy94PvTO27Cz5ZdHaC047SStqBEEIyApBrTiK/oHAuLlvjxCsqjSvc8Z1F5zy/PvfpiZZlBaxAJVSs0H31Zwlt15xZnqK1WTQRSmvDXtThJDZpNeKGq/Pz0VZRJQWH6hAsS0iSVZ8AUkbkJp5snmea2tldXPBD+ev2rzzmFHCdQRAKr/x2c/jR/SHTQHEf4PD/dTbX3/y3QKvP2Ay6FKs5kbuNXjSpuZIG4uKT7/+0d8+eqJ39xygcQh7F4jfZKYmHWWkEyNMKZswcgCK2pgHdIJd+w4Wtt3Pj3HcHaR4xYaiK86eGK89lzLKYbKxaO++skpDOAYe1MjSr874eXFkmw8jxHTaiNYqY0wU+MNVtSs27pg+eWSzhj3wLBmpSXxUigyOIw6XZ8e+sv4986I3UmoHYFJlpNioGpEMmDGm0fAlBytdHp/JoIuv8gEXSrKYbBaj3eURotZuHD2VHgy+pt5RVdsg8K3UieR2Sg37c60oRDsNh0ZS2wyVRtTAELQw5bl2u9/A+mw22SCPZ3j/ggtOO+mLH/9ItoZpkAb3/eCbeeefOtZs1Bt0otPtQcc82/cvgXDZQpgxdsfVZ8+cs7Te4WpJicMYEzWasvLqL39a9J/rzjUZ9O2om7CajEkWE/BhhD2VIcZx3NZdJSg2RpDMVFtBXieLyRAq+AnBtfVOV2MdUKw4Yr7CIqytd+woLtNGZi478aFSajbpF67avHrzToyRJCscR5au3XbyZfe88fnPgsCBVQfJfVDtAJUSqkoVRbVZTdV19lsfe1uSowXvQUezmAw0MslaLJBV1WY2jB3SG0Un7GMMIbRgxSaHK3xoKhTA5MpxBEoZ48u9yCjTisKm7fskWYlSbt62a8LTrdwckKQoLwFjZLMYk6ymCH+MSVZTdB81OFcbG/bgZr/BCKHOmSkGXcRuWo2HMbR6007UAbYmTCogcYp4DGMaQThUWbu9+ACLN3c9xpgxJmqE7IwUWVaiKB8Cz+/cW4aOrk0RDH7vgfK6BlcUfjTKmEbgu+VkoXDRDSiIig7gAGgT4NG9vgCO4AHCQPWqEUQhquYR7R7hczlBIbj/xgttFqOshklYUSk1GXSrNu1YtWmHSa/TaUVKWcLTHxawlkNfDtTop9jMd117rsvjCzuBKaV6rfjFj4tUlaYlW7m2WfyYMabTajqlJ8uKEqkAlVKmE4VN2/e5vT4StWM4jPD5e67d9us7K2e+tObbV+DPym9eWv3tK6eOG+LyhOnpHAVHHAqLcOvu0uo6e5QGIX8JAKP7a5/+hDHWCPy7X88988bHDxyuTkuyMIaUsKxsTZBlxWYxrty0Y96yDcDUFuVGR0leSwj2+6Ve3Tp3y8mE+G6kI+Hb/75iU6Qay2BnAWD48foC9XZ3TYMjEJC0mrYRUEQH5PeVHq4Cgy8u84QjnErp4jWFothKtUKrfSmijwdyxJZvKPL4/M20FlienTNS01NsUmSZp1Kq02qWr9/uD0gk5iK0GAFSrVf3ztHpcjlCfL7A/BWbMIq/cxfiLD26ZsuKGkmxgVa5O/cdrG1wwjbXvnvBiWsKd/sjK3wYA7OKMb9LFgqnbNksxii50+BQrW1wojZGJaBezu7ykEi5lhiplBoNOkgBa58WGLaEgxBMGeuWk3n9hafaneHZZjBGikrfnzkPYWQ26FWV/m0YDuKLsHMT3vC1508Z3Lub2+tvufFSxvQ6cfveA/NXbOqcmSIIXJucrLCIenXvLMtqlLaTWo1m/6GqVRt3MsZavT7PcRqBF0L+iIIAG37sAwM0t/gRQhuLigOy/NdtzAMAerV5yzZs2Vny6iezb3nsLa1G0GvFGD1+4Pz44fdVKNp6xgghSZaP5k0RTAKSNHpIH0JIsJy3JYDS7kB59eYd+/QhIfCgsCcEy4ridHtrGxx2l0fg+T75uVedO+nVB2/4/dOnZ7xyT3zlA8cRp8e7ZWcJamLdORpQSjFGu/Yd3LG3LEov2rigsRvHwar1W/egI3P7Mcaw2nt2zZYkOUqIR6cVd5UcXLSmECMU357IkDTaN7+LzWyUI2ucKqV6nfjzwjUBSQZWibhjcJ/uUbybEG4or65fW7grlj0rEmDPWrKmUBM5togxCUhy186ZndKT2ZFpDaAnpSZZOC6awYQRPlhR06aBwdVgQUViXcUIgzmIjq5TUViAOnXH1Wd36ZTul6SWu5CqUpNB99uyDRU19Z2zUsJWsiSAIlgmGGNGmVbU3H/jRZIcnvILOoV+MGueqBH0Wi1teyvO4f0LCIlKsYARpXTGL4txbDl6jB3xB2rP2zHzjhD8oHFv3rGP57i/RcsHxnPcxXc89+RbXydbzRjj2KsTwZu9dXdJAKy68KkZCCFUb3dF6V7a+o0Y43l+HLSciVLI10RpV9vgFASeNAp7IiuKw+2paXC4vf4ki2nCqAH333jRt2/cv3LWyyu/efn9p2679fIzRwzoMXpQ78zUpChWbFuBEWYUbdy+t/FfRwfY8f9YXeiIyg8fLxCMJVkGb3+zDwd6wMhBPWU1GmsvQghh/N7XcxGKM/EGwRgx1DkzpUfXbPAohD2MMabXidv2lM5esLpNEzu2MRCE0LD+BVazUVWjRbgppT8vXItxOyNdICz3Hihfv61YrwvD0Nw0HhwIyCMG9iCENFOzYDp3Sk/WaoRIegNjjOdJcWk5aot/Dpb8ocq62gZnxAxHjFVV7ZyZguKh/jZD0B195zVnu9zh3dECz9c2OL/5dVl2RqoSmfD4nwyMsRphYkABxfTJIyeNGuB0eVu+YYinLFtXtG1PaUaqrU1pIjDTRgzomWQxKZF1MjBQ5yxet734ACGk1XQxjFv8iX1MocML/g38zAFJ3r3/kEbgw2aT/rXAGOJ5rrrOLvB8kIoh1nMR4jmupt5Za3dFOgYsj0OVtRzHtU/uAz9/ZqptaL98FDVDGxwwf6wuFHheVanb46ttcLg9vmSr+ZSThjx155W/vv/4mu9enfP+Y4/ffvkZE0fkZadD8qqiqoqiCgLXo2t2QIqbI4cxJgjc9j2lCCHuqBPLYcktXLk5vt0OIwGqoRav2SorajM9A97PxJEDDDptFGmqUmo26hes3Dx3yXqOkPj2SVKoijE+eXi/6N8LXBcvvP+t1x8gbZze0QFcsN1yMvvk53r9ARzh+1LKTAb93KXry6vqohAiRQH41b76ZUljglXkw3iegzYWzQCKbKf0ZIvZoKgRKe5FjWZ3yUGX20tijkoA+8rmHfvcHl9UZRT3yMuO5YLhTwYe1ghDgqLia86bOrB3V7cvnDuaUr1OnL1gFaUU2I0SaBMYQhjjh265hOPCh/A5jrg83h/nr0pLtiiRI18tATMtt1PaoN7dvL4AjlKrxRGvL/DYGzNQvPN1og0v+DdYDwcraw9X1WuEiOrzXwuQghQbE1rzMzHGiqoCT0vLs2HP8nj9+w9Vie1lzyAY+wKB/j3yUmzmKP2YgdOjus7x29L1kiSLGmHMkN4P3Hzxj+8+vPrbV2a//fC9N1xw8oh+KTYza6IjhEA3JPeBTjigV9dIO2M7QBkTBaHkYKXD5TmaEC9cCmN8qLJ20459em3UrLq2DD7KwZQynajZXXJo6679qAW/B2OsX48u/Xvkef3hyQCawDSCcM8LH9U7XDzHxZHDDkZ+5sQR0ZUPSplBpy0qPvDEm18RQpS4RhxUSgnGp540JCDJkULvjDGNwFXVNrz79dyWLcxbBWWMEFzb4Px89h8mgz7Sk2KMfQGpW07mmMG9UQvlGN5VapK1c2aqJMvh/bWMiRqhrKJmQ1Fx7FEJgjHG+I9VW6L48yilWlHok5+LOqBPIwrJEXvgxovCdqCgwOBbcui3ZRuMet1fro3qcQcUUIwa1Ou8U8fanWFIfKlKjQbdr0vWlxysjNKqOyxgSp89ZZSsRGt7oarUajb8/Meaj779nec4WVHa9yxtQnPBv7e03OkOT2LcgYOIfzHyn2inTMIYlAatRkDhvLlw2d37Dx2qrA1LMxfbTbCsqGOG9EZRXYVw8X1l5dMnjZz1xv0rv3np90+eeuzfl00ePSjFZm7kHqYUdBGgIyQhmwR4gwb26hrJY9kOgMVfU+8oPVSF2v2SEUJN3vVl67bV1Nk1QhSzD8myApUXrfyhVKU0OgUs0FMuWLmp5eBVSnmOO/+0sX6/FLVbEtNpxX1llf964HWotIyXv50jhDE2sHe3Yf0KPF5/FOVDUdUki/GNz3/66uclAs/JSiSnZpsBwv6cU0bbLNHaUKkqtZgM7838bc/+QzzXtjcA9YrPv/9tWUWNNvIK4gjx+vxnThxh0GvVcCWOqkoJwf0KukhSxEgWNDL+dt6KGHVH0MIrqusXrSk06nWRikslRclItfXunoM6jEsA3NFnTxk1YWT/sO5oKF0+WF7TkZvo3xnAmvXAjRdZjHqlRWCLIcRzxOHy1tY7Bb5tFUzwsc6ZMqpzZmogQucwAKXMYjL89/mPFq8phM7v7XuW2PFnHT9otbtLDkXqD9ERgMnq9vr0WrFVsrxjDEWlVpPBbNKH/S0UcC9fv93j8+u1YvucvcCCN3pwbxR144Ctf+TAnqMG9YKfMIZUSjFGjeRNUePicOXe3XOsJoOiqvHq50EI8fq8ew+UD+jV9WjkDQzv9xWbophWQJvz+Yv35GSlht39Q0EZ4wi+54WPl63bZjTow9rilDGNICxcteXeGy5otpnCPy8+fdxrn8yGnINIMklVVZvZOGfxuqvvfeXDZ+7QCLyiqBxHjn75UEo5jrvugqlL128jGEcRp4who15306P/IwRffMbJjDFVZe0oK2fsiJRwQjClND83a+qYwd/NW5FkMYWd4RARc3q8tzz+zm8fPUEwicJ7EQpZUQWeW7hqy7tfz02K2uISmqZcde4kFGGNwJwZPbj3R9/+HunGlFKzUT/791V3X3deXueM6DRZCCGVUoHn3pv5W2VNfYrNEnZ4BGOfXxrUu7vVbIjxqdsHhhDB+OFbLjllw0ORFshfvQjrOIIQoqo0v0vWdRee8tKH36e04PNhDHEcRqjNKbTATJWSZLn0rAnPvTsrNTkMU1DTLRghGKvokjtf+OLFu6eMHUQZBU6zdj9XdDS/bvGBw8cmnx9jxHOc1xfweAOjB/XCGHl9gWOQ2BUjwBbPSk8y6LRhCVLAJFqxcXu7EyExxgFZzs5I6dejC4rBVQjTCEx7jBHPkSMN+8gnksby9M5ZqVGsorYCyNSKD5Sjo+BuY4xxhNhdnjVbdkViJSIE+wNyQZdOp508tE9+bv+eef16dInyZ0DPvL4FXc6cOCIQubEsBEe37tq/90B5Mzc1vOe0ZOs15091urzRvV+KqiZbzd/8uuzMGx4rLi3neQ5OpzF08WGMUUoVVW15JEQczpk6enDvblDjG+UiBGONwF//wOtPvvW1Sinkt8M8iT4AOB0mVVjiUoTQLZefEZ3LS6XUYjQsW7ft34+/QwiGzJLod5QVReC57cUHrrv/teiRaeikfM7U0T3ysimlYYUrrMRxw/qk2MyyEt7NA+k+dpfn/pc+xQhRFi0JH4ZXuGv/W1/+YjUbI9XawIeeNn4Y6oDMvlCAO3r04N7nnTKmIQK9fELqHw2Az+fOa87pnJkaCIQxzSGFvt1XvvmyMzqlJ7dq9AuCIMnKhbc/+/pnPxFMwPOnquG66h4xtnZEskMEPygX+w9V8TzX0QF+jiOUodoGR/fcrM9f/M/8T5/+8qX/ZqYm1Ttc/ImRokIwlmW5T/dcFK5dGyhoDpdn257SdpefgTzr3zPPZNBFCfCHAkjp26qZYYQopQLP9+zaOSDLkXK12gqGGMa45GAlOgrGeHh1azbvOlheE4l4ByoexwzpwxiTZIXSVppSK4pKKTt5eD+LyRDFlOQ5zu50/7FqC2qxcUN22y2Xn9EtN9Mb1eGPmmT/ig3bJ15x38sf/dDgcHMcIYRAFXhjfqWqgnwN/h20SajOwBg3C+yBLiJqhAdvujjQWi0GZYwQYtBrn3rr6ylXPTh/xSaMcXCeBG+qqKoS8neY1XAkxxFQVkIvC07mUYN6nTVphN3pjqKUK6qaZDV//N386+5/zeML/Kl5ABstY4wxylhj/AVjgedXbdox/cYn7E63GDmdCJRvq9n43+vPj7K1gXMiJytt5KBeHp+fRFDUIJL608LVj70xg+cIIVhR/xxjkL8LVkpZefUVd78oyQqJwE8FWnunjORp44aio6MwigXgjr7/postRr2sqons/fgCCihSkyx3XHO20x0mnnKUV85Mtf3n+vNaJV6jlGp4TuC5e57/6LTrHl68prBxIROMEVJDNpPgNgKCox050Y1PCNuQJMuVNfU8x3VcAyyCMccRu9ODEXrw5kuWffXCOVNHU8qmjB28eMZz08YPq6lzYIyPe7yKIUYIaXTCt1hnsFVt21N6uKpOo2mnkw3YzsF733EWA0OIUiYrqqKqA3t1VSmNV1sXxhDP82XxYIyfv2KTEtmBTxnjef7k4f2AtIAQTKKC5zmCcY+u2b275/j8UpRyfJ7n5q/YhFpUeYHcTbKYnrzjiuhRdoCiqmaT3ucP3P/Sp2MvvvvhVz9fW7gbGgVBfiUkXoT+HWPscHk3bCt+68s5Ey679+Nv58N1gtfkOKJSeuakEReeNrbe7oyuEINwTbaa12/dc+4tT029+sEPZ/0OOlnwpjzH8SF/h93tcFXdb0s33PP8Rydfes+hylrUIq+YIfTovy+1GPXRI4CqqiZbTV/+tGjyVff/vnxj04YFGXIYY0yaCCccLs/z7307/cYnahuc+ghNGQA8RxqcrjuuPju/SxalLKrbAyGELj9rAlWjcWiC7H/+vVm3PPZWVZ0dXgKMkWAMH4sQMm/ZxlOvfXj/wSp9ZAJHaNR09uRRqcmWVmNPRw9CCGW0R16na8+fane6+bh2xUwANRVQXHfBKQN65nl8/ji6vSH758ZLpk0YOaAhqgKNGjOdUZLVvGx90fQbnzjtuoc/nDWvuLS8sdHMEWuZ4zgiCFyD011db2+rwXwEV3+93VXX4OJbY8xuN3iO8wUCPr80bfywx2+/vF9BF9TEoq+qNCst+Ye3Hnr2nW+efW8WR4iuvYHzowfGWJaVtGTr6MG9UDip1kg3tmXX0bQyUinVacVRA3uieGcGMfan/4fjCCZY1AgIoaH98uPYvpMxxnNcVa3dH5Cauua07SkYQxxHApK8bP02nTY8YR/GWJblzLSkYf1bqXgMhUJVnuMmjOi/ZvNOkyH8B4KGPeu37S2vrs9KS2o2fpiTF5x20sKVmz/+bn5aik2Wo2XbqirlOC7FZi6vrnv+/e/e/OKXvOz0Xt06d8vN6pSebDbqeY5TKfX6AnUNzkOVtSWHKkvKKitq6gMByef2dsvN+r9LmquYkNT94r3Xry3cXV1n12nF6F50RVWNBh1CbNWmnUvXbUu2mQu6dOrZNbtLdnpqkkWnFTFCsqK6PL6q2oay8pqSQ5UHDlXVNDgoZYqiFpeW52SlsZBoP+xZBXnZD958yV3PvJ+eYo2SdqSoNNlq3r7nwHm3Pn3S0D5nTRo5YkCP7IwUg07LEGtwuPceKF+8ZuvsBav27D9sNhm0ohBV6nN2p+ekIX3uuuacSE7+ICAyctrJQwf0ytu576Bep40ksCllVrPxo1m/L1y15YJTTxo/sn9ep3SDQasoam29Y9OOfT8tXLNoTSFHiFGvjfS2MQImQdPNl5/BjlWLICgPu+vac2b9ttzh9EQndkygrQD/nF4r3nfjhZfd9aJep0Vxkj4YI8ywwJO3Hr15wuX3+gOyIPBRioAYQ6qqmg16xtiy9UWLVhcmWUx5nTO652ZmZ6TYLEZRo+E5IslKRXV98YHyHXvLausdpjbWdPBNN2MY4+p6h8vj46NyYLUPhBCEWK3d2TU74+FbL7nsrAmoKR0XEpE4jlDGMEL333TRkH4Ftz7+1sGKGpvFFCkbokOBMfb6pdGDe2emJVHGWmp/EFZcu2V3pLZLCMqBCG7Zeit4C0mWszNSICU4Lgpm8F5N+WWNnt6SgxWbd+xbt3XP+q17otfLtRUCT+odbrvLkyFq2jNgRjlMCneVFJeWR4qYQArVhJFdk63m2FOoYC+ePHrgyx//EJHXBTGNwFfX2Zeu23rxGeMhmf+IWxNMKXvp/n/t2HtwQ1GxzWKMPhuhllIjCDqbqFK6/1DVrpJDqkoZYmD5MoSgmIxgzHOcRsNrBMGgE90avrj0MKXNk/IIxiqlGam2j56988wbHoPMwehxJfi4RoMOYyTLysai4tWbd8LKginBEGMUogxY4HlRw1uMBoHnq+oaiooPTBo9sNl0hQDzv688c8XG7T8uWJ2aZI4q+1W9TosQW7a+6I9VhUaD1mo2akUNYszt9dc7XJKk6HViktUEHTIiXYcQ4pckq9nw3lP/FkWh1UAYxkhVmVbU3H39eZfd9YJRr4syxVVKbVZTTZ3jhQ++e+3TH81GvSgKqkrdXp/H6+c4zmzQoXABviB4nquqsz9w08X5uVlgikUZW7wQTD2546rp/3nuw9QIKYcJtBuwuM6dOmb8iHkrNm43Ry4xbSsIwSqlBXmd3n/6tgtvexbCcNENMLi1yaCHgpSiPaWbtu+llLFgPhVGGGOeI1qNph22elDwI4RQTZ3DL0kmgz4ShVb7wPOc2+NjDN1w0amP3HppWrIVxt9ym0MIKao6deygRV88d/Ojb/22dEOyzYwiE1x0EAjGsqKMH9EPQU+8I+UBBPidbl9R8QGtqGk2NthSGUM+f8AXkIKtslvewh+Q+3TPNRp0YXWLGEEZA6olCAWBsFcUtfjA4Y1Fe9ds3rVpx76Sg5V2pxtixjpteyR0WAC7gMfrr2twZrSrQS1rbMyzxecPGHRaGm4jAzaFk4f3QwjF3gIY3ueQvvldszPKKiJmDzDEMEbzlm285IzxLT8BxpgxajLovnr1ntOvf3RvWbnNbGy10gbEP0II3jYY8cHkxz//yRopYmRF5TiurLympt6RnmJt7nggRFXpSUP7fPjM7Vfd87JeK4LnIPoYQPwTgg16nRGj0DFg1PifoFdIpRSpCsZ42+5SFM5+hdPfeeLW/Qcrt+8ts5kNUV4CbdqwCMYqVR1OTwNzwYMY9VpiIBCqjDJ4Qoiqqoqifvr6XQV52TFKVshIOP/UsV/PWTp3yfpkqynKIFWVCgIPpbCyogQkGWMk8FyS1YRYNJGPEOI54nR7B/Xqds+/zqO0/Su3HSCEUMauv/DUT39YuK+sQtdi80ngKMEQ4wh5+JZLTrvu4fhy18JCPn388LcevfnGR/5nNupj4elrXMgY63WiQa9tGcaC9JR2GOpHuPpr6h1tIidqFWAz1dY7BvTq+uzdV08ePQg1+fYjReJ4jlNVmp2R8tO7jzz2xoyXPvxeFAVRIxxLbgqVUoNOO25YPxRuH6SMcRjv3FdWXlUX2sMQwoUBWfa4/BwhBXmdLjr9ZITQ029/bdBrm9k3IM+GDyxA4XSL6Ah+bI5wBGPU1Fi2uLR8/bY9Kzfu2FhUDNQ6CCFRI4gawWYxYoQpi2ZmtQOEYJ9fqmtwoXbxRXOEY4wtWl0YhTBKpdSg1500rC9qi18EzCO9Thw9pPfu7+brtJqwXZkoZXqtuHrzLrvTbTUbW+ousDg7Z6b+/P6j593ydNGe0hSbRVGUWB4WKutiOUzguTq7c19ZRXqKFWZX6AEQdDj/1LGU0hsefENWVL1OjMUTxhhiseWOAGXF7pJDNJyUhey5ZKvp2/89cMa/Hi05WGUzG6PTjAQTBXmeQ026GqWs1aAYz5GAJEuy+ukLd00ZM1hRaVs4RTBC6LUHb9i8Y1+Dw63TaqJsGkH9DFJGUFOBQ/QbcIQEJEWnFT94+jajXkcpxZE7ScYdGCOqMr1OvP/GCy//z4t6nYhimGAJxA7wb40d2uecqaNnzV2eZDHGkReL44iiqlefNwVh/O/H3+Z5XidqYnHbMIQYZfFtttKU3IcYQqimwcFY+zO0m18aY68vICvKPTdcsOTL5yePHgQkM60WGQf9mY/ffvlXr95jMugcLu8xI6TEGAcCUtfOGX0LgJCr+WhZYyujvV5/AArqwLXu9vhq7U6TQXf59Infv/XQim9euu//LsjJSvWH4z6jjIoaYVi/AhRzSjylDPI4cVOeFMboYEXN7Pmr/vPsBydd8t+TLr77mntf/XDWvB17yxBC0KlWrxMJbuTu7Yg+IrKi1DtdCLV5WlLKMEbFB8qLdu+PlEVFMPYHpG45mY0cKW33qU4ZMyiKHwI43Q5W1KzatBNF6LMCce4undJ/++iJU04aXFXbwHBzZ9VRgmDiD8jw1cIq7yD7L5w27vu3HkqymurtLp7j4mhrMsY0Gr6sorqq1o7C8YYGX8KcDx7vW5Bb0+AQeC4WBw8LQfQjMUICzzndPo1G+Ob1+847ZYzaNqnf2G8tJyvtsxf+gzGWJDmW00MTYqIDAquyon76wn8G9u6mqq2QAXQEYG8875SxJw/v54xA4J/A0QAjxBB64MaLTAZdfNkwUZNZe/W5k7998wGLSQ+kn8clk/2IedPgcMfLv4ExDshKz66d5374xFN3XmnQa8FlF6M3mBCMMFZVdfrkUX98/uyIAT2q6xxxoUaJ5da+gDSsX75W1ABJTjPAGDYWFUNepaKo9XaXPyCNHNjzzYdvXPXNyx8+c/up44boRI2q0orq+pZvFJIH01OsjWSfkT88ayqCgsMgDdnp9i5bt+3Jt74+9dqHR1/4n0vveuHNL37ZsqNEpRQa0ut1ImpqXxtTNfdRgDHW4HCjtpfyQyrf4tVbo2S6YoL9AWnUwJ4agY/eLaYl4K2OHdIno7HHbvjDIKln3vKNUS7FEUIpTUu2zn7nkUf/fZksqw6Xl+NIfLfdoj2lKLIWyHFEUemk0QMXf/ncaScPrW1w+iUZ6heO/taMIYHn6u2uRlaGcAoQ1yT753385LlTRlfXOVS1eaeDdoPnOIZQVZ1jQM+uCz59Ztr4YYqqtkO74ghRVXXcsL5fv3ovx3Euj0/guaN/QRhjUEoEgZv52r3TTh6qtmt4cQFDjBD80C2XROmdmEC7AU6+nt06X3Pe1OhVrO0DLORTxw1d8uXzZ0wcXmd3ev1S+1rrtgqMULBguNmvjriZw+WJl1wlBHu8/jFDeo8a1EuSFeBpadMVMEIcx6mq2j0367ePnrjp0mn1dldYV2R8gRFGCI8b3g+FNb9Y4w64c99BVaU19Q6zSX/9haf+9tETCz9/5v8umZaVngyUsZCr5XB5wqoOfknu0SUb2PXDRhOAejZo3COEdu47+OGs36+4+8WR5995+r8effJ/Xy3fsN3nl6xmQ4rVbNA3WvYg7Dvo5TR/EIQZYxBTaCvAWlqwclN0ImGM8ckj+qG2KxZQkpeeYh3SN9/nC0QinIfaimXri3x+KUprVwivEkIevvWS+Z88NWnUAIfT43B7gSP5KHV2hphOq9mz/zCK2j4OOHFzs9Jmv/3wh8/c3jkzpbbe4QtIUCh4lDoxx3GKShu9DhFeNShASRbTzNfve/XBG7Sips7uRAiF3VliAbw9hFC9w0UZu///Llz4+dN9C3JVtXmiZZseRFXpqeOGzP3wiYK87Oo6BzuKERKMeZ5TVLWqzj6wV9ffP3l62vhhUMHRvuEdPWD/GTes79mTR9md7uM4kr8rYFu767pzOqWnRGfdaR9gIXfJTv/uzQe/eOnunl2z6xxOp8cLpHYkZvM4LJrYQQjky7u9Pk+DIyDJzcfQeDQC6lx/FNrUNoExpNHwhbtKKKWCwLU7fMBxHKVUK2reeOSmAb263vvCJ15/wKDXdly2v6yoVrNhxIAeKJyfnzJKMIH+LkP7db/0zPHnTh2TlZ4Mv22sUyAEIaQyihBqcLpbvlKCsSwrg/p0Q03M8KgxkEMpY42EDRxGCNldno3bihev3bpiw/ade8vsLg/BWKsVDXqtyaCHSP+xyX6Audj4n8YMMegLwPn8gbZeDfjmahocW3aWmIx6TDCPW+xfGFHK0pKtIwf2RG1pqBoEJANOHjPw1yXrBJ7DEcj7DXrtgUNVW3buGzWoV8sQexBBPpxh/Qt+ef+xXxeve/fruSs2bvd4/TqtqBU1HCEMMXAdR+FSC75DjDFGmDImy4rH69+6e3+kVIMgONJY/HLlOZOmTx756Q8Lv/jxj+3FZYwxvVbUaHiCSdB1HV2dwk3/hxFWKfX5A367a/3WPTdeMi3KaoXCOYbQrVecefr4Yc+//+0P81fVNjgMOq0oaiBXOYrnHO4KNQ6UUX9A9vr8FqPh4tPH3XXtuf175qFGuuKjUu4hMjK0X/7iL597+u2Zn3y/oLbBaTRoRY0ALzzaCBFC+M8R+vyS1x/olJb83+vPu+Pqs7WipjFFqU3jIY3l180+K3yFdqg4cJUHbrpo3rKNiDGej9gdFCq/j0aQQAVKpEEe/fXhDWCoV4xw/VaNPTAeIKu6Hac3A7BPZqTYbrvqrHte+Dgz1SYr4Zt0Y4wp157pGlzIF04bd/aUUd/NW/n57D/Wb93T4HFrBEEUBYHnCQgOhqIsZVi/8D+GGNTlBmRZkhWCsc1iHNAzb9ywfueeMgYdma/GB89HCHn9gXjFGxhjGoE/VFHrcHltlmh7WauAvQbYFfrmd7n+wdeLSw8nWTuk0o8Q7PEFBvbqmtc5A4V1wmOMEGKMffHy3dMnjQRppFIKTpWWM8Dp9oYhgESMIwRa8cLpUBOPm4Rb6aGqlZt2LFy5eW3h7rLyGkmWRY1Gp9UkWU0IkqRiyJNqB2AnwsHZ1JRICOkFahO1GSQJMoR4wvnr7JW19rbeCNSdXxev37e7VGs1qeEyXDiO+F3eCeOGZGeksHYVPoDeNmFkf4xwRU09xiRsJgLPcb4Gx8w5S0cN6tWq0gvZ4xjj0ycMP33C8A1FxT/8vuqP1VuKSw83ePwYI0HgBZ5r0txR09pCrCngrVKmqoqiqLKiUko1gpCRaps6dvBFp4/T67SMtVIXHtQ/LCbD7VdN/7+LT1uwcvOPC1av3LTjUGWtJMmEEI3A8zwXwukcvGCjQqAoKvRrllUVMWTUa/M6Z1x73tQrzpkE7bmjDAAmiarSvM4Z7z7579uumv7Fj4t+WbR2/8FKWVEFgRcFnue5RvKeI6sYKKWyokiyoiiKqNF0y8k47eRhl501AfJpQG+OS+AcPpPZqH/+nmuvOnfyB7Pm/bpo3cHKGlVloobXCEKTZRVmhIqsSLIiK4qo0fTo2uncqaOvOndKp/Rk1F6lxOH2uOodsqI2Y//FmKiqog9IbQ3HQcpF7+45V583+eW3vtZawq8ghBDHcX6Hq6XBFzt8AclX76hFSFHDpHNyHOdvcB7N9VVKaxucqqpEEPyc3+5yur1RrkAprW1wyoocTvBzfoc7+ulhAWnpN1x82sffzS/auV+MkISEMZEkSasV25FXH1zIGkG49Mzxl545vnBnydylG5as3bqr5FBdgzMgyxhjnpDgboJJyGZCG+kmFZWqqgoSVitqkm3mvOz0AT27jhjQY2jf/G65mSGjbSH4YWsIBGSE4hM4glzl2gZn6eGqoxT8qJFYFCuqOmJgjz8+f+amR976+Y81UI0TX8oBgklAkof1K4Am6y31XHjzffJzITwfauI3HzP0H/L4SItYnKqoZpO+d7ecUJcmpXTb7tLFa7YuWLW5cGdJTb0TY6TTikaDjmA9yNo4GvdBywamQ2MygaLKqqoqahOBPOY5otEIokawmIwGvdZi1JuNerNRbzLoTAa9xWQgBDfSHLUlvRkOTrGZ77vjCoMuIkW/x+c/uTHmgtoxfUCAFXTp9Py919bWOyIRW2KCAwG5S3Y6is2vENT2CMZD++YP7ZuvKJcX7Tmwoah48469u0sOlVfX250eX0BSVJWqjc8GnJU8R3Ra0Wy0piVbunbO6N09Z2DvbgN65qUmWdr0aBCVoJRpRc2ZE0ecOXFEg8O9Zee+DduKt+0p3VdWUVPvdLg8kiQrKqWMIoYQRgQTniO8wJtNepvZmJWW1D03q3+PvEG9u/XOzwGWJ3glsQwAVl/v7jnP3n31AzdeuHrLrqXrtm0s2rv/UGW9w+XxSrKiwJfFGHEcJ/CcQafNSkvumpMxuE/3ccP6DutXAPkolFKEcJxTJsFgYKx395xXH7jhoZsvXrFh+5K127bs3FdWUWt3un1+WVZUKC3GGHGE8Dyv12rSM2x5nTOG9s0fP6LfiIE9oT8npPK1VSmB93j5WROH9ivQiprmyRMYMcoEgTfpdSGHx3ZljBlj9/zrfKNeSwiJtGlDlgzoVW3dgWGvG9Yv/767roy0SDHBPl9gQM+84PFteQSEEEqymh68+SIagVEUE+zzBwb07Bp2/PBvm8X0wE0XQmC02XtoevwuYU+POjbMGDPotG89dstvS9a3LMsKjkBVqcWk10AT17ZvUhxHGEPQNWpAr64DenW9/8YLy6vrdu07uGPfweLS8oMVNdV1dpfb5w80rmaEGCGcqOG1Go3RoEuxmdOTrTmdUrtmZ3bvktWlU1pasjX0FoqqEty8A25jJ3WgRjnn5qd+X77BYjTEhbiA47h6u+vzF/9z0enj2uEfC4tgUe+jr3/5wgffGXSiIPBxFIccRxoc7i9f/u/5p46NMmbYc5tMuvBgCGGETrnmoZUbdxgNuqDCSDD2+gO9u+es+e5VhJA/IG0oKl64cssfqzbv2HvQ5fEKAq/TilDFEIn/px1otG6a3MuqqsqKKskKaIscx+m1osVkSEkypyVbs9NTMtNs2RkpGSm21CRLss1sNuqNel0caQD+HgDXR7N54nR76+2uBqfb6fYGAjKkiPI8r9NqzEa9xai3mo1mo77V68QCmIoIH0EXzxhqcLjqHW6Hy+32+gOSDBknGoE36LWgutnMxhBJjxBCQD3bVp8fTNHQkTc43VW19qraBrvT4w9IlFJB4A06rc1izEixpaVY9VrxKG96lCOsqXdU19mr6xwOtycgyVSlsOhsZkNasi0jxWo06EJHCL6BjhthAicm2mdvtBu0MdQbJtjo80s+f0CSFUj05jgiagSdKEbakEEmRllZTRY/hqPVNmmd0YERooxuLz6AjqJ7WzNAaAQh9Pjtl/ctyL3tiXfdHr/JqIuX219RVJvFOKRvdxRVfQMPRPRLgRnt9QcwwaEeZoyxrKh5nTOWb9g+e/7KpeuKiksP+wOSKGp0oibFZoGmIUevzRwh6SmVFVWSZUi0FHjeZNRnpSfnZqZ16ZzePSezW05mdmZqVlqSzWyMhRYetBH4D/Cct2OElLJWmxS0++KhAPq86MdAHmVbr9xUAo4Yo5QxjDDHEZCsXVB6K6OiTa16juIZg1MxyO4ADwLFnNHPBaUhSCzYPtUcaKOCl+IIsZmNNrOxZ9fsSKeAixIe/BjkxoeOECHEcSQ1yZKaZOmTH/EUcIDFa4Rq1FaN+ChcHTDOVg+LvZyqJeBjddz1WaPciYZWF0j0avj2LW0ExAkx1EUdzRcMRdCh9OdaRhgTTDDWaTWRZHxw6SGQ9LGtZT76r48GDDGe4yBPuFU/MKS2xeRobSL4u+C0kwryOl1732tFew4k2+IQ8ieEeHz+gd0652aloaNrPAOqoqKqXl+gmatfpdSo1y5fXzR3yXqvP6DXinqtaNTrwJl/lDScwQ/PGJMVNSDJkiwzhrSikGIz52Sl9erWuXf3nB5ds/M6Z2SnJ+tCbK+QwYfMpGBWH268PuBoBhkEIThGJr6jREdLF4wRDiEZaNSLGv8behwKvst2b0aRx4BDLQUWHEToGBrdqX9+zVb11zbcvelSjamFwf/9edvG/ACCcZsYqzpmhI1v5ghSRdw44YPlBnFBxxUixXecYdHRH6t96Y3N0HEvoc3RnXig2VpGIanCTXS9fx7ZvlV8hOCP7/7IKBMFfl9ZhT8ga0UhepgfI4QJiZ34muc4RVUH9Oy64NOn/+/hN39csProQ/4E44AkD+7TnRBylLEJkJ0+v+SXpJZJKxhjvyRrRY1BpwWNMixbbYwICntKqSQrfkmSFVXg+WSrqXd+Tt/83IG9uvbJz+me2yk9xRp2nAyx4H4Hcyle8uCfiWDO/HEewxFpff+IW8eIOCqvCSRwDBCcrvGatU2CnyGEUXw7PjHGBIGvqK4/VFnTPTcrUrwEFIK9B8rLq+vHDesLPsBYliWQlidZTd+++cAjr33x4offGXQ6gW+dyTwKMEIj+vdARxGbCAkoYo2GD0gyDuc5gF5b7bbvcUjFkSQpkEemE8WsNFvv/NyhfbsP6ZvfJz83OyOl2Yng9MZNW19CxieQQAIJ/NMQtPgZQlgUBITiplQAb4bD5dmz/3D33CzIEo90sMVkmHr1Q7dfNf32q6ejkCS+6OCaqoqfuOOKnl2zb3/qfa8/YNBp2ydQFVU1GfUDe3dFbcxRR02mc1MyBQ5I8uI1W2fOWeJweaMT1LQJUNGBGJIUxe+XZEURRU2n9OQBPfNGDOw5ckDP3t07W83G0FMaJX3MsZ8EEkgggQT+3mjqzocQQkinDV+z0W5gjGVV3bandNr4YZFsaAhIpyZZenTNvuOxt7YXl77ywA1Ggy5GZzuUcKgqvfSsCfldOl197yslZZVJ1ta7qDUDwdgXkLrlZHbLyURtKcyA/ngcR8B03rnv4Pe/r/xp4eodxWWUMVNIenD70Bg/xlil1BeQ/AGJEJKeYh01sOeYIX3GDOnVr0eeLUTYN2ZONWWFJCR9AgkkkEACoQhx9SNk1GshkzBeV4cs36I9B1DUKk+gc+lbkLt03dYvflq8bc+B95+6rV+PLlA726oIxrix8dGw/gV/fP7sNfe9unDl5tQks9qWWjhMcECS++TnihqBxpZmGOTtQRz2eP2/r9j49S9Ll63f1uBw67SiyajHGLU7OR83FdkriuL1e2VZMei1BXmdRg/uPWFk/xH9e2SmJQUPhvT4xvzw45E5lUACCSSQwF8FQYufIYTMRn18+XCAv2/P/sOKovI8F70scmCvrqpK01NsRXtKT7nmoTceufH8U8cyxiiLiRoCGh9lpNp+fu+Ru5754L2vf7NZjEDAFMtQoeZtcJ9uCCHKWBSxD179YI+WfWUVX89Z+t28FbtLDiGETAZdis1CKW21Bib8MJoqLyVJ8foDlNJkm3lcv/yJIwdMHDWgX0EXTVPtdWjF0TFLj28Vf2aftncihfpaEhlYCSSQQAJxxxFZ/UkWU3y3WkaZqBEOVtYcrqrL7ZQWKbE/SIdn0Gv9Aclo0Adk+cr/vrR19/7Hb7scvNwxhfw5QinjOf7NR27K79LpwZc/FQRe1AixmN2UUY1GGNS7G4osb/4U+RxmjC1aXfjZ7D8Wrtxc2+DQ67RWswExFGym1yaEyHvZ4wsghDJTkyaPHXTq2CEnD+/XNScjeCQ8CzCKdHQlT1iElqoFEczH/DP7NH4TqSWzeiIrO4EEEkig3ThC8KfYzBjHjWwHNeX32Z2eXSWHQPCHze+DTbx7bmZmalJ5dR2INLNR/+w7s4p2H3jvqX+nJlnCEui2BCGYMaaq7LYrz+qek3nDQ284XN5WGX6AVCfFau7ZtTMKF5UI5upzHHa6vbPnr/r0hwXrtxUriho08dvh1ccYEUwQRpKseFx+hFB2RsrpE0acPn7YycP7ZaTa4DAw7gnGGIdpB9BxaCbjQ4v4w8pdWZYDgUAgEJAkKRAI+P1+WZYlSVIURZZltZHov5HMBGNMCOE4juM4QRB4ntdoNBqNRhRFrVYLfxFFkeO4hJhPIIEEEogjjujOl5pk4biItM/tA8ZYlpVtu/efctLg6Pl9Rr2uR172/kOVWo1GpZQxlJZsmbt0/eSrHvj42TuG9M2PPd0PiP2njR827+OnrvzvSzv2liVbTVHS/TBGkiR3KeiSnmJt1iUFPPaEEIRwWXn1lz8t/uqXJcWlhzUCb9TrgNqpHSY+IYRgLCuKy+dRVZqZmjRt3NDTJwyfOGpAkGkZtA3I0Ts2xn2obR3sDhn6NlRV9fl8TqfT4XDY7Xb4r8vlcrvdbrfb7/eD1JdlmYYgFrc/6AEAUAVA8Ot0OqPRaDQazWazzWazWCw2m81sNhsMBr1e3+plE0gggQQSaIYjuvOlJlu0Gk1bW0VFB0OM48jW3aUohvy+gb27/rJ4bVDSyIqaZDXtP1g57fpH33j4xotOHxf7fYHhp29B7vxPn772vld/W7ohNcmiUjXswxFMJFnpk5+LMQ6qF40RdEIQQlt37//42/k/zF9VUVNv1GmTrSbKUDui+MDXplLq8foCkpJsNZ128rDpk0dOGjWwU1NvX+CIJKSxMrCtt2gTgpIehyD4W4/HU19fX1tbW1VVVVNTU1dX19DQ4Ha7fT6fJElgwQfPApkNf8EY8zyP2ujwD3UwSJLk9/tZC8CVJUm6+OKLTz311BjTMBNIIIEEEggiyNXfaPEb9FqfX4LeX3G5AYT5d+4rk2VFEPhIYX5wOQzu071Z1buiqAadVqX0mvteXb9tz2O3XWbQaVFsEgUYflJs5h/efuiuZz5456tfkywmglFYzYYx1r9nF4QQQ41+dcgqWLVp5/sz5/66ZL3D7TUbdKk2i0qp0kavfmPyP0Y+v+T1BfQ6cVj/gjMnjjxz4vDuuVlwTIi871hJFhT2TX1Jm9QsWW5oaKiqqjp06NDhw4erq6vr6+vdbjfI+FDPPCFEFMVmn6BZWl/YVIAoCHoXmtjvGSGkWXSA4zie5wVBSE1NHT58OIprJkECCSSQwD8EQVc/QgglWU1JFtMBT3UcA8kMMY3AH6qoPVhZ27VzRqTEfqC361fQJclilGQl2FkZJFNAkgnGhypr/QEZBH+MgKY+HCFvPHxjXnb6Q69+rhPFlux+lFFRFHp3z4Wfg8j/Y9WWt2fM+WP1Fn9ANhv1KVZzOxL3oL2ErKoul4cylt+l07STh557ypgRA3o03Zoxyjpa3kcS9rW1tYcPHy4rKysrK6uoqLDb7T6fT1GUoIznOC7Uox4U5y0T7tqEZkEExpjaBPCjcBwniqLJZALfflJSUkpKSlJSks1mgx9qNBqUEPwJJJBAAm1H0NWPGUN6rZiZZttXVq4VNXGz+Fkjf9+ufQe7ds6IlN8HUYDsjJS87IzC3fsNOhGS6RRFbXC4BvXp/vhtl586bkg7BgDkuKpK77zmnC6d0m985M1m7H4YI1lRkyzm7rmZIPLnLd/41he/LFm7VaXUZNDrtKKqtlnkQ7aELxDw+gJJFtO08cMuOn3cqeOGBil9VJVC/B51DGluWGGvKEplZWVpaem+ffvKysqqq6s9Ho+qqmBMcxyn0+mCp6PGQoY4dD0O1TaCYl5RFPDVazQag8FgsVhSUlJSU1PT0tJSU1OTkpIglh/p6RJSP4EoaJawcnwHk0ACJxT+zOqnlHIcyclKl9Vt8d1Rgb9vy86SKPx9CGh6OdK/V976bXvMBh3ByO50m4y6R2+77K5rzgFWwfbldwfT/c6ZOrpTRsqV/32prLzGZjZAuh/GWJLkLvm5malJ85ZtfOOzn5au38YYMht1GGGVUlVtgw4EUXxFVe1OD0aoZ7fO504dff5pY6FeADU2IEeEdFR+fjAWHpS1qqpWVFTs27evuLi4tLS0trbW7/cjhMBtHjTo4cS4iHl0pKSnlMqyDGIeIQRi3mazpaenZ2RkZGRkpKenJyUlGY3GsBt0M+9ClLKCBBBqW+rJX6Vioh0PdVweLbon7K/ytlFrD9ISfy3VKpjYdGyShJrN3hjv2HLTa/XKUdBs7v0p+EEk5+dmMUrjSN6HEGKM8RxXuLMEReXAh+cc1q/g0+8WyAp1uNynjhv69F1X9S3IRTGz90cBpPsN718w/5OnL//Pi2u27Eq2mqBVtiDwlbUNp//r0SVrtzHGzEY9RkilFLWlwgFIBv0B2eP1W82GMyeOuOysCVPGDNLrRBTi0u8geR+sPgh+YLvdvnfv3p07d+7du7e6utrn82GMoXDOaDSiYyLpVVUlhOh0uuTk5LS0tKysrOzs7MzMzOTkZLPZ3HI2h243QQH/F9ouTwT8tbZgQKv+m3Y81KJFi0pKSmDCn3POOXq9/hh4iVqdq0GR06HDaHnTtt7xL7roYnzSoEXEHZNqqTbNXtgDg0VVzX54NFcOxZ+CH4R9j66duPg1lQFQxrQaYce+g5DXFunbgLd/SN/ukqwYdOKzd1913QWnIIRUlRISrXk5pZTF1vQa2P1yslLnfPDY9Q+8PnfpeqNeRykVeL623nmostZkaLTy2/SAHEcYQx6vT5KVrp0zzp4y+vLpE/rk58JvVRWqA+Lv0g915gdz4srKynbs2LFr164DBw44HA7GmCAIgiB0hLBHTTOPMaYoCkh6jLFOp0tNTc3IyOjcuXNOTk5GRkZycrIois3ODR3GcbTS/k7w+/3z58+XZbnV14gxDgQCw4YN69at2/GKm4RO4EgHYIwdDsfChQtj3NAlSRo7dmynTp0KCwvXrl2r1WpFUZw2bVpHF3/CUN1u95YtWxoaGpxOp9frVVUVIQSqdnp6evfu3Tt37oyOVaAK1ldb1xSMraioqKioSBTF6LIAY6yqqtFonDp16nHUOEPNnuhHwtNVVlZ++eWXdXV1kyZNmjx5cod+DrfbvWDBAriFLMtpaWnjx4+PPjxwRUDGFcdxJpPJYDBgjJuNU1XVOXPmuFyu6OoLrPThw4f36tUreIUQwY8RQii/SyejXns0nW3DPowg8OXVdSUHK/sW5EbM78MYIZSblXbnNefcesUZuZ3SYVuIYiIzxiB3DyEUo9MGEywrqtmon/XG/Wff9OTCVZtB9vM8MQl62hYrP9SrzxEypG/3K8+ZdM6U0UlWE2qqwu8gE7/ZRJckqaSkZNu2bTt37iwvL/f7/cHAOeoAYR8EY8zj8TDGNBqNxWLJzMzMycnJycnJzs5OSUmB/LvQg0OteYw7vH7hHwVY0oFAYN68eV6vN1R9D376UBlACHE6nampqcdR8AfHI0kSIQRKQFvC5XLNmTOnpQHU7CIIIUKI1+vNz8/v1KmTKIpGoxHIoI7B08E7LCsre/vtt4OZp6FJBrBM+vXrd+GFF6alRaQxjSOC68vv97csw4kEGNiuXbt++OEHs9kcnDyhG0jwyiDM0tPTJ0+efByXc+iTarWtZ3/PnDmzsLDQYDB89dVXeXl5HbQEQCQVFRXNnDnTYDCAdWQwGAYOHGi1WsPeEWNcWlq6cuXKPXv2OBwOWZYRQnq9vnPnzhMmTOjXr1/oWZTS5cuXV1dXC4LQUj8L1ardbndKSkokwd8odzNTk8oqakRNmGu1G5Dft233/r4FuZTRsMTyMACLyfDifdchhKCePsq3gAM4jDcWFVfV2qeNHxbdjcaaCHcJz63fuuf1z37atL1YJ2pgQjOGGGtDvIQjJCDLDU632ag/e/LI6y44ZdLogU0epCYTP95V+EGHD0x0SZKKi4sLCwt37NhRVVUly3Iz474jhH3oYERRHDp0aNeuXXNzczMyMpoZVcGZF+q077jxJBAWhBD4LrBHS5J0vEeEUJN0qa2t3bRpU2lp6f79+2+44Ya8vLwY1Xee57VaLVwECKNCr8xC8lLb3TWjfSCEGI1GMJQZY5AwK8symG6CIGzYsKGkpOTuu+/OysrqONkPVy4uLi4qKtq9e7dWq7399tuP8po6nY7neXi3Pp8vvl7howGldNu2bbt37965c2dubu7VV1/d6iwKDj70Lx30LbZu3Wo0Gg0Gg6IogiC4XK7t27ePGTOm2R0ZY36///vvv1++fLksy2DEwxQCN9LmzZvPPPPMc889N5ZoEWzOoBCADWCxWEIPOELwU8r0OjE/L2vvgXKdqAlPdtNeMMY2Fu295MzxrR5JKUVR+8kCbwzHkXqH69WPZ7838zeXx/fIrZfef+OFKJzpz0I49nfuO/jqJ7O/m7fC6wsYDbpY2v+EghBMMPH5Ax6fPzMt+cqzJ11z3pT+PfPgt41RiXib+KFRH4yxoih79+7dtGnT9u3bKysrVVUNkt12nHHfDIQQj8czaNCg6667rtk4UULSHz+Iojhp0iQwoIPO5w0bNhBCJEnKzc3t3r170A7w+/05OTktLwLfMS6fr+WUCP4cY7xhw4aPP/7YbDZHlyJGo/GUU04J/hNjXF9fv2XLFkEQAoFAQUFB586dgw8lSVJ6ejqKsDOyIzNIYhl8s5G3Clh9Xq932rRp48ePB33r4MGD8+fP379/v8ViqaurmzVr1h133BHppihEDrW6vwdfb8sjZ82aVVRUJAhCjx49Yh8/HJmfn3/GGWdotdpgvGDHjh0VFRU8zxNCxo0bF1S8VFU1mcI3eYlFRMWIsE8a9HJ9/PHHdrtdURT49K3i4osvVhSlrq5u4sSJYO5HCTYFb9TWiAlsknv27AHuE5PJ5HK5EEKFhYVjxoxpJvXBvb9hwwaQboIgWK1WnU7X0NBgt9vBYTB79uy0tLSxY8fSJnK50047zePxNIvOg7qwa9eu7du363Q6t9udn58/bNiw0Mc8wrcGtvjAXt1+/mNtXL5WyJWZRhC27t6PYgjGR1HWgiIcITTj58XPvjtr9/5DNrPRYjI88toXO/eWvfX4LSaDLjRro9ExwOFDlbWvf/bTF7P/aHC6rWajzRJT857QUWGMPF5/QJLzu3S6fPqEy86a2DkzBYFXHzGuAxL1m7n0Dxw4sGnTpi1bthw+fFhRFI1Go9VqwaMY6vw8NqCUdunSBSyqoFKSkPTHC/DmtVrteeedF/rz+vr69evXg5jv27fv9OnTW54buuKCX/MoxxOqrTa7ePCfEPw2GAyBQCD0KZo9lNVqvfjii0N/vnfv3nXr1omi6Pf7R4wYcdJJJ7UcQLNLBccTaTChCJ3SwdNbXjMSYBNPTk5OSUmBn2RkZPTt2/fJJ5+sra01Go379u2rq6tLTk4ONftCRxjqzg27rFqOMPSJ4IfAhIExblMKG5w7YMCAAQMGhP78s88+O3DggCAIhJCLLrooukc9XoojivCkoe8NY6zX6xVFCQQC8KRR7gu/yszM/O9//xsUE5GOb/Y+2zQHYIS7d++ur6/nOC4tLa1v376zZ8/W6XTFxcVOpxP03aB6Ryk1GAxjx46dP3/+2WefPWjQoJSUFFEUHQ7Hr7/+umjRIr1er9Ppfvvtt+HDh0MsieO4SZMmhb27oiirVq3iOI5SqtForrjiimbpGkcIfsjvG9o3X4h3fh9jTNQIxQfKq+sdaUmW9vlVgiJ88459j73x5bzlG/WimJpkURSVUpqabJn567J9Bys+e+E/3XOzFJUSjDAhHEdcHt87X815Z8bcw9V1FqM+2WpWVDX2Ij3QVNxen6yoA3t1vfb8qRdOG2c1G1CHefWbufQbGhq2bNmyfv36kpISv98PPPagcR97eR8coSAIubm5MMJEtP7EQdDfA6vM4/HAP8ErHlTUUJNRgo60LyEiWF1dnZKS0tIWbylNg7cLlfHBfd/pdNbW1kqSBPmeLWtHg8yMwZ+EVTtCI82EEChRQU2JS6EPFRxMs9NhPPX19S6Xy2AwpKSkBJ0iLd8hnF5fX+90OjHG0B4CHSlvWoWiKMEnUlVVp9P17t174cKFoigqiuLz+Zq9SRihJElQcKvValNTUwVBaHnf4IerqqpyOBwIIbPZnJqaGhTw8Lxw96BhEDw9Rm8HnIua3nloxqjH49FoNM2kb+jp8M/y8nJKaadOnYJjiOTGaNa7K/RXhBBKaWVlpcvlIoRYrdbU1NTQmdasLUiUWRR6O4QQx3HNEvuDjxwcSSAQqK6uVhTFZrNZrVbUxjlQWFiIMVYUpVOnTkOGDPnpp584jrPb7Tt37hwxYkTLFzhhwoTRo0dnZWUFr2C1Wi+77LKqqqrt27fr9fqamppDhw517doVnrGlc5dSynHc+++/f+DAAZvNZrfbL7300s6dOzfXuY98xRghNKBnXnKS2ecLxDG9nzEmCFxNnWPH3gNpw/tTxrg2uk0oYxxH3B7fix99//YXczx+f5LFxCgLtt1TFDUlybxl5/4pVz340bO3Txw1EH7+1S9LXvzgu6LiA2aDPsVqVlU1dioejhCGkNPjZYwN79/j/y4+7dxTxogaAXWYVz/UxKeU7ty5c82aNUVFRQ0NDUBmZzQaj6O8B8BUtlqt2dnZKFFSf4IhuLxhWwld7cF/hh7zv//9r7a2VpblK664olevXt9+++3atWsrKysvu+yy0047zeFwvPPOO9BucdCgQRdccEFwxyGELFmyZN68eWCyX3/99RCkBwFWVVX1448/7tq1y+/3q6oKxn1+fv6FF15osVjmz5+/YsUKSqlWq4XN95NPPtFqtV6vt3fv3pdddlmz7bWZUI/+UEHvdPAZ9Xp9aWnp7NmzS0tLJUniOC47O/vss8/u2bNn2H189erVS5curaysDAQCGGOtVpuXl3fGGWfEnoUQvHXoUgV9hVKq1+ttNlvoYRhju93+22+/bdu2zel0wjsxm83Dhg2bNm0amGtBuxNjvGHDhnnz5lVVVUHShkajsVqtJ5100tSpUyVJeuutt6CHFugNlZWVTz75JChMl1xySd++fVt9imZ+gmZyFD5B0EFNCNmwYcMPP/zA83xWVtaNN95YXFz87bfflpaW2my2xx9/XKvVLly4cNGiRQaDQZKkf/3rXyCKYKr4/f4333zT5XIpilJQUHD11VfDleG/S5YsWbx4MUxR+BZJSUlTp04dNWpUfX39u+++K8uyx+PBGIuiuGvXrqeeegohpKrq9ddfn52d3exJ4Z8//vjj+vXrtVotIeTmm29OTk5GCAUCgbfeesvpdEqSdM011+Tn5//4449r1qxxOp2UUlEUCwoKzjvvvIyMjFZlP2hmfr9/9+7dGo3G5/N16dIlKysrJSXF4XBgjAsLC0eMGNHshSOEYCRB/QNjDBXRQ4cO3bp1K2y8DQ0NoR8i9CKUUp7nV65cuW7dOqvV6nK5+vfvP3Xq1Jaf+0iLH2PGWGZacq+unVdt2mk08HGULgQTSZY3Fe0bP7x/my4bTOL7bemGh1/7vHDXfvDtt3TUK4pqMertLvc5Nz/1+kP/1zs/56FXPl+6bptW1DRx7LdN5DvcHoLJycP63XjZ6WeMHw6KUUeI/GYmPrhn165dW1ZWpqoqyHuE0DHOVIoEjLEsy9nZ2RB5Sgj+vy4wxnV1dWVlZZIkeTyemTNn/vzzz8nJyYIgwGdVFOXw4cOBQMDn80EpWihcLtehQ4dMJpPf7wcJBNP4wIEDr7zyisPhEEVRkiSYMG632+v1XnDBBQihurq6PXv2JCUlBfPFDh8+jBDyer1JSUnxejTUlOW0evXq77//3uFwQJI/Y2zPnj0vv/zy7bffHpSCMAxFUT755JOVK1cKgqAoCm4qOti8efO2bdv+7//+b8iQITHOefCKB5WSffv2FRYWGo1Gu90+evRog8EQvC/GeN++fW+//XZdXR2YoaBaSZL0ww8/7Nq167bbbgMeAniuuXPnzpw5E7QBqBiUZbm6uho+kKqqZWVlNTU1wdB7IBAoKSmBkoegByi+kGW5rKwMjJN9+/a98soriqIAHygM2+FwHDhwwGw2QwPP0HMZY5WVlXa7XZbl4NeH1/Lll1/OmzdPr9cHybxlWa6oqOjbty9qqmYCbwrof9BXDARkMH7UEvX19WVlZVAjpygK/JDn+erq6urqakmSDh8+vHz58oULF2q1Wo7jNBqNLMvr1q3bt2/fPffc06rsZ03JlTU1NTqdTq/Xd+/eHWPcs2fPZcuW6XS63bt3ezyelvsna1HdCr8NVcIiBW5g6dXW1n7zzTc6nQ4qCC6//PKw42xePwNd8kYN6rlk7VaC9XEUMhAF31hUjKK26TviFMYYYxxHqursj73+5Rc//kE4LjXJoipqpPC8oqqiRqMo6h1PvcfxnM8vJVlNjLH2ifwpYwbdcvmZU8cOhl+pKiXxDuQHRT58nj179qxcubKwsNBut4NLP+iyi+NNjx6qqhYUFKAEde5fH1qtlud5vV6/Zs2aXbt2GQwGv9/v8/kgVR5jDJKSUgrmYyhgT9RoNM1slBkzZrhcLpPJRAiZOnVqSkqK3W4vLCxMSUkBf2lSUlLXrl0RQuBLRwhlZmaCbRRjflYsgJXl8/lmzpwJuY2EkOrqasaY0Wj0+XwzZsx49NFHIWoGg58xY8ayZcusVquiKGPGjOnVq5fX612+fHl5eTkh5OOPP87Ozk5PT29134cb7d+/X1VVr9e7e/fu5cuXg4c/Pz9/+vTpQaMWIVRTU/Pmm296vV5RFFNSUsaMGZOSkrJ///5ly5aJorhz584ZM2bccMMNoCiUlpZCoZ0syzk5OcOHDyeEVFRUbNy4cciQIfAUoJS7XK6gPyAjIwNj7PP5OojPAPKNeJ6XZXnWrFmBQADefDD/n+d5URSbBQiCEAQBqi5hjoG/etOmTfPnz09KSgoEAr179x40aBCl9ODBg9u2bYP8A4g2QncxMIr0en1mZiZCCOylSKMNDgZcDsFPFlwLc+fOra2thRJlxtjhw4c5jrNarfX19bNmzbrttttieSeFhYWgveXl5cGoBg4cuHz5cp7nGxoadu3a1VKJDDupMMYHDx6E1yKKYlpaWtgjg6qS2+02mUx2u/26665LT08P691pLvghzD92aF+B/yG+/XkpZaKo2V58wBeQdKImUjV/EEDVhzH+9rflD7/6RcnBSqiPD/r2I98IUgE0jDGLUR97Bl+TyPcSgqeMGXTrFWdNGTMIhbAFdJzI9/v9mzZtWrFixd69eyVJ0mq1J4JLPxIYYxqNpkePHsd7IAnEBzDHtm/fLklSQUHBuHHjDAYDeB3BJI00FVv+ihBSVlZ24MABEDxXXnnlxIkT4VfTp093Op3w96lTp55yyikLFy78+uuvTSYTOFfz8vIgKIDiFz+C4LQgCFdeeeWQIUN4nt+7d++HH37ocrl0Ol1lZeX27duHDBkC9y0qKlq6dKnNZpMk6frrrx82bBhcZMSIEc8991xtba3X6/3999+vvPLK6IIfQhiLFy9evHgxa+pMAUr81KlTL7300qAfG/77/fffO51OjUaTnZ192223QT7B8OHDc3JyPvzwQ4vFsm7dusmTJ4OqtHHjRsge0Ov1t99+O/gCEULnnnsu2IJarRZKBp577rnS0lKEUOfOne+99174XnBM3JNywEUBRrOiKHq9fsKECfn5+WD0o3BTJRQsXDPPDRs2QA/utLS0O++8Mzhmt9sNQj05OfnBBx/0+XxPPvmk0+kMBAIDBw68/vrrQdRFST+KPp8RQvX19WlpaVdddVVBQQHGePPmzR9//LEsywaDAaobQCGINAdg1u3cuRPyT/v06QPDKCgoSE5OBu6TLVu2gKIWBXALn8+3adMmrVYbCAS6dOmSkZGBWiwQeOQ//vhj8+bNVqvV6XQOGzZs3LhxkWI6zQU/eLOH9c/PzkipqXdAI93og4sRjDGNwB+sqN1bWt6vRxfGKI5K38sRUl5V98DLn838dZlOK6QkmVsV+aGnw19iZCIihGCEnB4vQmjSqAG3XTW9uciPq1EbKvLr6+tXrly5evXqiooKUDmDBlYc7xhH4Cb+KagES5j7fw9ASLJfv37//ve/Q32J7fi+DocDCBwRQkGnLrivQaQhhDiOaxZFhpL3uBQUhAJjHAgEzjvvvDFjxiCEGGM9e/Y8++yzP/roI3CVl5SUDBkyBG66aNEijuM8Hs/kyZODUp8xZjAYpk6d+tFHH+n1+u3bt/t8Pp1OF132Q5oeuJExxlAFhxAqKSlZvnz5uHHjUNNmXVFRUVhYqNPpVFW9+uqrg68IITRq1KhFixYdOHBAVdUNGzYEfSRBbSwQCICFoKpqqCnfLFkdtzGx/yjBcdyNN94I3njUQpxHR2j2n9PphGGrqgrmEDxpUNdBTa26Q0+PyywihFx11VXAeIMQGjJkyN69e3/77TeLxeLxeEpLS6MIfvisJSUlVVVVWq1WEAQIJ8E36tatGxBK7t69u9WJBJ6POXPmVFdXWywWu90+ceLEYIZN8DCQJuXl5d9//z0k3EBKIIq8fltY/BhTSq0m4/ABPb79bblNFNrUoiY6gMZn8459/Xp0oYxFEvvQjOeXRWtve/Ldiup6m8UYmsQXX0CdiNvjUykdN6zfHVefDT0AO0jkB/NZMMaHDh1aunTp+vXr7XZ7kGUvmN16wgJ2tIKCAkEQ2pTolMCJDLAFzznnHI7joC8zaq9dmJ6eLggCbHM///yzy+UaO3YsuDpDCwpwCLEdalG/Hi+A+3fgwIGhxQ49evQwGAygnUCqFMdxTqdz//79Go0mEAh07twZvMcgYgkhBoMBYvZ2u726ujo3NzfKUEHbGDx4MCQDer3egwcPFhcXq6p68ODB999/v7q6+vzzz4chgSah1WqTk5Ohbjs4TkJIcnJySUkJz/Pg7EUIZWRkyLJsMpk8Hs+rr746bdq0AQMGBHeP47seIco+atSovn37BmcRarv6CJMhLS1t27ZtVqu1pqbmpZdemjZtWp8+fcDWD335YQ33ds+ioGFTUFAQWhPRq1ev33//nTXlK7R6ncLCQlVVZVnOzc0FjQ3exsiRI9etWycIQm1t7Z49ewYMGBBptJDguXnz5nnz5pnNZqfT2b9/f6gFaCb1wVD8/PPPA4GAwWBwOp3XXnutzWaLMh/CcGSCSD7lpMGz5i6Lb7ceYMNdt23PledMinJlyihPuOUbikoPVnXOTPFLcqQjjwYYY44Qt9cnycqoQT3vuOac6ZNGog4W+fAZiouLFy9evGXLFq/XG/Tqn7AmfksQQvr163e8R5FAPAFlGhA+BHO8fQqoqqppaWnjxo2bM2cOJK7PnTt3yZIlBQUFU6dODSUNPQaApwATMJi+B+nfQP6DQgoFa2pqPB4PJNZ8//33P/74Y/ANgDkkiiIYW0DDEv2+kiQNHDhw9OjRwR+WlJR88MEHDQ0NNpttzpw5vXr16tOnD2rKauR53uFwPP30082uA25zSL2EVzdmzJhly5ZVVFSYzeba2toPPvggNTV12LBhp556qslkOpavNywopXl5eSwkVbndO9uUKVPWr1/vdrsNBkNZWdlbb72VkZExcuTIKVOmBNMyOgKMMVEUgz4SeBadThcsc1OjZoxB9eb27dshQ9bj8bz33nvBSpNAIAC0SJTSLVu2NCNLCAKkfklJyYcffiiKIhjxV155ZXAah46WEDJnzpydO3dC/d7YsWOHDx8eXQsMI/ihbH3CyAFpyVafX+IIidJLt02gjIkaYcuOfdCBl0UofgedYNywfv/74he5Awx9EPk+f8DjCwzq3fWOq8++8PRxHCHHQOTv3Llz4cKF27Ztk2VZp9P95UQ+7ERJSUkQ4E/4+f8eAAGp0WiOPr4Ol7rwwgsxxn/88YeiKJBxXVhYWFhYeM4555x55pknpk/L6/UCKRZqilY0OwBWsc/ni2XBQuJOkGCAMda1a9eLL7749ddfh5yy5cuXg+D3er3BAgqo9Wp5U6iqgHdrNptvv/32jz/+ePfu3YIgQC7FL7/8sn79+ptvvhk4tY6j3Q96VbsVx6CriTGWmZl5++23f/rppwcPHtRoNHq9vra29ttvv92wYcMtt9ySlpYWWnYfX7R7isLLP3DgQEVFBSQq1tfXV1RUhIZdQGuBtM1AIBBaqxm8CMdxZWVlb7zxBsTIMMY33XRTampq2OrE/fv3//zzzyaTyefzpaWlXXTRRa3qf2EEP8aYUtYpPXnUoJ6/LFpnNRlUGk/Bv6+soqy8Ji87nVGGSfgkRoRQr26dzUa9oqoE4zjuEzzHBWS53uHqkZd96xVnXn3uZK2oQU25hB0n8rdt27Zw4cIdO3ZA8QkkQv+FRD4ANNahQ4cmCvn+gYhxN4Q468UXXzxs2LBly5YBCwVUr3377bedOnUaPHgwbGcdOtqWtlF0gJ8DIaQoysiRI1NSUliLUitQfIFfJZbJHyrGGGM5OTlms1mSJEiCA6suGMY2mUyTJ09udlkW0tUNNelVGRkZ99133+rVq9esWQMRBJvNBtb/Qw89FCTTjT42iCrCYcc9YNdytPCk3bt3f/jhh1euXLl27VqoSLTZbPv37//kk0/uvvvu4z7sSNi6dWsgENBoNJA0GhTtwa+JEBIEoaampri4uG/fvqHfK1i78frrr0OLB1VV//3vf0PooZmTHyEkSdLnn38OskaW5csvv9xkMrWq/IVvhwXcvWdOHPnjwjXx3NwZ4nmuweHesmNfXnZ69G49OZmpnTNT9+w/rBM1cTEROI6oKq21O7NSk+6+7rwbL5lmsxhRkCcgrnOIhZRjbtu27ffff9+5cydjDKyfEz+QHwUY48GDB6NEId8/A8E8KYyx2+1GR9KXRvJ5gkbbrVu3bt26ORyOuXPnLlq0CLq8rFixAuZPM0TK9z5msNlsED/2+/0DBw5sxq/SEm2a/MHXFeQGCCJYuY4xPv/882O5GmgMY8aMGTNmTElJyYwZM8rKyoxG4+HDh3fs2DFkyJCWpnDL1XqMpSYU9IPmBGEL+Dlshi1fCwCqISZOnDhx4sTt27d/+eWX9fX1FouluLi4tLS0W7duLU85vrMIgkFFRUVQ+p+amnrzzTeHRg04jtu6deuMGTOMRqOqqoWFhcEsSNQk9ffs2fO///0vEAiAP/iWW24J5kzQEApncPL/9NNPJSUlVqu1oaFh8uTJ/fv3Dx4ZOrDgWY3jDDt6kIKnjBucnZ4ciKG3d+zACFNK1xbujnYMRiqlgsD37p4bkOSwXoE2AaJNDU4PpfSmS09f9vWL9994oc1iVFUave1vO8CaqC4xxjt27HjllVdee+21nTt3arVaoOD4y1n5QQQtj969e6MTwEpIoEMBq16n00HCuUajOXjwIFCkgSsSIQQ9YFgLYnZYcYqiKIpisVguueSS3Nxcn8/H8zxU9MHkAb0B5hVw5B2fR0WIMZaWlpaamirLsiiK8+fPh4R8eARZlhVFgZUbi9aOm5hug9oMvJC1a9dCSxVFUTIyMuAdQgcdsP8WLlyImvQDuGlQtQruG+BQAWEpy3LXrl2vvPJK1CTw6urqgsMI0oAGownBh0UIbdmyZfbs2bNmzfrwww9rampQB4tMs9kMk4RSunv37uB4CCG7du0CZtJm2gk8KaUUnrRPnz4XXngh8EFRSoNPGnSdgm7abCoeS8AwDh8+fOjQIai+6927d1ZWVnoTMjIyUlNTR44cabFYYKbt2LEj2FLr/9s78+goqnyP31tVvXcn6Q4NhISGEAgQByKLIFFkQLbAoEREhwcEBRU3fIq47854HD3q8OIZh8eZwSUDqKOIIgFkxAUFyQtBE3ESs7BMQgiQ9Fa91/L++E3q1ITOBt2S5ff5g6Pd1VX3VqX7e3+/+1tA9cvKytavXw8v9uvX74knnhg7diwhBBJDlIU4HFxRUbFnzx6LxQLdqJcvX64+Uk2rexLd4qeUipLU35Z0bc64dz76hy3JEqvYfujWU/JjlSwTlmkzw+TfSRS/Gr51xxcXE2AI2/ken5/IZOG1Vz565+LLR2eQOFffo5RWVVUVFRWVl5eDlU96Qrh+h4CfPzs7G4JTUPh7PbIsazSalJSUhoYGs9nsdrvfe++9JUuWGAwGt9u9ffv2mpoayEdSPgJ/JNXV1SNHjlTyrEKhEDRQgWpipOULDv8N7N+/PyMjA9IBfsncMwBcslOmTNm8eXNycvLx48c3bty4ePFiu92uHCPL8pkzZ5KTkzv5lw8/BTAXt9u9f//+Tz/91Gg0gjxcffXVpCW9cPDgwadOnTIajVD1NicnB0INAJ/PFwqFbDYb6GJDQ0MwGExPT1dur8vlUiQTuq/C7TUYDBCT2NjYWFxcPGnSJEhkAJn58MMPa2pqGIYZOHAgCEZc9dLhcIDr22AwfP3111lZWVlZWbIsHz16dMuWLa2eOEzn+PHjWq120KBByg1XUh4opcpMdTqdTqfzeDw6na62traqqmrEiBHK1nj8ZtQWZWVlgUBAq9VCELSi6PCuJEkmk2n48OElJSVms7mxsbG2tnbUqFGwmP7uu+82bdrEMAxEpM6bN08UxYqKCqalVxCUiBgyZAil1O/3v/POO7DQkWV5/PjxNTU1MPFWX0nYDFIXnoou/IRAAD5Z8ptpW3Z8EUPBkmVZr9P8fKy+4WzToP7JchvuYnhxXNYwrVYjyRdoInMsGwiF/IHQlHGjH7vzpjlTJ5D4S/7Jkyd37doFRTYUx34Mr3UJgfivDl2gSHeDqujSu/D1nDx5cnFxMSHEaDR+++231dXVSUlJTU1NDQ0NNpsNEpTVZ2toaHjttdccDsewYcPAhi4vLz9z5oxOp+N5XsmPJ4SMGDEC3AlGo7GkpKSurk6SpEmTJl133XUdrizbmdEFzBp+K2fMmHHkyJGKioqkpKTS0tLKysr09HSr1UopDQQCZ86caWpqeuKJJ+x2ezv7XMqv8759+8rLywkhwWDwzJkzLpcLJuv1evPy8kD2YGl18803v/rqq/DTX1hY+PnnnzscDr1eLwiC1+utra297LLLbr/9dlidfP3117t3787Ozk5LS0tISIDy3lAoxmazgTcOGDlyJBQJZhjmrbfe2rt3bzAYXLVq1dChQ10uF8/zUEwmJydHp9N1/oa3M/Go78K9dTgcDofj2LFjUK7/9ddfdzgckN8oimJiYiK4fODXEnRr586dZWVlY8aMSU1NNZlMsHwxm80+ny81NRVyB2RZ5jguIyOjrq4O6gEXFBRAubo1a9ZYrdaoT6qtuSivdPWbohxACAE/vyAI/fr1y8jIgJWW8hH4j+zs7JKSElgKl5WVjRo1ClR/48aNoBpwB7Zs2aLujQSr6vT09GeeeYZSum3btoaGBovFIoqiXq/fuXPnxx9/fP6oGIbheX7OnDnLly/vWPhZlpFlMnXiZWNHph+tOmHU62JSyE+WZQ3HnXN6fvjnsUH9kyVJZtkoN5GhVJSkcVkZjhR7w1mntot1hP69ne90ZzgGPbjqhlsXzWJZRpJkeOviZ6FGaukE1dTUVFRUdODAgWAwaDQaIXyvp1v5ChDPnJmZqeTqXOoRIZ1CluVwOAz/KmXJ1UQiEfArQsyRGvjJnjBhwuTJkw8cOJCQkGAwGJqamqA3zIABA5YsWVJYWOj3++Ek8Kny8nKe50+ePFlbWwuvcBxHKXU6nddee21OTg7sdEqSNHDgwClTpuzatctqtWq1WlDHqBu3UScFtLWwhixqKHIXddYwZvU90Wq1d99994YNG44eParX64PBYHl5uSJFYGPV1ta2I/ySJEG1Wkrp6dOn6+rqlDtJKeV5PjExMT8/f/bs2cr+iCzLWVlZq1evfvvtt3me1+v1ygfhmHA4XFlZGQ6HwWKGnKCysrLvv/8eDoC2AjqdLj8/XwnskmV56tSpX3755blz58CzcuLECaihSwipr6/3er1QXmbixIkd3nBCiCAIEJYYtXkb1KZt64nAE7/pppteeeUVnufBCVpTUyNJUjAYXLhwIcMw27dv5zjO7/cTQliWhTrHgUDg8OHDJSUlykzD4XBSUtKKFSvU1X9nzZoF2yiwtDp27BjsNHU4F/qf2QfQjIqoCk+ppwBfk3A4HDW0BQZTX19fXV3NMIzH48nOzgYHj/rXEgaclZVlNBph4+zIkSPXX389uP2VkEDlhMqfmdxS6EL5jx9//FGWZXigctstg+UW1C+2bfETIkqiRsP9dv416176q9lokDpd7r59KKWCIBw88s/caROjJgrKLWl1TU4vpZR0JZmQYSgl1OXxmY36B1ctenDVDf2sCaQlaD8m41dQCkP6/f69e/fu27fP7XYbjUYIYO41kq8giuJVV10Fq/J2LC2kW6HRaAYNGsRxnMViAe9oK+x2O6SK2Wy2qI+VYZjbbrvNZrMVFxdDixeTyTRu3Lgbb7xx4MCBe/fu9fv98JtICJFlOTs7u7m5uaamxu12Q9CWTqcbMGAABKMpp4Xf3Jtvvhk2v+EMFovFYrF0OCmdTpeammo0Gs1ms3q/QA1UHTYYDEpCuXpGdrs9HA4Hg0HoHaCMJzExcd26dfv37//uu++gOx9pSVFLSkpyOBypqamkba+4yWQaMmQIVJpTttg1Go3RaOzfv/+IESPGjx+flJSkXjfAdSdNmjRkyBBI/HG73eDGZ1nWaDTa7XaI/4KflPnz5x86dOjUqVN+vx8E1WQyZWRkzJ07Ny0tTVa1Wk5MTLz33nu3bNly8uRJONJqtcKZwSccCoVGjx7tcDg6s5S3Wq2pqalwP8/fizEYDPBELBbL+U8EFiKZmZlr16794IMP6uvrwXthtVqvueaa3Nzcb7/9NjU1leM4xYHPcdzChQsPHz7c2NgYDAZh/BaLZdSoUXPnzlWvvWRZTktLu+eee/7+979DIwaNRpOYmNjOjJS5UErVhf/S0tL0ej0hBHIo1Gi12tTUVK1Wm5CQoC4dqKBsTyQmJlosFp7no66o4HHbbLbx48f//PPPBoMhHA43NDQMHTrUbrenpaUpm0GtlgtAOByGWliEkCFDhkCnjPa1BhSqVe+r9rItYSaN55yTFz3g9Qe4GHXpZRiG9/mvnnDZ7jd/f/7CGWLsCSF/+3jfk398x+3x6bSazjgbKKUsw/gCwUhEmPfrK56697+yR6WrTxhDFN++LMsHDhzYuXPnqVOnoMFDr3Hsq6EtfXifffbZDouVIt2NVht+XXpXjdfrPXv2rCiKNptN6R+q3lBUf1wQBJfLBeavyWRSV6I9H6fT6XK5CCEmkwla9nV+Um2NGSwzeBcyy9XvKuYg/c9ytuq/bZfL5fF4BEGAdPmkpKROOrpCoVAkEhFFEX4loIGN8m5Up7ryYiQSaW5uht42er3eYrFElRmfz+f1ekVRBJE7v56d+n+hvyIIP7Tse/XVVysqKiBifPz48Z0J2enw76TDJ6KMp6Ghwev1arXalJQUSHVTH9bq4x6Px+fzSZIErYchOT7qTEVRPHXqFLRmsFqtUe9bJ+dywZ9SnAH0Pztit3Ww4nI4348S19/YDsosgKH8wAv/+6e/7exnTeh8j7v2LkmJIEpmo+HQB38caLdKssy0LNwkSWZZpvGc65GXN2399CuTUafluM6oPssygiC6vb4xI9OfuPu3N8zOIS3b+TG/fcqXpKKiAjqOa7Va2CTrfVY+wDCM1+tduHBhXl4ehvX1TaLmELfjWmz1RxL1xahn7g60NSR1MtUFnJOcl1Wlpq1b1Or1qGNoa2BR1+hut/v55593u91Dhw597LHHfskguPPH087T79JMu+FfUXemg5U1+Nlvvzm3cPu+mKg+IUSWiZbjzja5jvxUkzttoizJhKXw2FiWfvpF8UMv/bX2ZIMtySJJcoeqD9LudPPWBPMza5b+9y3Xm40GSZZJrPP0iMq3f+7cue3btx86dEiSJKVJRmyv1a2A6BvoLIK2ft+EaalCQ9pVL/W7rSyktj7S+TN3lQ5t0LZQvOVEpVWdseFaXVehM59V3zf1uqrVPVGPrdWLUc+pnJC25GGWlpbW1dVpNJrc3FwItvjFvtStxtPOyEm0mbZzG5WD1c8rpmPvVXQg/AzDSJKUNdwx79eT3i/62ppoiYnCUUoFUTzQss0viCLHssFQ+On/+dufCndoNFxyUqe8CxzH+vzBcETImzXlmfuWjRqWRhTffkwfurLoFkVx7969u3fvdrlcJpOJ9qKg/baAoNCcnJzk5GRcVvdluqrKnT84tnp/AQNo5+MXcJKLv26HZ+jqg1CrrNlsXrp0aUpKClRS+uW/0V163Bc2U6R9OrGXRggh5L78BR//42CsRE6SZY1Gc+j7CkmSKaEsy/xQcWzNc28c/L4iOclCCOlQ9VmGESXpXLN79HDHM2uWqn378YvbLy8v/+ijj2pra6GzTi/27auRJMloNM6cOZPgIhpBejLw/VVnVCJ9k46Fn2UYSZImjsmc/+srtn12AAreXeRVZVky6LQ/VZ8853T3T07a+O6up9YX+vxBuy2xw/a7lFKWYdy8T6fVrLtt0SN3LE60mOKUqie3VN1qbm7+6KOPDh48SAiBhJleb+gDYO5PmzYtJSUFzX0E6QXILdVF8evcZ+lY+EmL0b925Q07v/w/KRYNe2SZcBzj9PBFX5b8UFFb8M7HtkSLxWzoUPU5lg1HIk4PP+2KX73w4IpJY0eS+MTtE1W0yBdffPHJJ59AoxGIHY35tbotYO7PnTuXoLmPIL2CVokMSB+ks80TQQVveeS1dz/9ypZoiVmgHyH+QNBiMp5fYaAVDKWUoU43b7clPrr6pruWzmcojXfc/smTJ99///0ff/yxF6fqtQOY+zNnzly2bBma+wiCIL2DTln8gCyTR+9YvPOL4lipPiGEEmIxGTsUVI5lA6FwIBjKm5Xz+7X5GY4UWYamxfEy9AVB2Llz5549e0KhEMTt9zXVJ4SIopiQkDBv3jxM3EcQBOk1dFY4GYaRZGlUxuBbF81yeXiOi5mnqH1BZRjKMEyTy2O3Jb750gNb1z+S4UgRRInS2AejKvWSqqqqXnzxxW3btsmyDI0u+kIQXyugaubMmTOV7iCXekQIgiBIDOisq58QIskyJeRss3vK4rVOD6/hulY//wLgWNYfDIUjwrLrpz/338tS7DZJkgklTBxECAz9cDj8ySeffPbZZ5IkQdnnmF+oR0ApjUQiycnJTz/9NBQdQ+FHEATpHXTB1Q976v2Tkx5bfdNdz/7Jbk2Moc+/9bUYSgg95/JkDk19Ye2K62deSeIWxAfmLBj6W7durampMZlMkLIf82v1FCil4XA4Ly8PHB64u48gCNJr6ILFTwiRZSLLkiTJc1Y+eej7SovJIMZh51sx9FcumvX8/cttSRZRkpj4VGYAVRNFcceOHbt27YJ20X1Z8klLU4exY8fef//9qPoIgiC9jK4JP2lRypLyn6/Nf1yv66AvUFcBdW9yeTPTU19cd8uCGZNJPA19Qgil9MSJE5s3b66srARDvw8G8bUCXCBPPfVUSkoK7u4jCIL0MrosqAzDiKI0cUzmmuULml1eTeyi/DiWDUUEl9d3642zvtz80oIZk0VJkmU5TqH7UNxxz549L730Uk1NDXQPQ9WHmL4FCxZAxR5UfQRBkF5Gly1+8u/C9cQfDE1f+nDlsXqjQX+RekkpZRja7OYdKfY/rLtl0dyrSTwNfai639TUtHnz5sOHD6Ohr8AwTCAQyMjIeOSRR7CwF4IgSK/kQoSftLTr/fbwT7mrnjLoWzdU7hIsy4QjAu8P3JQ79Q8PrUwdkBy/HX3FcV1SUrJ169bm5ua+U3K/k8iy/Pjjjw8ePBid/AiCIL2SCzTpWIYRRemqCVnrVi1qdnkuzOFPCeE41u31GfTaDc+vKXzlodQByaIosfHpDw2O63A4vGXLljfeeIPneZPJJIoiqj7AsqzP58vLyxs8eDA6+REEQXorF2jxE+j0IMuSJOWufPrgkX8mWIxdat4D3nWnh5991fj1T94xYmiqsu9+YeNpf6jg3v/Xv/711ltvVVdXm81mEq1tdp+FYRifz3f55Zffd999cK8u9YgQBEGQuHDhwk9aIvxrTjRcs/ThYCjMsWwnz8ZxLO8LsCzz6OqbHrr9RqgQEI8dfaJy73/zzTfvvvtuIBAwGo19PGGvFVCux2KxPPnkk1arFZ38CIIgvZiL0lqI8M8YklLw1J28P9iZgnqUUpZlzjV7MtNTP9343CN3LKaESFJcQvdJi3s/FAq9/fbbf/nLXzBNvy0kSVq5cqXVakUnP4IgSO/mYuWWZRlBFBfNuWrdqkVnnR1s9sPBTje/avHsfYV/uGpCliiKENJ/kcOICjgk6uvrX3755X379pnNZozePx+WZb1eb15e3mWXXYblehAEQXo9F+XqB2RZliSZUnLDPS/s/rrElmQRhCgmNcexbo8vKdH80kO3Lrt+BmlJDbjIq7c1JEIIpbS4uLiwsNDv96N7Pyqg+lOmTFm9ejWqPoIgSF8gBsJPoH8PJU1O78z8x2pOnjab9OpAP4ZSQmmT0zP1il+98dw9o4alxS9hj7QY+rIsf/jhh0VFRVqtluM4NPTPB0rzpqenP/zww1qtlmAnHgRBkD5AbISftJjvP1WfnLXiiUAopNVwkiQTQliWCYUjoVBkTf51z9+/XKvhBFHk2JjV+2sFqL7X6920aVNpaSlG77cFtCJMSEh49NFH+/XrhwF9CIIgfYSYCT8hRBAljmX2Hfwh7+7faTUalqUMw7g8fH9b0von71g4K4e0CHOsrtgKOPmxY8c2btx4+vRps9mM7v2oUEpFUWQY5qGHHkpPT0cnP4IgSN8hlsJPCAFr/t1Pv1r56B8tZqPT7Z02eeyG5+/NcKSIosQw8XLvK5v6Bw8eLCwsjEQier0eVT8qlFJZliORyL333pudnY2qjyAI0qfgYnw6lhVE8be/meb08Gt/t+GBW/N+v3aFhmPjl6ZPVJn627Zt27Fjh16v1+l0qPpRgRsVDAZvv/12VH0EQZA+SIwtfkCSZYbSIz/VjMvKIL+Iez8YDG7atOnQoUMWiwXq9MXpcj0aUH2/35+fnz99+nRRFNm4BVsgCIIg3ZO4CD9pscIlWabxjBUH1T979uyGDRtqamosFgsa+m2hqP7SpUtnzpyJtj6CIEjfJF7CT+Js6Cvnr66u3rBhg9PpxEz9doB9/UAggKqPIAjSx4mj8McVkK7i4uI333xTFEWtVouZ+m0Bqh8MBpctWzZjxgxUfQRBkL5MzxN+JYB/9+7d77//vk6nY1kWVb8tGIYRBEEUxVtvvTUnJwdVH0EQpI/Tw4RfCeB/7733ioqKsD5P+zAMEwqFNBrN6tWrMYYfQRAEITFP54sroPqCILz55pvffPMNBvC3D8uyfr/farXeddddGRkZqPoIgiAI6UHCD7rl9/v//Oc/l5WVJSQkYChfO0D3nYyMjDvvvNNut2PmHoIgCAL0DFc/qL7L5Xr99dcxba99YCuE5/kpU6bccsster0ebX0EQRBEoQcIP+hWY2NjQUFBQ0ODyWRC1W8LlmXD4bAoinl5efPnzyfxT6pEEARBehbdXfhBt+rq6tavX4/J+u3DsizP88nJyStWrBgzZoyS/nCpx4UgCIJ0I7q18Cvd9goKCnie1+l0mLYXFYZhJEny+Xzjxo3Lz8+32Wxo6CMIgiBR6b7CD9JVVVVVUFAQCoWwRE9bQPS+Vqu97rrrcnNzCbr3EQRBkLbppsIP0lVZWVlQUBCJRFD1o8IwjCiKfr9/5MiRS5YsSU9PR/c+giAI0j7dUfghX//nn39ev369KIoajQZVvxWUUkqp3+/X6/Xz5s3Lzc2F8oVo6CMIgiDt0+2EH1S/rKzsjTfeYBgGdq8v9aC6EZRSqMcXiUSys7MXLVo0ePBgoqppiCAIgiDt0E0L+Ph8PpPJ1NzcrNfrsRS/Avj2eZ4fNGjQggULcnJySMu2CKo+giAI0hm6ncVPWoxXn89XVFT01Vdf8TxvNBr7uOkP0/f7/QkJCdOnT58zZ47RaMQdfQRBEKSrdEfhJ6q49NOnTxcVFRUXF4dCIYPB0AflH6YcCAQMBsOVV145d+7c/v37EwzdRxAEQS6Ibir8hBBowAPadvz48c8++6y0tDQUCoHzv9e354HwPVEUQfInTpw4e/Zs2M6XJAnevdRjRBAEQXoe3Vf4gVby//nnn5eWlvp8Pr1ez3Fcr5R/EPVIJBIKhSwWy4QJE2bMmOFwOAhKPoIgCHLRdHfhB9Sb2adOndq/f39xcXFTUxPHcTqdjlLaC1YAoOiyLIdCIUEQ+vfvP2nSpKlTpw4YMICg5CMIgiAxomcIP6C2/j0eT0lJycGDB48fPx6JRHQ6nUajUY651CPtAoqcg4mv0WiGDRuWk5MzYcIEs9lMUPIRBEGQmNKThB9Qy78sy5WVlYcOHSovL29qaiKE6HQ6juNIt18BKFouCEIoFKKU9uvXb+zYsZMnTx4xYgQcg5KPIAiCxJyeJ/yAWv4JITzPHz16tLS0tKqqyul0EkK0Wi3HcaCa3WQRoKi4LMuRSCQSiRBCbDZbZmbmhAkTRo8ebTKZ4EiUfARBECRO9FThV4DsPmUF4Ha7KyoqysvLq6urz507F4lEOI7TaDQsy16SRYBa7EVRjEQiUITYbrcPHz58zJgxo0aNslgsUeeCIAiCIDGnxws/AHKutpKDweCJEycqKiqqq6vr6+vdbrcoipRSWASwLKv+rPLvxQCXVgYgy7IkSYIgCIIgyzLHcYmJiampqZmZmSNHjnQ4HDqdDo4EvUcTH0EQBPkF6CXCr6AY9Gq72e1219XVHT9+/MSJEw0NDU6nMxAICIJAKWVZlmEY+PeCpRcuKkmSKIqiKIKQsyxrNBqtVuugQYOGDBkydOjQtLQ0xbgnqPcIgiDIpaC3Cb+CsgJopayCIDQ3N58+fbqxsfH06dNnz551uVw8zwcCAfDDn/9B9ccV94D6GJZlNRqNwWCwWCxJSUl2u33gwIEDBw4cMGCAzWZr5V2IOioEQRAE+WXotcKvRr2vf/4OuiiKPp/P6/V6PB6e571er8/nCwQCoVAoHA4LggCrAUIIaDzHcVqtVq/XGwwGk8lkNpstFktCQoLZbDaZTGqZB5Qawyj2CIIgyCWnTwh/K9TrgJiLcVxPjiAIgiAXSV8U/vNpP77vfFd/W8egzCMIgiDdHBR+BEEQBOlDYMo4giAIgvQh/h97VRzHj2RlzgAAAABJRU5ErkJggg=="
_ENG_LOGO_HUB = "iVBORw0KGgoAAAANSUhEUgAAAcwAAABcCAYAAADjyoVxAABtkUlEQVR4nO2ddZxc1dnHv+dcG9vZWY27EoMkSAjuEty9QJEalBZqlLb0hboXKjjFSinF3V2SYBHirpuszo5fOe8fd2ZWk2xCEkI7v89nyDJy77nHnvPY7xG0QQIegB7tv5eQfFkg9wRvAmBQQgkllFBCCf+9sEHOUngzlccdTnz1jPz7RdkoOrzRq1fYzJp/AHU+QgZAgVKfR8NLKKGEEkooYedCCECA8jIg7s1ZuW9RV5ckLyNF4Q+rfOBQJdTjCDEO5QG4/i+Rn1vjSyihhBJKKGHnwQMUoCEkKDVHKHFitmXlUkD6ArGqKmy4welCyNEozwZ02rTPEkoooYQSSvhfggIchDSU8ubbWnpvGhqSEvBMN/DHdsLSoCQsSyihhBJK+N+FAAyUZwshR5tu4I+AJ/Ro/72k5C18rbJkfi2hhBJKKKGENniA43nsL6XkUoQw8VXQEkoooYQSSiihDQohTCm5VKLYG19YlsywJZRQQgkllNARAlAo9hZmrH8hyKeEEkoooYQSSugejqQkLEsooYQSSihhS9BLQT4llFBCCSWU0AOUBGYJJZRQQgkl9AAlgVlCCSWUUEIJPUBJYJZQQgkllFBCD1AK+PkCQUrJ1mUAqSJ3vtpGEn0hBEJsbcZR4b5qB3L3C6Tccrs8z9tRDSj2TVsfdX5YAUrhKYVS3ib7Ytv6eFug8LyOjRBSbnI2KaW2ed4Ury8EUspN9A+AQCkPz+v5vbZ3f21ubD4L/PX62e+5pefd3BwXQiCF2GSSfU/Xh8zPk25mOKqbedW5DW3j34NxU6p4n886/7Y3/isEZnFRAio/LsrzOgxu28bmF2ApbAa72oBsCsrzyGSzbHW6rBBomoama2j5PnI9r2fPLcDO5fBcd2tuCEIgNYmm6WiaREqB56ntKLwESrlkEplNf0UBUmIGg2xvTg4pJVIIbMfGzmQhl8tX9ek8NgqEBNNEDwQwDA08hdupH7LZLHhqJ2RCCwzTpP3em02nN12RSGqY5tZzmoj8+CvPI5vNorI5cBy6f0AFuoEIWFimiRRiC/NTkc3mtl9/KZCGga5tX2Ob8jwymU3PT2mY6LrcctcKgW3n8JxNrUGBYVl0K0+F8Ndvzu52aiIEejCA7MHhI5NKgttNnytA1zED1ibnkW3n8FyPHg+YKMwhDU1qSE0iBdt5D9k2CDPW/4shMTpBSomUEtdzyWVtVDYLrusPmqaBYSA1CUL4ney4YNv+51KAYaJbJobhl/r0eipEPgco5WEEQvSprcavJNPDiacUtmOTaE0Qj8dxkmlAIMIhAgETz3U3e8pVnkdVTW+qYuEe943y8vdMJGhpiZNOJMF2wApghUMIPuukFyjPRQ+E2X3sKKRQXfccBUIKnHSCT+YuxBOb1qK2BoVDWTqRANshUFnFiBEjGLPbSIYMGkDv2moi4SASRSqVoq6ujuXLV/DpvAUsWLSURGMLmCZmOIwgr3Wi0bdvb4Kmnu+XHSE1fcGtXJs1a9djeyqvGQgGDBhAyNQ69qHyN6xUMs6qtRsQm9GUOkKg65JcNovTmgQrQJ+Bg9ht9Ah2GzmM/v36UFEexTR0HDtHc3Mzq1evYf6ChXw6fyGrV60Dx0VEIliG3v08EZK+fftgGRLlfTa+FYVC1yT1GzbQ3JraiufcwnWVhxUIM6h/765amfIrSG2sW09jPImUcrMyU3kelTW9qYlF/DXY7nEF4LkOy1euLo5p8TMhcG2bmr79GTqgt9+XBcGYF5bCs/l03nwSaQdNE93uBQJwlWDUbqOJhS1fkyzcSCmElCSaG5m3cBlC0zs/LcpTVNbUEA0Heji//f3BztkkkwkSiSR2OuPv35aFFQohJbju5yM4v1ACs6BJesojm0xDNgPBIAP69WXkyOGMGjmC4UOHUFtbTWVVFaFgECkltm0Tb22lob6e1WvWsXDRYuYtWMSSpcuI1zeAp5DhEJZlodTnf4ppDyk1Mq3NjDnoRN5/8jY0z0H1UAAopbDtHK3xOPX19cydO5c33niT5158nZUr1qKVlWPqsovGA6BrklRjnJ/97R5+cNlx2I6DrmlbvKenFHYuR2trK42NjSxdvJjpM2bywouv8u4Hs0CzCIYCuFultbZB03XSDfVMu+Qanrr1J5v/st3KgVMP4q05a7FC5mbNRlu+r+YLSk8wad/9OefMUzj68AMZPXwQ2pYGw3NYsngRL7/8Cv/+zxO89s6HeLqFqSkyIsrrb7/E1BG9sB0XrQdm5q2FUh5C6thNq5g09QgW1mUIGpB2g7zyxsscPGEAjusV7+26Lrqu8/oT93PIKV/DqqhEbWG8pJQozyXbHCc2YAgnn3wCp514DHtPmkB1LLLFNjbVb2DmzA948smnefSpF1i9vgkrHGw7qAkBbg4v1Ju333iZSUOiOI7XI7P8puC4Hpapc+3XL+TXtz5NqLIcZxvnZQFSamQSccbscxQfvP4AlvLwEMX16noeuqbxnUvO4nd3v0CoYtP31HWNVH0T1/zxDn595ak4joOWX4NKeUip0bpxMXvsdSTLGm3/EJHvL03XSdfXc+F1P+OuG670FYXOmmSugSl7H8T7CxsIBI1u1odACo+MbfLSm69x2B4Du23n+88/zL7HXYxZ3nGe6Jog1Zzh5n8+xKUnTiWbs3u2h3guufwe0tzczOLFS/jow4949fW3eO+DWbiuIFgW3uY95LPgC2GSFfgbVs52yDa3gGUycY/xHH3U4Rx6yIGM3W03amurut24NnUGTWZyrFy5iukzPuC5F17h5dffZOPK1WAaBCJhUDvW/9Vj5BsvhEY4YCIwt/ICQcrLy+k/YAB7TJzIueedR8P61fzjzjv4+e9vpSFpEwx3J8BE/r6+/0nTtM36ZAqQgB4MEgwGqa2tZfTo0Rx73HH85MfX8uIzT/LDH/2MmXNXEiiPbKWpNw/PBaOM8885FQDHdbtdhI7joBtlnHPK0bw582ZkJIDnbcP9hECTgnR9I8Mm78tPf/Q9zjzhUPR2k8p13fxG1XWmCQGapjNs5G4MG7kbl331a7z78jOc/+VvsrwhgTQFpmmi6zqaru9Qq6wWMIvXL/6bd1Vomuxgmiu4L3p0XU0nm2rF00Jc+p0f8L0rL2FY/5ri50p5uO6mDyuaJqmoruWIo4/hiKOP4cfXreSGH/2Qm+59BitS1mXcDNNE1w30z7h76Xljk74D/McKkEIghIbW4f2t978K2kyUbWvQL3S8pQNDwXLmuR6anm9JXqgq1+mxH7Uwxz2l0PJtcF3XV2C2sE+aZsCf44bRI/MvQDgcpqKigoEDBzJhwgROOeVk8Bzeef0Vbrjxlzz3+kcEY+U7XWju8gJT0zRc1yXV0ERZVSUnXnQeF5x7JvtNnUIo4AuPjO2RSueKjnR/TDou+MLJq/C5pmmMGDGM3UYO40vnnsGqtXU88eQz3HPfg0x/fwZISTAS2YVMtQpXKTTl9VjDLP6yGIDjP0dV7/58+9qfcOzRh3PGOZcye0UDgVBgExPfD9zxPIUUW9cP7X3EUjM58vhT2X/qFE4/7SyeeWc+wbLQVplWhJRkUwn6T9iPo6eOBeUVF29n+BuJ4oRTTua6395Bs+2hi630xgmBhiLdlODsb1zNTb/6IVUhHZTCdT1fqEjZ7tRf8Iu3HeaFEG3vex5C19n3sKMY1DvKknUtaFbbpuZ6asdpmEJusq8V/viK/L39v3t2bU3XSMeb6T1kLLfc+hdOOHB3gPxG5m/oQkh0veN8KBwvCsKjYNlxXJfqPgM5+qC9+PMdjyJEtMs9i0LAU59Jw3RdD6Ft3iT6WeAp3+yr2mmY29rmwhgVfqqUP149CkXoTkgJ0a5VPYMoBIBs6dqdoDxf2Lqui+jmcNvt/truXm17iM7UQ47k2QP344pLL+Hme54jECvftoP3NmKXTSsRQviLsaUFPMVlX7uUd994nnvv+AtHHHogCkFLIk08mcVxHKT0tSBd1/wgF00W/Zwyv6m1/xwgnc4ST2aIJzPU1tby9csv4rWXn+ZfD/6DvfecTLqxCcd22k5mnzM6R2W2fxU2EE+1e3kenuchBB36wDfV2oyetB/PPHE/g6sC2Dlnk6c/IQqv7u/b/lW4p79Ztt1TAI5tE6rqx3333srgmhC5nNtjLQZAkxIvY3PiiSdSbgpcT21ywUqp4SnoO2oiR0zdHTeZQtuqwA6BJiAdT3H1L/7AAzf9hKqQ5pvP2mncAl84+P1ceOa2ede+j/zvK7x0HNvximqe53m4rt9vbncv148i3dL4b+ka3Zne/SdtP8b5Vw96SGoamZYWBo+fyssvPMEJB+6O4zi+FpJfg37bfEFY6B8pJVq7/hGCoo9N1zSU8kgk011NiIX2bqYfYAv90O5V7PcddCDu0qdb0bddrtXd9bZ3g3cUNjNe7edEhxf5cJNOe4jrOHhamJv+/ncOnDSUbDKzQw6Zm8IuKTClpuG5HumGJg474jDefPkpbrn5t4waPTIv4LIAvhlLK4Ssb8N92g2Gbdu0JDO4ruKMU47ntZef4qa//oGqWJR0U7NvLtspof/bhsJGrcl2r8JEE6LDpiCEwDAMbNum/8iJ3PTz7+KlEqhtCHroPNHb7tk1aE43DGzHoaL/bnz9opNxW1vRZE8PIwLXyaHF+nDWKcfk3+rY3s4nVeV5gM45p58Abhavp2oTvpkw3RTnkmtv4Lff/ZK/UBUdzL9FE1VeeDY31jN3zlzefm8G73/wMYuXriCeTBf7RUhfyxNStpMFgvJYDF3XMA0dXdO6vnRti8JeSlk8DHb+vaEb6LpGsCLWY5PYliCkxE0nqRw0lqceu58xAyuxHd/32f4enuv6Gqam4TlZVi5fxvQZH/DO+zOZNXc+6+rqUeQjuaUfWVvYKLepXWLT/dD5ZZl+v4QsY9ORwiXsUHiuTUNDA/UNjcVXY1NL/lAqu6xpTdfxHAfMcr5/1SWQS6LEzlNodjmTrK7rpFoThEMhfvWn33DF1y9Dk9CSyBS1yB0BkT/dAsSTGTRN4xtfuZijjzqcb11zLU898gRWRQyRX9S7Cgomnttu/j0vvD2HQCiYP61LYuVRJk7ehzNOO4FowJ987YW+oeu4nse0s85n6p9v5+256wiGTNwtBMcUTHzxulVccfV1NOYEugQhdWqqK5m8176cfcZJlFmiyz21vDZ01FFHcd1v/0FB0drSdiU1SaalhUlHHss+u/VFeV6XTbXzgaZg+jrs2OMZPuj3LKnPYhnaFjUKqWlk4s1MOuo0bv7p1/BcB6FpHQRB4bkE8PIzj3Pbnf/knY/mUN/YTNp2kFISDgaoqunFhHHjOXba0Zx03FH0qgiDkn4bhEQjx21/uZkhNWXYrttFoAkhsHNZqvoN58tnH4fs5JX3lEIKwcvPPMH02UuxglaX5/PNdxpeqommtJPPy/tsc1goF1sF+Outf2Vs/xiO42J0ssS4noemadSvXc7f/nY7Tzz/KktWriWezOApMA2DaDTK0CFD2f+A/Tn1lOPZd+Ju/nNt5RpTeZPy/FkzeOy5tzACgS26UjxPYRoab85ahgyYO0zTLKErPM9FSo01Cz7i4Gnnk5ABCrHaUtPp33cAl3/zSi46+bDi2BagSQlKMeWAg+hbW87alIOp7Zwzzy4lMHVdJ9XUxMgxu3HPnX9jnz33oDWVQymFvhPNogWh3JJIM2jQQB7/z338/Fd/5Mc/+RmaaaAbxq4REETbJHn71Zd4+JEXwIhC0VelwLuJP/z9CB556A5G9akobrAACIFyPYQZ5aSjD+TtD+5CRAJ+YM1mbwoIyCZaeOzRx4inJWj595UH3k3c+fCXePbBvxALCJRq8ycL6ZthBg0eTO+qMlbGXSx9y5NdCsCVnH76qRgCHFd1LB2gPOKJNNGyEO0DlhzXI1Q9iFOO3p9f3/IEsiq2mZy2Qp/YyFANv/v19VhC4SLROglLJQReupmrvnElN9/9GAgdEQyg6xqWaYJSpDM5li9byvL583jioYe4fthILv3q17jm0pMJGJqfgqgy3PTzn/lj1t3JQdPAbabPXidz0dnHIfP37uAT0wQP/eN2bn3oSdBjfgh+t88m0aMRdE1+ps1F1zVSDQ2c8JXvccYhE/wAq04ROJ7roWmSV596iIu/+n2Wr94IwRC64a+fQj82NjWycf063n/9Nf7wx5s49Iij+NmN1xEJh/IpVD2D53loUmPmOy/xg+9dC0YM7C3M40J/h8KYwU358EvYkXAdmw0bNpCQITTacufXLF/K+2+9T/SlFzn14PH+4St/QBZ5m3SsujeDeleyZsFGpG7g7gSJucsITF3XSTU0cNDhh/LP+26nT201zYk0hq7zeVnrdV0nk/HNv9d97ypGjBjGxZd8nWw2ixGw8sm4uwbC4QiaUUGoogyn2C6Bpks+ffsZvnnt73n27ht9YdjOhCGEAhSTJu0BmsJTPe9rISWxWIyUJdpOeEKgSXj/iQf59wvnc9mJ+3WKZPWvb5WXEQoFUc2t+NNwc0whklwmTXjgSE4/7iDA1ziBYhRXcuNyrv31vfz+tz/BbBd1U+A4OfvMU/njXU9ge5ufTZomSTc0ccyXL+Hg8f1xXbeTVcP3RwonzZcvuIC7H36JYHUvhMoTErQLbNE0iaZZyGAQBGxYt5wbrr6SR/79AOn6DJqh4SpBMBbbpC/XT1PQqSgv20yrIRyJohvVhCqiOJs68Cg+c9oEQuDaOfSK/vzw25cXtdf28DwPqUneevbfTDvtUtJamHBNTTHSstA/An+NGYbh51S7Di89+iBvv/cue4zog1lWhruVkc2WFULXKwhWVeBu7mAExYOftwP9mCVsHgX3kCGNDgLTCIdJ1dXx0BPPcurB4/0IqnYHZAUIyyAcCsBOPOjsEj7MgrCcdtLxPPXYP6muqqIlkckLy88XPquLpCWR5sxTjufxRx4gHApiZ3Lb7GfZEfADGNxOLwc752CUxXj3rddY02Lng2Ha+TPz4em9evfG3AbNucs9HacYUTtr3uL8Pbr9IaqHJAxSSpxkikOOOpZh1UG8dqZLX0jBnJnvctttd7NgbQLa+Wy1fBTk7lMPYcr4odipLQQJeC5Y5Xz5gjP9zb1T+1zXD96566bfcPfDLxGu7YPnODgFYdDuuwXh4PeLi2EGCVVXMveTj1he14JmaMXPHWdzL2eL0cSe53/P2dy1tkM0oSYl2dYE+x5+DHsPr8lHKrfTvvPBT/G6pVz8te+S1iOEQr6/vHPEuaJddKzjpzgEq2twkk28O30W6OY22NkKgWdqy69CoNTnICw3GezSzWtXjp3YHugcOFiIDRBSYts5oGv8lwBUziaZysAWyB+2Jz73HV/XdVKNjRxx7FE89MCd6IZJOpPbqSbYzUHlNQdN02huTXP4wfvz8EP3Yhk6ruN0sK3vmvC1rVw2QzyR7fpxvvmGZaJLsVVMtVtCtwEm+c2pZWMjjS0JpKbRDVdPp984YJZx7hkn+SkuHTlNQMDjz7xILr6aZ154A2jvAxN+OLsZ5YyTjoTMphldhJRkM2l6jRjLQXuN9iO1231XKV9zStav4Gd/ugtZVoVr2z1erEp5OI5LIBjC6Akt2i4IgQKlM+3YI/30mk7PUIhcvuOvN7NoeT3hkInj9PwQ5joOUtMJhILb5JRyHAfPS5FKp8ls6ZVKk83ZW32P7QE7l8Nz0pttZyqdxvPS2I7zubTxc4MQ2Okkrudx2IEHAnTwuBdStOrXrGTx6o1I09hpqX+fqwqnaRqpeJw99prMg/fd7idA55ytDP/fvvCUylNu+dqJaRqYmsDxICNsmhNpjjh4f+64/S+cc87FWOEQsOvufUJIcF0i0Ri1lX5bO4ixfMMzyRQ5z0N2E926mavno2IFmlawjvrsIJ6STBy/W/4WnYJUFMyZ/QkbG5NYscrN5lEJKcmmEwwYuy9H7z8eIdoSp5VSaJok07yGJ196D7Qgjz72JNdceGwHrafgfzzx5JP40e/uJO543RqBpRB4GZsJe0ykOuAz17Q3N/rJ35JXn3mCpasaCFZV427DZvbF9ZUJXNdBRquYMnlsPl2gvZ1MoekabrqRBx59AREq26bE8m3ieM4Pd3lFJSNG7UYgEtvCvZXPlBNvYu2Gxu1Gi7fFZubb2btvf0aMHkUgXLZJ64GmSTKJOL0r/VzUXf1ovk0oWGCUC3mTrCYEsVgN53/vp1x26oF+DEs7t4jtuJimweOPPsKGhiTBysqdRmDwuQlMIQRONkd1TTUP3HM75eXlJFOZHRYFuykU1H/y6QFByywGk2RyLmvr6vn40yX0qqlk8tjhKKVoTqQ569QTWHD9tVx/7fUEqyu37C/ZwZCahq7rfvCFaAv68TwXO9nCKaedRnXQ595tn8pR2JxWrVyFk3UIhCVeDyMolefS1NSEkxY4haAfAM/hsLMu5/TD9iwKtfzN8BRoAu65/98oYbDpOgo+NCnIpm2mnXgiMUt28CkWAj3effVFPl2yDitazofvvMHHSxuYPLSqGEEspMRViv5j9uTIqbvzr+c/wioPt/P1+hAC8BSjdhudvz60P7upvHB49c33EOiI/zG/lxACJ5cj2m8Yg/v3zr/X9rmnPKTQWD5vFvOWrUezwjstorwwJ44+5UvMPfH8LX7fp/8zePLeP3PKJT8hULH5g9v2QuGw950bf8/VN/Rs/sj8s+3svXFHouDO6jNid95461U8ZD4VTaFrBn37D6C6okCp2PGoYJoGy2a/y09+eyd6Wfm2MXhtIz5XgWlnMvz5rr+z28ihtCTSXSLtdhQKSf4IMA3dj1gE0jmHhctW8/GnS5gxawGfzF/G0pXrqG9s4bm7f46uSbI5haFptKayXPf9b/Pe+zN57qlnPxeapvZojbfg5DYS35htc4JLDT1Sxllfv45f/+AyP3m+U86SylPfvfPedEBDih4kHOR3yUBZjPPOO5t4rpDCISivqmHK1P04/YSj8oFAvn5ZSBI3DJ3n/nUbDzz5DmY0ugXfnMB1bGR5L846dVqHe/t/SxDw6GPP4Hl+dGq8bjWPPPMak79xih98kj8cKNcDXefcM07kX8+8hye64TdVCoTBwH61+b7p8KG/yFWOhUuWowwj74P9H4IA13GprqqmqtyADhw2bZaJZUsWkUznCFiRLQZcb29IqRXHfHPQda2YH/p5QNN1/nvE37bAnzdmMMIee+ze7Tdc10XmCQvA30syqVZefu4prv7+/7Gu1cUKmDvVYvO5CExN10jXN3L+pRdx9mkn0JLM7DRhCRTNrAB1DXFenbuYN2bM5r2P5rFg2WqaWhJ4nkfQMklnc5x0xFT2n7wbral2gT7K52f88+9/wb4zZhJvTSB1fafT6BVyDS/86hXsddjJ6KY/gQSCcHmMiRMnMmH0kG5/q5SHFIJM02oeee4tRDjco8jfQhBCWW1//nrrLZv8nmr33UIAw3P//gdnf+VHqGAETW1eOEtNkm1pYcIRx7Lf2L6odlR4hb9TG5fx1KvTEeEQtuMiDMHjTzzJT75+KobW5pPV8uN92LQTGDbodyyt98mq2wd8KBRoOpWR8sKTdngYKYBsmoamBOxASrVdFQIBnkckVE5I0MYzWYTfIxvrmmEzLEw7Gj1Zg67r5St0fH6juDX1P/974Ssv7XvCZzXqmHNfCBBcsXAuv/zVH1m4eDVmZdVOH7+dLjClENiZLH0GD+Rn119L1u6arL0joYDV6zbyzodzefGtj5j+yQJWr9+Yr1xgYOgasWi4ra2uyxnTfMdze4JtKSWpTI4RQwdx7fev4eorryFYXbVNPq3PgkLXHXrkNA49svvvqDyRgeikMTmOi2EY3PrnPzF/6UaClRVbrSVv6vtSasX75TJJPpg+nfvuvZ/b7n8MVw9gGmKL1UOkAOVKzjzjVHSUn3uZn7Geq9B0eO2FZ1m2uoFAVTWu42KGQ8yb8Tbvz1vHAWP6FPlZC1yqoaqBnHrU/vz61qeQVeVdczKFQHZzeCuMvHJdbNujS9je/wLyjyyNTZDE51XyrON8bk59n+h9y9HXBcvGpugCdzQ8z83P/y3NI5XXmv9b55ufDNv56bw841OByc33lStGT9yXt95/m3/8/Sa+8d1fkpEWuiZ2WqTzTo+uEZrESSS5+ttXMKBvLzJZe7umZ/juyEKV+44h7OA/8CU/+AMXXPNbHnn+bTY2tRAMWBi6RjZnI4QocncmUhlGDenPoftOJJ1zurRT1zWSGZvLvnwB4/fZk0xr4nNLNfHcfFpBh1ee47Q9FVshjcFVGIbBm888xA9/cxdm+baRGBeSiNsTkbdR4/ljkIo38e8H7uevt9+HowcJmNpmq1cUrmtn0oQHjOT0aQcCoi33EkD6C+jhx55FaBYShRCg6xaqZQOPPPVi/nk7xNeBgjNPPwXTVHQN3vQ1KDubbft+2ye+0DRMggHd16C2ure+4Mjn2TrZrG8Z6NwBeXL+SNDqGX3TDoBPjaej69pmX5blV4gpj4Q+F1o8KbUetVPX9f9iYemju1SaAu93R1rPPGc2Bl/66rf51+0/R+YS3YjbHYedqmFKKckm04wYP4aLLjiHdHb7RcQW8vEMXUPXDYTwo6ls28lXTPAFYTRssfeEUbw5cw5SClKpDDVVMabsMZr99hzL/Y+9wqr1GwkFLNKZLMcfNoWKsqBvNu62jJRLeSTA9759Jeedc+HnpnkIKZCbIB1w84WihaAo0AD+/Y9buOybPyYtAxjCzw3eWnQ+IBRItGWRDBtitf35/d9v54wzTuK0877GuriDaWmb1TALc+Wos45meE2oSKUFvsasSUnjijk889pMVMDCtm0/5cTPeuDJx5/kxm9fQLiDWdY3o+6x/6HsPW4Ib81dR6BdnUzf5GhT19LkN6KTnchToBlBetVWwCer82P9v2OYVfi5cc2tTcQdqCikxrTjxQXo068Wocstpwttz7Z1oMZ7GzMQwNuMj1l5HoZpMXv6m0hryzR62wsFpq1Xn32C92Yv831wm7i3lJJsOsO+hx7DwXuO6sjS9QVHgV6zbtlcLvvmj0gLC03kK/1ISWVlJZMm78P5551ObTTQRkUpBEIpcrbLsWdewhUvv8bv7nh6s3VFtyd2usD00ikuufhLVJaX0ZLIfOZ8y8JmFw1ZADS1pti4biOO61JbFaM6FiFre+Rsx6+cAEw7ZG8eeuYNDtpnPEfsN4m9JoxkxKA+zF+2lt/e9jCGrmO7HtGyMCcevi+up5CbIO7WdY1k1uak445mwuSJzJ41BysU2umpA6KLybV7KCfHu2+9wU03/Y0Hn3wFLRjB6IF5tDs42TSfzJ5HzlPoukHffv3o16sKoNPi9s2/Uw49jscfUBxy3EVkvdDmlRDlgRHmrDNO8hXDdocBTyk04JUXXyCecaksL8dzHQpiTwaqWb9kDu/OXs7hEwfno2klflqEh26Vc+YJR/DWB3/vUCdT+TZXli9b41+pQ38KvziurjFht5E8/OwHCBHe6j77QkMpNF2nYeMG6hvSVPQKdiheVRjvYaPGUFkWpNnZhpJq24giNd7bL/OD7/1g8xSB7WFamOHgTovmLVAZPnzPrfz1wadBi8GmNnpdA6eZb/yq1heYnleMmO0p2gpwd/Nhu0PtZq4AiM0Kak95Wz3GvgCEVLyRZ556FkfLUyEWOR8VD9x5J3+54588+/h9jOzXjtZTCN8M6ym+8fXLufXB50i7ivaB+jsKO01gCiHIZrJUDxzAWWecSs7xPrN2qZQiFPSL4r7y3mwefOo1ZsxaQH1THM/ziEUjHL7fJL53+RnUVJSTydmksw677zaUtx/6A31q/OCOdM7fbP/55Ks0x5P0qo7RFE9w4F7jGTdyMOmss1mziGP7WuYF557JNdM/QObraO4MFA74q1YsY0NTa748Utu0sXM5WluaWb5qNZ989DFvvTedj2bNQ9mKQHk5SnlbLSxV3szbvHYp044+lrq0wDA0yiJlHHj4Mfz+V9czpFe0ndAU6LqObdtMPvh4fnntpXz9h3/ZpM+3kHvZd+wUjjlgdxCqgzm2oCEfevL5zD/qdAQdtRkhBJ7rUl5ZDXTUggs1PU8+/VR+9Ie7aW2Xk6n8JDDmzp1bLADcoV35/z3qyEO5/s8P4G4FjeB/A5RS6IZBcuM65i9dyYheo/yc5XxAlZACT0HvYWPYZ/xQnnl3MWYkuFP9hFYw6FPjVfeAGg9fM/08/JjhSBTdzFMZbkJg6rpGql4SDlrbdhMB6bRd+LMDFCBMC9OyQG3GqOkpMAwCgUDbRTshk82Au21WeKnplMfKu3DJAuiGzrIPX+H//vQP7v/Nt/x4g7yCJaVEIRg8diITRw/ijdlrMHpQOOKzYqcJTCklbjLJtHNOZ2Df2mJFkG2FUgrT0JmzYDk3/OUBXnjzQxzXJRgwMTSf/m3dhkZuuucxZsxawKN/+zHlZX7una5pVMbKiKeyPhGOFGSyNq+/P5uAZaDwTa3TDt4bU5dksrnNtlXTJDlXccLx07jxl7+nNZnsIrh2FDxXoWmCn/3wGm554GWC7RagUuC6Dsp2wLb92RwIYAUjyLDYDmkwPg+kdARSKlpb4zz2j78yd9EK3nj2QXpFNDzaBI+er47ylW99l/v+/QzvLdhAIGh0meSalGTTOY478SQqAx1zL9ujsqqGyqqetLINUmq4nqLfbpM5aj8/J9MsD+O6Hsrz0IMB5n7yIcsbMgypCnTQlAvlhvY85GgOnDSC1z5ZQyhidcnn3HxjfPYgz3W/kMZcqemQauKNd2Zy/L4jO83xghYe4LKLzuKZ13+A0CJbzfUp8/7pbbF6qDzNnuO4n3tu9OZQpDLcAmWhk6ea3GrkfTAtGxuBTtYS4R9sNFlG3+oKhLuC7sSdEH5wVLg8Sm11tMt1Ct9u3LixQCi2Te3sTFzQ/hGkFWLWR5/g4ke6t3kAfA1T0yOMGNKPNz5YhghbXZ5he2OnRagoAF3npOOO/czX8pQiYBp8unglR114Lc++PoNQwEQKQTKVoSmeIGfbWKZBv17VvPfRPP52/9METL0Ywmw7DpqUCCkIWQbzlqzm08UrCAUscjmb6oooh+03EcdtM8duSgAKIchkcwwbPID995uCm0zu9OAfgchHwoq8edZ3npumRSASIVhZSai6kkAomI8k3D6bSXtWFqlplPXqx6J3nuVbN/zFZ09pt1kWKrbLYAXXXfMVVDbRTS07v+6lLO/Nuad3k3vZ6d6FgtXdvTZ5XvE8FDrn5utkqnbja1oBEisX8uQr7wOqU5qN8M1iZpRf/PT76E4CR3Wkz9sU/ILoOngu6eYWnB18Et5R8DyFMCVPPPUMWSWKm1gBmibxPMUJZ1/MCYdNIlXfjGkaPbp2oaZnJpEgk8z8l6dT7FgoBegaq1atIKN8l00HYZSPJJ44YRTKznUMqMtDSImbyzFgyHAGVfsm0/Z5tyh/7BcvWeF/fwc9y6bmQWHf8Svb9CTi+LNjp+zqQghy2Sx9B/Zn770nk3W61jLc6us5LoP61jB8cD//JCkEpx97AL/7waX88rsXs9uwgSQzWVzPIxwO8vr0WWRybjEfrzAIKp928PbMObS0JjEMg1Q6y+TxIxk+qA/JdBY3b7bU88WYu4PK08oddcSh4Hm7TNZBQagUCL53pKlYKUU2axOIVfDQbX/l7U/X+paFdvfUpJ/7ePSp53LA5JFkkukONHZSk+QSCcbvexBTx/QtBvi0R+F07m6BZNvNRw53PqFLzV/2hx57PMMHVpPN2EUt0lMCobncdvvd5BB+aku732qahud5TDn6dP72q++Sa9xIOmvni5lrm4z281yXdEM9tjK5/Kor2WNoNXbW/sIFcXiehxUpY+E7r/LY63OLwXRFFAKhzAi3334Lk3brQ2JDQ5GJSuu2f/xI0EwqSap+I5MPOoqvXngSbjq50yjr/tuglIdhWaxYuojl6/2CBO0P/FL6f59y2qlYpkfOURi63mFcDF3iZTOcdMpJWLLAE9x2D6lJBA4fzvoUdGvbLGpCFAMRO78MQ8fLJtljz4lo+IUPOijKeZ6wdDq70wLwdsps9IN9MkzcfRy9e1UX0ze2FQJfQ6yIhrn0rGMYPXQAT932U+765bf4+nnH8a0LT+Lhv1xH/15VZLM2uiZpakmQTGeQsqOpVOSjH9+YMSef8wM52+Hw/SZiaJKySIDycIBIyCSeTGE7Trdtl1LiKdh3770wo+XYu7A5aMdCIXQTL76eX990hx9Y0H4iC+EfLgJRvn/llyGbRLVjZpHkcy9PPxkduvVJFEPytS2F5Hcfli+EL8TDNYM56cj9UO0sAp7nEohGmfPy09zx2NtITeti2iscAi65+sc8et+fGd4rSqq+nnRTM5lkqo1AuzVBurGRVH0TVjDCaV/6Mq+/9gJ//+31RM18jMMXS176EBKpUtz4s1+T8QSCjtq8zBdZrxk8hhdfeILzTjucXGsLqfoG0q2Jtv5JJkk3N5OqryeTyjJmwmR+8/fbeeu5Bzhx//E4mcwmg+02BT/B3TfHdq3e093LY0uVYL6IUEphmBbpdSt47f1PikQrBUjpH/xGTD6EX/7oG9hNG0g2NpFJJP1xaWmhtW4tU064gGu+fArKUx0YlJTnIVA0r1rAux8vRIYC2xQ45bkOzc3NZJtbSDW3kG5uId3cTLqpida6tYzc5xh+fOUF+TJyHQrgIjSB8DIsW7UWdGOnZAftFB+mEAJch8kT9/A3xE61zTqjkEepSZmnN+v6ZV3TSOccjjtkH445aC9qK8poSWYQCBzXpV9tJYdM2Z07H36eoGVhmjpGnomnqF0qhWkarNvYzCfzlhC0LDJZm17VMY49aC/W1zczc85iPpy7iI8/XcrsBcu4/Rff4sC9xpJIZTu0SwhB1nYZNmwIgwYNYNGixQSCwS8w0fa2w3VczGiEZx66n/e/dRn7jKgtEghAOy3z9PPY/0+389actQRDJq6CXC5NuP9wTp12EEAHU1GBG/aFx//N0699iBXefP9KIcmms5xy3pc4aOJIn+u0HVMTwBmnn8Sf/vEkTrvF5nkCI6D40Xe+zxFTn2V4bTjPDNMu8EhKXE9x0rlf5vBjj+Phfz/Gi6+/w4KlK0km00jdoLyqiuFDhrDPvlM48tCDGDHI51/10o3YrtoZFqQdAs91saIx5rz8BNf/+UF+edVZ2LaDYbRtJzK/div7DePef/+Hr7/5Gg8//gwfzppH3YYGPKWwQhH69evHuHFjOfSQgzhov70I6H6ntCRSPvVhD1GYJ0eddD4fTj6iR9Gknuei6SbLZ73HWZddg2OEkdsQ8bmrwjegZrnnvoe4/KT9uhzO/DFSXHXdL5gweV/ue+gJFqxYS85R9OrVm8OOOJLLLjqLsCGKUa0FuEqhS8kT/3mENevjW82nXdiDw9EqTjrpBNLCKobuCSmJRcvZc8pUzjvnNKrCfjUS2aGMnK9tblj6KR/NX4EWtDabRrS9sFMEplIKdJ1xY/3qFZs7Vbueh6FpRCydVNYhErJoTWW79RO5niIYsJBCEE9lu+RJFvJ2crbDqKH9KQtZJPLqu8wnxFq6ZNb8pazd2EhZKEjOdghYJt/55e3Mmr+MDQ1N2I4fTNSaTDNr/jIO2mts53qmPjG141BeFmLksCEsmjsPEQptj+77AkKhGxap+jX88s938ejN30d4LhTYM4UfHKJZUb53xcW8deG1qEgQTSlyyTRHnH40I3tFOuRegn/QUblWfvqTn/LOJ4tABjafPKpp4DazsFly0J03dEhQl1KiFEw+4HD2HDeEdz9dX8zJVMpDD4RpWD6bMy+4gpceu4uKgOhUCNsnhnddl0hFLy687HIuvOxyPNctWiFM0+zQHM9x8KTMm/a3U1d/TnBdj0B5iF9d911GjxjKhdP2xrFtNF3vQIeolIdCMuWAg5lywMEA5LI5FApNN9A7+c5yuRyGYfbIL9wexZzfqlomVtVu1W9jbgNC/fdZhDzXxSov592n/8O/X7+UMw4ah+24GO1S+aT098FDjzmBQ485wf9dp3zPQhWi4nXz6zLTtIZf/+VeZDiy1dplocpN7ZAxPPzIQ5v9bnslpwDH9VnK/nH3PdQ3pAlWBXYKl/dOMcm6rosRDjNw4AA8Nu3E9TxFWchCofjNHY9y6Lnf5aV3PiEastoI09tB4A+u47odFpimSbK2y7wlqzANPzJz6sQxeMrDdRVBy8ybjfzrvffxPHJ5M7Gua7S0Jnn5nY9oTaYIBQMELBPbcXGzNrMXLCveu0v78+HZQ4cOAc/9oioQ2wWO42KWl/HUP//B9MUb8mbMtvEraJnHnH4u+00cSiaRQRMKtBBnnXEK4PsTC/Bc3y/86Yy3+HjROoI1/QlVVhKqrtrkK1xZQaC8D+++8QormjIdimf7vjcXGYhx9olHQCbVweTkui7BWAUfvvAfjj3jMta2+MQVnX2imtZWBNp1/Tw5y7LywlK18x2rIpH050nSv/2gUFLHkjkuOfcc7nj0DXTDJwxx8kQZ4G+MUuSZqPLvm5aJZVnomswXj/b7ztcipE8bt432tS0FgrV/OY5f1DqZymzHftnVoKGLDN++4tusaEhj6Bq243TQomV+LRTmpRSiGPfQmS7YdV0UGlJ4XHPVt5m7dGOeAH3b9fLuxsYtzpf2wtIfW9txMAyDWW8/zy9u/ifGNrKUbQt2uMAsbExlZRFqa2rw43O6ihKlFOGgyRvT53DUhT/kR7+7mzkLl3PGFTfy69seJmiZvvDrXJKJTqcfpbBMncUr1jFvyUpMwyBomew1YRSalETCAT5dvIJ4axLD8AkKZs5ehGUaRZXfzod7J5JpdF1j4phhXH7WsTxw8w+5+sun5vMy2yIri4E1eb9J/769/GT77axGeHlTdeeXUjuOU8XP0+xmQuf/3cwv0Q0Lp3E1v73pLj+U3XXarqEUjuMgAhVc/bULIJMgk85QO2p3jt5/HK7rtIWcuy624+EpeOzxJ0mlHITqjgqw48u2baRu0LJsAU++PN3X/nJ28ZrK8/898ZSTKSszyDodS1O7jkOwqpL3nvkXBx52Ms+9M7voEy0s6kIf+IEKsjgf/FeB9ksU8101KdGCMbae7WxLEcHbNgM2L1Q2f03leWBYaF6SS849m6v+7y/Ec6qoQTv5Q4SX9z/52rnq0Ed+kW4/YMP1PKTmB0+FA+YmKevaz7/P2g+q3X+3L9R2a+emDgFbXoO+NqiHI6yZN51pp17IorXNGLpfmq4QBOiX4NOKh7/CwUXLj1dhrrue53+PDNd+86v85b7nCVZsuUqT2swcc/PR7J1f+bj/dgcqFxBIKTF0ndeffZTjz7icFkfza/H2uDc/G3aKSdZ1PSLhEJFIuNsBdj2PSNDklgef5eqf34quSSpjZaj8Irru9//gg7mL+fOPvkqvqvJN0tSBb9vWpWDGrAU0tyQIBEyGDOxD315VPPjMm/zziVeYvXAFz975M3pVRVld18j8pStxHJeNjS2EAhZDB/Zlnwmj2H+vcUweN5xB/XoRNP2uyuRcsrlcPm1D5KMjJT5LmIUEBgwYQJ5Mdbv1oRASQwjolIYh9YJfcAfoswosK+inp7Q3keX/joRDm9WiHdfXMh//5718fPVl7DGwosPnhWue/OWvc8Q//s2Lb37MBRdfRG24axqCpmmgEjzy9GuIUM+qqviPIIEc/37kSb5x2oFY7U2q+b8HjpvK6Ufuy52Pv4tZHulwKHMdl2BFJUvmTOfYo4/j/Asv5KqvX8TEUYO73KmjgBEIKboc6JYvns8/7rqL2csaMSyjh5qUQko9318dz7iFPjQNfZvmWygcLkZGFlD4OxS0tnhN5XkI3SSgu/zp+mt5/pmn+d7VV3LaCUcQsTqVklMd+Z0LLhOEKJa6yqVaeeG55/j1Lf9GD5ehuphKBeFw2LcofcbjfuE5w6HgZ7tQt9fWCXZjVm4bL62H46UwLavrGsw/fCQS2qJ533NcguUx5r7zAgccfCw33vgTLjjtKMxOLGv+/M0fIdpT0bW7wUfvv8GPfnQDT788g0DFlop0+5cLBLvZQ7YByrX5aOZ0br/tDm7/5xM40sKyjJ0aJ7LDBaafe+cRDAYxjE2Uv1K+GaCxpRXHcaksj5CzfQYYTUqqK8t5/IV3WLB0NbfccCX77D6SeDJb5IjtDu9+9Kl/cpKSZCrDiZf/hNkLlqOUYtjAvkTyi+TjeUtoaIxz4N4TOHDvcRy09wTGjRpCLOIzW2RyLjnbxnE8LMskYGoETP+3WdujuaWZlngr9fUNNDc2Yts277w7AxnYxjDrbvoGIbCzaVauW4/mOnhCFgWV6ypMw6/PuT0dYwqF0DWWLJ7HjA9qyOZybWZvpZCaTvPapdibS1hWCk03STeu4fobf8+1l5+KnWtXIg2B57mYVpADp+7NS+8tICKzPPHUs36QTbvIVV3XWTZ3BrOWbcAIhHvs4PdP2CFmvvEit/3zCXqXGXmqQ7/RruuiGwax2j5IoTpQ8BXgOg6BcBnKc7jnpj/wz/vv45BDDua4Yw5n6l4TGTKwL5WxaLdsUC1NTaxcsZzp02fw3PMv8dIb79Fc34JeHvVTVrYwRXz+Vo1MaxNvz/gIzXM6kE27nm9RWb2xCaHrPeZv9aeVx4czZ6C19CLneEWt1/M8TNPkw7mLEfqWBXHBTxmqqmD+x+9z0dnn8rMJe3D8sUdw2MH7MX7MSGqrqwiYepf1mstm2Lihjjlz5vLKq2/wzPMvM2fuQtADWCGr6yFE2cx4/11a+0Rw3M+WvqWUh66bLJ/7aTf5wNt4zTzfbjrRzJvTP8RQbofx8jyFaeqs2tDij9dm+lYphdB1Vi5dyIwPPibXYe34h6hk4yqyjtqixcJ1XYLRcjauWcyl557Dn/88ldNPOZ7DDpzCyKGDqK4sL9a0bY9kopW1q1cz/f33efSxp3j6pTfJZF2ClbEtBvkoBUKXLJg7mxm1kmzO7rFv2lMeuUyGRDzOilWrmD9vPu/P/IiP58zHSeUwY+VYQu30oEphxvrvUG1WSkkmlWa3MaN57/Xn0Q3/RNB+4bieRzRk8fxbH3Ha128gHArkgy8U6UyWSCiIpmm0JlMELJNff+8SLjzlMJIZuxg5CRRZexzX5Yjzv8/8JasIBi1s2/FNvqEAjS2t9KquYPojf6a8LMSCZWtIZ7LssdswDE3geJDOZFHgVzHJj288mWH5ihUsWLCYufPms2DBIlauXMXa9XXEW1qIJ1M42azPC6npmKFgh41GiM9We09KSSjQ/YlfCkE6myXnetvdb+p5LspjExux6BEXsBCCXC6LEJtme9QNX6vMpZKwKe1RapjhbasuoZSHnUxt+gtmANMytnBtgaZruHaOXCIJHhjlUXrV1tK/T29i0QiBUAgpFJlUini8hdVr17G+biOZeAKkjh4JYxp+qsrWPEXBT9p9q0Bo2jbldPqm766j4hNhiK2mr5T5yjjpZAoyWbACVFRX0adPH/pUVxAKhzEsAyebJZVMUrdxI+vWrqO+oQlyDgSDBPJrZ1ObodvJB/fZIIpFCbYnNj9eAqHJHo+X57l5U2X3a1DTtR6veyF9n3K6NQE5G1kWpXfvXvTv25vK8jICoRCagEw6TTKRYO369axdt55EUxykhhGJoGuii2tsc3BdNz/Hts4MrVzH5wMuCOZAADMY+FyZsnaOwEynGTlyOO+/+SKmFcBz3S5+x4BlsmzlOg4977vkcg4522bU0AFMHDuMux5+gWgkhGkYZG2bRDLNNy88iRu+dQHg09oVGEZCQZO5C1dwxJd+4HNfahpu3rGvlGLE4L6ce+KhfOP8E3E9D9PwT72pdAZN14kE/I27NZVh/oKFvD/9A955dzqfzJrDspWrSLe0gOOA1EDXEIZfJkjTpV8DkrbF0t4EVfi38NztzR2dTVXdQik/aKa7laHYrLb9WbG5626fIrjKD8ASIl+KbBPf3cwm2oMWdMtmUrz01viVRBu7j+u6fqUU2/ZZjdqiXfw5YhgYRls0aGHz26Yn2A7j0PNrqm4FaU9R8N16nsK2bVzb9je+omXAF8joOpppYOgGPjHUln2nO2Ke7wgay+05Xv6lNs94szUojI/ruuRy3cxfhO96MQx0w8hH1vq5nFt9NyG26SBfMNe3WT52vkbZGTvcJFtIOE2l0uRyNoFAkM6PLITAth361FbSr1c185f60a11DU385MrzmDxuBNf84nYy2RyRcJBYNMLv73yEhcvW8PcbrqC2qpx4wqfS0gTMXbyCZCpDOBSgoTlOWTjE0QftydnHH8xh++5BrCxEayqL5ykc6RGydMojQerqG3nppRm8+NKrvPXOu8xfuIRcS9yvZGxaGJZJKBYrcl26rodj29h2Djvt5G0Q+IndhoGu+7mfIl/AWCmFlw84cWwbz3F8jdQwMAOBzWs3QqDrm9nwd2DW7va49havUWRe2lF5cGr7RdJ10h5M00RYVpcNsngQ2oy2sXW33f49s6PmjS/4/L91XccwjG4EiCpqTn5B5Z5de2eV4vqs2J7t9C+1/a7Xfnw2NX9RCpUP0vpM83cbgxIL/bcrZbLvlKAfKSWJZJLWRIJYLNblc4GfhhANW4waNoBPFiwlEg6ysbGFGbMWcukZRzFm+EC+eePfWbJiHZahU1NZzrOvz+Doi6/j7l99m91HD6Ex7pvcPpizmGRjMzWVgzhz2kF86ZTDmTR2GAJoTWVpTqSJRoJIoLGllRdfeo/HnniKV159g5XLV/pE5QELMxAgVF1V1AIdxyGdSvmnMUALBIiWl1NVVUVNr17U1NRSWV1FLBYjFI4QCASwLBNDN9Dy1GkIP0Uik8mQTLQSb4nz0Ycf8Oarr6IbxhZ9GiXseuiRheB/GKX+2bVRGp+eY6domLquk2hNULe+jqGD+pNVXRNRC2eQCaOG8K8nX0Pmtc45C5dzwqF707e2kj41lSxatgZN86nJqmJRFi9fw/GX/YTbfnYVRx0wiaztsmZ9Pd/8yplcccGJDBvYG9eD5tYUhm5QFrJwFMyY+TEP/+cxHn/qORYtWOibWUMhAmWRNpo0pcjlcr5v0vOwIhEGDhrMkGFDGTp8OAMGDqKmVy/KysqwLAupafncTj/nTsvn3TmuSzqVIt7SQsPGeuo3bmTjhg001G8kHo+zZvXqfJJ3adKWUEIJJeyq2GkaZjaVYtmKlUzdZ3K3gqEgQMePGoxp+mQDhq4zY9ZC/nr/0/z0pvvJZHOYhk5jSwIpBUHLIlZeRmsyxdEX/pA7fvVtzj/5MH7z/Uvo36uSnOPRGE8SCgapKAvRFE9y76OPc899D/L62+9ix1shFCRQVpbPlfPbYts2diYDShGtqmLExEmM330CI0fvRu8+fQiHwyAEbj7fz3VdkslkMX9JKUVDfT1rV69m5coVrFq5krp162hqaCSZTODlcr6/IG+jF6aJYfSsokMJJZRQQgmfD3Yil6zL7Dmfwhknd+uqk0Jgu4oRg/tRVRGlNZkmEg7y5sw5vPDmB4RDARzXJRS0+OV3LyYYsPjBb+6kbkMjVZVRrrzoJCaNG4HnKSpjURrjSSLhMJXRMCtWreXeBx7i3vv/xcK5n/rRlpEwoerKYoCB6ymymQzYNuFYObtPnMiee+/DmHHjqKmtRdM0craNncuRSCT85/IfDpGn2hN5Jpg//ubXLFqwwL+eY/ucmLru+3Ly/oL2eZolk0gJJZRQwq6Pncgla/DBhx/jeKrbfDWfvNyhX68qhg3sy/RP5mMENXQpCQRMEsk0Jx+1H9d97Wx2G9YfgAG9q/nXU69zxZdOYvLYoTieIpHKEA4FCVk68xct5dbb7uaBBx+mbtVqCAYJVvjJ84VoPM/zyKXTIASDhw1j3/33Z8+996ZPv/5+IeNsllQqBfngJSFEt2TwCt95vmrlSubMmoVhGASCQRB+GoQqCccSSiihhC80dorA9DwPGQjw8ay5rF6znr59exe5Wzt8z/WwAga7jx7CWzPnUBbxk11rq2L86OvncM7xB+F4+FVJhODIAyZzzIGTcRU0xpOEQyFikSDzFy3lL3+7jXvvf4iWDRuRZRFC1VVFiqeC2TSbTKIHAkzeZx8OPvwIxk2YQDgSIZvJkE75AUSbEpCdUfDVLpw/HxwHPRj8L+EMLaGEEkooAXaihmlaJvXr1vHOu+9xzhknk3E9tE0kvU8cMxwp/Vy3RDLN+ScexjnHH0RzIoNW5KSEVDqL7bhYpkFlNMzSlau5+eZbuPuef9JUtwE9WkaouqpYSLgg+DKtrZjhMAccdhiHH300I0aN9pP/02niLS1dqMJ6Aj/S12HWxx91KdZaQgkllFDCFx87RWBC3t/nKR59/GnOOeNkuuNyElLgKpiw21CikTA52yEYsJg5ZyHxZKZDWRrX9ZBSUBkN0dgc53d/uImb/3obG1evRo+WdxCUPruFIJNMYpgmBxx+OMccfzxDh4/AzUewAu0Ih7cOSikMy2Lt6tUsWrAAIxD43BNsSyihhBJK2L7YaQLT8zz0SJgXXn6VBUtWMHTwQLKdzLJCCLJZh2EDezN0YG8WLF1NJpujqSVBKp0hVl6GbfvllcojAWxPcfe9D/LzX/+RRXPmopWVEaqubhOUQiA1jUw6DUoxecoUTjj5FEaNGYPrOCQTCQRstTbZGcrzsEyTme+9R6qlhUA0utPKzewQSEGBm0OhNl9zsoRdA+1IsgtkCbvEtUrYaRDteGZVac3uEOw0gamUwjRN4nUbuPf+f3Hjj79LyvU6cJEWchajYYs9dhvGe+98zInHH8LPr7mQWDRCJpMjELAIGJK335vJj37yM1596VWwrA6CEnwh6LoOdmsrg0eO4tSzzmTyXvugUNtNUBag6TrxeJw333gdaZpbXUy1pxCa9OnKPM+PMpJt1R5UB1qrbb2BX2bJS2dRea5IoWlIy095KSzIkhBtB9FzSsId1m9CoBwHN+sTaghd98dsW+aDEKicjZsvfiBNA7GNlVB2JISmgfLaUSrmBfxWcJwWUaRj7MqoC2yftbWjoRRuMuWTjUmJDJhb/k0JW42dJjDB593UwhHu+sd9fO2yi6isqsRxOvLKInzhut/kccQiYW685kt4nu+vjJUFqdvYwPd+9Xv+dstd2JkMwYpYkYUH8hytUpJJJAhHyjj1orM4ato0gqEQyXw6yPYSlIVnikajvPDsM6xdsYJAJLL9zbH5xew0t4Kmo4UCCAmebeNmcyjHQ4aCSPMzbGxCoHI5lNAJjRqCFvKpsrxkgtSy9QC4yXSxekZBiP5PQwiUbePmnC0KTaVUh8PHdm2D66LHYgQH9wYBTkMj6ZUbEXpPS0i1u5bjYvTpTaBvBQLIrV1Ppq5l66+1o5CvfmQ3xxG6gRYwULbjrwMEemQrS3UJgZdO42X9Mez4hD6ZrgwF0Exjhx2EPzPyWQhlE4chTQ0vkyG1aNUuMVz/bdjh5Oudoes6qfoGfnj9D7nxJ9+jJZHZZMWLcMAgnsxgmiYBQ/LPhx/juh/dyNL5CzBjMTRN6xCJKqX0iwdn0kzaZwpnn38Bg4YO9ckCXG+7Ckrw56mUAtu2+fH3v8eGdevQTXP7BvwIAZ6L50D1SUdSc/ReWDUxhCZwkylyGxrILF3FxideI7WyYdu0gbywNPoNYsSvv0n57oPQw0H06goSr7/CB9O+A5FqgkN7IzSBl0ySXrZ+u5YT+6JBSImTSFIx7Sj6n3swXiLlk1VL0WYaU8rXeDyFCAXIzJnNst/9C8wt8AZvTTs0DaelhdhxxzPhjmvQdMHG+x5kzpV/Q49FUVvhGhCaht3YQr+rv8GoH50BQrHyx79h8Z+exqgs2zbtbXsin8+tlEbtmcdSffgkjPIQys6R29BIy1szWP/Im6D1bA0IKXBSWWpOP47qA3fDTWYpmDOFAC+TJbNsJY0vv09qZSNaWYgeE97uKHRWhIUAO4eo7Mukp/9EeEA5mUXz+OjYq8nlBELSVXEuYZuxUzVM8DUyI1rGzX+9hXPOPp2RwwaTythdygh5StHcmiZWFmTF6rV87wc/5V//fAgsk1B1dZFhpwBN00gnEpSVl3PmJZdw6BFH4nke8ZYWNE3b7sIS/MoTZWXlPHj/faxbsWKH+C6FACerGPx/32bwV44B120zt+WfSQ+YeA3raJ33KkalgXLbKmYIKdtMV+TNqqKj6UpoAteB/t+6kNojdifX0Ezyk7mk18fJLlmI8hRG/4GMf+DnmBUBktPf5eMzrkfpAfA8/+Td3izmtfm9CveDdvds/13VzkyZLw3V/vfdPUMR+e8X7imk6Pjbdp0opGwzrXX+/20cGJWzCY4cSd9Tj8NON/v1DXNZ3HTO/4phoIUC4LlIq4zWGsmyXz8AQiJk/r55YmohREeBtInn7m78IF/xxnbxECiv4xwsjoEqFEMQ7UzuskufKc/Dc1yEUB21qnbj1qXd+c+6E6oiX6xAuZ4/Zvlx2lo/qRDg5DyG/OwaBl9+JMp2wPPwsg5mn95U7juEDY+8hqv0nlXHEAKVcyjbaxL9Lj6SXH0cLRxEC5i+iTOTxbMdBm1Yz+Lv/o665z5BjwTb+qTLPBJ5t0k3c7Xb7xdMy6qr9trpMOr3mX/4Ik+52fk3Xi6Hl82hcvaW7725dVXCJrHTBaZSCt0waGlo5Jrv/5in/nNfF0XFcV0CpokVNLn3nw/z/R9cz9qVKwlWVnYwvwLFSZBuaWH8nnty4aWXMmDgIJ+NR6ntXueuAM/zCIZCLFq4kGceewwzFNr+gT5S4iaThPecQv8LDsdtasbesIG6f71Ian0jWjBMcOggqo/Zx19I7fsxr5naLa0I0yxuAk4iiXIVejSMf/QUeJkMbtYjOKg3bmsKsknmXXod8eWrkXolwlPQGMfLOSjbwctksFtaUFoWITW0SBCVy+Fk/IUqAybSNIp+Fc9x/e+FA23+tnTWl/mmUfS3eKk0Xs5BBkyEriMkqJyDk84gLQstaLYtboFvSsu5frmtoIWbziBME2npxVO1kBIvk8FJ28iQhTR0lG3jpDLIYBDN2jZTm1IKaVkkP/6E5X+5By+bAc/BGj6CiskjUEphr11C49uzQWhI0yKzaAHoBrgOTjLtX8M0EULh5Nx25kQBroPdkkYYhu9H1Hwh4ybTKA/0aJ4Qw3XRwiGS773LzIMu8k2MqRRaJJTXLhVOa9IXUJqGNHWcdAYtHPLv25xEhkIIXfqm3WiY+gcfpvnpZ/zP4wn0spDf766D05rttt3KtnHSOfSywrxq6383kUQJDT0SwEtlcB0PGTBwszlA5J9xC1qhlHiZNNaw0fQ+ZT/cllYyS5ew6Ht/IRu3CY8dRdnoGt/C4m6FABACL5XGbmjBS2VJfjyblk+WIAyL6J4TCA2sQpbXMOJX3yQx9ypSdWmkIRFC4uWyOMksMhhAmjrKdXDiaWQggBboGMvQNg9zvvvE0FCui9OaROgGejjQThCDchyU4+b7R8NtzSAsCy1k4bYmUJ5vfu5gzcrHM3QVthIvm8VJtWur7eCkum9rCZvGTheYkK/+HYvx7GNP8ue/38FVX/0yLYk0uq7jOE6x1NY13/sR9/3jAWQwQKiqqoOgBJCaRi6bRaA4/YILOPHU0wCIx+P5yiA7xmRY4Ix1bJu7bruVTCaDFQxud9+lEAJle4RGDUUKFxEJsuF3D7H4zw8iKPf9tZrGyr/UoOkKLRwonuJVNgtWiL6Xn0TVgRMwa2PgueTWbqDh+Teoe/xtMCzIZQmOG8/wC48h1KcMz3FBGAz6/ldQukXTsy+TzQTpe96RSM3Dy+UwBw5l1O+/hwgGya1cxrLfPkj5UUfS9/T9EcplwwOPUf/GPNB0el96LlVTR5JdtpQVf3gQO5nDGjqcQVeehh7UaH33fVbf/RJIjbKp+9D7tIMI9K7EqIz6GkVznNSipay//xni89aghwKAwkll6XX+aVTvPxpcm9U3P4A+Yhz9zjsCM6xY+M1f0bqsCZwM1qAh9Dn3aMomDMWsKsONt5L4+FPW3P0EqZVN6GFr60/ZnocMBYi/+TbNL7yCMAzcXDM1519C1QETUK5HdtFCFlzzS4QeBddFmAZSl5j9BzL8ytPRgzqJ6TOoe/FTBn7zXMonDaPxkSdY8vuHMfsPov+3jicysh9mbQVawMBLpcmt20jD82+w4Yl3wDB9gZROExq9G/0vOBKhQ+Ld6ax58A20oIHSQgy54SuE+kRx1q9lxV8fp/a8i6g6cBxSF7TO+JAVf3oIO2EjTR03kabsgP3ofdxeIBSNjz/HxlfmIHSJOWgow688rdju9c/NYdi3zqNsdD9wcjQ8+RKr734edBPwNVmnNUX5wfsz4LITCPQqx22Ns+ZvD5Czowy4+ChQHg2PPs3aR9/vqL1tak1o0tegghbCs2n9aB6OZ5Gav5SNmu773rfWBimlf+iKhmh9820W3HgPUgthDRvN7g/9HKtcoVX3pmKfUSQefA+9KooTT2D07ceAq46hfNJIzJoobiJJcvYC1v3jCVoX1fnP43oITeImUphDhjDovGMoGz8EozyMl0qRXrKCun89S9P0RWhlYT/wsTVF7bmn0ff0fRGey+qb78Er68/AS6ehh01ydXWs+fu/aHh9LlpZeNNPqwoHlhRG//4MOPdYyicOx6yO4rYmSc6ex9p/PEli8YYe9X0Jn5PABF9Ds6Jl/PAHP2HPibuz/5Q9aUmkKY8Eef7l1/nGld9h8afzCVZ11SqhzQRbXVvLpV//BhP33JNEaytqB2qVBSjPIxIt545b/sai2XMIlO/YNBIvk0VoGm5rkqqTjiW1PkNiwSpy6xuwmxPY6zfghoIIXQK+T0OWVTH6lh9Tc/CYfGCIA0JDCEXNiYdQtf/DzL/2DpxsDnPIUAZ/9Rwyq5fi2Q4YAXpfcDJWsAbqlpNU/Rl41vHEly1C2Q56rz4MuOIcpB4h9fEbLP/VPbhZRc2Jh6GHdUR8PRtf/BBZ058+F5xAbGxvvPh4Nv7nBRo/XEbZfnvT94JpSM0l89EMPMcD16Xi2MMZ+tXTSdetxk3nEKaFHgsTO3gvep10IHPPv5bGj1ZjlAV9U9rkPehz7iE4jU3olZXEDpmCETLAbsWIBnATScoPOYgxf7mG0IBKwMPe2II+fiQVB+1FzXH78enF19Mydx1ayNx6oZnXMrVgEKFpiCaFHrLIF3lEGAZGOIYIRXzflxS4LXG0iip6nXEUuqmIDOtFzZfOo2LP4QgMsrNm4jlpjD79GXrdZchMM7mmBMoDozIKSlFzymGUjfwbi37xb7TyKCpnY/TvT5+LTkQzBBtlK6vvfRmCBmgmVdMOoXxEJdmVa4hM3ZeKqePxUimcZJryqRMJVIeZ842bIGDhZW3C48bQ78snglA4i+aw4bmPEZqOVlHpt1v3KNttAL0vv4iysQNxWxO4aZuKAyYhybH0by9gVEZxWlopP/JwJtzxfTTdA01HaBqxfcbROm81od2GYEZDZOd84JfME6HNLAIPGbBIzppL08wl9DpkJNaocYz6zdeYd/XfkbFyUC7KcfMuAhByK/cBpXxrDGH0WJjM4uVk1jUTrO0DCrRI0Lf6pNMEx+/O2NuuJTK8FgC7vhmtPErFAXtSe+IBzLvspzS8txQ9GsSNJwlPmcLYW75HqE8UNA2VyyFMk/L9JtLr9MNZ+oPfsOr+tzAqylC2gzV0IFVHTMFpamHwj68kNHYEUvjuj9CooVRMncC8i66l7pX5aMFNPKcm8ZJpQrtPZMxtPyAytMYP5KpvRotFqThwT2pPOJB5l15Pw4zlHbXcErrF5yYwlVJIXSOTynDu+Zfw+stPM3hgP358w6/4+S9/j6sUoequWqXIh/GnW1oYN2kSl19xJTW1tUVf5Y6oxt4erusSi8V48vHHeP7JJwlEy3aYsFSehwyZxN/5gNS6ViJ9Q4iRIxl107XYza3YdQ0k5i5k439eouGtuYhgACEFdsph8HUXU3PIbjjNCRIzZrLs9w9CuJohP7yEyOAqas49jcTH81l66wtk5s5izld+Sp+LTiI8sBIvGWfpj+7GTihSc+diZz/h00wL/b50DFrYJLt8CUuvexJlhnDr6xDhCMlP5hCfvYLoqF6ExoxAtwTGwAEEasLk1tejRYNEJwyl4cOlRCeNwksksJsbaHj9E4RhgK5IvD+dT859g8a3ZuOmffNu5YnTGP6DszGqetP/y8fR9JU/oQiC9E2PdmMzdlOK2P4TaHzkCRremU9wUG9yzSlkRS0jfvENgjUhcqtXs/yGv7HhtU8o23sqI268lODAIQz/v0v5+KzrUdu6T+TNogDKdTsK3cJnrpuvZi/zjjgHpzmBaycxhw3DWbeKBd/8BQSjeBtXI40gXqKZJd/5BY0vvUt2YxzluJgDBzHyd1dTtltfep13Euv+9TKJNWnfXJvLYTc04xkCJ5FtZ13xzaq59R6EyjGtJmafdiWy33AGf/M0vHQDsUOmUjb8QeIrEn5QVyaD3dAMQuFl7LZrtWu3NWgg8Tff4sPv/ILyww6j35kHkmtS1JxyBKvvexUvZyMjFQz+7peQKoObVtQ/9BCr7nuRyN77Muy6L+E2NmErDy9rb9EaJITASWTo943zqdxrCOgGKpuj5qzTcJrjLLz+foRlEtxjEiNvuARNy7L8JzdR/+6SnlsQ8gcgvTqKFg7S+4SDKRvdBzeZw4gJMis3IIQCI8Kwn11BeGAMu24DK391K+uffp/guImM/OXXCA3qw/Abv0Li5GuxMzYiWsXwn3+DYJWF0xKn7p7/sPaxt4jsNYUh3z4DPWQx5CdfJ/7BQuLLW/IWohxOvBU7niHQN8aSq66nedZqas89jX5n7AdGgEHfOZ/G967DdVRXn63wzfoEowz7+RWE+5eTW1/Hyp//nbrnZxKaMJmRv/wqwYH9GH7jV2g99VocR/0vx/H1CJ+bwIQ8d2w4yMrlyznj/C/Tp3cvnnjoYaxYJUY+4rU9RL7CRyaR4IjjT+D8iy9GCkkikdjhWiX4wrI8FuPVl1/i3ttvxwx28iFsbyiFNC2c9SuZd/nPGH79JZSNGeBXW6mpxKgoJzRmOL3OOJoVP/0jy295Hhk00Gt7U33YHrjxFHhplt3wdxo/XgdkIFLB+L9egZuyqTpmP1bd9xrZFStYNWcJlSceTWSYDk6WDQ89Q+uqZoxImR8VuyFFn/OnIQwdp76ONXf+ByUCoGkY5WGcpgZa562kfGwfrCGDsHrHsEYNw6wMEn9nIeHJEyjbYzjSmkl03GCEJrFXrya1fKPvw9Jg47+fwurXj/K9xmP2qkQIDxG2UI6HyuYIDB2IUWbhur7vtWBK08vLaHr0KWZf+luUNPxACs8hNu04wkOq8FyPxMwPqHvufbRojJY336b5kyOwKkMEx+5G+YSBNMxcjR7eBi1zWyAEaAIhLVSikXmX/JiGD5YipYE0LfRojMzCBaxcsITyfXenas/dMcoC2IksbtqvzSoiZYSG9KF12YJiLq7QNITWLjCncDtNgm4gRZYF3/8d61+ZDbxP7OAp1B4wHJVzMGsrUEuaiz4woWkgVEdBJgRIgQyFsNcsZ/4VvyGxtoGWj9dRfdQUgjU6eiyKUR4mtaaZ8D57EBlSheeCvWIhi396O3ZG0vLRQoK7jWTgOVPbrru57pISJ56g8rSTGfV/FyG8LI0vvIXsPYjomH70vvxLuIk0C355F1VjRxPbdxxewzrcRBI/TLQHQ6JJ7OYkFScdx+Qjj0DoBmZ1Ocpx0WtqiL/wPI3vLUYICO0+nvJx/fByLum5c1j32JvIUJTWGdNpfO9wgn2iWEOGUbHPSNY8PZ3Kgw6ibHgNru2QmT+XJTfchSstWj5YQGjsaAacsTcEKqk+fBItNz+bD9ISgECPhdlw212suPsZdDPE0v+7hdjU8UQGxggMG0ZkZB+aPlyF1pk5TQq8VIayg6cQHdPHb+usWax/8m1kuJz4++/TOP1w+taWYQ0bTmzSMOpeW4gRDX7+0dC7MD5XgQngOi6BaDkzZnwIjkOwqgbPc7v4A4WUeI6DY9ucd8mlHH/KKaSSSWzP3vEm2HyFkVgsxqsvvcQtN/0ZzTB8H+MOTnZSnocWCpL6+GNmnXE1ZRNGEdl9OJGxwynfdwJWLIASJv2/fjYbnn6P5KomAv0rMaK+MHPWbCBT1+qb83ImmaWrcJI5jJCOUVONHjaxcxKjKoLQCyXHhL/xtYBm+QniRkVZvkEgdB0jFkMZgXyYvUC4OVren0vfk/dGKy+nbPxwzHHD0DSbdfc/Tf9BI4iMHUbZ+BEE+lSAgNaP55FrzqBXBXDiafpddTlDv3UKRiyCFrL8a3seubp6lAfSMnwzW4c+Fwjp0vDCOygtgFldBgpy9U1Y/XshdQ2nOUHZAQez75zDO0TvCk3DrOpFaFAv6t9dDsLKPyDFZ91h8BRa0CLx0Uzic9dgVddSUHO9VIrg+N0Z/eerKRvZxw/Qyke32o3NOK1ZjKjVFiizJbVAKaRhYK9bR3JxHWZlDV4mh9ua9CkquwkU2ey1TIPk4mVkmzJYlTUoqXCTWUSvcFsUrOthVFcgdYnQJZnlq3Ft36ws6l3Si1ai5P6IHqwfpTxEoIy+FxyL8Gxya9aw+Jpf45QNZo+Hf06gKkC/b11CrjlL+OA9wM6RnD2P1oXr0AKBnh+CPA+tLIJZXelHCmeyOC2tNDz2LMt+dS+up6E8D7NfLdIy8FpaCO4+mSmfPJofBxBCgpSY1TWEhvYBbKx+vRAShNRJLVyGqzR/PW10Sc5bBtq+4OF/T3Sc3kJ5pBevQTMjGJUR7Hia9Mp6yoZWoYSJWVmOcpdDFx3THwOrXy80U8dpiROavDdTZj3WaQ1IjOoaQkN7w0uf9qyf/ofxuQtM8P2ZgWAwn2bV1bwppcTO5dA1jW9cfQ0HHHwwrfF4jyuJfNa2aZpGMBTi8Uce4b677kQ3jGLFk50B5Th4toNQEH//I1re/RCUhzlkOOMf+Dmh3hoiGCbQp4LE0vp8qocqCrd84U5/09W0NsYg180/g/Ij8tpbE928P8iQ/vXanzpdz/++nqfgEiBNncSHn2LHs1gVBrHDpmKMHoazYQNNr86k8qz1VI0cTO0JB6CZEpwcLdPnoKSBSqcIjhnP0GvORtNyZBYtYtXN/yK5eB0iVsvIX1+BFdzU4cTPzfOyTj660I8sBPAcx39mXSe3eg2pVRuRpplPY1F4OQehCTIb4ggjP4+8Tik43XAebzdIgZfO+D5mzzfbCgmupzPou1+mfFw/7MY4dXc+yIYXPsSOZ+h3xcXUHjHe90lvDQTgekUz8WfSIoppLXmTc1EgddRGvYwfUSs8Dy0S8nMoPYXrOGhlkfy3t9C/QoBjI8srsarKUErgbNxILuHgNizi06/8mnF3XYsZkgz+8RV4WRvl5Nj4nxexkx5GTLSlWW0GyvPQIkEaH3ualXe+iFYWwctkyK6tI7NqIzIQQAZMvKSfvoPngaFjr6sjuXQd0jKL/eDlbKSukV7ThBB+RGoB0jTyxeP96H5hFLZghXKcroc0IZFhCy9no5SH5+JHmyvlp9TkchSITbrrO2W7/oFD17HXrSe5bH03bZWk17bkc7i32FX/09glBCawyQhTIQSZdJpIOMyV3/kue0yeXPRX7kgo5dfKDIXDZDMZbvvLX3jpmacxQ6GdollCW0h+2YH7U3vQSOr+9QKp1fV4GRvleHiZXJEaTNk2bjKDDJjk1q8nUxfHKqtE792H2ORhrHlyBqCo2X8SekADIcksX4XdkkEri+TNnJvtkHy+mIsMhxB45FoSflqIaSAsk8yy5SSXb8SM9Kbi0P3RqqpIfvAO6TVraZ21iJqph1F78gGAh9PUROsny5BBCy+dJDh8IFJzQDNIvPceK+/+FyAomzoNPRr0BWFPKeiU8jeB+ctxsy7SNMitXMbss3+IR4Fpx0MQJjyiH7n6JrRwECeepPaCsxhwrt/GdX+/j7WPTEcvD+04M1V7lUIIlGOjxyoIDqjBbc1AOs7yX91JsqERqGZQVTl4LuzYc+Jngp92opNZvJxcSxYrIgmNH0vVgePY8MpsrMHDqD1uH1Q6gwhugZlHkQ9eSeKksig7hzViBNWHTmDdYzNoefsF5n49xIQ7roFkAi0Wo+XFV1h598vosVjPiRsUCEMnt2YtTdM/QA+U+zEEpoEejfhWJtdDGjrpRctxUja6ZeBuXMfcc67GUeALfw8IEB42EDceRzOCpBYsw7UV0s0RmTieYK8IybXrEWaEigMmoDI5ZNAgOX85SsiipUEIcLMO1dMOYe2Db5Cua6L88MOJjumLm3NR8SZSy+qQptF1P1IKoWv5trropoFTt4a5Z38Hpzh5PCBIePgAnMYmZNAqBf1sAbuMwOwOQghcx6G2thdf+9ZVjBq9G/HmZjR9xzW7ICgNwyBSVsb8T+dy9223sXT+fAJlZTu3AHSe9kyrrGbQtV+j15lHkVq8iszqepSCyO5jCPQOIcNhUnM+JrFkPTIUxm1uYN39L1Dxu0txW1oZ8tOrCI55CSKV9D79MJTroRkO6+57HiUMyFOAFUgIOiwaBcLQcBoacTIOFjrGwGGMve8XeK6g+dkXWXnPKxjlEZyGZuIfLqZiQt987qdG/IN5eDi0fjQP1z4czfIjJdMLl5Ba2YA0w6icJLtuI0oYqHSSyNT9GHzZedgqzKBvnYN0sihpdGVZKWi+bicCAk+hhYMkZn5A/Ruf0mfa7pTtfwAT7voZG174AISG1beW2MFTCPfS+Oj4b2NnfROW2ac35fuMR+HS9Hgsr7FuJZTa/Ek9v/kq1+sSJCQ0DSeRINfUihxegReIMuTHl1P/5qfUnnkyVXsOxG5MYJR3YgsqXLNA3tD+doV7dRb6Xrv32/+kwFAkVPf36NxufA2teC1PISyL3MplrPvPW4z49gnYDTDyph/TZ+4yQmNHYUYt3FQWI7yZyFj/yghNR7U0UffIG1T+4kKcRo9hv/ou0f1eJ1ufJDJxrO8WrozhpjMERo6gYu8R1L+7BLMnDEWFeeR5PtmEHkYvCxctMG3rQSFDQTJz57LhuQ8ZeN7+hCZOYsI/f0Hdk+/huWD1qSW6357ERlfy8QnfxM0ESH70MfWvzqXvCXuAGMBut/6YDc/NIDJ5ElV7D0EEg2QWzWXD8x/6B7eWVv92UuIm0wR3n8ik5/9KfO5qyiaNQdMUelWUdffcT3JVC1rY6rR+fe1TCwVJzZ7Nxhc/ZsBZ+xLacy/G3/8LNjz9HkpJzD61lO+/J9GR5Xwy7SrSDVkwZEnL3Ax2eYHp2Db7H3Iwu+8xiWQyidQ0PNf17e/bMaSrg6CMRGhoaOA/D/6T5556ipxtE4xGd35B6Lw50a7bQOvspQR6l1NxcD+koeUT9zO4WZvM/E9ZdN2tuLZE0z20aJgN9z6EWRNl8GXTsIYNY9gvxoNyceJJnPoGlvz8b2x8fb6f4J5nhpGhIHpZCJENtvWtUgjdxGmoY9UtjzHi2rPQoxEqDt0Xs6wSb+1ClG2jhEAIl5YZc5HfOBbdVeiWIP7hfIQIkpqzECejsCrL0KJRNnw8DztpY1gKQkGSH37I2n++xqCLD0OPRRn51+shl2Tt3U9TfuiBhIbE8BpCHfpGhoLosQg4BsJo79tUKDSkzLLku79B8G1qDhlPzTkn0ev8k30zoRB4tkP8xZdxcy7I/FLwXJzWFEJTOC3JIptSjyGEb9qSfv5dN4MKmo5eHkGPGOiFza74ew1hJ1l104NE//4dzFiUvpefz4ArLRpfeJmGN+dTe/RkBI4/DwpjZBjosTI0Q6CFzHZ9IdCjYfRYGaol1ME/K8Mh//2s72dE5a9lmeixMoTw8qxS+d90bjdtmrFeFkaPRVDJcIHuFS1ssvp3txKojtDn5CkYkTA1QwdQ/9ATrJ0fZ9CVJ/tCdgsHUOUqtEiQujsfwKyOMPDiozBqezHwWxeC9IkwMouWsuHZeVQcMZXg6NGM/9dvmHPetTTOWObzzW7Kj6lABoMYsQhKSKSlF1NTutO2lALNgmU/+SPSgF7H7knVScdQc9pxRbYpz7ZJvPuuH+ktNTRslv7gt2jG1VQfPI7YIftTNe1wvGwWN5EiNXcOC7/9WzJN2TwhRf5enkKPWKy5+R4Ce06l95lH4CZSeLkcG+97mCW//w8yHALPN8vKcBA9EsIJWYVHQzMUy677PVL7Fr2OnkT1adOoPfO4vOvbb2vr22/jpO2tn+v/g9jpXLLbAs/zGDVmDMefeBLjJkzAME2ymQy2na/OsBUVIwooaopKIaTEsiwM06SxoYG3Xn+N559+mo1r12KGw0gpP7/6lnlmHD1aTnjMEILD+mPVxJABAy+ZIrVoOU1vfEiuJYtezCX0TUNuMkt43GhiU8dj9YqB55Jds47mNz4gubQOrSxc1NoUgrJJYzHLA6hclvjMT3HSToeISy+TIzBiCOFhfRD4/tHs8hW0zl+FMA1wHGRZjNiUsf4VlUPLu59gJ3MIzSA6ZQJGxEQIQWLWPNKrG9t8OK6LpyRVRx1A+cShqGyGlrdm0DRjGZVH7ose1HDjLTS9+QlKSZTjEB6/G8EBVaA8Eh/NJbO+pSOXrpSoXBalNKJT9iC6+zCMyjKwbbLrN5JesJzE/OU+C5AusVttRt16I/1O25fWN99g1nn/h+NoPebjVK6LUVNDcEA1CnDjLaSXrGvbiPJjqcUqiO09BqEJnI0baJm5EPS2dgspcJIZwmN3o/rofdDDBtmly6l75DWskaOJjOwNrkvrzFmk17UgUOi9ehGdOAIhILd6NfFZy33CdGlQvu8E9LCJl07S8u5sPMfXFCMTxxHoXQ6eQ+sHc8g2pMDzCAwbQmRUPwSK1PzFJJfWIaRAlsc6tvuDhSA10HTKp+yOHjZQmTTN787GzXn+3PEcci0JImNGExraG2dDHY3TP2HY737JoPP2AylYevUNrLzvLYyKyGa0wXZzevcxVEwdj1lTDnaOzIo1tLw3i8Ti9VQcui9WdQhhWNjr1tD07qd+2lJ3QlkIlO0QGjea8MAqFILssuW0fpqfz5sS5PlxVLaibO8JRCeO9IPNXJdcXT2pRStIfLoMJ5n196XiPNQp33+SPw+jIbx0mtSCpTS+9gF2wvZzeAXYja0M+OFVDLvqOFCKpVf9iBUPzqDqiClYVRGyK1fTMnOBT1yhSd9EbwSITh6NZhl4qQTxDxbk4/ckysnhOVCeb6tRHQXH9tu6cAWJT5fjpLL+eO3y0uDzxRdCYALk0mmEpjF67FgOOOhgJkycSFV1NVL45OeO4+AVg1jo3t9V4FAUAk3TMEwTXdPIZrOsXLGC9995h/feepMNa9ciLQvTsnaNupb5Beplcr5Ppv2ISYkWDiJ12eUUXWCB8TJ2h2vJYMA/dXfanLxU2g+QEAIZCnRJT0AKVCaLl3X8JiiFtEyf2q4Qrem6+aomAAItHERovp/Op3VT+VN9oCMdWt6X5yaSRROhME2fCiyRRHn+82gF+jjhB8x4OT+gQgsFO2mZbc8L+Xvbbof3ha4hAxZC11DZDHrfIUx+9RYCRpyPT/gmTR+u9v2nPfVf5n3JhTZ1W50kn4fppjJ+l+l6WxBH+68Vxi6d82WFkD6rSy7r5y0i/Ko1hgbkfdiprD9MpoEM5gkUULiJdF778eeKfz3wUhk8O08tGAr4m68QqGwWt0hzaLVxF7tux3bnWZfaj62QEhkOIjTpR/tOmMTALx3EhodfILWyHmFaVB1/JAMvm4ZeFsbduJKPj/s2qfos0pCbFlJ+r/hF5lPpfB+09WlhTrutSZ8II691d9e3ncej/TzqMJ+3MNZC+BV8/N/mD6r5lBwZMDtaGArzMJHuaOaXEi0SQuapD4UmfYH5gysZ+s1p4Hosv+5XrPjHa0hD9wPVjELf09ZOpXBTGf//pWz7vHhv8JIpf7y31NYSNold2iTbHoFwGKUU82bPYd4nn1BZU8PosWMZO2ECQ4cNp6a2llAohN7Ov1nY1AsTRuBrlrlcjng8ztrVq1k4fx6zP5nF0iWLySWTaJZFIBr1w8p3BWEJRdOsXmZ0cxBQbX6Lzp94ns+XGgp1eb87IaBFwkWTXbdCwlOIgIUeLCxG3/ZWvLdSoGnosbIO9yIfEKFHI+3e7+ofA9DLy/zrCv9+yvPa3isQeOe/L0NBtHD+/e5I19tfNxLu2ndKofIBFkqBFjJpfvZ1Wt96naYPV/rBHlszBwqmUdPM/283bVIKdD3fRyIfSNWN6a/92AmK3xPF8Wz/zHlWoQqTLmOCQC9v1+/F/vNNsppou1+bSdbCCAbp0q+barfYxD08D2EF6fu18+l3ySm+QJcSLWDipbPk1q5h0TV/ILm2tYfUbH57/HEPdXlfuR5aJJx/ps79sKlLFq63Fb/J/04p0MIhtEh382oT87us8zzsZv0q5ZuJK8t9hiPTPxTp0bB/oO1uzuTN4kWS/Q5xCP61/b7pQVtL2CS+MBpmAYU0Etu2cbP+idoKh6mqrqZ3nz7U1NZSUVlFpCyCZVr+idRxSafTxFtaaGioZ8P69dTV1dHU2IjK5UDXMS0Lqfl5VjstqKeEXQcFUvhECmGYaJFAqUj2Z4JvjpWRGNUnHER0/FDMmnKEUNiNzSTnLqb+6TdIrWoq8Zi2g5A+o1HViUfT65iJoBQbH3ycjW/M3zbO4xK2K75wArOA9n5L1/NwbNvPY+pUjgjoenrSNDTDQNf1fIkb9fn5KEvYdVAsedS91lfCViJPz+a0pkFKpOHnBCvb8VM0QtteLea/GsKnJyy4UmQw0GYWL+FzxRdWYHZGUYCKguO6eyd/YdLt1PSQEkr4X0XhEFIw70LeNy4+Wz3S/3a0ywIo9dOugy+MD3NLKAnAEkrYBaFUFz9wT5h3/ufheaWA1V0QpdCoEkoooYQSSugBSgKzhBJKKKGEEnqAksAsoYQSSiihhB6gJDBLKKGEEkoooQeQwFbWCSqhhBJKKKGE/zk4EsVchFD4tV5KKKGEEkoooYQ2eAihUMyVCKZTJFQqoYQSSiihhBLawSffFUyXnsdtKJVji6XPSyihhBJKKOF/DgKlcp7HbdKJr54B6r48HYe9xZ+WUEIJJZRQwv8G7Dxf5n1OfPUMAUiqqsKGG5wuhByN8mx8BqCSxllCCSWUUML/IhTgIKShlDff1tJ709CQ9NNKGhpapZLTUGoOQhr4wtKlFAhUQgkllFDC/w48fNknENJAqTlSyWk0NLSCn1biATLbsnJpzspNQanbgAxCaghRytMsoYQSSijhfwNCSITUgAxK3ZazclOyLSuXkpeV7c2uBeGJHu2/l5B8WSD3BG8CYHS9cgkllFBCCSX818AGOUvhzVQed/jxPUA72fj/SvC8S026kFMAAAAASUVORK5CYII="


@app.route("/send", methods=["POST", "OPTIONS"])
def send():
    if request.method == "OPTIONS":
        return ("", 204)
    # optional shared-secret guard
    want = os.environ.get("SEND_TOKEN")
    if want and request.headers.get("X-A2Z-Token") != want:
        return jsonify(error="Not authorised"), 401

    d = request.get_json(force=True, silent=True) or {}
    to = (d.get("to") or "").strip()
    subject = d.get("subject") or ""
    html = d.get("html") or ""
    if not to:
        return jsonify(error="No recipient (to) provided"), 400

    # For a proposal send, build the subject/body from HIS OWN build_email so it
    # matches his desktop tool exactly (warm tone, service list, all reg links).
    if d.get("attach_proposal") and d.get("proposal") and not d.get("html_override"):
        try:
            esubject, ebody, ehtml = _email_for(d["proposal"])
            subject = subject or esubject
            html = ehtml or _text_to_html(ebody)
        except Exception:
            pass  # fall back to whatever html/subject was supplied

    # Sender mailbox: the platform may choose which mailbox to send from (info@, payroll@, etc).
    # Falls back to MAIL_FROM. If ALLOWED_FROM is set (comma-separated), the chosen mailbox
    # must be on that list - a light guard so only your real mailboxes can be used.
    frm = (d.get("from") or "").strip() or _mail_from()
    if not frm:
        return jsonify(error="Sending mailbox not configured (MAIL_FROM)."), 500
    allow = [a.strip().lower() for a in (os.environ.get("ALLOWED_FROM") or "").split(",") if a.strip()]
    if allow and frm.lower() not in allow:
        return jsonify(error="That sending mailbox (%s) is not on the allowed list." % frm), 400

    message = {
        "subject": subject,
        "body": {"contentType": "HTML", "content": html},
        "toRecipients": [{"emailAddress": {"address": to}}],
    }
    cc = d.get("cc") or []
    if cc:
        message["ccRecipients"] = [{"emailAddress": {"address": a}} for a in cc]
    bcc = d.get("bcc") or []
    if bcc:
        message["bccRecipients"] = [{"emailAddress": {"address": a}} for a in bcc]

    # optional proposal PDF attachment
    if d.get("attach_proposal") and d.get("proposal"):
        try:
            pdf = _pdf_bytes_for(d["proposal"])
            name = d.get("attachment_name") or ((d["proposal"].get("company") or "Proposal") + " - proposal.pdf")
            message["attachments"] = [{
                "@odata.type": "#microsoft.graph.fileAttachment",
                "name": name,
                "contentType": "application/pdf",
                "contentBytes": _b64.b64encode(pdf).decode("ascii"),
            }]
        except Exception as e:
            print("=== SEND: PDF ATTACHMENT BUILD ERROR ===", flush=True)
            traceback.print_exc()
            print("=== END ERROR ===", flush=True)
            return jsonify(error="Could not build the PDF attachment", detail=str(e)), 500

    # optional engagement PDF attachment (mirrors attach_proposal; additive)
    if d.get("attach_engagement") and d.get("engagement"):
        try:
            pdf = _engagement_pdf_bytes_for(d["engagement"])
            name = d.get("attachment_name") or ((d["engagement"].get("company") or "Engagement") + " - engagement.pdf")
            message.setdefault("attachments", []).append({
                "@odata.type": "#microsoft.graph.fileAttachment",
                "name": name,
                "contentType": "application/pdf",
                "contentBytes": _b64.b64encode(pdf).decode("ascii"),
            })
        except Exception as e:
            print("=== SEND: ENGAGEMENT ATTACHMENT BUILD ERROR ===", flush=True)
            traceback.print_exc()
            return jsonify(error="Could not build the engagement attachment", detail=str(e)), 500

    # optional invoice PDF attachment (added Aug 2026)
    if d.get("attach_invoice") and d.get("invoice"):
        try:
            lp = _logo_tmp(d.get("logo_b64") or (d.get("invoice") or {}).get("logo_b64"))
            ipdf = _invoice_pdf_bytes(d["invoice"], lp)
            iname = d.get("attachment_name") or ("Invoice %s - A2Z Accounting Solutions.pdf" % (d["invoice"].get("no") or ""))
            message.setdefault("attachments", []).append({
                "@odata.type": "#microsoft.graph.fileAttachment",
                "name": iname,
                "contentType": "application/pdf",
                "contentBytes": _b64.b64encode(ipdf).decode("ascii"),
            })
        except Exception as e:
            print("=== SEND: INVOICE PDF BUILD ERROR ===", flush=True)
            traceback.print_exc()
            print("=== END ERROR ===", flush=True)
            return jsonify(error="Could not build the invoice PDF", detail=str(e)), 500

    # inline logo: swap the text header for the real logo image (added Aug 2026)
    lb = d.get("logo_b64")
    if lb:
        b = str(lb).strip()
        if b.lower().startswith("data:"):
            b = b.split(",", 1)[-1]
        try:
            _b64.b64decode(b)
            ok_logo = True
        except Exception:
            ok_logo = False
        if ok_logo and _SHELL_TEXT_HEADER in html:
            html = html.replace(_SHELL_TEXT_HEADER, _SHELL_LOGO_HEADER, 1)
            message["body"]["content"] = html
            message.setdefault("attachments", []).append({
                "@odata.type": "#microsoft.graph.fileAttachment",
                "name": "a2zlogo.png",
                "contentType": "image/png",
                "contentBytes": b,
                "isInline": True,
                "contentId": "a2zlogo",
            })

    # branded emails (engagement etc.): embed the high-quality A2Z logos inline.
    # cid:a2zlogo = A2Z Accounting (header); cid:a2zlogo2 = A2Z Practice Hub (footer). Additive.
    if d.get("a2z_logo"):
        try:
            if "cid:a2zlogo" in html:
                message.setdefault("attachments", []).append({
                    "@odata.type": "#microsoft.graph.fileAttachment",
                    "name": "a2zlogo.png", "contentType": "image/png",
                    "contentBytes": _ENG_LOGO_ACC, "isInline": True, "contentId": "a2zlogo",
                })
            if "cid:a2zlogo2" in html:
                message.setdefault("attachments", []).append({
                    "@odata.type": "#microsoft.graph.fileAttachment",
                    "name": "a2zhub.png", "contentType": "image/png",
                    "contentBytes": _ENG_LOGO_HUB, "isInline": True, "contentId": "a2zlogo2",
                })
        except Exception:
            pass

    try:
        token = _graph_token()
    except Exception as e:
        return jsonify(error=str(e)), 500

    try:
        r = _rq.post(
            "https://graph.microsoft.com/v1.0/users/%s/sendMail" % frm,
            headers={"Authorization": "Bearer " + token, "Content-Type": "application/json"},
            json={"message": message, "saveToSentItems": True}, timeout=60)
    except Exception as e:
        return jsonify(error="Send request failed", detail=str(e)), 502

    if r.status_code in (200, 202):
        return jsonify(ok=True, sent_to=to, attached=bool(message.get("attachments")))
    return jsonify(error="Microsoft 365 rejected the send", status=r.status_code, detail=r.text[:500]), 502


# ---- Engagement agreement PDF (added; reuses _build_workbook_ltd + GEN; nothing else changed) ----
def _engagement_pdf_bytes_for(d):
    """Build the engagement PDF and return raw bytes (for /send attachments used by the accept flow)."""
    tmpdir = tempfile.mkdtemp()
    wb_path = os.path.join(tmpdir, "wb.xlsx"); out_path = os.path.join(tmpdir, "engagement.pdf")
    wb = _build_workbook_ltd(d); wb.save(wb_path)
    wb2 = GEN.safe_load_workbook(wb_path)
    GEN.build_engagement(wb2, out_path, ref=d.get("ref"), acceptance=d.get("acceptance"), eng=d)
    with open(out_path, "rb") as f:
        return f.read()


@app.route("/engagement", methods=["POST", "OPTIONS"])
def engagement():
    if request.method == "OPTIONS":
        return ("", 204)
    d = request.get_json(force=True, silent=True) or {}
    tmpdir = tempfile.mkdtemp()
    wb_path = os.path.join(tmpdir, "wb.xlsx"); out_path = os.path.join(tmpdir, "engagement.pdf")
    try:
        wb = _build_workbook_ltd(d); wb.save(wb_path)
        wb2 = GEN.safe_load_workbook(wb_path)
        GEN.build_engagement(wb2, out_path, ref=d.get("ref"), acceptance=d.get("acceptance"), eng=d)
    except Exception as e:
        print("=== ENGAGEMENT PDF BUILD ERROR ===", flush=True); traceback.print_exc()
        return jsonify(error="Engagement PDF build failed", detail=str(e)), 500
    fname = (d.get("company") or "Engagement").replace("/", " ").replace("\\", " ")
    fname = " ".join(fname.split()) + " - engagement.pdf"
    return send_file(out_path, mimetype="application/pdf", as_attachment=True, download_name=fname)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8000))
    app.run(host="0.0.0.0", port=port)
